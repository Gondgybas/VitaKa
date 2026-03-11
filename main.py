# -*- coding: utf-8 -*-
"""
VitaKa - Программа для учета комплектующих на заводе
Версия: 1.0
"""
from __future__ import annotations

import tkinter as tk
from tkinter import ttk, messagebox, filedialog, simpledialog
import pandas as pd
from openpyxl import Workbook, load_workbook
from datetime import datetime
from pathlib import Path
import os
import sys
import json
import copy
import time
import subprocess

# ─────────────────────────────────────────────────────────────
#  КОНСТАНТЫ
# ─────────────────────────────────────────────────────────────
SETTINGS_FILE = "vitaka_settings.json"
DEFAULT_DB_NAME = "vitaka_components.xlsx"
SHEET_COMPONENTS = "Комплектующие"
SHEET_LOG = "Лог"

STANDARD_COLUMNS = ["id", "тип", "диаметр", "длина", "количество", "вес_единицы"]
STANDARD_HEADERS = {
    "id": "ID",
    "тип": "Тип",
    "диаметр": "Диаметр",
    "длина": "Длина (мм)",
    "количество": "Количество",
    "вес_единицы": "Вес/ед. (кг)",
}
LOG_COLUMNS = ["дата_время", "операция", "комплектующее", "изменение", "комментарий"]
LOG_HEADERS = {
    "дата_время": "Дата и время",
    "операция": "Операция",
    "комплектующее": "Комплектующее",
    "изменение": "Изменение",
    "комментарий": "Комментарий",
}

COMPONENT_TYPES = ["Винт", "Болт", "Гайка", "Шайба", "Шпилька", "Заклёпка", "Штифт", "Другое"]
DIAMETER_OPTIONS = ["М2", "М2.5", "М3", "М4", "М5", "М6", "М8", "М10", "М12", "М16", "М20",
                    "3", "4", "5", "6", "8", "9", "10", "12", "16", "20"]


# ─────────────────────────────────────────────────────────────
#  НАСТРОЙКИ
# ─────────────────────────────────────────────────────────────

def load_settings() -> dict:
    try:
        if os.path.exists(SETTINGS_FILE):
            with open(SETTINGS_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
    except Exception:
        pass
    return {}


def save_settings(settings: dict):
    try:
        with open(SETTINGS_FILE, "w", encoding="utf-8") as f:
            json.dump(settings, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f"Ошибка сохранения настроек: {e}")


# ─────────────────────────────────────────────────────────────
#  РАБОТА С EXCEL
# ─────────────────────────────────────────────────────────────

def get_db_path(settings: dict) -> str:
    folder = settings.get("db_folder", os.path.dirname(os.path.abspath(__file__)))
    return os.path.join(folder, DEFAULT_DB_NAME)


def initialize_db(file_path: str):
    """Создать файл Excel с нужными листами, если не существует."""
    if os.path.exists(file_path):
        return
    wb = Workbook()
    ws_comp = wb.active
    ws_comp.title = SHEET_COMPONENTS
    ws_comp.append(["id", "тип", "диаметр", "длина", "количество", "вес_единицы", "доп_параметры"])

    ws_log = wb.create_sheet(SHEET_LOG)
    ws_log.append(["дата_время", "операция", "комплектующее", "изменение", "комментарий"])

    wb.save(file_path)


def load_components(file_path: str) -> list:
    """Загрузить список комплектующих из Excel."""
    try:
        initialize_db(file_path)
        df = pd.read_excel(file_path, sheet_name=SHEET_COMPONENTS, dtype=str)
        df = df.fillna("")
        items = []
        for _, row in df.iterrows():
            item = {col: row[col] for col in STANDARD_COLUMNS if col in df.columns}
            # Дополнительные параметры хранятся как JSON в колонке "доп_параметры"
            extra_raw = row.get("доп_параметры", "") if "доп_параметры" in df.columns else ""
            try:
                item["доп_параметры"] = json.loads(extra_raw) if extra_raw else {}
            except Exception:
                item["доп_параметры"] = {}
            items.append(item)
        return items
    except Exception as e:
        print(f"Ошибка загрузки комплектующих: {e}")
        return []


def load_log(file_path: str) -> list:
    """Загрузить лог из Excel."""
    try:
        initialize_db(file_path)
        df = pd.read_excel(file_path, sheet_name=SHEET_LOG, dtype=str)
        df = df.fillna("")
        return df.to_dict("records")
    except Exception as e:
        print(f"Ошибка загрузки лога: {e}")
        return []


def save_all(file_path: str, components: list, log_entries: list):
    """Сохранить все данные в Excel, заменяя оба листа."""
    try:
        folder = os.path.dirname(file_path)
        if folder and not os.path.exists(folder):
            os.makedirs(folder, exist_ok=True)

        # Подготовить DataFrame для комплектующих
        rows = []
        for item in components:
            row = {col: item.get(col, "") for col in STANDARD_COLUMNS}
            row["доп_параметры"] = json.dumps(item.get("доп_параметры", {}), ensure_ascii=False)
            rows.append(row)
        df_comp = pd.DataFrame(rows, columns=STANDARD_COLUMNS + ["доп_параметры"])

        df_log = pd.DataFrame(log_entries, columns=LOG_COLUMNS) if log_entries else pd.DataFrame(
            columns=LOG_COLUMNS)

        # Если файл существует – сохраняем с остальными листами
        if os.path.exists(file_path):
            with pd.ExcelFile(file_path, engine="openpyxl") as xls:
                other_sheets = {s: pd.read_excel(xls, s) for s in xls.sheet_names
                                if s not in (SHEET_COMPONENTS, SHEET_LOG)}
        else:
            other_sheets = {}

        with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
            df_comp.to_excel(writer, sheet_name=SHEET_COMPONENTS, index=False)
            df_log.to_excel(writer, sheet_name=SHEET_LOG, index=False)
            for name, df in other_sheets.items():
                df.to_excel(writer, sheet_name=name, index=False)
    except Exception as e:
        print(f"Ошибка сохранения Excel: {e}")
        messagebox.showerror("Ошибка сохранения", f"Не удалось сохранить файл:\n{e}")


# ─────────────────────────────────────────────────────────────
#  ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ
# ─────────────────────────────────────────────────────────────

def next_id(components: list) -> int:
    if not components:
        return 1
    try:
        return max(int(c.get("id", 0)) for c in components) + 1
    except Exception:
        return len(components) + 1


def component_label(item: dict) -> str:
    """Читаемое название комплектующего для лога."""
    parts = []
    if item.get("тип"):
        parts.append(str(item["тип"]))
    if item.get("диаметр"):
        parts.append(str(item["диаметр"]))
    if item.get("длина"):
        parts.append(f"x{item['длина']}")
    return " ".join(parts) if parts else f"ID={item.get('id', '?')}"


def diff_items(old: dict, new: dict) -> str:
    """Описать разницу между двумя версиями комплектующего."""
    changes = []
    fields = {
        "тип": "Тип",
        "диаметр": "Диаметр",
        "длина": "Длина",
        "количество": "Количество",
        "вес_единицы": "Вес/ед.",
    }
    for key, label in fields.items():
        ov = str(old.get(key, "")).strip()
        nv = str(new.get(key, "")).strip()
        if ov != nv:
            changes.append(f"{label}: {ov} → {nv}")

    # Дополнительные параметры
    old_extra = old.get("доп_параметры", {})
    new_extra = new.get("доп_параметры", {})
    all_keys = set(old_extra) | set(new_extra)
    for k in sorted(all_keys):
        ov = str(old_extra.get(k, "")).strip()
        nv = str(new_extra.get(k, "")).strip()
        if ov != nv:
            if not ov:
                changes.append(f"{k}: добавлено «{nv}»")
            elif not nv:
                changes.append(f"{k}: удалено «{ov}»")
            else:
                changes.append(f"{k}: {ov} → {nv}")

    return "; ".join(changes) if changes else "Без изменений"


# ─────────────────────────────────────────────────────────────
#  ДИАЛОГ ДОБАВЛЕНИЯ / РЕДАКТИРОВАНИЯ КОМПЛЕКТУЮЩЕГО
# ─────────────────────────────────────────────────────────────

class ComponentDialog(tk.Toplevel):
    """Диалоговое окно для создания или редактирования комплектующего."""

    def __init__(self, parent, title: str, item: dict = None):
        super().__init__(parent)
        self.title(title)
        self.resizable(True, True)
        self.grab_set()
        self.result = None

        self._item = copy.deepcopy(item) if item else {}
        self._extra_rows = []  # список (key_var, val_var, frame)

        self._build_ui()
        self._populate(self._item)

        self.update_idletasks()
        w, h = 480, 520
        x = parent.winfo_rootx() + (parent.winfo_width() - w) // 2
        y = parent.winfo_rooty() + (parent.winfo_height() - h) // 2
        self.geometry(f"{w}x{h}+{x}+{y}")
        self.minsize(400, 400)

    # ── построение интерфейса ──────────────────────────────

    def _build_ui(self):
        outer = ttk.Frame(self, padding=10)
        outer.pack(fill="both", expand=True)

        # Прокручиваемая область
        canvas = tk.Canvas(outer, highlightthickness=0)
        scrollbar = ttk.Scrollbar(outer, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)

        self._scroll_frame = ttk.Frame(canvas)
        self._canvas_window = canvas.create_window((0, 0), window=self._scroll_frame, anchor="nw")

        self._scroll_frame.bind("<Configure>", lambda e: canvas.configure(
            scrollregion=canvas.bbox("all")))
        canvas.bind("<Configure>", lambda e: canvas.itemconfig(
            self._canvas_window, width=e.width))
        canvas.bind("<MouseWheel>", lambda e: canvas.yview_scroll(-1 * (e.delta // 120), "units"))

        form = self._scroll_frame
        r = 0

        # Стандартные поля
        ttk.Label(form, text="Тип *").grid(row=r, column=0, sticky="w", padx=5, pady=4)
        self._type_var = tk.StringVar()
        cb_type = ttk.Combobox(form, textvariable=self._type_var, values=COMPONENT_TYPES, width=28)
        cb_type.grid(row=r, column=1, sticky="ew", padx=5, pady=4)
        r += 1

        ttk.Label(form, text="Диаметр").grid(row=r, column=0, sticky="w", padx=5, pady=4)
        self._diam_var = tk.StringVar()
        cb_diam = ttk.Combobox(form, textvariable=self._diam_var, values=DIAMETER_OPTIONS, width=28)
        cb_diam.grid(row=r, column=1, sticky="ew", padx=5, pady=4)
        r += 1

        ttk.Label(form, text="Длина (мм)").grid(row=r, column=0, sticky="w", padx=5, pady=4)
        self._len_var = tk.StringVar()
        ttk.Entry(form, textvariable=self._len_var, width=30).grid(row=r, column=1, sticky="ew", padx=5, pady=4)
        r += 1

        ttk.Label(form, text="Количество *").grid(row=r, column=0, sticky="w", padx=5, pady=4)
        self._qty_var = tk.StringVar()
        ttk.Entry(form, textvariable=self._qty_var, width=30).grid(row=r, column=1, sticky="ew", padx=5, pady=4)
        r += 1

        ttk.Label(form, text="Вес/ед. (кг)").grid(row=r, column=0, sticky="w", padx=5, pady=4)
        self._weight_var = tk.StringVar()
        ttk.Entry(form, textvariable=self._weight_var, width=30).grid(row=r, column=1, sticky="ew", padx=5, pady=4)
        r += 1

        form.columnconfigure(1, weight=1)

        # Разделитель для доп. параметров
        ttk.Separator(form, orient="horizontal").grid(row=r, column=0, columnspan=2,
                                                       sticky="ew", padx=5, pady=8)
        r += 1
        ttk.Label(form, text="Дополнительные параметры:", font=("", 9, "bold")).grid(
            row=r, column=0, columnspan=2, sticky="w", padx=5)
        r += 1

        self._extra_container = ttk.Frame(form)
        self._extra_container.grid(row=r, column=0, columnspan=2, sticky="ew", padx=5)
        self._extra_container.columnconfigure(0, weight=1)
        self._extra_container.columnconfigure(1, weight=1)
        r += 1

        ttk.Button(form, text="➕ Добавить параметр", command=self._add_extra_row).grid(
            row=r, column=0, columnspan=2, sticky="w", padx=5, pady=4)
        r += 1

        # Кнопки OK / Отмена
        btn_frame = ttk.Frame(self)
        btn_frame.pack(fill="x", side="bottom", padx=10, pady=8)
        ttk.Button(btn_frame, text="Сохранить", command=self._on_ok, width=14).pack(side="right", padx=4)
        ttk.Button(btn_frame, text="Отмена", command=self.destroy, width=10).pack(side="right", padx=4)

    def _add_extra_row(self, key: str = "", value: str = ""):
        row_frame = ttk.Frame(self._extra_container)
        idx = len(self._extra_rows)
        row_frame.grid(row=idx, column=0, columnspan=3, sticky="ew", pady=2)
        row_frame.columnconfigure(0, weight=1)
        row_frame.columnconfigure(1, weight=1)

        key_var = tk.StringVar(value=key)
        val_var = tk.StringVar(value=value)
        ttk.Entry(row_frame, textvariable=key_var, width=16).grid(row=0, column=0, sticky="ew", padx=(0, 2))
        ttk.Entry(row_frame, textvariable=val_var, width=16).grid(row=0, column=1, sticky="ew", padx=2)
        ttk.Button(row_frame, text="✕", width=3,
                   command=lambda f=row_frame, kv=key_var, vv=val_var: self._remove_extra_row(f, kv, vv)).grid(
            row=0, column=2)
        self._extra_rows.append((key_var, val_var, row_frame))

    def _remove_extra_row(self, frame, key_var, val_var):
        self._extra_rows = [(k, v, f) for k, v, f in self._extra_rows if k is not key_var]
        frame.destroy()

    # ── заполнение данными ─────────────────────────────────

    def _populate(self, item: dict):
        self._type_var.set(item.get("тип", ""))
        self._diam_var.set(item.get("диаметр", ""))
        self._len_var.set(item.get("длина", ""))
        self._qty_var.set(item.get("количество", ""))
        self._weight_var.set(item.get("вес_единицы", ""))
        for k, v in (item.get("доп_параметры") or {}).items():
            self._add_extra_row(k, v)

    # ── валидация и сохранение ─────────────────────────────

    def _on_ok(self):
        тип = self._type_var.get().strip()
        if not тип:
            messagebox.showwarning("Внимание", "Поле «Тип» обязательно для заполнения.", parent=self)
            return

        # Валидация числовых полей
        qty_str = self._qty_var.get().strip()
        if qty_str and not self._is_number(qty_str):
            messagebox.showwarning("Внимание", "Количество должно быть числом.", parent=self)
            return
        weight_str = self._weight_var.get().strip()
        if weight_str and not self._is_number(weight_str):
            messagebox.showwarning("Внимание", "Вес/ед. должен быть числом.", parent=self)
            return
        len_str = self._len_var.get().strip()
        if len_str and not self._is_number(len_str):
            messagebox.showwarning("Внимание", "Длина должна быть числом.", parent=self)
            return

        extra = {}
        for key_var, val_var, _ in self._extra_rows:
            k = key_var.get().strip()
            v = val_var.get().strip()
            if k:
                extra[k] = v

        self.result = {
            "id": self._item.get("id", ""),
            "тип": тип,
            "диаметр": self._diam_var.get().strip(),
            "длина": len_str,
            "количество": qty_str,
            "вес_единицы": weight_str,
            "доп_параметры": extra,
        }
        self.destroy()

    @staticmethod
    def _is_number(s: str) -> bool:
        try:
            float(s.replace(",", "."))
            return True
        except ValueError:
            return False


# ─────────────────────────────────────────────────────────────
#  EXCEL-СТИЛЬ ФИЛЬТР
# ─────────────────────────────────────────────────────────────

class ExcelFilter:
    """
    Фильтр в стиле Excel: клик по заголовку столбца открывает
    выпадающее меню с чекбоксами для выбора значений.
    """

    def __init__(self, tree: ttk.Treeview, refresh_callback):
        self.tree = tree
        self.refresh_callback = refresh_callback
        self.active_filters: dict[str, set] = {}   # column_id → set of selected values
        self._all_data: list[tuple] = []            # все строки (values)
        self._window_open = False
        self._last_click = 0.0
        tree.bind("<Button-1>", self._on_header_click)

    def set_data(self, rows: list[tuple]):
        """Запомнить исходные строки (до фильтрации)."""
        self._all_data = rows

    def get_filtered(self) -> list[tuple]:
        """Вернуть строки, прошедшие активные фильтры."""
        if not self.active_filters:
            return self._all_data
        result = []
        columns = list(self.tree["columns"])
        for row in self._all_data:
            ok = True
            for col_id, allowed in self.active_filters.items():
                if not allowed:
                    continue
                try:
                    idx = columns.index(col_id)
                    val = str(row[idx])
                except (ValueError, IndexError):
                    val = ""
                if val not in allowed:
                    ok = False
                    break
            if ok:
                result.append(row)
        return result

    def _on_header_click(self, event):
        now = time.time()
        if self._window_open or now - self._last_click < 0.35:
            return
        region = self.tree.identify_region(event.x, event.y)
        if region != "heading":
            return
        col = self.tree.identify_column(event.x)
        col_id = self.tree.column(col, "id")
        self._last_click = now
        self._show_menu(event.x_root, event.y_root, col_id)

    def _show_menu(self, x: int, y: int, col_id: str):
        self._window_open = True
        win = tk.Toplevel(self.tree)
        win.overrideredirect(True)
        win.geometry(f"+{x}+{y}")
        win.lift()

        def on_close():
            self._window_open = False
            win.destroy()

        win.protocol("WM_DELETE_WINDOW", on_close)

        columns = list(self.tree["columns"])
        try:
            col_idx = columns.index(col_id)
        except ValueError:
            on_close()
            return

        # Уникальные значения столбца из всех данных
        unique_vals = sorted({str(row[col_idx]) for row in self._all_data})

        currently_selected = self.active_filters.get(col_id, set(unique_vals))

        frame = ttk.Frame(win, relief="solid", borderwidth=1)
        frame.pack(fill="both", expand=True)

        ttk.Label(frame, text=f"Фильтр: {col_id}", font=("", 9, "bold"),
                  padding=4).pack(fill="x")
        ttk.Separator(frame).pack(fill="x")

        # Поиск
        search_var = tk.StringVar()
        search_entry = ttk.Entry(frame, textvariable=search_var, width=24)
        search_entry.pack(padx=4, pady=4, fill="x")

        # Список с чекбоксами в прокручиваемой области
        list_frame = ttk.Frame(frame)
        list_frame.pack(fill="both", expand=True, padx=4)
        canvas = tk.Canvas(list_frame, height=200, highlightthickness=0)
        sb = ttk.Scrollbar(list_frame, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=sb.set)
        sb.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)
        inner = ttk.Frame(canvas)
        cw = canvas.create_window((0, 0), window=inner, anchor="nw")
        inner.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.bind("<Configure>", lambda e: canvas.itemconfig(cw, width=e.width))

        check_vars: dict[str, tk.BooleanVar] = {}
        check_widgets: dict[str, ttk.Checkbutton] = {}

        def rebuild_list(filter_text=""):
            for w in inner.winfo_children():
                w.destroy()
            for v in unique_vals:
                if filter_text.lower() not in v.lower():
                    continue
                var = check_vars.get(v)
                if var is None:
                    var = tk.BooleanVar(value=(v in currently_selected))
                    check_vars[v] = var
                cb = ttk.Checkbutton(inner, text=v, variable=var)
                cb.pack(anchor="w")
                check_widgets[v] = cb

        rebuild_list()

        def on_search(*_):
            rebuild_list(search_var.get())

        search_var.trace_add("write", on_search)

        # Кнопки "Выбрать все" / "Снять все"
        btn_row = ttk.Frame(frame)
        btn_row.pack(fill="x", padx=4, pady=2)
        ttk.Button(btn_row, text="Все", width=6,
                   command=lambda: [v.set(True) for v in check_vars.values()]).pack(side="left")
        ttk.Button(btn_row, text="Сброс", width=6,
                   command=lambda: [v.set(False) for v in check_vars.values()]).pack(side="left", padx=4)

        ttk.Separator(frame).pack(fill="x", pady=2)

        def apply_filter():
            selected = {v for v, var in check_vars.items() if var.get()}
            if selected == set(unique_vals) or not selected:
                self.active_filters.pop(col_id, None)
            else:
                self.active_filters[col_id] = selected
            on_close()
            self.refresh_callback()

        ttk.Button(frame, text="Применить", command=apply_filter).pack(pady=4)

        # Закрыть при клике вне окна
        win.bind("<FocusOut>", lambda e: on_close())
        search_entry.focus_set()


# ─────────────────────────────────────────────────────────────
#  ВКЛАДКА "КОМПЛЕКТУЮЩИЕ"
# ─────────────────────────────────────────────────────────────

class ComponentsTab(ttk.Frame):
    def __init__(self, master, app):
        super().__init__(master)
        self.app = app
        self._filter: ExcelFilter | None = None
        self._build_ui()

    def _build_ui(self):
        # Панель кнопок
        btn_bar = ttk.Frame(self)
        btn_bar.pack(fill="x", padx=6, pady=(6, 2))

        ttk.Button(btn_bar, text="➕ Добавить", command=self.add_item).pack(side="left", padx=2)
        ttk.Button(btn_bar, text="✏️ Редактировать", command=self.edit_item).pack(side="left", padx=2)
        ttk.Button(btn_bar, text="🗑️ Удалить", command=self.delete_item).pack(side="left", padx=2)
        ttk.Button(btn_bar, text="🔄 Сбросить фильтры", command=self.reset_filters).pack(side="left", padx=8)

        self._filter_status = ttk.Label(btn_bar, text="", foreground="blue")
        self._filter_status.pack(side="left", padx=4)

        # Таблица с прокруткой
        tree_frame = ttk.Frame(self)
        tree_frame.pack(fill="both", expand=True, padx=6, pady=4)

        self.tree = ttk.Treeview(tree_frame, show="headings", selectmode="browse")
        vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(tree_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        vsb.pack(side="right", fill="y")
        hsb.pack(side="bottom", fill="x")
        self.tree.pack(side="left", fill="both", expand=True)

        self.tree.bind("<Double-1>", lambda e: self.edit_item())

        # Строка состояния
        self._status_var = tk.StringVar()
        ttk.Label(self, textvariable=self._status_var, anchor="w").pack(
            fill="x", padx=6, pady=(0, 4))

    def _setup_columns(self, extra_keys: list[str]):
        """Настроить столбцы таблицы с учётом доп. параметров."""
        cols = list(STANDARD_COLUMNS) + extra_keys
        self.tree.configure(columns=cols)
        for c in cols:
            header = STANDARD_HEADERS.get(c, c)
            self.tree.heading(c, text=header, anchor="w")
            width = 120 if c not in ("id", "тип") else (40 if c == "id" else 90)
            self.tree.column(c, width=width, minwidth=50, anchor="w")

        # Устанавливаем фильтр после конфигурации столбцов
        if self._filter is None:
            self._filter = ExcelFilter(self.tree, self.refresh)
        else:
            self._filter.active_filters.clear()

    def _collect_extra_keys(self) -> list[str]:
        keys = []
        seen = set()
        for item in self.app.components:
            for k in (item.get("доп_параметры") or {}):
                if k not in seen:
                    seen.add(k)
                    keys.append(k)
        return keys

    def refresh(self):
        """Перерисовать таблицу (с учётом фильтра)."""
        extra_keys = self._collect_extra_keys()
        self._setup_columns(extra_keys)

        all_rows = []
        for item in self.app.components:
            row = [item.get(c, "") for c in STANDARD_COLUMNS]
            for k in extra_keys:
                row.append(item.get("доп_параметры", {}).get(k, ""))
            all_rows.append(tuple(row))

        if self._filter:
            self._filter.set_data(all_rows)
            visible_rows = self._filter.get_filtered()
        else:
            visible_rows = all_rows

        self.tree.delete(*self.tree.get_children())
        for row in visible_rows:
            self.tree.insert("", "end", values=row)

        total = len(self.app.components)
        shown = len(visible_rows)
        self._status_var.set(f"Показано: {shown} из {total}")

        if self._filter and self._filter.active_filters:
            self._filter_status.config(text="⚠ Активны фильтры")
        else:
            self._filter_status.config(text="")

    def reset_filters(self):
        if self._filter:
            self._filter.active_filters.clear()
        self.refresh()

    # ── CRUD операции ──────────────────────────────────────

    def add_item(self):
        dlg = ComponentDialog(self.winfo_toplevel(), "Добавить комплектующее")
        self.wait_window(dlg)
        if dlg.result is None:
            return
        new_item = dlg.result
        new_item["id"] = str(next_id(self.app.components))
        self.app.components.append(new_item)
        self.app.add_log(
            operation="Добавление",
            component=component_label(new_item),
            change=f"Добавлено новое комплектующее: {component_label(new_item)}",
        )
        self.app.auto_save()
        self.refresh()

    def edit_item(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showinfo("Выбор", "Выберите строку для редактирования.")
            return
        row_values = self.tree.item(sel[0])["values"]
        item_id = str(row_values[0])
        item = next((c for c in self.app.components if str(c.get("id")) == item_id), None)
        if item is None:
            return

        old_item = copy.deepcopy(item)
        dlg = ComponentDialog(self.winfo_toplevel(), "Редактировать комплектующее", item)
        self.wait_window(dlg)
        if dlg.result is None:
            return

        change_desc = diff_items(old_item, dlg.result)
        item.update(dlg.result)
        if change_desc != "Без изменений":
            self.app.add_log(
                operation="Изменение",
                component=component_label(item),
                change=change_desc,
            )
        self.app.auto_save()
        self.refresh()

    def delete_item(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showinfo("Выбор", "Выберите строку для удаления.")
            return
        row_values = self.tree.item(sel[0])["values"]
        item_id = str(row_values[0])
        item = next((c for c in self.app.components if str(c.get("id")) == item_id), None)
        if item is None:
            return
        label = component_label(item)
        if not messagebox.askyesno("Удаление", f"Удалить «{label}»?"):
            return
        self.app.components = [c for c in self.app.components if str(c.get("id")) != item_id]
        self.app.add_log(
            operation="Удаление",
            component=label,
            change="Комплектующее удалено.",
        )
        self.app.auto_save()
        self.refresh()


# ─────────────────────────────────────────────────────────────
#  ВКЛАДКА "ЛОГ"
# ─────────────────────────────────────────────────────────────

class LogTab(ttk.Frame):
    def __init__(self, master, app):
        super().__init__(master)
        self.app = app
        self._build_ui()

    def _build_ui(self):
        btn_bar = ttk.Frame(self)
        btn_bar.pack(fill="x", padx=6, pady=(6, 2))
        ttk.Button(btn_bar, text="💬 Добавить комментарий", command=self._add_comment).pack(side="left", padx=2)
        ttk.Button(btn_bar, text="🔄 Обновить", command=self.refresh).pack(side="left", padx=2)

        tree_frame = ttk.Frame(self)
        tree_frame.pack(fill="both", expand=True, padx=6, pady=4)

        cols = list(LOG_COLUMNS)
        self.tree = ttk.Treeview(tree_frame, columns=cols, show="headings", selectmode="browse")
        vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(tree_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        vsb.pack(side="right", fill="y")
        hsb.pack(side="bottom", fill="x")
        self.tree.pack(side="left", fill="both", expand=True)

        for c in cols:
            header = LOG_HEADERS.get(c, c)
            self.tree.heading(c, text=header, anchor="w",
                              command=lambda col=c: self._sort_by(col))
            widths = {"дата_время": 140, "операция": 90, "комплектующее": 150,
                      "изменение": 300, "комментарий": 200}
            self.tree.column(c, width=widths.get(c, 120), minwidth=60, anchor="w")

        self.tree.bind("<Double-1>", lambda e: self._add_comment())
        self._sort_col = "дата_время"
        self._sort_asc = False

        self._status_var = tk.StringVar()
        ttk.Label(self, textvariable=self._status_var, anchor="w").pack(
            fill="x", padx=6, pady=(0, 4))

    def _sort_by(self, col: str):
        if self._sort_col == col:
            self._sort_asc = not self._sort_asc
        else:
            self._sort_col = col
            self._sort_asc = True
        self.refresh()

    def refresh(self):
        entries = sorted(
            self.app.log_entries,
            key=lambda e: e.get(self._sort_col, ""),
            reverse=not self._sort_asc,
        )
        self.tree.delete(*self.tree.get_children())
        for e in entries:
            self.tree.insert("", "end", values=[e.get(c, "") for c in LOG_COLUMNS])
        self._status_var.set(f"Записей в журнале: {len(entries)}")

    def _add_comment(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showinfo("Выбор", "Выберите запись лога для добавления комментария.")
            return
        row_values = self.tree.item(sel[0])["values"]
        dt_val = row_values[0] if row_values else ""

        # Найти запись в логе по дате
        entry = next(
            (e for e in self.app.log_entries if e.get("дата_время") == dt_val),
            None,
        )
        if entry is None:
            return

        current = entry.get("комментарий", "")
        new_comment = simpledialog.askstring(
            "Комментарий", "Введите комментарий:",
            initialvalue=current,
            parent=self.winfo_toplevel(),
        )
        if new_comment is None:
            return
        entry["комментарий"] = new_comment
        self.app.auto_save()
        self.refresh()


# ─────────────────────────────────────────────────────────────
#  ДИАЛОГ НАСТРОЕК
# ─────────────────────────────────────────────────────────────

class SettingsDialog(tk.Toplevel):
    def __init__(self, parent, settings: dict):
        super().__init__(parent)
        self.title("Настройки")
        self.resizable(False, False)
        self.grab_set()
        self.result = None
        self._settings = dict(settings)
        self._build_ui()
        self.update_idletasks()
        w, h = 480, 160
        x = parent.winfo_rootx() + (parent.winfo_width() - w) // 2
        y = parent.winfo_rooty() + (parent.winfo_height() - h) // 2
        self.geometry(f"{w}x{h}+{x}+{y}")

    def _build_ui(self):
        frame = ttk.Frame(self, padding=12)
        frame.pack(fill="both", expand=True)

        ttk.Label(frame, text="Папка для Excel-файла:").grid(row=0, column=0, sticky="w", pady=4)
        self._path_var = tk.StringVar(
            value=self._settings.get("db_folder",
                                     os.path.dirname(os.path.abspath(__file__))))
        path_entry = ttk.Entry(frame, textvariable=self._path_var, width=40)
        path_entry.grid(row=0, column=1, sticky="ew", padx=(4, 0), pady=4)
        ttk.Button(frame, text="...", width=3, command=self._browse).grid(row=0, column=2, padx=4)
        frame.columnconfigure(1, weight=1)

        ttk.Label(frame, text="Файл будет сохранён как:").grid(row=1, column=0, sticky="w", pady=2)
        self._file_label = ttk.Label(frame, text=DEFAULT_DB_NAME, foreground="grey")
        self._file_label.grid(row=1, column=1, columnspan=2, sticky="w", padx=4)

        btn_frame = ttk.Frame(frame)
        btn_frame.grid(row=2, column=0, columnspan=3, sticky="e", pady=8)
        ttk.Button(btn_frame, text="Сохранить", command=self._ok).pack(side="right", padx=4)
        ttk.Button(btn_frame, text="Отмена", command=self.destroy).pack(side="right")

    def _browse(self):
        folder = filedialog.askdirectory(
            title="Выберите папку для Excel",
            initialdir=self._path_var.get(),
            parent=self,
        )
        if folder:
            self._path_var.set(folder)

    def _ok(self):
        self._settings["db_folder"] = self._path_var.get()
        self.result = self._settings
        self.destroy()


# ─────────────────────────────────────────────────────────────
#  ГЛАВНОЕ ПРИЛОЖЕНИЕ
# ─────────────────────────────────────────────────────────────

class MainApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("VitaKa — Учёт комплектующих")
        self.minsize(800, 500)

        self.settings = load_settings()
        self._restore_geometry()

        # Данные
        db_path = get_db_path(self.settings)
        initialize_db(db_path)
        self.components = load_components(db_path)
        self.log_entries = load_log(db_path)

        self._build_menu()
        self._build_ui()

        self.protocol("WM_DELETE_WINDOW", self._on_close)
        self.bind("<Configure>", self._on_resize)

    # ── интерфейс ─────────────────────────────────────────

    def _build_menu(self):
        menubar = tk.Menu(self)
        self.configure(menu=menubar)

        file_menu = tk.Menu(menubar, tearoff=False)
        menubar.add_cascade(label="Файл", menu=file_menu)
        file_menu.add_command(label="Настройки…", command=self._open_settings)
        file_menu.add_separator()
        file_menu.add_command(label="Открыть папку с Excel", command=self._open_db_folder)
        file_menu.add_separator()
        file_menu.add_command(label="Выход", command=self._on_close)

    def _build_ui(self):
        self.notebook = ttk.Notebook(self)
        self.notebook.pack(fill="both", expand=True, padx=4, pady=4)

        self.tab_components = ComponentsTab(self.notebook, self)
        self.notebook.add(self.tab_components, text="Комплектующие")

        self.tab_log = LogTab(self.notebook, self)
        self.notebook.add(self.tab_log, text="Журнал изменений")

        self.tab_components.refresh()
        self.tab_log.refresh()

        # Строка состояния внизу
        status_bar = ttk.Frame(self, relief="sunken")
        status_bar.pack(fill="x", side="bottom")
        self._status_var = tk.StringVar(value="Готово")
        ttk.Label(status_bar, textvariable=self._status_var, anchor="w",
                  padding=(6, 2)).pack(side="left")
        self._db_path_label = ttk.Label(
            status_bar, text=get_db_path(self.settings), anchor="e",
            foreground="grey", padding=(6, 2))
        self._db_path_label.pack(side="right")

    # ── лог ───────────────────────────────────────────────

    def add_log(self, operation: str, component: str, change: str, comment: str = ""):
        entry = {
            "дата_время": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "операция": operation,
            "комплектующее": component,
            "изменение": change,
            "комментарий": comment,
        }
        self.log_entries.append(entry)
        if hasattr(self, "tab_log"):
            self.tab_log.refresh()

    # ── сохранение ────────────────────────────────────────

    def auto_save(self):
        db_path = get_db_path(self.settings)
        save_all(db_path, self.components, self.log_entries)
        self._status_var.set(f"Сохранено: {datetime.now().strftime('%H:%M:%S')}")
        self._db_path_label.config(text=db_path)

    # ── настройки ─────────────────────────────────────────

    def _open_settings(self):
        dlg = SettingsDialog(self, self.settings)
        self.wait_window(dlg)
        if dlg.result is None:
            return
        old_path = get_db_path(self.settings)
        self.settings = dlg.result
        save_settings(self.settings)
        new_path = get_db_path(self.settings)
        initialize_db(new_path)
        # Перезагрузить данные из нового места, если файл там уже есть
        if os.path.exists(new_path) and new_path != old_path:
            self.components = load_components(new_path)
            self.log_entries = load_log(new_path)
        else:
            # Сохранить текущие данные в новое место
            self.auto_save()
        self.tab_components.refresh()
        self.tab_log.refresh()
        self._db_path_label.config(text=new_path)

    def _open_db_folder(self):
        folder = os.path.dirname(get_db_path(self.settings))
        if os.path.exists(folder):
            if sys.platform == "win32":
                os.startfile(folder)
            elif sys.platform == "darwin":
                subprocess.Popen(["open", folder])
            else:
                subprocess.Popen(["xdg-open", folder])
        else:
            messagebox.showinfo("Папка", f"Папка не найдена:\n{folder}")

    # ── геометрия окна ────────────────────────────────────

    def _restore_geometry(self):
        geom = self.settings.get("window_geometry", "1100x650+100+50")
        try:
            self.geometry(geom)
        except Exception:
            self.geometry("1100x650")

    def _on_resize(self, event):
        # Сохраняем геометрию только для главного окна
        if event.widget is self:
            self.settings["window_geometry"] = self.geometry()

    def _on_close(self):
        save_settings(self.settings)
        self.destroy()


# ─────────────────────────────────────────────────────────────
#  ТОЧКА ВХОДА
# ─────────────────────────────────────────────────────────────

if __name__ == "__main__":
    app = MainApp()
    app.mainloop()

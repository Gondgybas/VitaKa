# -*- coding: utf-8 -*-
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import pandas as pd
from openpyxl import Workbook, load_workbook
from datetime import datetime, timedelta
from pathlib import Path
import os
import json

DATABASE_FILE = "production_database.xlsx"
DATA_PATH = Path(__file__).parent  # –Я–∞–њ–Ї–∞ –≥–і–µ –ї–µ–ґ–Є—В —Б–Ї—А–Є–њ—В


def initialize_database():
    if not os.path.exists(DATABASE_FILE):
        wb = Workbook()
        materials_sheet = wb.active
        materials_sheet.title = "Materials"
        materials_sheet.append([
            "ID", "–Ь–∞—А–Ї–∞", "–Ґ–Њ–ї—Й–Є–љ–∞", "–Ф–ї–Є–љ–∞", "–®–Є—А–Є–љ–∞",
            "–Ъ–Њ–ї–Є—З–µ—Б—В–≤–Њ —И—В—Г–Ї", "–Ю–±—Й–∞—П –њ–ї–Њ—Й–∞–і—М", "–Ч–∞—А–µ–Ј–µ—А–≤–Є—А–Њ–≤–∞–љ–Њ", "–Ф–Њ—Б—В—Г–њ–љ–Њ", "–Ф–∞—В–∞ –і–Њ–±–∞–≤–ї–µ–љ–Є—П"
        ])
        orders_sheet = wb.create_sheet("Orders")
        orders_sheet.append(["ID –Ј–∞–Ї–∞–Ј–∞", "–Э–∞–Ј–≤–∞–љ–Є–µ –Ј–∞–Ї–∞–Ј–∞", "–Ч–∞–Ї–∞–Ј—З–Є–Ї", "–Ф–∞—В–∞ —Б–Њ–Ј–і–∞–љ–Є—П", "–°—В–∞—В—Г—Б", "–Я—А–Є–Љ–µ—З–∞–љ–Є—П"])
        order_details_sheet = wb.create_sheet("OrderDetails")
        order_details_sheet.append(["ID", "ID –Ј–∞–Ї–∞–Ј–∞", "–Э–∞–Ј–≤–∞–љ–Є–µ –і–µ—В–∞–ї–Є", "–Ъ–Њ–ї–Є—З–µ—Б—В–≤–Њ", "–Я–Њ—А–µ–Ј–∞–љ–Њ", "–Я–Њ–≥–љ—Г—В–Њ"])
        reservations_sheet = wb.create_sheet("Reservations")
        reservations_sheet.append(
            ["ID —А–µ–Ј–µ—А–≤–∞", "ID –Ј–∞–Ї–∞–Ј–∞", "ID –і–µ—В–∞–ї–Є", "–Э–∞–Ј–≤–∞–љ–Є–µ –і–µ—В–∞–ї–Є", "ID –Љ–∞—В–µ—А–Є–∞–ї–∞", "–Ь–∞—А–Ї–∞", "–Ґ–Њ–ї—Й–Є–љ–∞", "–Ф–ї–Є–љ–∞",
             "–®–Є—А–Є–љ–∞", "–Ч–∞—А–µ–Ј–µ—А–≤–Є—А–Њ–≤–∞–љ–Њ —И—В—Г–Ї", "–°–њ–Є—Б–∞–љ–Њ", "–Ю—Б—В–∞—В–Њ–Ї –Ї —Б–њ–Є—Б–∞–љ–Є—О", "–Ф–∞—В–∞ —А–µ–Ј–µ—А–≤–∞"])
        writeoffs_sheet = wb.create_sheet("WriteOffs")
        writeoffs_sheet.append([
            "ID —Б–њ–Є—Б–∞–љ–Є—П", "ID —А–µ–Ј–µ—А–≤–∞", "ID –Ј–∞–Ї–∞–Ј–∞", "ID –Љ–∞—В–µ—А–Є–∞–ї–∞", "–Ь–∞—А–Ї–∞", "–Ґ–Њ–ї—Й–Є–љ–∞", "–Ф–ї–Є–љ–∞", "–®–Є—А–Є–љ–∞",
            "–Ъ–Њ–ї–Є—З–µ—Б—В–≤–Њ", "–Ф–∞—В–∞ —Б–њ–Є—Б–∞–љ–Є—П", "–Ъ–Њ–Љ–Љ–µ–љ—В–∞—А–Є–є"
        ])

        # рЯЖХ –Ы–Ш–°–Ґ –Ф–Ы–ѓ –Ы–Ю–У–Ш–†–Ю–Т–Р–Э–Ш–ѓ –Ш–Ч–Ь–Х–Э–Х–Э–Ш–Щ –Ъ–Ю–Ы–Ш–І–Х–°–Ґ–Т–Р –Ь–Р–Ґ–Х–†–Ш–Р–Ы–Р
        changelogs_sheet = wb.create_sheet("MaterialChangeLogs")
        changelogs_sheet.append([
            "ID –ї–Њ–≥–∞", "–Ф–∞—В–∞ –Є –≤—А–µ–Љ—П", "ID –Љ–∞—В–µ—А–Є–∞–ї–∞", "–Ь–∞—А–Ї–∞", "–Ґ–Њ–ї—Й–Є–љ–∞",
            "–Ф–ї–Є–љ–∞", "–®–Є—А–Є–љ–∞", "–°—В–∞—А–Њ–µ –Ї–Њ–ї-–≤–Њ", "–Э–Њ–≤–Њ–µ –Ї–Њ–ї-–≤–Њ", "–Ш–Ј–Љ–µ–љ–µ–љ–Є–µ", "–Ъ–Њ–Љ–Љ–µ–љ—В–∞—А–Є–є"
        ])

        wb.save(DATABASE_FILE)
        print(f"–С–∞–Ј–∞ –і–∞–љ–љ—Л—Е '{DATABASE_FILE}' —Б–Њ–Ј–і–∞–љ–∞!")


def get_database_path():
    """–Я–Њ–ї—Г—З–Є—В—М –њ—Г—В—М –Ї –њ–∞–њ–Ї–µ —Б –±–∞–Ј–Њ–є –і–∞–љ–љ—Л—Е –Є–Ј –љ–∞—Б—В—А–Њ–µ–Ї"""
    settings_file = "app_settings.json"
    try:
        if os.path.exists(settings_file):
            with open(settings_file, 'r', encoding='utf-8') as f:
                settings = json.load(f)
                return settings.get("database_path", os.path.dirname(os.path.abspath(__file__)))
    except:
        pass
    # –Я–Њ —Г–Љ–Њ–ї—З–∞–љ–Є—О - —В–µ–Ї—Г—Й–∞—П –њ–∞–њ–Ї–∞
    return os.path.dirname(os.path.abspath(__file__))


def load_data(sheet_name):
    """–Ч–∞–≥—А—Г–Ј–Ї–∞ –і–∞–љ–љ—Л—Е –Є–Ј Excel —Б —Г—З—С—В–Њ–Љ –њ—Г—В–Є –Є–Ј –љ–∞—Б—В—А–Њ–µ–Ї"""
    db_path = get_database_path()
    file_path = os.path.join(db_path, "production_database.xlsx")

    try:
        if os.path.exists(file_path):
            df = pd.read_excel(file_path, sheet_name=sheet_name)
            return df
        else:
            print(f"вЪ†пЄП –§–∞–є–ї –±–∞–Ј—Л –і–∞–љ–љ—Л—Е –љ–µ –љ–∞–є–і–µ–љ: {file_path}")
            return pd.DataFrame()
    except Exception as e:
        print(f"вЭМ –Ю—И–Є–±–Ї–∞ –Ј–∞–≥—А—Г–Ј–Ї–Є –і–∞–љ–љ—Л—Е –Є–Ј {sheet_name}: {e}")
        return pd.DataFrame()


def save_data(sheet_name, df):
    """–°–Њ—Е—А–∞–љ–µ–љ–Є–µ –і–∞–љ–љ—Л—Е –≤ Excel —Б —Г—З—С—В–Њ–Љ –њ—Г—В–Є –Є–Ј –љ–∞—Б—В—А–Њ–µ–Ї"""
    db_path = get_database_path()
    file_path = os.path.join(db_path, "production_database.xlsx")

    try:
        if os.path.exists(file_path):
            with pd.ExcelFile(file_path, engine='openpyxl') as xls:
                sheets = {s: pd.read_excel(xls, s) for s in xls.sheet_names}
        else:
            sheets = {}

        sheets[sheet_name] = df

        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            for s, data in sheets.items():
                data.to_excel(writer, sheet_name=s, index=False)

        print(f"вЬЕ –Ф–∞–љ–љ—Л–µ —Б–Њ—Е—А–∞–љ–µ–љ—Л –≤ {sheet_name}")
    except Exception as e:
        print(f"вЭМ –Ю—И–Є–±–Ї–∞ —Б–Њ—Е—А–∞–љ–µ–љ–Є—П –і–∞–љ–љ—Л—Е –≤ {sheet_name}: {e}")
        messagebox.showerror("–Ю—И–Є–±–Ї–∞ —Б–Њ—Е—А–∞–љ–µ–љ–Є—П", f"–Э–µ —Г–і–∞–ї–Њ—Б—М —Б–Њ—Е—А–∞–љ–Є—В—М –і–∞–љ–љ—Л–µ: {e}")


class ExcelStyleFilter:
    """–§–Є–ї—М—В—А –≤ —Б—В–Є–ї–µ Excel –і–ї—П Treeview - –≤—Л–њ–∞–і–∞—О—Й–µ–µ –Љ–µ–љ—О –њ—А–Є –Ї–ї–Є–Ї–µ –љ–∞ –Ј–∞–≥–Њ–ї–Њ–≤–Њ–Ї"""

    def __init__(self, tree, refresh_callback, columns_config=None):
        """
        tree: ttk.Treeview –≤–Є–і–ґ–µ—В
        refresh_callback: —Д—Г–љ–Ї—Ж–Є—П –Њ–±–љ–Њ–≤–ї–µ–љ–Є—П –і–∞–љ–љ—Л—Е
        columns_config: dict —Б –љ–∞—Б—В—А–Њ–є–Ї–∞–Љ–Є —Б—В–Њ–ї–±—Ж–Њ–≤ (–Њ–њ—Ж–Є–Њ–љ–∞–ї—М–љ–Њ)
        """
        self.tree = tree
        self.refresh_callback = refresh_callback
        self.columns_config = columns_config or {}

        # –•—А–∞–љ–Є–ї–Є—Й–µ –∞–Ї—В–Є–≤–љ—Л—Е —Д–Є–ї—М—В—А–Њ–≤
        self.active_filters = {}

        # –Ш—Б—Е–Њ–і–љ—Л–µ –і–∞–љ–љ—Л–µ (–і–Њ —Д–Є–ї—М—В—А–∞—Ж–Є–Є)
        self.original_data = []

        # рЯЖХ –Ч–Р–©–Ш–Ґ–Р –Ю–Ґ –Ф–Т–Ю–Щ–Э–Ю–У–Ю –Т–Ђ–Ч–Ю–Т–Р
        self._filter_window_open = False
        self._last_click_time = 0

        # –Я—А–Є–≤—П–Ј—Л–≤–∞–µ–Љ –Ї–ї–Є–Ї –Ї –Ј–∞–≥–Њ–ї–Њ–≤–Ї–∞–Љ
        self.tree.bind('<Button-1>', self.on_header_click)

    def on_header_click(self, event):
        """–Ю–±—А–∞–±–Њ—В–Ї–∞ –Ї–ї–Є–Ї–∞ –њ–Њ –Ј–∞–≥–Њ–ї–Њ–≤–Ї—Г —Б—В–Њ–ї–±—Ж–∞"""
        import time

        # рЯЖХ –Ч–Р–©–Ш–Ґ–Р –Ю–Ґ –Ф–Т–Ю–Щ–Э–Ю–У–Ю –Ъ–Ы–Ш–Ъ–Р
        current_time = time.time()
        if current_time - self._last_click_time < 0.3:
            return

        # рЯЖХ –Ч–Р–©–Ш–Ґ–Р: –Э–Х –Ю–Ґ–Ъ–†–Ђ–Т–Р–Х–Ь –Т–Ґ–Ю–†–Ю–Х –Ю–Ъ–Э–Ю
        if self._filter_window_open:
            return

        region = self.tree.identify_region(event.x, event.y)

        if region == "heading":
            column = self.tree.identify_column(event.x)
            column_id = self.tree.column(column, "id")

            self._last_click_time = current_time
            self._filter_window_open = True

            # –Я–Њ–Ї–∞–Ј—Л–≤–∞–µ–Љ –Љ–µ–љ—О —Д–Є–ї—М—В—А–∞
            self.show_filter_menu(event, column_id)

    def show_filter_menu(self, event, column_id):
        """–Я–Њ–Ї–∞–Ј–∞—В—М –Љ–µ–љ—О —Д–Є–ї—М—В—А–∞ –і–ї—П —Б—В–Њ–ї–±—Ж–∞"""
        try:
            column_index = list(self.tree["columns"]).index(column_id)

            # –°–Њ–±–Є—А–∞–µ–Љ —Г–љ–Є–Ї–∞–ї—М–љ—Л–µ –Ј–љ–∞—З–µ–љ–Є—П
            all_unique_values = set()
            visible_unique_values = set()

            visible_items = list(self.tree.get_children(''))

            if not hasattr(self, '_all_item_cache'):
                self._all_item_cache = set()

            for item_id in visible_items:
                self._all_item_cache.add(item_id)
                values = self.tree.item(item_id)["values"]
                value = values[column_index]
                visible_unique_values.add(str(value))
                all_unique_values.add(str(value))

            for item_id in self._all_item_cache:
                if item_id not in visible_items:
                    try:
                        values = self.tree.item(item_id)["values"]
                        if values:
                            value = values[column_index]
                            all_unique_values.add(str(value))
                    except:
                        pass

            # –Ю–њ—А–µ–і–µ–ї—П–µ–Љ –≤—Л–±—А–∞–љ–љ—Л–µ –Ј–љ–∞—З–µ–љ–Є—П
            if column_id in self.active_filters:
                currently_selected = set(self.active_filters[column_id])
            else:
                currently_selected = all_unique_values.copy()

            # –°–Њ—А—В–Є—А–Њ–≤–Ї–∞
            selected_visible = sorted(list(currently_selected & visible_unique_values))
            selected_hidden = sorted(list(currently_selected - visible_unique_values))
            unselected = sorted(list(all_unique_values - currently_selected))

            unique_values = selected_visible + selected_hidden + unselected

            # –°–Њ–Ј–і–∞—С–Љ –Њ–Ї–љ–Њ —Д–Є–ї—М—В—А–∞
            filter_window = tk.Toplevel(self.tree)
            filter_window.title(f"–§–Є–ї—М—В—А: {column_id}")
            filter_window.geometry("320x600")
            filter_window.configure(bg='#ecf0f1')
            filter_window.transient(self.tree)
            filter_window.grab_set()

            x = event.x_root
            y = event.y_root + 20
            filter_window.geometry(f"+{x}+{y}")

            # –Ч–∞–≥–Њ–ї–Њ–≤–Њ–Ї
            header_frame = tk.Frame(filter_window, bg='#3498db')
            header_frame.pack(fill=tk.X)
            tk.Label(header_frame, text=f"–§–Є–ї—М—В—А: {column_id}",
                     font=("Arial", 12, "bold"), bg='#3498db', fg='white', pady=10).pack()

            # –Ъ–љ–Њ–њ–Ї–Є —Б–Њ—А—В–Є—А–Њ–≤–Ї–Є
            sort_frame = tk.Frame(filter_window, bg='#ecf0f1')
            sort_frame.pack(fill=tk.X, padx=10, pady=10)
            tk.Label(sort_frame, text="–°–Њ—А—В–Є—А–Њ–≤–Ї–∞:", font=("Arial", 10, "bold"), bg='#ecf0f1').pack(anchor='w',
                                                                                                    pady=(0, 5))
            tk.Button(sort_frame, text="вЦ≤ –Я–Њ –≤–Њ–Ј—А–∞—Б—В–∞–љ–Є—О (AвЖТZ, 0вЖТ9)",
                      command=lambda: self.apply_sort(column_id, 'asc', filter_window),
                      bg='#3498db', fg='white', font=("Arial", 9), relief=tk.RAISED).pack(fill=tk.X, pady=2)
            tk.Button(sort_frame, text="вЦЉ –Я–Њ —Г–±—Л–≤–∞–љ–Є—О (ZвЖТA, 9вЖТ0)",
                      command=lambda: self.apply_sort(column_id, 'desc', filter_window),
                      bg='#3498db', fg='white', font=("Arial", 9), relief=tk.RAISED).pack(fill=tk.X, pady=2)

            tk.Frame(filter_window, height=2, bg='#95a5a6').pack(fill=tk.X, pady=5)

            # рЯЖХ –Я–Ю–Ы–Х –Я–Ю–Ш–°–Ъ–Р
            search_frame = tk.Frame(filter_window, bg='#ecf0f1')
            search_frame.pack(fill=tk.X, padx=10, pady=5)
            tk.Label(search_frame, text="рЯФН –Я–Њ–Є—Б–Ї:", font=("Arial", 10, "bold"), bg='#ecf0f1').pack(side=tk.LEFT, padx=5)

            search_var = tk.StringVar()
            search_entry = tk.Entry(search_frame, textvariable=search_var, font=("Arial", 10), width=20)
            search_entry.pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)

            def clear_search():
                search_var.set("")
                search_entry.focus_set()

            tk.Button(search_frame, text="вЬЦ", font=("Arial", 8), bg='#e74c3c', fg='white',
                      command=clear_search, width=3).pack(side=tk.LEFT, padx=2)

            tk.Frame(filter_window, height=2, bg='#95a5a6').pack(fill=tk.X, pady=5)

            tk.Label(filter_window, text="–§–Є–ї—М—В—А –њ–Њ –Ј–љ–∞—З–µ–љ–Є—О:",
                     font=("Arial", 10, "bold"), bg='#ecf0f1').pack(pady=(5, 5), padx=10, anchor='w')

            # –§—А–µ–є–Љ —Б–Њ —Б–њ–Є—Б–Ї–Њ–Љ
            list_frame = tk.Frame(filter_window, bg='white', relief=tk.SUNKEN, borderwidth=1)
            list_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=(0, 5))

            scrollbar = tk.Scrollbar(list_frame, orient=tk.VERTICAL)
            scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

            canvas = tk.Canvas(list_frame, bg='white', yscrollcommand=scrollbar.set, highlightthickness=0)
            canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            scrollbar.config(command=canvas.yview)

            checkboxes_frame = tk.Frame(canvas, bg='white')
            canvas_window = canvas.create_window((0, 0), window=checkboxes_frame, anchor='nw')

            # –Я—А–Њ–Ї—А—Г—В–Ї–∞
            def _on_mousewheel(e):
                canvas.yview_scroll(int(-1 * (e.delta / 120)), "units")

            def _bind_to_mousewheel(e):
                canvas.bind_all("<MouseWheel>", _on_mousewheel)

            def _unbind_from_mousewheel(e):
                canvas.unbind_all("<MouseWheel>")

            canvas.bind('<Enter>', _bind_to_mousewheel)
            canvas.bind('<Leave>', _unbind_from_mousewheel)

            trace_id = None

            # –Ю—З–Є—Б—В–Ї–∞ –њ—А–Є –Ј–∞–Ї—А—Л—В–Є–Є
            def cleanup_and_close():
                try:
                    _unbind_from_mousewheel(None)
                except:
                    pass

                try:
                    if trace_id is not None:
                        search_var.trace_remove('write', trace_id)
                except:
                    pass

                self._filter_window_open = False

                try:
                    filter_window.destroy()
                except:
                    pass

            filter_window.protocol("WM_DELETE_WINDOW", cleanup_and_close)

            # "–Т—Л–±—А–∞—В—М –≤—Б—С"
            all_selected = (len(currently_selected.intersection(unique_values)) == len(unique_values))
            select_all_var = tk.BooleanVar(value=all_selected)
            checkbox_vars = {}

            def toggle_all():
                state = select_all_var.get()
                search_text = search_var.get().lower().strip()
                for value, var in checkbox_vars.items():
                    if not search_text or search_text in value.lower():
                        var.set(state)

            select_all_frame = tk.Frame(checkboxes_frame, bg='#e8f4f8')
            select_all_frame.pack(fill=tk.X, pady=2)
            tk.Checkbutton(select_all_frame, text="вЬУ –Т—Л–±—А–∞—В—М –≤—Б—С",
                           variable=select_all_var, command=toggle_all,
                           font=("Arial", 10, "bold"), bg='#e8f4f8',
                           activebackground='#d1ecf1').pack(anchor='w', padx=5, pady=5)

            tk.Frame(checkboxes_frame, height=2, bg='#95a5a6').pack(fill=tk.X, padx=5, pady=2)

            checkbox_frames = {}

            # –°–Њ–Ј–і–∞—С–Љ —З–µ–Ї–±–Њ–Ї—Б—Л
            for value in unique_values:
                is_checked = (value in currently_selected)
                is_visible = (value in visible_unique_values)

                var = tk.BooleanVar(value=is_checked)
                checkbox_vars[value] = var

                bg_color = 'white' if is_visible else '#f8f8f8'
                cb_frame = tk.Frame(checkboxes_frame, bg=bg_color)
                cb_frame.pack(fill=tk.X, padx=2, pady=1)

                display_text = f"{value} рЯФТ" if not is_visible else value
                cb = tk.Checkbutton(cb_frame, text=display_text, variable=var,
                                    font=("Arial", 9, "italic" if not is_visible else "normal"),
                                    bg=bg_color, fg='#888' if not is_visible else 'black',
                                    activebackground='#e0e0e0' if not is_visible else '#f0f0f0')
                cb.pack(anchor='w', padx=10, pady=2)

                checkbox_frames[value] = cb_frame

            # рЯЖХ –§–£–Э–Ъ–¶–Ш–ѓ –§–Ш–Ы–ђ–Ґ–†–Р–¶–Ш–Ш
            def filter_checkboxes(*args):
                try:
                    if not filter_window.winfo_exists():
                        return
                except:
                    return

                search_text = search_var.get().lower().strip()

                # –°–Ї—А—Л–≤–∞–µ–Љ –≤—Б–µ
                for cb_frame in checkbox_frames.values():
                    cb_frame.pack_forget()

                # –Я–Њ–Ї–∞–Ј—Л–≤–∞–µ–Љ –љ—Г–ґ–љ—Л–µ
                for value in unique_values:
                    if value not in checkbox_vars:
                        continue

                    var = checkbox_vars[value]
                    cb_frame = checkbox_frames[value]

                    if not search_text:
                        cb_frame.pack(fill=tk.X, padx=2, pady=1)
                    elif search_text in value.lower():
                        cb_frame.pack(fill=tk.X, padx=2, pady=1)
                        var.set(True)
                    else:
                        var.set(False)

                # –Ю–±–љ–Њ–≤–ї—П–µ–Љ "–Т—Л–±—А–∞—В—М –≤—Б—С"
                if search_text:
                    checked = sum(1 for v in checkbox_vars.values() if v.get())
                    select_all_var.set(checked > 0)
                else:
                    checked = sum(1 for v in checkbox_vars.values() if v.get())
                    select_all_var.set(checked == len(checkbox_vars))

                # –Ю–±–љ–Њ–≤–ї—П–µ–Љ canvas
                try:
                    checkboxes_frame.update_idletasks()
                    canvas.configure(scrollregion=canvas.bbox("all"))
                except:
                    pass

            trace_id = search_var.trace_add('write', filter_checkboxes)
            search_entry.focus_set()

            def on_frame_configure(event=None):
                try:
                    canvas.configure(scrollregion=canvas.bbox("all"))
                    canvas.itemconfig(canvas_window, width=canvas.winfo_width())
                except:
                    pass

            checkboxes_frame.bind("<Configure>", on_frame_configure)
            canvas.bind("<Configure>", on_frame_configure)

            # –§—Г–љ–Ї—Ж–Є–Є –і–ї—П –Ї–љ–Њ–њ–Њ–Ї
            def apply_value_filter():
                selected_values = {value for value, var in checkbox_vars.items() if var.get()}
                if not selected_values:
                    messagebox.showwarning("–Я—А–µ–і—Г–њ—А–µ–ґ–і–µ–љ–Є–µ", "–Т—Л–±–µ—А–Є—В–µ —Е–Њ—В—П –±—Л –Њ–і–љ–Њ –Ј–љ–∞—З–µ–љ–Є–µ!")
                    return
                cleanup_and_close()
                self.apply_filter(column_id, selected_values, None)

            def clear_filter():
                if column_id in self.active_filters:
                    del self.active_filters[column_id]
                self.update_column_headers()
                cleanup_and_close()
                self.refresh_callback()

            # –Ъ–љ–Њ–њ–Ї–Є
            buttons_frame = tk.Frame(filter_window, bg='#ecf0f1')
            buttons_frame.pack(fill=tk.X, padx=10, pady=10)

            tk.Button(buttons_frame, text="вЬУ –Я—А–Є–Љ–µ–љ–Є—В—М —Д–Є–ї—М—В—А", command=apply_value_filter,
                      bg='#27ae60', fg='white', font=("Arial", 10, "bold"), relief=tk.RAISED, borderwidth=2).pack(
                side=tk.LEFT, padx=5, expand=True, fill=tk.X)

            tk.Button(buttons_frame, text="вЬЧ –°–±—А–Њ—Б–Є—В—М —Д–Є–ї—М—В—А", command=clear_filter,
                      bg='#e74c3c', fg='white', font=("Arial", 10)).pack(side=tk.LEFT, padx=5, expand=True, fill=tk.X)

            tk.Button(buttons_frame, text="–Ю—В–Љ–µ–љ–∞", command=cleanup_and_close,
                      bg='#95a5a6', fg='white', font=("Arial", 10)).pack(side=tk.LEFT, padx=5, expand=True, fill=tk.X)

        except Exception as e:
            print(f"вЭМ –Ю–®–Ш–С–Ъ–Р: {e}")
            import traceback
            traceback.print_exc()
            self._filter_window_open = False

    def apply_sort(self, column_id, direction, window):
        """–Я—А–Є–Љ–µ–љ–Є—В—М —Б–Њ—А—В–Є—А–Њ–≤–Ї—Г"""
        column_index = list(self.tree["columns"]).index(column_id)

        items = [(self.tree.item(item_id)["values"], item_id) for item_id in self.tree.get_children()]

        try:
            items.sort(key=lambda x: float(str(x[0][column_index]).replace(',', '.')),
                       reverse=(direction == 'desc'))
        except:
            items.sort(key=lambda x: str(x[0][column_index]).lower(),
                       reverse=(direction == 'desc'))

        for index, (values, item_id) in enumerate(items):
            self.tree.move(item_id, '', index)

        window.destroy()
        self._filter_window_open = False

    def apply_filter(self, column_id, selected_values, sort_order):
        """–Я—А–Є–Љ–µ–љ–Є—В—М —Д–Є–ї—М—В—А"""
        self.active_filters[column_id] = selected_values

        column_index = list(self.tree["columns"]).index(column_id)

        # –Я–Њ–Ї–∞–Ј—Л–≤–∞–µ–Љ –≤—Б–µ –Є–Ј –Ї—Н—И–∞
        for item_id in self._all_item_cache:
            try:
                self.tree.reattach(item_id, '', 'end')
            except:
                pass

        # –Я—А–Є–Љ–µ–љ—П–µ–Љ –Т–°–Х –∞–Ї—В–Є–≤–љ—Л–µ —Д–Є–ї—М—В—А—Л
        visible_items = set(self.tree.get_children(''))

        for col_id, values in self.active_filters.items():
            col_index = list(self.tree["columns"]).index(col_id)

            items_to_hide = set()

            for item_id in visible_items:
                try:
                    item_values = self.tree.item(item_id)["values"]
                    value = str(item_values[col_index])

                    if value not in values:
                        items_to_hide.add(item_id)
                except:
                    pass

            for item_id in items_to_hide:
                self.tree.detach(item_id)

            visible_items -= items_to_hide

        self.update_column_headers()

        if self.refresh_callback:
            self.refresh_callback()

    def update_column_headers(self):
        """–Ю–±–љ–Њ–≤–Є—В—М –Ј–∞–≥–Њ–ї–Њ–≤–Ї–Є –Ї–Њ–ї–Њ–љ–Њ–Ї (–і–Њ–±–∞–≤–Є—В—М/—Г–±—А–∞—В—М –Є–љ–і–Є–Ї–∞—В–Њ—А —Д–Є–ї—М—В—А–∞)"""
        for col in self.tree["columns"]:
            current_text = col.replace(" рЯФљ", "")

            if col in self.active_filters or current_text in self.active_filters:
                new_text = f"{current_text} рЯФљ"
            else:
                new_text = current_text

            self.tree.heading(col, text=new_text)

    def reapply_all_filters(self):
        """–Я–µ—А–µ–њ—А–Є–Љ–µ–љ–Є—В—М –≤—Б–µ –∞–Ї—В–Є–≤–љ—Л–µ —Д–Є–ї—М—В—А—Л"""
        if not self.active_filters:
            return

        # –Я–Њ–Ї–∞–Ј—Л–≤–∞–µ–Љ –≤—Б–µ
        for item_id in self._all_item_cache:
            try:
                self.tree.reattach(item_id, '', 'end')
            except:
                pass

        # –Я—А–Є–Љ–µ–љ—П–µ–Љ –Ї–∞–ґ–і—Л–є —Д–Є–ї—М—В—А
        visible_items = set(self.tree.get_children(''))

        for column_id, selected_values in self.active_filters.items():
            column_index = list(self.tree["columns"]).index(column_id)

            items_to_hide = set()

            for item_id in visible_items:
                try:
                    values = self.tree.item(item_id)["values"]
                    value = str(values[column_index])

                    if value not in selected_values:
                        items_to_hide.add(item_id)
                except:
                    pass

            for item_id in items_to_hide:
                self.tree.detach(item_id)

            visible_items -= items_to_hide

        self.update_column_headers()

    def clear_all_filters(self):
        """–Њ—З–Є—Б—В–Є—В—М –≤—Б–µ —Д–Є–ї—М—В—А—Л"""
        self.active_filters = {}

        for item_id in self._all_item_cache:
            try:
                self.tree.reattach(item_id, '', 'end')
            except:
                pass

        self.update_column_headers()

        if self.refresh_callback:
            self.refresh_callback()



class ProductionApp:
    def __init__(self, root):
        self.root = root
        self.root.title("–Ю–Ю–Ю –Т–Є—В–∞-–Ъ–∞")
        self.root.geometry("1400x800")
        self.root.configure(bg='#f0f0f0')

        # –°–Њ–Ј–і–∞—С–Љ –≤–µ—А—Е–љ—О—О –њ–∞–љ–µ–ї—М —Б –Ј–∞–≥–Њ–ї–Њ–≤–Ї–Њ–Љ –Є –Ї–љ–Њ–њ–Ї–Њ–є –љ–∞—Б—В—А–Њ–µ–Ї
        header_frame = tk.Frame(root, bg='#2c3e50', height=50)
        header_frame.pack(fill=tk.X, side=tk.TOP)
        header_frame.pack_propagate(False)

        # –Ч–∞–≥–Њ–ї–Њ–≤–Њ–Ї –њ—А–Є–ї–Њ–ґ–µ–љ–Є—П
        title_label = tk.Label(
            header_frame,
            text="рЯП≠ –°–Є—Б—В–µ–Љ–∞ —Г—З–µ—В–∞ –њ—А–Њ–Є–Ј–≤–Њ–і—Б—В–≤–∞",
            font=("Arial", 16, "bold"),
            bg='#2c3e50',
            fg='white'
        )
        title_label.pack(side=tk.LEFT, padx=20, pady=10)

        # –Ъ–љ–Њ–њ–Ї–∞ –љ–∞—Б—В—А–Њ–µ–Ї (—И–µ—Б—В–µ—А—С–љ–Ї–∞)
        settings_button = tk.Button(
            header_frame,
            text="вЪЩпЄП –Э–∞—Б—В—А–Њ–є–Ї–Є",
            font=("Arial", 11, "bold"),
            bg='#34495e',
            fg='white',
            activebackground='#1abc9c',
            activeforeground='white',
            relief=tk.FLAT,
            padx=15,
            pady=5,
            cursor='hand2',
            command=self.open_settings
        )
        settings_button.pack(side=tk.RIGHT, padx=20, pady=10)

        # –Ш–љ–Є—Ж–Є–∞–ї–Є–Ј–∞—Ж–Є—П –њ–µ—А–µ–Љ–µ–љ–љ—Л—Е toggles
        self.materials_toggles = {}
        self.orders_toggles = {}
        self.reservations_toggles = {}
        self.balance_toggles = {}
        self.writeoffs_toggles = {}
        self.details_toggles = {}

        # рЯЖХ –Ш–љ–Є—Ж–Є–∞–ї–Є–Ј–∞—Ж–Є—П –і–∞–љ–љ—Л—Е –і–ї—П –Є–Љ–њ–Њ—А—В–∞ –Њ—В –ї–∞–Ј–µ—А—Й–Є–Ї–Њ–≤
        self.laser_table_data = []

        self.notebook = ttk.Notebook(root)
        self.notebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        self.materials_frame = tk.Frame(self.notebook, bg='white')
        self.notebook.add(self.materials_frame, text='–Ь–∞—В–µ—А–Є–∞–ї—Л –љ–∞ —Б–Ї–ї–∞–і–µ')
        self.setup_materials_tab()

        self.orders_frame = tk.Frame(self.notebook, bg='white')
        self.notebook.add(self.orders_frame, text='–Ч–∞–Ї–∞–Ј—Л')
        self.setup_orders_tab()

        self.reservations_frame = tk.Frame(self.notebook, bg='white')
        self.notebook.add(self.reservations_frame, text='–†–µ–Ј–µ—А–≤–Є—А–Њ–≤–∞–љ–Є–µ')
        self.setup_reservations_tab()

        self.writeoffs_frame = tk.Frame(self.notebook, bg='white')
        self.notebook.add(self.writeoffs_frame, text='–°–њ–Є—Б–∞–љ–Є–µ –Љ–∞—В–µ—А–Є–∞–ї–Њ–≤')
        self.setup_writeoffs_tab()

        self.laser_import_frame = tk.Frame(self.notebook, bg='white')
        self.notebook.add(self.laser_import_frame, text='–Ш–Љ–њ–Њ—А—В –Њ—В –ї–∞–Ј–µ—А—Й–Є–Ї–Њ–≤')
        self.setup_laser_import_tab()

        # рЯЖХ –Э–Ю–Т–Р–ѓ –Т–Ъ–Ы–Р–Ф–Ъ–Р
        self.details_frame = tk.Frame(self.notebook, bg='white')
        self.notebook.add(self.details_frame, text='–£—З—С—В –і–µ—В–∞–ї–µ–є')
        self.setup_details_tab()

        self.balance_frame = tk.Frame(self.notebook, bg='white')
        self.notebook.add(self.balance_frame, text='–С–∞–ї–∞–љ—Б –Љ–∞—В–µ—А–Є–∞–ї–Њ–≤')
        self.setup_balance_tab()

        self.load_toggle_settings()

        # –Ч–∞–≥—А—Г–Ј–Ї–∞ –љ–∞—Б—В—А–Њ–µ–Ї –Є –Њ–±—А–∞–±–Њ—В—З–Є–Ї –Ј–∞–Ї—А—Л—В–Є—П
        self.load_toggle_settings()
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)

        self.material_logs_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.material_logs_frame, text="рЯУК –Ш—Б—В–Њ—А–Є—П –Љ–∞—В–µ—А–Є–∞–ї–Њ–≤")
        self.setup_material_logs_tab()

    def load_settings(self):
        """–Ч–∞–≥—А—Г–Ј–Ї–∞ –љ–∞—Б—В—А–Њ–µ–Ї –Є–Ј —Д–∞–є–ї–∞"""
        settings_file = "app_settings.json"
        default_settings = {
            "database_path": os.path.dirname(os.path.abspath(__file__))  # –Ґ–µ–Ї—Г—Й–∞—П –њ–∞–њ–Ї–∞ –њ–Њ —Г–Љ–Њ–ї—З–∞–љ–Є—О
        }

        try:
            if os.path.exists(settings_file):
                with open(settings_file, 'r', encoding='utf-8') as f:
                    settings = json.load(f)
                    print(f"вЬЕ –Э–∞—Б—В—А–Њ–є–Ї–Є –Ј–∞–≥—А—Г–ґ–µ–љ—Л: {settings}")
                    return settings
            else:
                print(f"вЪ†пЄП –§–∞–є–ї –љ–∞—Б—В—А–Њ–µ–Ї –љ–µ –љ–∞–є–і–µ–љ, –Є—Б–њ–Њ–ї—М–Ј—Г—О—В—Б—П –Ј–љ–∞—З–µ–љ–Є—П –њ–Њ —Г–Љ–Њ–ї—З–∞–љ–Є—О")
                return default_settings
        except Exception as e:
            print(f"вЭМ –Ю—И–Є–±–Ї–∞ –Ј–∞–≥—А—Г–Ј–Ї–Є –љ–∞—Б—В—А–Њ–µ–Ї: {e}")
            return default_settings

    def save_settings(self, settings):
        """–°–Њ—Е—А–∞–љ–µ–љ–Є–µ –љ–∞—Б—В—А–Њ–µ–Ї –≤ —Д–∞–є–ї"""
        settings_file = "app_settings.json"
        try:
            with open(settings_file, 'w', encoding='utf-8') as f:
                json.dump(settings, f, ensure_ascii=False, indent=4)
            print(f"вЬЕ –Э–∞—Б—В—А–Њ–є–Ї–Є —Б–Њ—Е—А–∞–љ–µ–љ—Л: {settings}")
            return True
        except Exception as e:
            print(f"вЭМ –Ю—И–Є–±–Ї–∞ —Б–Њ—Е—А–∞–љ–µ–љ–Є—П –љ–∞—Б—В—А–Њ–µ–Ї: {e}")
            messagebox.showerror("–Ю—И–Є–±–Ї–∞", f"–Э–µ —Г–і–∞–ї–Њ—Б—М —Б–Њ—Е—А–∞–љ–Є—В—М –љ–∞—Б—В—А–Њ–є–Ї–Є: {e}")
            return False

    def open_settings(self):
        """–Ю—В–Ї—А—Л—В–Є–µ –Њ–Ї–љ–∞ –љ–∞—Б—В—А–Њ–µ–Ї"""
        settings_window = tk.Toplevel(self.root)
        settings_window.title("вЪЩпЄП –Э–∞—Б—В—А–Њ–є–Ї–Є –њ—А–Њ–≥—А–∞–Љ–Љ—Л")
        settings_window.geometry("700x300")
        settings_window.configure(bg='#ecf0f1')
        settings_window.resizable(False, False)

        # –Ч–∞–≥–Њ–ї–Њ–≤–Њ–Ї
        header = tk.Label(
            settings_window,
            text="вЪЩпЄП –Э–∞—Б—В—А–Њ–є–Ї–Є —Б–Є—Б—В–µ–Љ—Л",
            font=("Arial", 16, "bold"),
            bg='#ecf0f1',
            fg='#2c3e50'
        )
        header.pack(pady=20)

        # –Ч–∞–≥—А—Г–ґ–∞–µ–Љ —В–µ–Ї—Г—Й–Є–µ –љ–∞—Б—В—А–Њ–є–Ї–Є
        current_settings = self.load_settings()

        # –Я—Г—В—М –Ї –њ–∞–њ–Ї–µ —Б –С–Ф
        path_frame = tk.LabelFrame(
            settings_window,
            text="рЯУБ –Я—Г—В—М –Ї —Б–Є—Б—В–µ–Љ–љ–Њ–є –њ–∞–њ–Ї–µ —Б –і–∞–љ–љ—Л–Љ–Є",
            bg='#ecf0f1',
            font=("Arial", 11, "bold"),
            fg='#34495e'
        )
        path_frame.pack(fill=tk.X, padx=30, pady=15)

        path_info = tk.Label(
            path_frame,
            text="–Т —Н—В–Њ–є –њ–∞–њ–Ї–µ –і–Њ–ї–ґ–љ—Л –љ–∞—Е–Њ–і–Є—В—М—Б—П —Д–∞–є–ї—Л:\nвАҐ production_database.xlsx\nвАҐ laser_import_cache.xlsx",
            bg='#ecf0f1',
            font=("Arial", 9),
            fg='#7f8c8d',
            justify=tk.LEFT
        )
        path_info.pack(anchor='w', padx=10, pady=5)

        path_entry_frame = tk.Frame(path_frame, bg='#ecf0f1')
        path_entry_frame.pack(fill=tk.X, padx=10, pady=10)

        path_var = tk.StringVar(value=current_settings.get("database_path", ""))
        path_entry = tk.Entry(
            path_entry_frame,
            textvariable=path_var,
            font=("Arial", 10),
            width=50
        )
        path_entry.pack(side=tk.LEFT, padx=5)

        def browse_folder():
            folder = filedialog.askdirectory(
                title="–Т—Л–±–µ—А–Є—В–µ –њ–∞–њ–Ї—Г —Б —Д–∞–є–ї–∞–Љ–Є –±–∞–Ј—Л –і–∞–љ–љ—Л—Е",
                initialdir=path_var.get()
            )
            if folder:
                path_var.set(folder)

        browse_button = tk.Button(
            path_entry_frame,
            text="рЯУВ –Ю–±–Ј–Њ—А...",
            font=("Arial", 10),
            bg='#3498db',
            fg='white',
            command=browse_folder,
            cursor='hand2'
        )
        browse_button.pack(side=tk.LEFT, padx=5)

        # –Ъ–љ–Њ–њ–Ї–Є –°–Њ—Е—А–∞–љ–Є—В—М/–Ю—В–Љ–µ–љ–∞
        buttons_frame = tk.Frame(settings_window, bg='#ecf0f1')
        buttons_frame.pack(pady=20)

        def save_and_close():
            new_path = path_var.get().strip()

            # –Я—А–Њ–≤–µ—А—П–µ–Љ —З—В–Њ –њ–∞–њ–Ї–∞ —Б—Г—Й–µ—Б—В–≤—Г–µ—В
            if not os.path.exists(new_path):
                messagebox.showerror(
                    "–Ю—И–Є–±–Ї–∞",
                    f"–Я–∞–њ–Ї–∞ –љ–µ —Б—Г—Й–µ—Б—В–≤—Г–µ—В:\n{new_path}"
                )
                return

            # –°–Њ—Е—А–∞–љ—П–µ–Љ –љ–∞—Б—В—А–Њ–є–Ї–Є
            new_settings = {
                "database_path": new_path
            }

            if self.save_settings(new_settings):
                messagebox.showinfo(
                    "–£—Б–њ–µ—Е",
                    "–Э–∞—Б—В—А–Њ–є–Ї–Є —Б–Њ—Е—А–∞–љ–µ–љ—Л!\n\n–Я–µ—А–µ–Ј–∞–њ—Г—Б—В–Є—В–µ –њ—А–Њ–≥—А–∞–Љ–Љ—Г –і–ї—П –њ—А–Є–Љ–µ–љ–µ–љ–Є—П –Є–Ј–Љ–µ–љ–µ–љ–Є–є."
                )
                settings_window.destroy()

        save_button = tk.Button(
            buttons_frame,
            text="рЯТЊ –°–Њ—Е—А–∞–љ–Є—В—М",
            font=("Arial", 11, "bold"),
            bg='#27ae60',
            fg='white',
            width=15,
            height=2,
            command=save_and_close,
            cursor='hand2'
        )
        save_button.pack(side=tk.LEFT, padx=10)

        cancel_button = tk.Button(
            buttons_frame,
            text="вЭМ –Ю—В–Љ–µ–љ–∞",
            font=("Arial", 11, "bold"),
            bg='#95a5a6',
            fg='white',
            width=15,
            height=2,
            command=settings_window.destroy,
            cursor='hand2'
        )
        cancel_button.pack(side=tk.LEFT, padx=10)

    def create_filter_panel(self, parent_frame, tree_widget, columns_to_filter, refresh_callback):
        """–°–Њ–Ј–і–∞–љ–Є–µ –њ–∞–љ–µ–ї–Є —Д–Є–ї—М—В—А–∞—Ж–Є–Є –і–ї—П –ї—О–±–Њ–є —В–∞–±–ї–Є—Ж—Л"""
        filter_frame = tk.LabelFrame(parent_frame, text="рЯФН –§–Є–ї—М—В—А—Л", bg='#e8f4f8', font=("Arial", 10, "bold"))
        filter_frame.pack(fill=tk.X, padx=10, pady=5)

        filter_entries = {}
        row = 0
        col = 0
        max_cols = 4

        for column_name in columns_to_filter:
            filter_container = tk.Frame(filter_frame, bg='#e8f4f8')
            filter_container.grid(row=row, column=col, padx=5, pady=3, sticky='w')

            tk.Label(filter_container, text=f"{column_name}:", bg='#e8f4f8', font=("Arial", 9)).pack(side=tk.LEFT)

            entry = tk.Entry(filter_container, width=15, font=("Arial", 9))
            entry.pack(side=tk.LEFT, padx=5)

            filter_entries[column_name] = entry

            entry.bind('<KeyRelease>', lambda e, tree=tree_widget, filters=filter_entries, cb=refresh_callback:
            self.apply_filters(tree, filters, cb))

            col += 1
            if col >= max_cols:
                col = 0
                row += 1

        buttons_container = tk.Frame(filter_frame, bg='#e8f4f8')
        buttons_container.grid(row=row + 1, column=0, columnspan=max_cols, pady=5)

        tk.Button(buttons_container, text="рЯЧСпЄП –Ю—З–Є—Б—В–Є—В—М —Д–Є–ї—М—В—А—Л", bg='#95a5a6', fg='white',
                  font=("Arial", 9),
                  command=lambda: self.clear_filters(filter_entries, tree_widget, refresh_callback)).pack(side=tk.LEFT,
                                                                                                          padx=5)

        tk.Button(buttons_container, text="рЯФД –Ю–±–љ–Њ–≤–Є—В—М", bg='#3498db', fg='white',
                  font=("Arial", 9), command=refresh_callback).pack(side=tk.LEFT, padx=5)

        return filter_entries

    def apply_filters(self, tree, filter_entries, refresh_callback):
        """–Я—А–Є–Љ–µ–љ–Є—В—М —Д–Є–ї—М—В—А—Л –Ї —В–∞–±–ї–Є—Ж–µ"""
        active_filters = {}
        for col_name, entry in filter_entries.items():
            filter_text = entry.get().strip().lower()
            if filter_text:
                active_filters[col_name] = filter_text

        if not active_filters:
            refresh_callback()
            return

        all_items = []
        for item in tree.get_children():
            all_items.append(tree.item(item)['values'])

        for item in tree.get_children():
            tree.delete(item)

        columns = tree['columns']
        for item_values in all_items:
            match = True
            for col_name, filter_text in active_filters.items():
                try:
                    col_index = columns.index(col_name)
                    cell_value = str(item_values[col_index]).lower()
                    if filter_text not in cell_value:
                        match = False
                        break
                except (ValueError, IndexError):
                    continue

            if match:
                tree.insert("", "end", values=item_values)

    def clear_filters(self, filter_entries, tree, refresh_callback):
        """–Ю—З–Є—Б—В–Є—В—М –≤—Б–µ —Д–Є–ї—М—В—А—Л"""
        for entry in filter_entries.values():
            entry.delete(0, tk.END)
        refresh_callback()


    def create_visibility_toggles(self, parent_frame, tree_widget, toggle_options, refresh_callback):
        """–°–Њ–Ј–і–∞–љ–Є–µ –њ–µ—А–µ–Ї–ї—О—З–∞—В–µ–ї–µ–є –≤–Є–і–Є–Љ–Њ—Б—В–Є –і–ї—П —В–∞–±–ї–Є—Ж—Л"""
        toggles_frame = tk.Frame(parent_frame, bg='#fff9e6')
        toggles_frame.pack(fill=tk.X, padx=10, pady=5)

        tk.Label(toggles_frame, text="рЯСБпЄП –Ю—В–Њ–±—А–∞–ґ–µ–љ–Є–µ:", bg='#fff9e6', font=("Arial", 10, "bold")).pack(side=tk.LEFT,
                                                                                                       padx=5)

        toggle_vars = {}

        for option_key, option_text in toggle_options.items():
            var = tk.BooleanVar(value=True)
            toggle_vars[option_key] = var

            cb = tk.Checkbutton(
                toggles_frame,
                text=option_text,
                variable=var,
                bg='#fff9e6',
                font=("Arial", 9),
                command=refresh_callback
            )
            cb.pack(side=tk.LEFT, padx=10)

        return toggle_vars

    def auto_resize_columns(self, tree, min_width=80, max_width=None):  # вЖР None –≤–Љ–µ—Б—В–Њ 400
        """–Р–≤—В–Њ–Љ–∞—В–Є—З–µ—Б–Ї–Є–є –њ–Њ–і–±–Њ—А —И–Є—А–Є–љ—Л –Ї–Њ–ї–Њ–љ–Њ–Ї –њ–Њ —Б–Њ–і–µ—А–ґ–Є–Љ–Њ–Љ—Г"""
        try:
            import tkinter.font as tkfont

            try:
                font = tkfont.Font(font=tree.cget("font"))
            except:
                font = tkfont.Font(family="Arial", size=10)

            for col in tree["columns"]:
                # –Ш–Ј–Љ–µ—А—П–µ–Љ —И–Є—А–Є–љ—Г –Ј–∞–≥–Њ–ї–Њ–≤–Ї–∞
                heading_text = tree.heading(col)["text"]
                heading_width = font.measure(heading_text) + 40

                # –Ш–Ј–Љ–µ—А—П–µ–Љ –Љ–∞–Ї—Б–Є–Љ–∞–ї—М–љ—Г—О —И–Є—А–Є–љ—Г –Ј–љ–∞—З–µ–љ–Є–є
                max_content_width = heading_width

                for item_id in tree.get_children():
                    try:
                        col_index = tree["columns"].index(col)
                        value = tree.item(item_id)["values"][col_index]
                        value_str = str(value)
                        value_width = font.measure(value_str) + 30

                        if value_width > max_content_width:
                            max_content_width = value_width
                    except:
                        continue

                # –Я—А–Є–Љ–µ–љ—П–µ–Љ –Њ–≥—А–∞–љ–Є—З–µ–љ–Є—П
                if max_width is not None:  # вЖР –Ф–Њ–±–∞–≤–Є—В—М –њ—А–Њ–≤–µ—А–Ї—Г
                    optimal_width = max(min_width, min(max_content_width, max_width))
                else:
                    optimal_width = max(min_width, max_content_width)

                tree.column(col, width=int(optimal_width))

                print(f"рЯУП –Ъ–Њ–ї–Њ–љ–Ї–∞ '{col}': {int(optimal_width)}px")

        except Exception as e:
            print(f"вЪ†пЄП –Ю—И–Є–±–Ї–∞ –∞–≤—В–Њ–њ–Њ–і–±–Њ—А–∞ —И–Є—А–Є–љ—Л –Ї–Њ–ї–Њ–љ–Њ–Ї: {e}")

    def save_toggle_settings(self):
        """–°–Њ—Е—А–∞–љ–Є—В—М –љ–∞—Б—В—А–Њ–є–Ї–Є –њ–µ—А–µ–Ї–ї—О—З–∞—В–µ–ї–µ–є"""
        settings = {}

        if hasattr(self, 'materials_toggles'):
            settings['materials'] = {k: v.get() for k, v in self.materials_toggles.items()}

        if hasattr(self, 'orders_toggles'):
            settings['orders'] = {k: v.get() for k, v in self.orders_toggles.items()}

        if hasattr(self, 'reservations_toggles'):
            settings['reservations'] = {k: v.get() for k, v in self.reservations_toggles.items()}

        if hasattr(self, 'balance_toggles'):
            settings['balance'] = {k: v.get() for k, v in self.balance_toggles.items()}

        if hasattr(self, 'writeoffs_toggles'):
            settings['writeoffs'] = {k: v.get() for k, v in self.writeoffs_toggles.items()}

        try:
            with open('toggle_settings.json', 'w', encoding='utf-8') as f:
                json.dump(settings, f, ensure_ascii=False, indent=2)
        except:
            pass

    def load_toggle_settings(self):
        """–Ч–∞–≥—А—Г–Ј–Є—В—М –љ–∞—Б—В—А–Њ–є–Ї–Є –њ–µ—А–µ–Ї–ї—О—З–∞—В–µ–ї–µ–є"""
        try:
            with open('toggle_settings.json', 'r', encoding='utf-8') as f:
                settings = json.load(f)

            if 'materials' in settings and hasattr(self, 'materials_toggles'):
                for k, v in settings['materials'].items():
                    if k in self.materials_toggles:
                        self.materials_toggles[k].set(v)

            if 'orders' in settings and hasattr(self, 'orders_toggles'):
                for k, v in settings['orders'].items():
                    if k in self.orders_toggles:
                        self.orders_toggles[k].set(v)

            if 'reservations' in settings and hasattr(self, 'reservations_toggles'):
                for k, v in settings['reservations'].items():
                    if k in self.reservations_toggles:
                        self.reservations_toggles[k].set(v)

            if 'balance' in settings and hasattr(self, 'balance_toggles'):
                for k, v in settings['balance'].items():
                    if k in self.balance_toggles:
                        self.balance_toggles[k].set(v)

            if 'writeoffs' in settings and hasattr(self, 'writeoffs_toggles'):
                for k, v in settings['writeoffs'].items():
                    if k in self.writeoffs_toggles:
                        self.writeoffs_toggles[k].set(v)

            self.refresh_materials()
            self.refresh_orders()
            self.refresh_reservations()
            self.refresh_balance()
            if hasattr(self, 'refresh_writeoffs'):
                self.refresh_writeoffs()
        except:
            pass

    def on_closing(self):
        """–Ю–±—А–∞–±–Њ—В—З–Є–Ї –Ј–∞–Ї—А—Л—В–Є—П –њ—А–Є–ї–Њ–ґ–µ–љ–Є—П"""
        # рЯЖХ –Р–Т–Ґ–Ю–°–Ю–•–†–Р–Э–Х–Э–Ш–Х –Ґ–Р–С–Ы–Ш–¶–Ђ –Ш–Ь–Я–Ю–†–Ґ–Р
        print("\nрЯТЊ –°–Њ—Е—А–∞–љ–µ–љ–Є–µ –і–∞–љ–љ—Л—Е –њ–µ—А–µ–і –Ј–∞–Ї—А—Л—В–Є–µ–Љ...")

        if hasattr(self, 'laser_table_data') and self.laser_table_data:
            self.save_laser_import_cache()

        # –°–Њ—Е—А–∞–љ—П–µ–Љ –љ–∞—Б—В—А–Њ–є–Ї–Є –њ–µ—А–µ–Ї–ї—О—З–∞—В–µ–ї–µ–є
        self.save_toggle_settings()

        print("вЬЕ –Ф–∞–љ–љ—Л–µ —Б–Њ—Е—А–∞–љ–µ–љ—Л")

        # –Ч–∞–Ї—А—Л–≤–∞–µ–Љ –њ—А–Є–ї–Њ–ґ–µ–љ–Є–µ
        self.root.destroy()

    def setup_materials_tab(self):
        header = tk.Label(self.materials_frame, text="–£—З–µ—В –ї–Є—Б—В–Њ–≤–Њ–≥–Њ –њ—А–Њ–Ї–∞—В–∞ –љ–∞ —Б–Ї–ї–∞–і–µ",
                          font=("Arial", 16, "bold"), bg='white', fg='#2c3e50')
        header.pack(pady=10)

        tree_frame = tk.Frame(self.materials_frame, bg='white')
        tree_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)

        scroll_y = tk.Scrollbar(tree_frame, orient=tk.VERTICAL)
        scroll_x = tk.Scrollbar(tree_frame, orient=tk.HORIZONTAL)

        self.materials_tree = ttk.Treeview(tree_frame,
                                           columns=("ID", "–Ь–∞—А–Ї–∞", "–Ґ–Њ–ї—Й–Є–љ–∞", "–Ф–ї–Є–љ–∞", "–®–Є—А–Є–љ–∞", "–Ъ–Њ–ї-–≤–Њ —И—В", "–Я–ї–Њ—Й–∞–і—М",
                                                    "–†–µ–Ј–µ—А–≤", "–Ф–Њ—Б—В—Г–њ–љ–Њ", "–Ф–∞—В–∞"),
                                           show="headings", yscrollcommand=scroll_y.set, xscrollcommand=scroll_x.set)
        scroll_y.config(command=self.materials_tree.yview)
        scroll_x.config(command=self.materials_tree.xview)
        scroll_y.pack(side=tk.RIGHT, fill=tk.Y)
        scroll_x.pack(side=tk.BOTTOM, fill=tk.X)

        # –Э–Р–°–Ґ–†–Ю–Щ–Ъ–Р –Ъ–Ю–Ы–Ю–Э–Ю–Ъ –С–Х–Ч –†–Р–°–Ґ–ѓ–У–Ш–Т–Р–Э–Ш–ѓ
        for col in self.materials_tree["columns"]:
            self.materials_tree.heading(col, text=col)
            self.materials_tree.column(col, anchor=tk.CENTER, width=100, minwidth=80, stretch=False)

        self.materials_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # –¶–Т–Х–Ґ–Ю–Т–Р–ѓ –Ш–Э–Ф–Ш–Ъ–Р–¶–Ш–ѓ (–Ъ–Р–Ъ –Т –С–Р–Ы–Р–Э–°–Х)
        self.materials_tree.tag_configure('negative', background='#f8d7da', foreground='#721c24')
        self.materials_tree.tag_configure('available', background='#d4edda', foreground='#155724')
        self.materials_tree.tag_configure('fully_reserved', background='#fff3cd', foreground='#856404')
        self.materials_tree.tag_configure('empty', background='#d1ecf1', foreground='#0c5460')

        # рЯЖХ –Ґ–Х–°–Ґ–Ю–Т–Р–ѓ –Т–°–Ґ–Р–Т–Ъ–Р –Ф–Ы–ѓ –Я–†–Ю–Т–Х–†–Ъ–Ш –†–Р–С–Ю–Ґ–Ђ –Ґ–Х–У–Ю–Т
        print("\nрЯІ™ === –Ґ–Х–°–Ґ–Ю–Т–Р–ѓ –Т–°–Ґ–Р–Т–Ъ–Р –Ф–Ы–ѓ –Я–†–Ю–Т–Х–†–Ъ–Ш –Ґ–Х–У–Ю–Т ===")
        test_negative = self.materials_tree.insert("", "end",
                                                   values=("TEST1", "–Ґ–Х–°–Ґ", "1.0", "1000", "1000", "10", "10", "15",
                                                           "-5", "2025-01-01"),
                                                   tags=('negative',))
        test_available = self.materials_tree.insert("", "end",
                                                    values=("TEST2", "–Ґ–Х–°–Ґ", "2.0", "2000", "2000", "20", "20", "5",
                                                            "15", "2025-01-01"),
                                                    tags=('available',))
        test_reserved = self.materials_tree.insert("", "end",
                                                   values=("TEST3", "–Ґ–Х–°–Ґ", "3.0", "3000", "3000", "30", "30", "30",
                                                           "0", "2025-01-01"),
                                                   tags=('fully_reserved',))
        test_empty = self.materials_tree.insert("", "end",
                                                values=("TEST4", "–Ґ–Х–°–Ґ", "4.0", "4000", "4000", "0", "0", "0", "0",
                                                        "2025-01-01"),
                                                tags=('empty',))

        print(f"   –Т—Б—В–∞–≤–ї–µ–љ TEST1 (negative): —В–µ–≥–Є = {self.materials_tree.item(test_negative, 'tags')}")
        print(f"   –Т—Б—В–∞–≤–ї–µ–љ TEST2 (available): —В–µ–≥–Є = {self.materials_tree.item(test_available, 'tags')}")
        print(f"   –Т—Б—В–∞–≤–ї–µ–љ TEST3 (fully_reserved): —В–µ–≥–Є = {self.materials_tree.item(test_reserved, 'tags')}")
        print(f"   –Т—Б—В–∞–≤–ї–µ–љ TEST4 (empty): —В–µ–≥–Є = {self.materials_tree.item(test_empty, 'tags')}")
        print(f"рЯІ™ === –Я–†–Ю–Т–Х–†–ђ–Ґ–Х –Ґ–Р–С–Ы–Ш–¶–£: –Т–Ш–Ф–Э–Ђ –Ы–Ш 4 –¶–Т–Х–Ґ–Э–Ђ–Х –°–Ґ–†–Ю–Ъ–Ш? ===\n")

        # рЯЖХ –Ш–Э–Ш–¶–Ш–Р–Ы–Ш–Ч–Р–¶–Ш–ѓ EXCEL-–§–Ш–Ы–ђ–Ґ–†–Р –Ф–Ы–ѓ –Ь–Р–Ґ–Х–†–Ш–Р–Ы–Ю–Т
        self.materials_excel_filter = ExcelStyleFilter(
            tree=self.materials_tree,
            refresh_callback=self.refresh_materials
        )

        # рЯЖХ –Ш–Э–Ф–Ш–Ъ–Р–Ґ–Ю–† –Р–Ъ–Ґ–Ш–Т–Э–Ђ–• –§–Ш–Ы–ђ–Ґ–†–Ю–Т
        self.materials_filter_status = tk.Label(
            self.materials_frame,
            text="",
            font=("Arial", 9),
            bg='#d1ecf1',
            fg='#0c5460'
        )
        self.materials_filter_status.pack(pady=5)

        # –Я–µ—А–µ–Ї–ї—О—З–∞—В–µ–ї–Є –≤–Є–і–Є–Љ–Њ—Б—В–Є
        self.materials_toggles = self.create_visibility_toggles(
            self.materials_frame,
            self.materials_tree,
            {
                'show_zero_stock': 'рЯУ¶ –Я–Њ–Ї–∞–Ј–∞—В—М —Б –љ—Г–ї–µ–≤—Л–Љ –Њ—Б—В–∞—В–Ї–Њ–Љ',
                'show_zero_available': 'вЬЕ –Я–Њ–Ї–∞–Ј–∞—В—М —Б –љ—Г–ї—С–Љ –і–Њ—Б—В—Г–њ–љ—Л—Е'
            },
            self.refresh_materials
        )

        buttons_frame = tk.Frame(self.materials_frame, bg='white')
        buttons_frame.pack(fill=tk.X, padx=10, pady=10)
        btn_style = {"font": ("Arial", 10), "width": 15, "height": 2}

        tk.Button(buttons_frame, text="–Ф–Њ–±–∞–≤–Є—В—М", bg='#27ae60', fg='white', command=self.add_material,
                  **btn_style).pack(side=tk.LEFT, padx=5)
        tk.Button(buttons_frame, text="–Ш–Љ–њ–Њ—А—В –Є–Ј Excel", bg='#9b59b6', fg='white', command=self.import_materials,
                  **btn_style).pack(side=tk.LEFT, padx=5)
        tk.Button(buttons_frame, text="–°–Ї–∞—З–∞—В—М —И–∞–±–ї–Њ–љ", bg='#3498db', fg='white', command=self.download_template,
                  **btn_style).pack(side=tk.LEFT, padx=5)
        tk.Button(buttons_frame, text="–†–µ–і–∞–Ї—В–Є—А–Њ–≤–∞—В—М", bg='#f39c12', fg='white', command=self.edit_material,
                  **btn_style).pack(side=tk.LEFT, padx=5)
        tk.Button(buttons_frame, text="–£–і–∞–ї–Є—В—М", bg='#e74c3c', fg='white', command=self.delete_material,
                  **btn_style).pack(side=tk.LEFT, padx=5)
        tk.Button(buttons_frame, text="вЬЦ –°–±—А–Њ—Б–Є—В—М —Д–Є–ї—М—В—А—Л", bg='#e67e22', fg='white',
                  command=self.clear_materials_filters, **btn_style).pack(side=tk.LEFT, padx=5)

        self.refresh_materials()

    def clear_materials_filters(self):
        """–°–±—А–Њ—Б–Є—В—М –≤—Б–µ —Д–Є–ї—М—В—А—Л –Љ–∞—В–µ—А–Є–∞–ї–Њ–≤"""
        if hasattr(self, 'materials_excel_filter'):
            self.materials_excel_filter.clear_all_filters()

    def refresh_materials(self):
        """–Ю–±–љ–Њ–≤–ї–µ–љ–Є–µ —Б–њ–Є—Б–Ї–∞ –Љ–∞—В–µ—А–Є–∞–ї–Њ–≤"""

        print(f"\n{'=' * 60}")
        print(f"рЯФД –Э–Р–І–Р–Ы–Ю refresh_materials()")
        print(f"{'=' * 60}")

        # –°–Ю–•–†–Р–Э–ѓ–Х–Ь –Р–Ъ–Ґ–Ш–Т–Э–Ђ–Х –§–Ш–Ы–ђ–Ґ–†–Ђ –Я–Х–†–Х–Ф –Ю–І–Ш–°–Ґ–Ъ–Ю–Щ
        active_filters_backup = {}
        if hasattr(self, 'materials_excel_filter') and self.materials_excel_filter.active_filters:
            active_filters_backup = self.materials_excel_filter.active_filters.copy()
            print(f"рЯФН –°–Њ—Е—А–∞–љ–µ–љ—Л —Д–Є–ї—М—В—А—Л –Љ–∞—В–µ—А–Є–∞–ї–Њ–≤: {list(active_filters_backup.keys())}")

        # –Я–Ю–Ы–Э–Ю–°–Ґ–ђ–Ѓ –Ю–І–Ш–©–Р–Х–Ь –Ф–Х–†–Х–Т–Ю
        for i in self.materials_tree.get_children():
            self.materials_tree.delete(i)

        # –Ю–І–Ш–©–Р–Х–Ь –Ъ–≠–® –≠–Ы–Х–Ь–Х–Э–Ґ–Ю–Т
        if hasattr(self, 'materials_excel_filter'):
            self.materials_excel_filter._all_item_cache = set()

        df = load_data("Materials")

        print(f"рЯУК –Ч–∞–≥—А—Г–ґ–µ–љ–Њ –Љ–∞—В–µ—А–Є–∞–ї–Њ–≤ –Є–Ј –С–Ф: {len(df)}")

        # рЯЖХ –Я–†–Ю–Т–Х–†–Ъ–Р: –Э–Р–°–Ґ–†–Ю–Х–Э–Ђ –Ы–Ш –Ґ–Х–У–Ш?
        print(f"\nрЯО® –Я—А–Њ–≤–µ—А–Ї–∞ –Ї–Њ–љ—Д–Є–≥—Г—А–∞—Ж–Є–Є —В–µ–≥–Њ–≤:")
        try:
            print(f"   negative: {self.materials_tree.tag_configure('negative')}")
            print(f"   available: {self.materials_tree.tag_configure('available')}")
            print(f"   fully_reserved: {self.materials_tree.tag_configure('fully_reserved')}")
            print(f"   empty: {self.materials_tree.tag_configure('empty')}")
        except Exception as e:
            print(f"   вЭМ –Ю–®–Ш–С–Ъ–Р –њ—А–Њ–≤–µ—А–Ї–Є —В–µ–≥–Њ–≤: {e}")

        if not df.empty:
            show_zero_stock = True
            show_zero_available = True

            if hasattr(self, 'materials_toggles') and self.materials_toggles:
                show_zero_stock = self.materials_toggles.get('show_zero_stock', tk.BooleanVar(value=True)).get()
                show_zero_available = self.materials_toggles.get('show_zero_available', tk.BooleanVar(value=True)).get()

            print(f"\nрЯУЛ –Э–∞—З–∞–ї–Њ –≤—Б—В–∞–≤–Ї–Є —Б—В—А–Њ–Ї:")
            inserted_count = 0
            tag_stats = {'negative': 0, 'available': 0, 'fully_reserved': 0, 'empty': 0}

            for index, row in df.iterrows():
                # рЯЖХ –Ф–Х–Ґ–Р–Ы–ђ–Э–Р–ѓ –Ф–Ш–Р–У–Э–Ю–°–Ґ–Ш–Ъ–Р –Ч–Э–Р–І–Х–Э–Ш–Щ
                quantity_raw = row["–Ъ–Њ–ї–Є—З–µ—Б—В–≤–Њ —И—В—Г–Ї"]
                available_raw = row["–Ф–Њ—Б—В—Г–њ–љ–Њ"]

                try:
                    quantity = int(quantity_raw) if quantity_raw else 0
                except:
                    quantity = 0

                try:
                    available = int(available_raw) if available_raw else 0
                except:
                    available = 0

                if not show_zero_stock and quantity == 0:
                    continue
                if not show_zero_available and available == 0:
                    continue

                values = (row["ID"], row["–Ь–∞—А–Ї–∞"], row["–Ґ–Њ–ї—Й–Є–љ–∞"], row["–Ф–ї–Є–љ–∞"], row["–®–Є—А–Є–љ–∞"],
                          row["–Ъ–Њ–ї–Є—З–µ—Б—В–≤–Њ —И—В—Г–Ї"], row["–Ю–±—Й–∞—П –њ–ї–Њ—Й–∞–і—М"], row["–Ч–∞—А–µ–Ј–µ—А–≤–Є—А–Њ–≤–∞–љ–Њ"],
                          row["–Ф–Њ—Б—В—Г–њ–љ–Њ"], row["–Ф–∞—В–∞ –і–Њ–±–∞–≤–ї–µ–љ–Є—П"])

                # рЯЖХ –Ю–Я–†–Х–Ф–Х–Ы–ѓ–Х–Ь –Ґ–Х–У –° –Ф–Х–Ґ–Р–Ы–ђ–Э–Ю–Щ –Ф–Ш–Р–У–Э–Ю–°–Ґ–Ш–Ъ–Ю–Щ
                tag = None

                if available < 0:
                    tag = 'negative'
                    tag_stats['negative'] += 1
                elif available > 0:
                    tag = 'available'
                    tag_stats['available'] += 1
                elif available == 0 and quantity > 0:
                    tag = 'fully_reserved'
                    tag_stats['fully_reserved'] += 1
                else:
                    tag = 'empty'
                    tag_stats['empty'] += 1

                # –Ф–Є–∞–≥–љ–Њ—Б—В–Є–Ї–∞ –њ–µ—А–≤—Л—Е 5 —Б—В—А–Њ–Ї
                if inserted_count < 5:
                    print(f"   –°—В—А–Њ–Ї–∞ {inserted_count}:")
                    print(f"      ID: {row['ID']}, –Ь–∞—А–Ї–∞: {row['–Ь–∞—А–Ї–∞']}")
                    print(f"      –Ъ–Њ–ї-–≤–Њ (raw): '{quantity_raw}' (type: {type(quantity_raw).__name__})")
                    print(f"      –Ъ–Њ–ї-–≤–Њ (parsed): {quantity} (type: {type(quantity).__name__})")
                    print(f"      –Ф–Њ—Б—В—Г–њ–љ–Њ (raw): '{available_raw}' (type: {type(available_raw).__name__})")
                    print(f"      –Ф–Њ—Б—В—Г–њ–љ–Њ (parsed): {available} (type: {type(available).__name__})")
                    print(f"      –£—Б–ї–Њ–≤–Є–µ: available={available}, quantity={quantity}")
                    print(f"      вЬЕ –Ґ–µ–≥: {tag}")

                # –Т—Б—В–∞–≤–ї—П–µ–Љ —Б —В–µ–≥–Њ–Љ
                item_id = self.materials_tree.insert("", "end", values=values, tags=(tag,))
                inserted_count += 1

                # рЯЖХ –Я–†–Ю–Т–Х–†–ѓ–Х–Ь –І–Ґ–Ю –Ґ–Х–У –Ф–Х–Щ–°–Ґ–Т–Ш–Ґ–Х–Ы–ђ–Э–Ю –Я–†–Ш–Ь–Х–Э–Ш–Ы–°–ѓ
                if inserted_count <= 5:
                    actual_tags = self.materials_tree.item(item_id, 'tags')
                    print(f"      –Я—А–Њ–≤–µ—А–Ї–∞: —Д–∞–Ї—В–Є—З–µ—Б–Ї–Є–µ —В–µ–≥–Є —Н–ї–µ–Љ–µ–љ—В–∞ = {actual_tags}")

                # –°–Ю–•–†–Р–Э–ѓ–Х–Ь item_id –Т –Ъ–≠–®
                if hasattr(self, 'materials_excel_filter'):
                    if not hasattr(self.materials_excel_filter, '_all_item_cache'):
                        self.materials_excel_filter._all_item_cache = set()
                    self.materials_excel_filter._all_item_cache.add(item_id)

            print(f"\nвЬЕ –Т—Б—В–∞–≤–ї–µ–љ–Њ —Б—В—А–Њ–Ї: {inserted_count}")
            print(f"рЯУК –°—В–∞—В–Є—Б—В–Є–Ї–∞ –њ–Њ —В–µ–≥–∞–Љ:")
            print(f"   рЯФі negative (–Ї—А–∞—Б–љ—Л–є): {tag_stats['negative']}")
            print(f"   рЯЯҐ available (–Ј–µ–ї—С–љ—Л–є): {tag_stats['available']}")
            print(f"   рЯЯ° fully_reserved (–ґ—С–ї—В—Л–є): {tag_stats['fully_reserved']}")
            print(f"   рЯФµ empty (–≥–Њ–ї—Г–±–Њ–є): {tag_stats['empty']}")

        # –Р–Т–Ґ–Ю–Я–Ю–Ф–С–Ю–† –®–Ш–†–Ш–Э–Ђ –Ъ–Ю–Ы–Ю–Э–Ю–Ъ
        self.auto_resize_columns(self.materials_tree, min_width=80, max_width=200)

        # –Я–Х–†–Х–Я–†–Ш–Ь–Х–Э–ѓ–Х–Ь –§–Ш–Ы–ђ–Ґ–†–Ђ –Я–Ю–°–Ы–Х –Ч–Р–У–†–£–Ч–Ъ–Ш –Ф–Р–Э–Э–Ђ–•
        if active_filters_backup and hasattr(self, 'materials_excel_filter'):
            print(f"\nрЯФД –Я–µ—А–µ–њ—А–Є–Љ–µ–љ—П—О —Д–Є–ї—М—В—А—Л –Љ–∞—В–µ—А–Є–∞–ї–Њ–≤: {list(active_filters_backup.keys())}")
            self.materials_excel_filter.active_filters = active_filters_backup
            self.materials_excel_filter.reapply_all_filters()

        # рЯЖХ –§–Ш–Э–Р–Ы–ђ–Э–Р–ѓ –Я–†–Ю–Т–Х–†–Ъ–Р: –Х–°–Ґ–ђ –Ы–Ш –Ґ–Х–У–Ш –£ –Т–Ш–Ф–Ш–Ь–Ђ–• –≠–Ы–Х–Ь–Х–Э–Ґ–Ю–Т?
        print(f"\nрЯФН –§–Є–љ–∞–ї—М–љ–∞—П –њ—А–Њ–≤–µ—А–Ї–∞ —В–µ–≥–Њ–≤ –≤–Є–і–Є–Љ—Л—Е —Н–ї–µ–Љ–µ–љ—В–Њ–≤:")
        visible_items = list(self.materials_tree.get_children(''))
        print(f"   –Т—Б–µ–≥–Њ –≤–Є–і–Є–Љ—Л—Е —Н–ї–µ–Љ–µ–љ—В–Њ–≤: {len(visible_items)}")

        for i, item_id in enumerate(visible_items[:3]):  # –Я–µ—А–≤—Л–µ 3
            tags = self.materials_tree.item(item_id, 'tags')
            values = self.materials_tree.item(item_id, 'values')
            print(f"   –≠–ї–µ–Љ–µ–љ—В {i}: ID={values[0]}, –Ф–Њ—Б—В—Г–њ–љ–Њ={values[8]}, –Ґ–µ–≥–Є={tags}")

        print(f"\n{'=' * 60}")
        print(f"вЬЕ –Ъ–Ю–Э–Х–¶ refresh_materials()")
        print(f"{'=' * 60}\n")

    def download_template(self):
        file_path = filedialog.asksaveasfilename(title="–°–Њ—Е—А–∞–љ–Є—В—М —И–∞–±–ї–Њ–љ", defaultextension=".xlsx",
                                                 filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                                                 initialfile="template_materials.xlsx")
        if not file_path:
            return
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "–Ь–∞—В–µ—А–Є–∞–ї—Л"
            ws.append(["–Ь–∞—А–Ї–∞", "–Ґ–Њ–ї—Й–Є–љ–∞", "–Ф–ї–Є–љ–∞", "–®–Є—А–Є–љ–∞", "–Ъ–Њ–ї–Є—З–µ—Б—В–≤–Њ —И—В—Г–Ї"])
            examples = [["09–У2–°", 10, 6000, 1500, 5], ["–°—В3", 12, 6000, 1500, 3], ["40–•", 8, 3000, 1250, 10]]
            for example in examples:
                ws.append(example)
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                ws.column_dimensions[column].width = max_length + 2
            wb.save(file_path)
            messagebox.showinfo("–£—Б–њ–µ—Е", f"–®–∞–±–ї–Њ–љ —Б–Њ—Е—А–∞–љ–µ–љ –≤:\n{file_path}")
        except Exception as e:
            messagebox.showerror("–Ю—И–Є–±–Ї–∞", f"–Э–µ —Г–і–∞–ї–Њ—Б—М —Б–Њ–Ј–і–∞—В—М —И–∞–±–ї–Њ–љ: {e}")

    def import_materials(self):
        file_path = filedialog.askopenfilename(title="–Т—Л–±–µ—А–Є—В–µ —Д–∞–є–ї Excel —Б –Љ–∞—В–µ—А–Є–∞–ї–∞–Љ–Є",
                                               filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")])
        if not file_path:
            return
        try:
            import_df = pd.read_excel(file_path, engine='openpyxl')
            required_columns = ["–Ь–∞—А–Ї–∞", "–Ґ–Њ–ї—Й–Є–љ–∞", "–Ф–ї–Є–љ–∞", "–®–Є—А–Є–љ–∞", "–Ъ–Њ–ї–Є—З–µ—Б—В–≤–Њ —И—В—Г–Ї"]
            missing_columns = [col for col in required_columns if col not in import_df.columns]
            if missing_columns:
                messagebox.showerror("–Ю—И–Є–±–Ї–∞", f"–Т —Д–∞–є–ї–µ –Њ—В—Б—Г—В—Б—В–≤—Г—О—В –Ї–Њ–ї–Њ–љ–Ї–Є:\n{', '.join(missing_columns)}")
                return
            materials_df = load_data("Materials")
            current_max_id = 0 if materials_df.empty else int(materials_df["ID"].max())
            imported_count = 0
            errors = []
            for idx, row in import_df.iterrows():
                try:
                    if pd.isna(row["–Ь–∞—А–Ї–∞"]) or row["–Ь–∞—А–Ї–∞"] == "":
                        continue
                    marka = str(row["–Ь–∞—А–Ї–∞"]).strip()
                    thickness = float(row["–Ґ–Њ–ї—Й–Є–љ–∞"])
                    length = float(row["–Ф–ї–Є–љ–∞"])
                    width = float(row["–®–Є—А–Є–љ–∞"])
                    quantity = int(row["–Ъ–Њ–ї–Є—З–µ—Б—В–≤–Њ —И—В—Г–Ї"])
                    duplicate = materials_df[(materials_df["–Ь–∞—А–Ї–∞"] == marka) & (materials_df["–Ґ–Њ–ї—Й–Є–љ–∞"] == thickness) &
                                             (materials_df["–Ф–ї–Є–љ–∞"] == length) & (materials_df["–®–Є—А–Є–љ–∞"] == width)]
                    if not duplicate.empty:
                        material_id = duplicate.iloc[0]["ID"]
                        old_qty = int(duplicate.iloc[0]["–Ъ–Њ–ї–Є—З–µ—Б—В–≤–Њ —И—В—Г–Ї"])
                        new_qty = old_qty + quantity
                        reserved = int(duplicate.iloc[0]["–Ч–∞—А–µ–Ј–µ—А–≤–Є—А–Њ–≤–∞–љ–Њ"])
                        area = (length * width * new_qty) / 1000000
                        materials_df.loc[materials_df["ID"] == material_id, "–Ъ–Њ–ї–Є—З–µ—Б—В–≤–Њ —И—В—Г–Ї"] = new_qty
                        materials_df.loc[materials_df["ID"] == material_id, "–Ю–±—Й–∞—П –њ–ї–Њ—Й–∞–і—М"] = round(area, 2)
                        materials_df.loc[materials_df["ID"] == material_id, "–Ф–Њ—Б—В—Г–њ–љ–Њ"] = new_qty - reserved
                    else:
                        current_max_id += 1
                        area = (length * width * quantity) / 1000000
                        new_row = pd.DataFrame([{"ID": current_max_id, "–Ь–∞—А–Ї–∞": marka, "–Ґ–Њ–ї—Й–Є–љ–∞": thickness,
                                                 "–Ф–ї–Є–љ–∞": length, "–®–Є—А–Є–љ–∞": width, "–Ъ–Њ–ї–Є—З–µ—Б—В–≤–Њ —И—В—Г–Ї": quantity,
                                                 "–Ю–±—Й–∞—П –њ–ї–Њ—Й–∞–і—М": round(area, 2), "–Ч–∞—А–µ–Ј–µ—А–≤–Є—А–Њ–≤–∞–љ–Њ": 0,
                                                 "–Ф–Њ—Б—В—Г–њ–љ–Њ": quantity,
                                                 "–Ф–∞—В–∞ –і–Њ–±–∞–≤–ї–µ–љ–Є—П": datetime.now().strftime("%Y-%m-%d")}])
                        materials_df = pd.concat([materials_df, new_row], ignore_index=True)
                    imported_count += 1
                except Exception as e:
                    errors.append(f"–°—В—А–Њ–Ї–∞ {idx + 2}: {str(e)}")
            save_data("Materials", materials_df)
            self.refresh_materials()
            self.refresh_balance()
            result_msg = f"–£—Б–њ–µ—И–љ–Њ –Є–Љ–њ–Њ—А—В–Є—А–Њ–≤–∞–љ–Њ: {imported_count} –Љ–∞—В–µ—А–Є–∞–ї–Њ–≤"
            if errors:
                result_msg += f"\n\n–Ю—И–Є–±–Ї–Є:\n" + "\n".join(errors[:10])
            messagebox.showinfo("–†–µ–Ј—Г–ї—М—В–∞—В –Є–Љ–њ–Њ—А—В–∞", result_msg)
        except Exception as e:
            messagebox.showerror("–Ю—И–Є–±–Ї–∞", f"–Э–µ —Г–і–∞–ї–Њ—Б—М –Є–Љ–њ–Њ—А—В–Є—А–Њ–≤–∞—В—М –і–∞–љ–љ—Л–µ:\n{e}")

    def add_material(self):
        add_window = tk.Toplevel(self.root)
        add_window.title("–Ф–Њ–±–∞–≤–Є—В—М –Љ–∞—В–µ—А–Є–∞–ї")
        add_window.geometry("450x500")
        add_window.configure(bg='#ecf0f1')
        tk.Label(add_window, text="–Ф–Њ–±–∞–≤–ї–µ–љ–Є–µ –ї–Є—Б—В–Њ–≤–Њ–≥–Њ –њ—А–Њ–Ї–∞—В–∞", font=("Arial", 12, "bold"), bg='#ecf0f1').pack(
            pady=10)
        fields = [("–Ь–∞—А–Ї–∞ —Б—В–∞–ї–Є:", "marka"), ("–Ґ–Њ–ї—Й–Є–љ–∞ (–Љ–Љ):", "thickness"), ("–Ф–ї–Є–љ–∞ (–Љ–Љ):", "length"),
                  ("–®–Є—А–Є–љ–∞ (–Љ–Љ):", "width"), ("–Ъ–Њ–ї–Є—З–µ—Б—В–≤–Њ —И—В—Г–Ї:", "quantity")]
        entries = {}
        for label_text, key in fields:
            frame = tk.Frame(add_window, bg='#ecf0f1')
            frame.pack(fill=tk.X, padx=20, pady=5)
            tk.Label(frame, text=label_text, width=20, anchor='w', bg='#ecf0f1', font=("Arial", 10)).pack(side=tk.LEFT)
            entry = tk.Entry(frame, font=("Arial", 10))
            entry.pack(side=tk.RIGHT, expand=True, fill=tk.X)
            entries[key] = entry

        def save_material():
            try:
                marka = entries["marka"].get().strip()
                thickness = float(entries["thickness"].get().strip())
                length = float(entries["length"].get().strip())
                width = float(entries["width"].get().strip())
                quantity = int(entries["quantity"].get().strip())
                if not marka:
                    messagebox.showwarning("–Я—А–µ–і—Г–њ—А–µ–ґ–і–µ–љ–Є–µ", "–Ч–∞–њ–Њ–ї–љ–Є—В–µ –Љ–∞—А–Ї—Г —Б—В–∞–ї–Є!")
                    return
                area = (length * width * quantity) / 1000000
                df = load_data("Materials")
                new_id = 1 if df.empty else int(df["ID"].max()) + 1
                new_row = pd.DataFrame(
                    [{"ID": new_id, "–Ь–∞—А–Ї–∞": marka, "–Ґ–Њ–ї—Й–Є–љ–∞": thickness, "–Ф–ї–Є–љ–∞": length, "–®–Є—А–Є–љ–∞": width,
                      "–Ъ–Њ–ї–Є—З–µ—Б—В–≤–Њ —И—В—Г–Ї": quantity, "–Ю–±—Й–∞—П –њ–ї–Њ—Й–∞–і—М": round(area, 2), "–Ч–∞—А–µ–Ј–µ—А–≤–Є—А–Њ–≤–∞–љ–Њ": 0,
                      "–Ф–Њ—Б—В—Г–њ–љ–Њ": quantity, "–Ф–∞—В–∞ –і–Њ–±–∞–≤–ї–µ–љ–Є—П": datetime.now().strftime("%Y-%m-%d")}])
                df = pd.concat([df, new_row], ignore_index=True)
                save_data("Materials", df)
                self.refresh_materials()
                self.refresh_balance()
                add_window.destroy()
                messagebox.showinfo("–£—Б–њ–µ—Е", "–Ь–∞—В–µ—А–Є–∞–ї —Г—Б–њ–µ—И–љ–Њ –і–Њ–±–∞–≤–ї–µ–љ!")
            except ValueError:
                messagebox.showerror("–Ю—И–Є–±–Ї–∞", "–Я—А–Њ–≤–µ—А—М—В–µ –њ—А–∞–≤–Є–ї—М–љ–Њ—Б—В—М –≤–≤–Њ–і–∞ —З–Є—Б–ї–Њ–≤—Л—Е –Ј–љ–∞—З–µ–љ–Є–є!")
            except Exception as e:
                messagebox.showerror("–Ю—И–Є–±–Ї–∞", f"–Э–µ —Г–і–∞–ї–Њ—Б—М –і–Њ–±–∞–≤–Є—В—М –Љ–∞—В–µ—А–Є–∞–ї: {e}")

        tk.Button(add_window, text="–°–Њ—Е—А–∞–љ–Є—В—М", bg='#27ae60', fg='white', font=("Arial", 12, "bold"),
                  command=save_material).pack(pady=20)

    def edit_material(self):
        selected = self.materials_tree.selection()
        if not selected:
            messagebox.showwarning("–Я—А–µ–і—Г–њ—А–µ–ґ–і–µ–љ–Є–µ", "–Т—Л–±–µ—А–Є—В–µ –Љ–∞—В–µ—А–Є–∞–ї –і–ї—П —А–µ–і–∞–Ї—В–Є—А–Њ–≤–∞–љ–Є—П")
            return

        item_id = self.materials_tree.item(selected)["values"][0]
        df = load_data("Materials")
        row = df[df["ID"] == item_id].iloc[0]

        # рЯЖХ –°–Ю–•–†–Р–Э–ѓ–Х–Ь –°–Ґ–Р–†–Ю–Х –Ъ–Ю–Ы–Ш–І–Х–°–Ґ–Т–Ю –Ф–Ы–ѓ –°–†–Р–Т–Э–Х–Э–Ш–ѓ
        old_quantity = int(row["–Ъ–Њ–ї–Є—З–µ—Б—В–≤–Њ —И—В—Г–Ї"])

        edit_window = tk.Toplevel(self.root)
        edit_window.title("–†–µ–і–∞–Ї—В–Є—А–Њ–≤–∞—В—М –Љ–∞—В–µ—А–Є–∞–ї")
        edit_window.geometry("450x600")  # вЖР –£–Т–Х–Ы–Ш–І–Ш–Ы–Ш –Т–Ђ–°–Ю–Ґ–£ –Ф–Ы–ѓ –Ъ–Ю–Ь–Ь–Х–Э–Ґ–Р–†–Ш–ѓ
        edit_window.configure(bg='#ecf0f1')

        tk.Label(edit_window, text="–†–µ–і–∞–Ї—В–Є—А–Њ–≤–∞–љ–Є–µ –Љ–∞—В–µ—А–Є–∞–ї–∞", font=("Arial", 12, "bold"), bg='#ecf0f1').pack(pady=10)

        fields = [("–Ь–∞—А–Ї–∞ —Б—В–∞–ї–Є:", "–Ь–∞—А–Ї–∞"), ("–Ґ–Њ–ї—Й–Є–љ–∞ (–Љ–Љ):", "–Ґ–Њ–ї—Й–Є–љ–∞"), ("–Ф–ї–Є–љ–∞ (–Љ–Љ):", "–Ф–ї–Є–љ–∞"),
                  ("–®–Є—А–Є–љ–∞ (–Љ–Љ):", "–®–Є—А–Є–љ–∞"), ("–Ъ–Њ–ї–Є—З–µ—Б—В–≤–Њ —И—В—Г–Ї:", "–Ъ–Њ–ї–Є—З–µ—Б—В–≤–Њ —И—В—Г–Ї")]
        entries = {}

        for label_text, key in fields:
            frame = tk.Frame(edit_window, bg='#ecf0f1')
            frame.pack(fill=tk.X, padx=20, pady=5)
            tk.Label(frame, text=label_text, width=20, anchor='w', bg='#ecf0f1', font=("Arial", 10)).pack(side=tk.LEFT)
            entry = tk.Entry(frame, font=("Arial", 10))
            entry.insert(0, str(row[key]))
            entry.pack(side=tk.RIGHT, expand=True, fill=tk.X)
            entries[key] = entry

        # рЯЖХ –Я–Ю–Ы–Х –Ф–Ы–ѓ –Ъ–Ю–Ь–Ь–Х–Э–Ґ–Р–†–Ш–ѓ (–Х–°–Ы–Ш –Ш–Ч–Ь–Х–Э–Ш–Ы–Ю–°–ђ –Ъ–Ю–Ы–Ш–І–Х–°–Ґ–Т–Ю)
        comment_frame = tk.Frame(edit_window, bg='#ecf0f1')
        comment_frame.pack(fill=tk.X, padx=20, pady=10)
        tk.Label(comment_frame, text="–Ъ–Њ–Љ–Љ–µ–љ—В–∞—А–Є–є\n(–µ—Б–ї–Є –Љ–µ–љ—П–µ—В–µ –Ї–Њ–ї-–≤–Њ):",
                 width=20, anchor='w', bg='#ecf0f1', font=("Arial", 10)).pack(side=tk.TOP, anchor='w')
        comment_entry = tk.Text(comment_frame, font=("Arial", 10), height=3, width=40)
        comment_entry.pack(side=tk.TOP, fill=tk.BOTH, expand=True, pady=5)

        def save_changes():
            try:
                thickness = float(entries["–Ґ–Њ–ї—Й–Є–љ–∞"].get())
                length = float(entries["–Ф–ї–Є–љ–∞"].get())
                width = float(entries["–®–Є—А–Є–љ–∞"].get())
                new_quantity = int(entries["–Ъ–Њ–ї–Є—З–µ—Б—В–≤–Њ —И—В—Г–Ї"].get())
                reserved = int(row["–Ч–∞—А–µ–Ј–µ—А–≤–Є—А–Њ–≤–∞–љ–Њ"])
                area = (length * width * new_quantity) / 1000000

                # рЯЖХ –Я–†–Ю–Т–Х–†–Ъ–Р: –Ш–Ч–Ь–Х–Э–Ш–Ы–Ю–°–ђ –Ы–Ш –Ъ–Ю–Ы–Ш–І–Х–°–Ґ–Т–Ю?
                quantity_changed = (new_quantity != old_quantity)

                if quantity_changed:
                    comment_text = comment_entry.get("1.0", tk.END).strip()

                    if not comment_text:
                        response = messagebox.askyesno(
                            "–Ъ–Њ–Љ–Љ–µ–љ—В–∞—А–Є–є –Њ—В—Б—Г—В—Б—В–≤—Г–µ—В",
                            "–Ъ–Њ–ї–Є—З–µ—Б—В–≤–Њ –Є–Ј–Љ–µ–љ–Є–ї–Њ—Б—М, –љ–Њ –Ї–Њ–Љ–Љ–µ–љ—В–∞—А–Є–є –љ–µ —Г–Ї–∞–Ј–∞–љ.\n\n"
                            "–Я—А–Њ–і–Њ–ї–ґ–Є—В—М –±–µ–Ј –Ї–Њ–Љ–Љ–µ–љ—В–∞—А–Є—П?"
                        )
                        if not response:
                            return
                        comment_text = "(–±–µ–Ј –Ї–Њ–Љ–Љ–µ–љ—В–∞—А–Є—П)"

                    # рЯЖХ –Ч–Р–Я–Ш–°–Ђ–Т–Р–Х–Ь –Ы–Ю–У –Ш–Ч–Ь–Х–Э–Х–Э–Ш–ѓ
                    self.log_material_change(
                        material_id=item_id,
                        marka=entries["–Ь–∞—А–Ї–∞"].get(),
                        thickness=thickness,
                        length=length,
                        width=width,
                        old_qty=old_quantity,
                        new_qty=new_quantity,
                        comment=comment_text
                    )

                # –°–Ю–•–†–Р–Э–ѓ–Х–Ь –Ш–Ч–Ь–Х–Э–Х–Э–Ш–ѓ –Т –Ь–Р–Ґ–Х–†–Ш–Р–Ы–Р–•
                df.loc[df["ID"] == item_id, "–Ь–∞—А–Ї–∞"] = entries["–Ь–∞—А–Ї–∞"].get()
                df.loc[df["ID"] == item_id, "–Ґ–Њ–ї—Й–Є–љ–∞"] = thickness
                df.loc[df["ID"] == item_id, "–Ф–ї–Є–љ–∞"] = length
                df.loc[df["ID"] == item_id, "–®–Є—А–Є–љ–∞"] = width
                df.loc[df["ID"] == item_id, "–Ъ–Њ–ї–Є—З–µ—Б—В–≤–Њ —И—В—Г–Ї"] = new_quantity
                df.loc[df["ID"] == item_id, "–Ю–±—Й–∞—П –њ–ї–Њ—Й–∞–і—М"] = round(area, 2)
                df.loc[df["ID"] == item_id, "–Ф–Њ—Б—В—Г–њ–љ–Њ"] = new_quantity - reserved

                save_data("Materials", df)
                self.refresh_materials()
                self.refresh_balance()
                edit_window.destroy()

                if quantity_changed:
                    messagebox.showinfo("–£—Б–њ–µ—Е",
                                        f"–Ь–∞—В–µ—А–Є–∞–ї –Њ–±–љ–Њ–≤–ї–µ–љ!\n\n"
                                        f"–Ъ–Њ–ї–Є—З–µ—Б—В–≤–Њ –Є–Ј–Љ–µ–љ–µ–љ–Њ: {old_quantity} вЖТ {new_quantity}\n"
                                        f"–Ш–Ј–Љ–µ–љ–µ–љ–Є–µ: {new_quantity - old_quantity:+d} —И—В.\n"
                                        f"–Ы–Њ–≥ –Ј–∞–њ–Є—Б–∞–љ.")
                else:
                    messagebox.showinfo("–£—Б–њ–µ—Е", "–Ь–∞—В–µ—А–Є–∞–ї —Г—Б–њ–µ—И–љ–Њ –Њ–±–љ–Њ–≤–ї–µ–љ!")

            except Exception as e:
                messagebox.showerror("–Ю—И–Є–±–Ї–∞", f"–Э–µ —Г–і–∞–ї–Њ—Б—М –Њ–±–љ–Њ–≤–Є—В—М –Љ–∞—В–µ—А–Є–∞–ї: {e}")

        tk.Button(edit_window, text="–°–Њ—Е—А–∞–љ–Є—В—М", bg='#3498db', fg='white', font=("Arial", 12, "bold"),
                  command=save_changes).pack(pady=20)

    def log_material_change(self, material_id, marka, thickness, length, width, old_qty, new_qty, comment):
        """–Ы–Њ–≥–Є—А–Њ–≤–∞–љ–Є–µ –Є–Ј–Љ–µ–љ–µ–љ–Є—П –Ї–Њ–ї–Є—З–µ—Б—В–≤–∞ –Љ–∞—В–µ—А–Є–∞–ї–∞ –≤—А—Г—З–љ—Г—О"""
        try:
            # –Ч–∞–≥—А—Г–ґ–∞–µ–Љ –ї–Њ–≥–Є (–Є–ї–Є —Б–Њ–Ј–і–∞—С–Љ –њ—Г—Б—В–Њ–є DataFrame –µ—Б–ї–Є –ї–Є—Б—В–∞ –љ–µ—В)
            try:
                logs_df = load_data("MaterialChangeLogs")
            except:
                logs_df = pd.DataFrame(columns=[
                    "ID –ї–Њ–≥–∞", "–Ф–∞—В–∞ –Є –≤—А–µ–Љ—П", "ID –Љ–∞—В–µ—А–Є–∞–ї–∞", "–Ь–∞—А–Ї–∞", "–Ґ–Њ–ї—Й–Є–љ–∞",
                    "–Ф–ї–Є–љ–∞", "–®–Є—А–Є–љ–∞", "–°—В–∞—А–Њ–µ –Ї–Њ–ї-–≤–Њ", "–Э–Њ–≤–Њ–µ –Ї–Њ–ї-–≤–Њ", "–Ш–Ј–Љ–µ–љ–µ–љ–Є–µ", "–Ъ–Њ–Љ–Љ–µ–љ—В–∞—А–Є–є"
                ])

            # –У–µ–љ–µ—А–Є—А—Г–µ–Љ ID –ї–Њ–≥–∞
            if logs_df.empty:
                log_id = 1
            else:
                log_id = int(logs_df["ID –ї–Њ–≥–∞"].max()) + 1

            # –Т—Л—З–Є—Б–ї—П–µ–Љ –Є–Ј–Љ–µ–љ–µ–љ–Є–µ
            change = new_qty - old_qty

            # рЯЖХ –Я–†–Р–Т–Ш–Ы–ђ–Э–Ђ–Щ –§–Ю–†–Ь–Р–Ґ: —З–Є—Б–ї–Њ —Б —П–≤–љ—Л–Љ –Ј–љ–∞–Ї–Њ–Љ
            if change > 0:
                change_str = f"+{change}"
            elif change < 0:
                change_str = str(change)  # –Љ–Є–љ—Г—Б —Г–ґ–µ –µ—Б—В—М
            else:
                change_str = "0"

            print(f"рЯФН –Ы–Њ–≥–Є—А–Њ–≤–∞–љ–Є–µ –Є–Ј–Љ–µ–љ–µ–љ–Є—П: —Б—В–∞—А–Њ–µ={old_qty}, –љ–Њ–≤–Њ–µ={new_qty}, –Є–Ј–Љ–µ–љ–µ–љ–Є–µ='{change_str}'")

            # –°–Њ–Ј–і–∞—С–Љ –љ–Њ–≤—Г—О –Ј–∞–њ–Є—Б—М
            new_log = pd.DataFrame([{
                "ID –ї–Њ–≥–∞": log_id,
                "–Ф–∞—В–∞ –Є –≤—А–µ–Љ—П": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "ID –Љ–∞—В–µ—А–Є–∞–ї–∞": material_id,
                "–Ь–∞—А–Ї–∞": marka,
                "–Ґ–Њ–ї—Й–Є–љ–∞": thickness,
                "–Ф–ї–Є–љ–∞": length,
                "–®–Є—А–Є–љ–∞": width,
                "–°—В–∞—А–Њ–µ –Ї–Њ–ї-–≤–Њ": old_qty,
                "–Э–Њ–≤–Њ–µ –Ї–Њ–ї-–≤–Њ": new_qty,
                "–Ш–Ј–Љ–µ–љ–µ–љ–Є–µ": change_str,  # вЖР –°–Ґ–†–Ю–Ъ–Р –° –Ч–Э–Р–Ъ–Ю–Ь: "+5" –Є–ї–Є "-3"
                "–Ъ–Њ–Љ–Љ–µ–љ—В–∞—А–Є–є": comment
            }])

            # –Ф–Њ–±–∞–≤–ї—П–µ–Љ –≤ –ї–Њ–≥–Є
            logs_df = pd.concat([logs_df, new_log], ignore_index=True)

            # –°–Њ—Е—А–∞–љ—П–µ–Љ
            save_data("MaterialChangeLogs", logs_df)

            print(
                f"вЬЕ –Ы–Њ–≥ –Є–Ј–Љ–µ–љ–µ–љ–Є—П –Ј–∞–њ–Є—Б–∞–љ: ID –Љ–∞—В–µ—А–Є–∞–ї–∞={material_id}, –Є–Ј–Љ–µ–љ–µ–љ–Є–µ={change_str}, –Ї–Њ–Љ–Љ–µ–љ—В–∞—А–Є–є='{comment}'")

            # –Р–Т–Ґ–Ю–Ь–Р–Ґ–Ш–І–Х–°–Ъ–Ю–Х –Ю–С–Э–Ю–Т–Ы–Х–Э–Ш–Х –Т–Ъ–Ы–Р–Ф–Ъ–Ш "–Ш—Б—В–Њ—А–Є—П –Љ–∞—В–µ—А–Є–∞–ї–Њ–≤"
            if hasattr(self, 'material_logs_tree'):
                self.refresh_material_logs()

        except Exception as e:
            print(f"вЪ†пЄП –Ю—И–Є–±–Ї–∞ –Ј–∞–њ–Є—Б–Є –ї–Њ–≥–∞ –Є–Ј–Љ–µ–љ–µ–љ–Є—П –Љ–∞—В–µ—А–Є–∞–ї–∞: {e}")
            import traceback
            traceback.print_exc()

    def setup_material_logs_tab(self):
        """–Т–Ї–ї–∞–і–Ї–∞ –Є—Б—В–Њ—А–Є–Є –Є–Ј–Љ–µ–љ–µ–љ–Є–є –Ї–Њ–ї–Є—З–µ—Б—В–≤–∞ –Љ–∞—В–µ—А–Є–∞–ї–Њ–≤"""

        # –Ч–∞–≥–Њ–ї–Њ–≤–Њ–Ї
        header_frame = tk.Frame(self.material_logs_frame, bg='white')
        header_frame.pack(fill=tk.X, pady=10)

        tk.Label(header_frame, text="–Ш—Б—В–Њ—А–Є—П –Є–Ј–Љ–µ–љ–µ–љ–Є–є –Ї–Њ–ї–Є—З–µ—Б—В–≤–∞ –Љ–∞—В–µ—А–Є–∞–ї–Њ–≤",
                 font=("Arial", 16, "bold"), bg='white', fg='#2c3e50').pack()

        # рЯЖХ –Ш–Э–Ф–Ш–Ъ–Р–Ґ–Ю–† –Ъ–Ю–Ы–Ш–І–Х–°–Ґ–Т–Р –Ч–Р–Я–Ш–°–Х–Щ
        self.material_logs_status = tk.Label(
            header_frame,
            text="–Ч–∞–≥—А—Г–Ј–Ї–∞...",
            font=("Arial", 10),
            bg='#d1ecf1',
            fg='#0c5460',
            relief=tk.RIDGE,
            padx=10,
            pady=5
        )
        self.material_logs_status.pack(pady=5)

        # –Ґ–∞–±–ї–Є—Ж–∞
        tree_frame = tk.Frame(self.material_logs_frame, bg='white')
        tree_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)

        scroll_y = tk.Scrollbar(tree_frame, orient=tk.VERTICAL)
        scroll_x = tk.Scrollbar(tree_frame, orient=tk.HORIZONTAL)

        self.material_logs_tree = ttk.Treeview(
            tree_frame,
            columns=("ID –ї–Њ–≥–∞", "–Ф–∞—В–∞", "ID –Љ–∞—В–µ—А–Є–∞–ї–∞", "–Ь–∞—А–Ї–∞", "–Ґ–Њ–ї—Й–Є–љ–∞", "–†–∞–Ј–Љ–µ—А",
                     "–°—В–∞—А–Њ–µ", "–Э–Њ–≤–Њ–µ", "–Ш–Ј–Љ–µ–љ–µ–љ–Є–µ", "–Ъ–Њ–Љ–Љ–µ–љ—В–∞—А–Є–є"),
            show="headings",
            yscrollcommand=scroll_y.set,
            xscrollcommand=scroll_x.set
        )

        scroll_y.config(command=self.material_logs_tree.yview)
        scroll_x.config(command=self.material_logs_tree.xview)
        scroll_y.pack(side=tk.RIGHT, fill=tk.Y)
        scroll_x.pack(side=tk.BOTTOM, fill=tk.X)

        columns_config = {
            "ID –ї–Њ–≥–∞": 70,
            "–Ф–∞—В–∞": 140,
            "ID –Љ–∞—В–µ—А–Є–∞–ї–∞": 90,
            "–Ь–∞—А–Ї–∞": 100,
            "–Ґ–Њ–ї—Й–Є–љ–∞": 70,
            "–†–∞–Ј–Љ–µ—А": 110,
            "–°—В–∞—А–Њ–µ": 80,
            "–Э–Њ–≤–Њ–µ": 80,
            "–Ш–Ј–Љ–µ–љ–µ–љ–Є–µ": 90,
            "–Ъ–Њ–Љ–Љ–µ–љ—В–∞—А–Є–є": 250
        }

        # –Э–∞—Б—В—А–Њ–є–Ї–∞ –Ї–Њ–ї–Њ–љ–Њ–Ї –С–Х–Ч —А–∞—Б—В—П–≥–Є–≤–∞–љ–Є—П
        for col, width in columns_config.items():
            self.material_logs_tree.heading(col, text=col)
            self.material_logs_tree.column(col, width=width, anchor=tk.CENTER, minwidth=80, stretch=False)

        self.material_logs_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # рЯЖХ –Ш–Э–Ш–¶–Ш–Р–Ы–Ш–Ч–Р–¶–Ш–ѓ EXCEL-–§–Ш–Ы–ђ–Ґ–†–Р –Ф–Ы–ѓ –Ш–°–Ґ–Ю–†–Ш–Ш –Ь–Р–Ґ–Х–†–Ш–Р–Ы–Ю–Т
        self.material_logs_excel_filter = ExcelStyleFilter(
            tree=self.material_logs_tree,
            refresh_callback=self.refresh_material_logs
        )

        # рЯЖХ –Ш–Э–Ф–Ш–Ъ–Р–Ґ–Ю–† –Р–Ъ–Ґ–Ш–Т–Э–Ђ–• –§–Ш–Ы–ђ–Ґ–†–Ю–Т
        self.material_logs_filter_status = tk.Label(
            self.material_logs_frame,
            text="",
            font=("Arial", 9),
            bg='#d1ecf1',
            fg='#0c5460'
        )
        self.material_logs_filter_status.pack(pady=5)

        # рЯЖХ –¶–Т–Х–Ґ–Ю–Т–Р–ѓ –Ш–Э–Ф–Ш–Ъ–Р–¶–Ш–ѓ –°–Ґ–†–Ю–Ъ (–Ч–Х–Ы–Б–Э–Ђ–Щ = –Ф–Ю–С–Р–Т–Ы–Х–Э–Ш–Х, –Ъ–†–Р–°–Э–Ђ–Щ = –£–Ь–Х–Э–ђ–®–Х–Э–Ш–Х)
        self.material_logs_tree.tag_configure('increase', background='#d4edda')  # –Ч–µ–ї—С–љ—Л–є
        self.material_logs_tree.tag_configure('decrease', background='#f8d7da')  # –Ъ—А–∞—Б–љ—Л–є
        self.material_logs_tree.tag_configure('neutral', background='white')

        # рЯЖХ –Ш–Э–Ш–¶–Ш–Р–Ы–Ш–Ч–Р–¶–Ш–ѓ EXCEL-–§–Ш–Ы–ђ–Ґ–†–Р –Ф–Ы–ѓ –†–Х–Ч–Х–†–Т–Ш–†–Ю–Т–Р–Э–Ш–ѓ
        self.reservations_excel_filter = ExcelStyleFilter(
            tree=self.reservations_tree,
            refresh_callback=self.refresh_reservations
        )

        # рЯЖХ –Ш–Э–Ф–Ш–Ъ–Р–Ґ–Ю–† –Р–Ъ–Ґ–Ш–Т–Э–Ђ–• –§–Ш–Ы–ђ–Ґ–†–Ю–Т
        self.reservations_filter_status = tk.Label(
            self.reservations_frame,
            text="",
            font=("Arial", 9),
            bg='#d1ecf1',
            fg='#0c5460'
        )
        self.reservations_filter_status.pack(pady=5)

        # –Ъ–љ–Њ–њ–Ї–Є —Г–њ—А–∞–≤–ї–µ–љ–Є—П
        buttons_frame = tk.Frame(self.material_logs_frame, bg='white')
        buttons_frame.pack(fill=tk.X, padx=10, pady=10)

        btn_style = {"font": ("Arial", 10), "width": 15, "height": 2}

        tk.Button(buttons_frame, text="–Ю–±–љ–Њ–≤–Є—В—М", bg='#95a5a6', fg='white',
                  command=self.refresh_material_logs, **btn_style).pack(side=tk.LEFT, padx=5)

        tk.Button(buttons_frame, text="–≠–Ї—Б–њ–Њ—А—В –≤ Excel", bg='#3498db', fg='white',
                  command=self.export_material_logs, **btn_style).pack(side=tk.LEFT, padx=5)

        tk.Button(buttons_frame, text="вЬЦ –°–±—А–Њ—Б–Є—В—М —Д–Є–ї—М—В—А—Л", bg='#e67e22', fg='white',
                  command=self.clear_material_logs_filters, **btn_style).pack(side=tk.LEFT, padx=5)

        # –Я–µ—А–≤–Њ–љ–∞—З–∞–ї—М–љ–∞—П –Ј–∞–≥—А—Г–Ј–Ї–∞
        self.refresh_material_logs()

    def refresh_material_logs(self):
        """–Ю–±–љ–Њ–≤–ї–µ–љ–Є–µ —В–∞–±–ї–Є—Ж—Л –ї–Њ–≥–Њ–≤"""
        for item in self.material_logs_tree.get_children():
            self.material_logs_tree.delete(item)

        try:
            logs_df = load_data("MaterialChangeLogs")

            if not logs_df.empty:
                # –°–Њ—А—В–Є—А—Г–µ–Љ –њ–Њ –і–∞—В–µ (–љ–Њ–≤—Л–µ —Б–≤–µ—А—Е—Г)
                logs_df = logs_df.sort_values("–Ф–∞—В–∞ –Є –≤—А–µ–Љ—П", ascending=False)

                for _, log in logs_df.iterrows():
                    size_str = f"{int(log['–Ф–ї–Є–љ–∞'])}x{int(log['–®–Є—А–Є–љ–∞'])}"

                    values = (
                        int(log["ID –ї–Њ–≥–∞"]),
                        log["–Ф–∞—В–∞ –Є –≤—А–µ–Љ—П"],
                        int(log["ID –Љ–∞—В–µ—А–Є–∞–ї–∞"]),
                        log["–Ь–∞—А–Ї–∞"],
                        log["–Ґ–Њ–ї—Й–Є–љ–∞"],
                        size_str,
                        int(log["–°—В–∞—А–Њ–µ –Ї–Њ–ї-–≤–Њ"]),
                        int(log["–Э–Њ–≤–Њ–µ –Ї–Њ–ї-–≤–Њ"]),
                        log["–Ш–Ј–Љ–µ–љ–µ–љ–Є–µ"],
                        log["–Ъ–Њ–Љ–Љ–µ–љ—В–∞—А–Є–є"]
                    )

                    # рЯЖХ –¶–Т–Х–Ґ–Ю–Т–Р–ѓ –Ш–Э–Ф–Ш–Ъ–Р–¶–Ш–ѓ –Я–Ю –Ш–Ч–Ь–Х–Э–Х–Э–Ш–Ѓ
                    change_str = str(log["–Ш–Ј–Љ–µ–љ–µ–љ–Є–µ"])
                    if change_str.startswith('+'):
                        tag = 'increase'  # –Ч–µ–ї—С–љ—Л–є (–і–Њ–±–∞–≤–ї–µ–љ–Є–µ)
                    elif change_str.startswith('-'):
                        tag = 'decrease'  # –Ъ—А–∞—Б–љ—Л–є (—Г–Љ–µ–љ—М—И–µ–љ–Є–µ)
                    else:
                        tag = 'neutral'

                    item_id = self.material_logs_tree.insert("", "end", values=values, tags=(tag,))

                    # –°–Ю–•–†–Р–Э–ѓ–Х–Ь item_id –Т –Ъ–≠–®
                    if hasattr(self, 'material_logs_excel_filter'):
                        if not hasattr(self.material_logs_excel_filter, '_all_item_cache'):
                            self.material_logs_excel_filter._all_item_cache = set()
                        self.material_logs_excel_filter._all_item_cache.add(item_id)

                # рЯЖХ –Ю–С–Э–Ю–Т–Ы–ѓ–Х–Ь –°–Ґ–Р–Ґ–£–°
                total = len(logs_df)
                increase_count = len(logs_df[logs_df["–Ш–Ј–Љ–µ–љ–µ–љ–Є–µ"].str.startswith('+')])
                decrease_count = len(logs_df[logs_df["–Ш–Ј–Љ–µ–љ–µ–љ–Є–µ"].str.startswith('-')])

                status_text = (
                    f"рЯУК –Т—Б–µ–≥–Њ –Ј–∞–њ–Є—Б–µ–є: {total} | "
                    f"рЯЯҐ –Ф–Њ–±–∞–≤–ї–µ–љ–Є–є: {increase_count} | "
                    f"рЯФі –£–Љ–µ–љ—М—И–µ–љ–Є–є: {decrease_count}"
                )
                self.material_logs_status.config(text=status_text, bg='#d1ecf1', fg='#0c5460')
            else:
                self.material_logs_status.config(
                    text="вДєпЄП –Ш—Б—В–Њ—А–Є—П –Є–Ј–Љ–µ–љ–µ–љ–Є–є –њ—Г—Б—В–∞",
                    bg='#fff3cd',
                    fg='#856404'
                )

            # –Р–Т–Ґ–Ю–Я–Ю–Ф–С–Ю–† –®–Ш–†–Ш–Э–Ђ –Ъ–Ю–Ы–Ю–Э–Ю–Ъ
            self.auto_resize_columns(self.material_logs_tree, min_width=80, max_width=300)

            # –Я–Х–†–Х–Я–†–Ш–Ь–Х–Э–ѓ–Х–Ь –§–Ш–Ы–ђ–Ґ–†–Ђ
            if active_filters_backup and hasattr(self, 'material_logs_excel_filter'):
                self.material_logs_excel_filter.active_filters = active_filters_backup
                self.material_logs_excel_filter.reapply_all_filters()

        except Exception as e:
            print(f"вЪ†пЄП –Ю—И–Є–±–Ї–∞ –Ј–∞–≥—А—Г–Ј–Ї–Є –ї–Њ–≥–Њ–≤: {e}")
            import traceback
            traceback.print_exc()
            self.material_logs_status.config(
                text=f"вЭМ –Ю—И–Є–±–Ї–∞ –Ј–∞–≥—А—Г–Ј–Ї–Є: {e}",
                bg='#f8d7da',
                fg='#721c24'
            )

    def export_material_logs(self):
        """–≠–Ї—Б–њ–Њ—А—В –Є—Б—В–Њ—А–Є–Є –Є–Ј–Љ–µ–љ–µ–љ–Є–є –≤ Excel"""
        try:
            logs_df = load_data("MaterialChangeLogs")

            if logs_df.empty:
                messagebox.showwarning("–Я—А–µ–і—Г–њ—А–µ–ґ–і–µ–љ–Є–µ", "–Э–µ—В –і–∞–љ–љ—Л—Е –і–ї—П —Н–Ї—Б–њ–Њ—А—В–∞!")
                return

            file_path = filedialog.asksaveasfilename(
                title="–≠–Ї—Б–њ–Њ—А—В –Є—Б—В–Њ—А–Є–Є –Є–Ј–Љ–µ–љ–µ–љ–Є–є",
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                initialfile=f"material_changes_log_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            )

            if not file_path:
                return

            # –°–Њ—А—В–Є—А—Г–µ–Љ –њ–Њ –і–∞—В–µ (—Б—В–∞—А—Л–µ —Б–≤–µ—А—Е—Г –і–ї—П Excel)
            logs_df = logs_df.sort_values("–Ф–∞—В–∞ –Є –≤—А–µ–Љ—П", ascending=True)

            # –≠–Ї—Б–њ–Њ—А—В–Є—А—Г–µ–Љ
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                logs_df.to_excel(writer, index=False, sheet_name='–Ш—Б—В–Њ—А–Є—П –Є–Ј–Љ–µ–љ–µ–љ–Є–є')
                worksheet = writer.sheets['–Ш—Б—В–Њ—А–Є—П –Є–Ј–Љ–µ–љ–µ–љ–Є–є']

                # –Р–≤—В–Њ–њ–Њ–і–±–Њ—А —И–Є—А–Є–љ—Л –Ї–Њ–ї–Њ–љ–Њ–Ї
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width

            messagebox.showinfo("–£—Б–њ–µ—Е", f"–Ш—Б—В–Њ—А–Є—П –Є–Ј–Љ–µ–љ–µ–љ–Є–є —Н–Ї—Б–њ–Њ—А—В–Є—А–Њ–≤–∞–љ–∞:\n\n{file_path}")

        except Exception as e:
            messagebox.showerror("–Ю—И–Є–±–Ї–∞", f"–Э–µ —Г–і–∞–ї–Њ—Б—М —Н–Ї—Б–њ–Њ—А—В–Є—А–Њ–≤–∞—В—М:\n{e}")

    def refresh_material_logs(self):
        """–Ю–±–љ–Њ–≤–ї–µ–љ–Є–µ —В–∞–±–ї–Є—Ж—Л –ї–Њ–≥–Њ–≤"""
        for item in self.material_logs_tree.get_children():
            self.material_logs_tree.delete(item)

        try:
            logs_df = load_data("MaterialChangeLogs")

            if not logs_df.empty:
                # рЯЖХ –Ф–Ш–Р–У–Э–Ю–°–Ґ–Ш–Ъ–Р: –њ—А–Њ–≤–µ—А—П–µ–Љ —Д–Њ—А–Љ–∞—В –і–∞–љ–љ—Л—Е
                print(f"рЯФН –Ч–∞–≥—А—Г–ґ–µ–љ–Њ –ї–Њ–≥–Њ–≤: {len(logs_df)}")
                if len(logs_df) > 0:
                    first_change = logs_df.iloc[0]["–Ш–Ј–Љ–µ–љ–µ–љ–Є–µ"]
                    print(f"рЯФН –Я–µ—А–≤–∞—П –Ј–∞–њ–Є—Б—М '–Ш–Ј–Љ–µ–љ–µ–љ–Є–µ': '{first_change}' (—В–Є–њ: {type(first_change).__name__})")

                # –°–Њ—А—В–Є—А—Г–µ–Љ –њ–Њ –і–∞—В–µ (–љ–Њ–≤—Л–µ —Б–≤–µ—А—Е—Г)
                logs_df = logs_df.sort_values("–Ф–∞—В–∞ –Є –≤—А–µ–Љ—П", ascending=False)

                # –°—З—С—В—З–Є–Ї–Є
                increase_count = 0
                decrease_count = 0

                for _, log in logs_df.iterrows():
                    size_str = f"{int(log['–Ф–ї–Є–љ–∞'])}x{int(log['–®–Є—А–Є–љ–∞'])}"

                    # рЯЖХ –С–Х–Ч–Ю–Я–Р–°–Э–Ю–Х –Я–†–Х–Ю–С–†–Р–Ч–Ю–Т–Р–Э–Ш–Х "–Ш–Ј–Љ–µ–љ–µ–љ–Є–µ" –Т –°–Ґ–†–Ю–Ъ–£
                    change_value = log["–Ш–Ј–Љ–µ–љ–µ–љ–Є–µ"]

                    if pd.isna(change_value):
                        change_str = "0"
                    else:
                        change_str = str(change_value).strip()

                    values = (
                        int(log["ID –ї–Њ–≥–∞"]),
                        log["–Ф–∞—В–∞ –Є –≤—А–µ–Љ—П"],
                        int(log["ID –Љ–∞—В–µ—А–Є–∞–ї–∞"]),
                        log["–Ь–∞—А–Ї–∞"],
                        log["–Ґ–Њ–ї—Й–Є–љ–∞"],
                        size_str,
                        int(log["–°—В–∞—А–Њ–µ –Ї–Њ–ї-–≤–Њ"]),
                        int(log["–Э–Њ–≤–Њ–µ –Ї–Њ–ї-–≤–Њ"]),
                        change_str,  # вЖР –Ш–°–Я–Ю–Ы–ђ–Ч–£–Х–Ь –Я–†–Х–Ю–С–†–Р–Ч–Ю–Т–Р–Э–Э–£–Ѓ –°–Ґ–†–Ю–Ъ–£
                        log["–Ъ–Њ–Љ–Љ–µ–љ—В–∞—А–Є–є"]
                    )

                    # рЯЖХ –Я–†–Р–Т–Ш–Ы–ђ–Э–Р–ѓ –¶–Т–Х–Ґ–Ю–Т–Р–ѓ –Ш–Э–Ф–Ш–Ъ–Р–¶–Ш–ѓ
                    try:
                        # –Я—А–Њ–±—Г–µ–Љ —А–∞—Б–њ–∞—А—Б–Є—В—М –Є–Ј–Љ–µ–љ–µ–љ–Є–µ –Ї–∞–Ї —З–Є—Б–ї–Њ
                        change_num = int(change_str.replace('+', '').replace(' ', ''))

                        if change_num > 0:
                            tag = 'increase'  # –Ч–µ–ї—С–љ—Л–є (–і–Њ–±–∞–≤–ї–µ–љ–Є–µ)
                            increase_count += 1
                        elif change_num < 0:
                            tag = 'decrease'  # –Ъ—А–∞—Б–љ—Л–є (—Г–Љ–µ–љ—М—И–µ–љ–Є–µ)
                            decrease_count += 1
                        else:
                            tag = 'neutral'
                    except:
                        tag = 'neutral'

                    self.material_logs_tree.insert("", "end", values=values, tags=(tag,))

                # рЯЖХ –Ю–С–Э–Ю–Т–Ы–ѓ–Х–Ь –°–Ґ–Р–Ґ–£–°
                total = len(logs_df)

                status_text = (
                    f"рЯУК –Т—Б–µ–≥–Њ –Ј–∞–њ–Є—Б–µ–є: {total} | "
                    f"рЯЯҐ –Ф–Њ–±–∞–≤–ї–µ–љ–Є–є: {increase_count} | "
                    f"рЯФі –£–Љ–µ–љ—М—И–µ–љ–Є–є: {decrease_count}"
                )

                print(f"рЯУК –°—В–∞—В–Є—Б—В–Є–Ї–∞: –≤—Б–µ–≥–Њ={total}, –і–Њ–±–∞–≤–ї–µ–љ–Є–є={increase_count}, —Г–Љ–µ–љ—М—И–µ–љ–Є–є={decrease_count}")

                self.material_logs_status.config(text=status_text, bg='#d1ecf1', fg='#0c5460')
            else:
                self.material_logs_status.config(
                    text="вДєпЄП –Ш—Б—В–Њ—А–Є—П –Є–Ј–Љ–µ–љ–µ–љ–Є–є –њ—Г—Б—В–∞",
                    bg='#fff3cd',
                    fg='#856404'
                )

            self.auto_resize_columns(self.material_logs_tree)

        except Exception as e:
            print(f"вЪ†пЄП –Ю—И–Є–±–Ї–∞ –Ј–∞–≥—А—Г–Ј–Ї–Є –ї–Њ–≥–Њ–≤: {e}")
            import traceback
            traceback.print_exc()
            self.material_logs_status.config(
                text=f"вЭМ –Ю—И–Є–±–Ї–∞ –Ј–∞–≥—А—Г–Ј–Ї–Є: {e}",
                bg='#f8d7da',
                fg='#721c24'
            )

    def delete_material(self):
        selected = self.materials_tree.selection()
        if not selected:
            messagebox.showwarning("–Я—А–µ–і—Г–њ—А–µ–ґ–і–µ–љ–Є–µ", "–Т—Л–±–µ—А–Є—В–µ –Љ–∞—В–µ—А–Є–∞–ї—Л –і–ї—П —Г–і–∞–ї–µ–љ–Є—П")
            return
        count = len(selected)
        if messagebox.askyesno("–Я–Њ–і—В–≤–µ—А–ґ–і–µ–љ–Є–µ", f"–£–і–∞–ї–Є—В—М –≤—Л–±—А–∞–љ–љ—Л–µ –Љ–∞—В–µ—А–Є–∞–ї—Л ({count} —И—В)?"):
            df = load_data("Materials")
            for item in selected:
                item_id = self.materials_tree.item(item)["values"][0]
                df = df[df["ID"] != item_id]
            save_data("Materials", df)
            self.refresh_materials()
            self.refresh_balance()  # <-- –≠–Ґ–Р –°–Ґ–†–Ю–Ъ–Р –Ф–Ю–Ы–Ц–Э–Р –С–Ђ–Ґ–ђ!
            messagebox.showinfo("–£—Б–њ–µ—Е", f"–£–і–∞–ї–µ–љ–Њ –Љ–∞—В–µ—А–Є–∞–ї–Њ–≤: {count}")

    def setup_orders_tab(self):
        header = tk.Label(self.orders_frame, text="–£–њ—А–∞–≤–ї–µ–љ–Є–µ –Ј–∞–Ї–∞–Ј–∞–Љ–Є", font=("Arial", 16, "bold"), bg='white',
                          fg='#2c3e50')
        header.pack(pady=10)

        # ========== –Ґ–Р–С–Ы–Ш–¶–Р –Ч–Р–Ъ–Р–Ч–Ю–Т ==========
        orders_label = tk.Label(self.orders_frame, text="–°–њ–Є—Б–Њ–Ї –Ј–∞–Ї–∞–Ј–Њ–≤", font=("Arial", 12, "bold"), bg='white')
        orders_label.pack(pady=5)

        # рЯЖХ –§—А–µ–є–Љ —В–∞–±–ї–Є—Ж—Л –Ј–∞–Ї–∞–Ј–Њ–≤ –Э–Р –Т–°–Х–Щ –®–Ш–†–Ш–Э–Х (—Г–±—А–∞–љ–Њ —Ж–µ–љ—В—А–Є—А–Њ–≤–∞–љ–Є–µ)
        orders_tree_frame = tk.Frame(self.orders_frame, bg='white')
        orders_tree_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)

        scroll_y = tk.Scrollbar(orders_tree_frame, orient=tk.VERTICAL)
        scroll_y.pack(side=tk.RIGHT, fill=tk.Y)

        self.orders_tree = ttk.Treeview(orders_tree_frame,
                                        columns=("ID", "–Э–∞–Ј–≤–∞–љ–Є–µ", "–Ч–∞–Ї–∞–Ј—З–Є–Ї", "–Ф–∞—В–∞", "–°—В–∞—В—Г—Б", "–Я—А–Є–Љ–µ—З–∞–љ–Є—П"),
                                        show="headings", yscrollcommand=scroll_y.set, height=8)
        scroll_y.config(command=self.orders_tree.yview)

        # –Э–∞—Б—В—А–Њ–є–Ї–∞ –Ї–Њ–ї–Њ–љ–Њ–Ї –С–Х–Ч —А–∞—Б—В—П–≥–Є–≤–∞–љ–Є—П
        for col in self.orders_tree["columns"]:
            self.orders_tree.heading(col, text=col)
            self.orders_tree.column(col, anchor=tk.CENTER, width=100, minwidth=80, stretch=False)

        self.orders_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self.orders_tree.bind('<<TreeviewSelect>>', self.on_order_select)

        # –Ш–Э–Ш–¶–Ш–Р–Ы–Ш–Ч–Р–¶–Ш–ѓ EXCEL-–§–Ш–Ы–ђ–Ґ–†–Р –Ф–Ы–ѓ –Ч–Р–Ъ–Р–Ч–Ю–Т
        self.orders_excel_filter = ExcelStyleFilter(
            tree=self.orders_tree,
            refresh_callback=self.refresh_orders
        )

        # –Ш–Э–Ф–Ш–Ъ–Р–Ґ–Ю–† –Р–Ъ–Ґ–Ш–Т–Э–Ђ–• –§–Ш–Ы–ђ–Ґ–†–Ю–Т (–Ч–Р–Ъ–Р–Ч–Ђ)
        self.orders_filter_status = tk.Label(
            self.orders_frame,
            text="",
            font=("Arial", 9),
            bg='#d1ecf1',
            fg='#0c5460'
        )
        self.orders_filter_status.pack(pady=5)

        # –Я–µ—А–µ–Ї–ї—О—З–∞—В–µ–ї–Є –≤–Є–і–Є–Љ–Њ—Б—В–Є –Ј–∞–Ї–∞–Ј–Њ–≤
        self.orders_toggles = self.create_visibility_toggles(
            self.orders_frame,
            self.orders_tree,
            {
                'show_completed': 'вЬЕ –Я–Њ–Ї–∞–Ј–∞—В—М –Ј–∞–≤–µ—А—И—С–љ–љ—Л–µ',
                'show_cancelled': 'вЭМ –Я–Њ–Ї–∞–Ј–∞—В—М –Њ—В–Љ–µ–љ—С–љ–љ—Л–µ'
            },
            self.refresh_orders
        )

        # –Ъ–љ–Њ–њ–Ї–Є —Г–њ—А–∞–≤–ї–µ–љ–Є—П –Ј–∞–Ї–∞–Ј–∞–Љ–Є
        buttons_frame = tk.Frame(self.orders_frame, bg='white')
        buttons_frame.pack(fill=tk.X, padx=10, pady=5)
        btn_style = {"font": ("Arial", 10), "width": 15, "height": 2}
        tk.Button(buttons_frame, text="–Ф–Њ–±–∞–≤–Є—В—М –Ј–∞–Ї–∞–Ј", bg='#27ae60', fg='white', command=self.add_order,
                  **btn_style).pack(side=tk.LEFT, padx=5)
        tk.Button(buttons_frame, text="–Ш–Љ–њ–Њ—А—В –Є–Ј Excel", bg='#9b59b6', fg='white', command=self.import_orders,
                  **btn_style).pack(side=tk.LEFT, padx=5)
        tk.Button(buttons_frame, text="–°–Ї–∞—З–∞—В—М —И–∞–±–ї–Њ–љ", bg='#3498db', fg='white', command=self.download_orders_template,
                  **btn_style).pack(side=tk.LEFT, padx=5)
        tk.Button(buttons_frame, text="–†–µ–і–∞–Ї—В–Є—А–Њ–≤–∞—В—М", bg='#f39c12', fg='white', command=self.edit_order,
                  **btn_style).pack(side=tk.LEFT, padx=5)
        tk.Button(buttons_frame, text="–£–і–∞–ї–Є—В—М –Ј–∞–Ї–∞–Ј", bg='#e74c3c', fg='white', command=self.delete_order,
                  **btn_style).pack(side=tk.LEFT, padx=5)
        tk.Button(buttons_frame, text="вЬЦ –°–±—А–Њ—Б–Є—В—М —Д–Є–ї—М—В—А—Л", bg='#e67e22', fg='white',
                  command=self.clear_orders_filters, **btn_style).pack(side=tk.LEFT, padx=5)

        # ========== –Ґ–Р–С–Ы–Ш–¶–Р –Ф–Х–Ґ–Р–Ы–Х–Щ –Ч–Р–Ъ–Р–Ч–Р ==========
        details_label = tk.Label(self.orders_frame, text="–Ф–µ—В–∞–ї–Є –≤—Л–±—А–∞–љ–љ–Њ–≥–Њ –Ј–∞–Ї–∞–Ј–∞", font=("Arial", 12, "bold"),
                                 bg='white')
        details_label.pack(pady=5)

        # рЯЖХ –§—А–µ–є–Љ —В–∞–±–ї–Є—Ж—Л –і–µ—В–∞–ї–µ–є –Э–Р –Т–°–Х–Щ –®–Ш–†–Ш–Э–Х (—Г–±—А–∞–љ–Њ —Ж–µ–љ—В—А–Є—А–Њ–≤–∞–љ–Є–µ)
        details_tree_frame = tk.Frame(self.orders_frame, bg='white')
        details_tree_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)

        scroll_y2 = tk.Scrollbar(details_tree_frame, orient=tk.VERTICAL)
        scroll_y2.pack(side=tk.RIGHT, fill=tk.Y)

        self.order_details_tree = ttk.Treeview(details_tree_frame,
                                               columns=("ID", "ID –Ј–∞–Ї–∞–Ј–∞", "–Э–∞–Ј–≤–∞–љ–Є–µ –і–µ—В–∞–ї–Є", "–Ъ–Њ–ї–Є—З–µ—Б—В–≤–Њ", "–Я–Њ—А–µ–Ј–∞–љ–Њ",
                                                        "–Я–Њ–≥–љ—Г—В–Њ"),
                                               show="headings", yscrollcommand=scroll_y2.set)
        scroll_y2.config(command=self.order_details_tree.yview)

        # –Э–∞—Б—В—А–Њ–є–Ї–∞ –Ї–Њ–ї–Њ–љ–Њ–Ї –С–Х–Ч —А–∞—Б—В—П–≥–Є–≤–∞–љ–Є—П
        for col in self.order_details_tree["columns"]:
            self.order_details_tree.heading(col, text=col)
            self.order_details_tree.column(col, anchor=tk.CENTER, width=100, minwidth=80, stretch=False)

        self.order_details_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self.order_details_tree.bind('<Double-1>', self.on_detail_double_click)

        # –Я—А–Є–≤—П–Ј–Ї–∞ –њ—А–∞–≤–Њ–≥–Њ –Ї–ї–Є–Ї–∞ –і–ї—П –Ї–Њ–њ–Є—А–Њ–≤–∞–љ–Є—П
        self.order_details_tree.bind('<Button-3>', self.on_detail_right_click)

        # –Ш–Э–Ш–¶–Ш–Р–Ы–Ш–Ч–Р–¶–Ш–ѓ EXCEL-–§–Ш–Ы–ђ–Ґ–†–Р –Ф–Ы–ѓ –Ф–Х–Ґ–Р–Ы–Х–Щ
        self.order_details_excel_filter = ExcelStyleFilter(
            tree=self.order_details_tree,
            refresh_callback=self.refresh_order_details
        )

        # –Ш–Э–Ф–Ш–Ъ–Р–Ґ–Ю–† –Р–Ъ–Ґ–Ш–Т–Э–Ђ–• –§–Ш–Ы–ђ–Ґ–†–Ю–Т (–Ф–Х–Ґ–Р–Ы–Ш)
        self.order_details_filter_status = tk.Label(
            self.orders_frame,
            text="",
            font=("Arial", 9),
            bg='#d1ecf1',
            fg='#0c5460'
        )
        self.order_details_filter_status.pack(pady=5)

        # –Ъ–љ–Њ–њ–Ї–Є —Г–њ—А–∞–≤–ї–µ–љ–Є—П –і–µ—В–∞–ї—П–Љ–Є
        details_buttons_frame = tk.Frame(self.orders_frame, bg='white')
        details_buttons_frame.pack(fill=tk.X, padx=10, pady=5)
        tk.Button(details_buttons_frame, text="–Ф–Њ–±–∞–≤–Є—В—М –і–µ—В–∞–ї—М", bg='#27ae60', fg='white',
                  command=self.add_order_detail, **btn_style).pack(side=tk.LEFT, padx=5)
        tk.Button(details_buttons_frame, text="–†–µ–і–∞–Ї—В–Є—А–Њ–≤–∞—В—М –і–µ—В–∞–ї—М", bg='#f39c12', fg='white',
                  command=self.edit_order_detail, **btn_style).pack(side=tk.LEFT, padx=5)
        tk.Button(details_buttons_frame, text="–£–і–∞–ї–Є—В—М –і–µ—В–∞–ї—М", bg='#e74c3c', fg='white',
                  command=self.delete_order_detail, **btn_style).pack(side=tk.LEFT, padx=5)
        tk.Button(details_buttons_frame, text="вЬЦ –°–±—А–Њ—Б–Є—В—М —Д–Є–ї—М—В—А—Л", bg='#e67e22', fg='white',
                  command=self.clear_order_details_filters, **btn_style).pack(side=tk.LEFT, padx=5)

        self.refresh_orders()

    def clear_orders_filters(self):
        """–°–±—А–Њ—Б–Є—В—М –≤—Б–µ —Д–Є–ї—М—В—А—Л –Ј–∞–Ї–∞–Ј–Њ–≤"""
        if hasattr(self, 'orders_excel_filter'):
            self.orders_excel_filter.clear_all_filters()

    def clear_order_details_filters(self):
        """–°–±—А–Њ—Б–Є—В—М –≤—Б–µ —Д–Є–ї—М—В—А—Л –і–µ—В–∞–ї–µ–є –Ј–∞–Ї–∞–Ј–∞"""
        if hasattr(self, 'order_details_excel_filter'):
            self.order_details_excel_filter.clear_all_filters()

    def clear_material_logs_filters(self):
        """–°–±—А–Њ—Б–Є—В—М –≤—Б–µ —Д–Є–ї—М—В—А—Л –Є—Б—В–Њ—А–Є–Є –Љ–∞—В–µ—А–Є–∞–ї–Њ–≤"""
        if hasattr(self, 'material_logs_excel_filter'):
            self.material_logs_excel_filter.clear_all_filters()

    def on_order_select(self, event):
        self.refresh_order_details()

    def on_order_select(self, event):
        self.refresh_order_details()

    def refresh_orders(self):
        """–Ю–±–љ–Њ–≤–ї–µ–љ–Є–µ —Б–њ–Є—Б–Ї–∞ –Ј–∞–Ї–∞–Ј–Њ–≤"""

        # –°–Ю–•–†–Р–Э–ѓ–Х–Ь –Р–Ъ–Ґ–Ш–Т–Э–Ђ–Х –§–Ш–Ы–ђ–Ґ–†–Ђ –Я–Х–†–Х–Ф –Ю–І–Ш–°–Ґ–Ъ–Ю–Щ
        active_filters_backup = {}
        if hasattr(self, 'orders_excel_filter') and self.orders_excel_filter.active_filters:
            active_filters_backup = self.orders_excel_filter.active_filters.copy()
            print(f"рЯФН –°–Њ—Е—А–∞–љ–µ–љ—Л —Д–Є–ї—М—В—А—Л –Ј–∞–Ї–∞–Ј–Њ–≤: {list(active_filters_backup.keys())}")

        # –Я–Ю–Ы–Э–Ю–°–Ґ–ђ–Ѓ –Ю–І–Ш–©–Р–Х–Ь –Ф–Х–†–Х–Т–Ю
        for i in self.orders_tree.get_children():
            self.orders_tree.delete(i)

        # –Ю–І–Ш–©–Р–Х–Ь –Ъ–≠–® –≠–Ы–Х–Ь–Х–Э–Ґ–Ю–Т
        if hasattr(self, 'orders_excel_filter'):
            self.orders_excel_filter._all_item_cache = set()

        df = load_data("Orders")

        if not df.empty:
            show_completed = True
            show_cancelled = True

            if hasattr(self, 'orders_toggles') and self.orders_toggles:
                show_completed = self.orders_toggles.get('show_completed', tk.BooleanVar(value=True)).get()
                show_cancelled = self.orders_toggles.get('show_cancelled', tk.BooleanVar(value=True)).get()

            for index, row in df.iterrows():
                status = row["–°—В–∞—В—Г—Б"]

                if not show_completed and status == "–Ч–∞–≤–µ—А—И–µ–љ":
                    continue
                if not show_cancelled and status == "–Ю—В–Љ–µ–љ–µ–љ":
                    continue

                values = (row["ID –Ј–∞–Ї–∞–Ј–∞"], row["–Э–∞–Ј–≤–∞–љ–Є–µ –Ј–∞–Ї–∞–Ј–∞"], row["–Ч–∞–Ї–∞–Ј—З–Є–Ї"],
                          row["–Ф–∞—В–∞ —Б–Њ–Ј–і–∞–љ–Є—П"], row["–°—В–∞—В—Г—Б"], row["–Я—А–Є–Љ–µ—З–∞–љ–Є—П"])

                item_id = self.orders_tree.insert("", "end", values=values)

                # –°–Ю–•–†–Р–Э–ѓ–Х–Ь item_id –Т –Ъ–≠–®
                if hasattr(self, 'orders_excel_filter'):
                    if not hasattr(self.orders_excel_filter, '_all_item_cache'):
                        self.orders_excel_filter._all_item_cache = set()
                    self.orders_excel_filter._all_item_cache.add(item_id)

        # –Р–Т–Ґ–Ю–Я–Ю–Ф–С–Ю–† –®–Ш–†–Ш–Э–Ђ –Ъ–Ю–Ы–Ю–Э–Ю–Ъ
        self.auto_resize_columns(self.orders_tree, min_width=100, max_width=300)

        # –Я–Х–†–Х–Я–†–Ш–Ь–Х–Э–ѓ–Х–Ь –§–Ш–Ы–ђ–Ґ–†–Ђ –Я–Ю–°–Ы–Х –Ч–Р–У–†–£–Ч–Ъ–Ш –Ф–Р–Э–Э–Ђ–•
        if active_filters_backup and hasattr(self, 'orders_excel_filter'):
            print(f"рЯФД –Я–µ—А–µ–њ—А–Є–Љ–µ–љ—П—О —Д–Є–ї—М—В—А—Л –Ј–∞–Ї–∞–Ј–Њ–≤: {list(active_filters_backup.keys())}")
            self.orders_excel_filter.active_filters = active_filters_backup
            self.orders_excel_filter.reapply_all_filters()

    def refresh_order_details(self):
        """–Ю–±–љ–Њ–≤–ї–µ–љ–Є–µ –і–µ—В–∞–ї–µ–є –≤—Л–±—А–∞–љ–љ–Њ–≥–Њ –Ј–∞–Ї–∞–Ј–∞"""

        print(f"\nрЯФН refresh_order_details –≤—Л–Ј–≤–∞–љ")

        # –°–Ю–•–†–Р–Э–ѓ–Х–Ь –Р–Ъ–Ґ–Ш–Т–Э–Ђ–Х –§–Ш–Ы–ђ–Ґ–†–Ђ –Я–Х–†–Х–Ф –Ю–І–Ш–°–Ґ–Ъ–Ю–Щ
        active_filters_backup = {}
        if hasattr(self, 'order_details_excel_filter') and self.order_details_excel_filter.active_filters:
            active_filters_backup = self.order_details_excel_filter.active_filters.copy()
            print(f"   –°–Њ—Е—А–∞–љ–µ–љ—Л —Д–Є–ї—М—В—А—Л –і–µ—В–∞–ї–µ–є: {list(active_filters_backup.keys())}")

        # –Я–Ю–Ы–Э–Ю–°–Ґ–ђ–Ѓ –Ю–І–Ш–©–Р–Х–Ь –Ф–Х–†–Х–Т–Ю
        for i in self.order_details_tree.get_children():
            self.order_details_tree.delete(i)

        # –Ю–І–Ш–©–Р–Х–Ь –Ъ–≠–® –≠–Ы–Х–Ь–Х–Э–Ґ–Ю–Т
        if hasattr(self, 'order_details_excel_filter'):
            self.order_details_excel_filter._all_item_cache = set()

        selected = self.orders_tree.selection()
        if not selected:
            print(f"   вЭМ –Ч–∞–Ї–∞–Ј –љ–µ –≤—Л–±—А–∞–љ")
            return

        order_id = self.orders_tree.item(selected[0])["values"][0]
        print(f"   вЬЕ –Т—Л–±—А–∞–љ –Ј–∞–Ї–∞–Ј ID: {order_id}")

        df = load_data("OrderDetails")
        print(f"   рЯУК –Ч–∞–≥—А—Г–ґ–µ–љ–Њ –і–µ—В–∞–ї–µ–є –≤—Б–µ–≥–Њ: {len(df)}")

        if not df.empty:
            # рЯЖХ –Ф–Ш–Р–У–Э–Ю–°–Ґ–Ш–Ъ–Р: –Я–Ю–Ъ–Р–Ч–Ђ–Т–Р–Х–Ь –Ъ–Ю–Ы–Ю–Э–Ъ–Ш
            print(f"   рЯУЛ –Ъ–Њ–ї–Њ–љ–Ї–Є OrderDetails: {list(df.columns)}")

            # –§–Є–ї—М—В—А—Г–µ–Љ –і–µ—В–∞–ї–Є –њ–Њ ID –Ј–∞–Ї–∞–Ј–∞ (–Є—Б–њ–Њ–ї—М–Ј—Г–µ–Љ iloc[1] - –≤—В–Њ—А–∞—П –Ї–Њ–ї–Њ–љ–Ї–∞)
            try:
                details = df[df.iloc[:, 1] == order_id]
            except:
                # –Я–Њ–њ—Л—В–Ї–∞ –њ–Њ –љ–∞–Ј–≤–∞–љ–Є—О –Ї–Њ–ї–Њ–љ–Ї–Є
                if "ID –Ј–∞–Ї–∞–Ј–∞" in df.columns:
                    details = df[df["ID –Ј–∞–Ї–∞–Ј–∞"] == order_id]
                else:
                    details = pd.DataFrame()  # –Я—Г—Б—В–Њ–є DataFrame

            print(f"   рЯУК –Ф–µ—В–∞–ї–µ–є –і–ї—П –Ј–∞–Ї–∞–Ј–∞ {order_id}: {len(details)}")

            for index, row in details.iterrows():
                # рЯЖХ –Ш–°–Я–Ю–Ы–ђ–Ч–£–Х–Ь –Ш–Э–Ф–Х–Ъ–°–Ђ –Ъ–Ю–Ы–Ю–Э–Ю–Ъ (–Э–Р–Ф–Б–Ц–Э–Ю)
                try:
                    detail_id = row.iloc[0]  # ID –і–µ—В–∞–ї–Є
                    order_id_val = row.iloc[1]  # ID –Ј–∞–Ї–∞–Ј–∞
                    detail_name = row.iloc[2]  # –Э–∞–Ј–≤–∞–љ–Є–µ –і–µ—В–∞–ї–Є
                    quantity = row.iloc[3]  # –Ъ–Њ–ї–Є—З–µ—Б—В–≤–Њ
                    cut = row.iloc[4] if len(row) > 4 else 0  # –Я–Њ—А–µ–Ј–∞–љ–Њ
                    bent = row.iloc[5] if len(row) > 5 else 0  # –Я–Њ–≥–љ—Г—В–Њ

                    values = (detail_id, order_id_val, detail_name, quantity, cut, bent)

                    print(f"      вЬЕ –Т—Б—В–∞–≤–Ї–∞ –і–µ—В–∞–ї–Є: {values}")
                    item_id = self.order_details_tree.insert("", "end", values=values)

                    # –°–Ю–•–†–Р–Э–ѓ–Х–Ь item_id –Т –Ъ–≠–®
                    if hasattr(self, 'order_details_excel_filter'):
                        if not hasattr(self.order_details_excel_filter, '_all_item_cache'):
                            self.order_details_excel_filter._all_item_cache = set()
                        self.order_details_excel_filter._all_item_cache.add(item_id)

                except Exception as e:
                    print(f"      вЪ†пЄП –Ю—И–Є–±–Ї–∞ —З—В–µ–љ–Є—П –і–µ—В–∞–ї–Є {index}: {e}")
                    continue
        else:
            print(f"   вЭМ DataFrame OrderDetails –њ—Г—Б—В")

        # –Я—А–Њ–≤–µ—А—П–µ–Љ —Б–Ї–Њ–ї—М–Ї–Њ —Н–ї–µ–Љ–µ–љ—В–Њ–≤ –≤ –і–µ—А–µ–≤–µ
        visible_items = self.order_details_tree.get_children()
        print(f"   рЯУК –Т–Є–і–Є–Љ—Л—Е —Н–ї–µ–Љ–µ–љ—В–Њ–≤ –≤ –і–µ—А–µ–≤–µ: {len(visible_items)}")

        # –Р–Т–Ґ–Ю–Я–Ю–Ф–С–Ю–† –®–Ш–†–Ш–Э–Ђ –Ъ–Ю–Ы–Ю–Э–Ю–Ъ
        self.auto_resize_columns(self.order_details_tree, min_width=100, max_width=300)

        # –Я–Х–†–Х–Я–†–Ш–Ь–Х–Э–ѓ–Х–Ь –§–Ш–Ы–ђ–Ґ–†–Ђ –Я–Ю–°–Ы–Х –Ч–Р–У–†–£–Ч–Ъ–Ш –Ф–Р–Э–Э–Ђ–•
        if active_filters_backup and hasattr(self, 'order_details_excel_filter'):
            print(f"   рЯФД –Я–µ—А–µ–њ—А–Є–Љ–µ–љ—П—О —Д–Є–ї—М—В—А—Л –і–µ—В–∞–ї–µ–є: {list(active_filters_backup.keys())}")
            self.order_details_excel_filter.active_filters = active_filters_backup
            self.order_details_excel_filter.reapply_all_filters()

        # –§–Є–љ–∞–ї—М–љ–∞—П –њ—А–Њ–≤–µ—А–Ї–∞
        final_visible = self.order_details_tree.get_children()
        print(f"   вЬЕ –Ш—В–Њ–≥–Њ –≤–Є–і–Є–Љ—Л—Е —Н–ї–µ–Љ–µ–љ—В–Њ–≤: {len(final_visible)}\n")


    def on_detail_double_click(self, event):
        """–Ю–±—А–∞–±–Њ—В–Ї–∞ –і–≤–Њ–є–љ–Њ–≥–Њ –Ї–ї–Є–Ї–∞ –њ–Њ –і–µ—В–∞–ї–Є –і–ї—П —А–µ–і–∞–Ї—В–Є—А–Њ–≤–∞–љ–Є—П –њ—А—П–Љ–Њ –≤ —В–∞–±–ї–Є—Ж–µ"""
        try:
            region = self.order_details_tree.identify("region", event.x, event.y)
            if region != "cell":
                return

            # –Ю–њ—А–µ–і–µ–ї—П–µ–Љ –Ї–Њ–ї–Њ–љ–Ї—Г
            column = self.order_details_tree.identify_column(event.x)
            if not column:
                return

            # –Я—А–µ–Њ–±—А–∞–Ј—Г–µ–Љ #1, #2, #3 –≤ –Є–љ–і–µ–Ї—Б 0, 1, 2
            column_index = int(column.replace('#', '')) - 1

            # –Я—А–Њ–≤–µ—А—П–µ–Љ —З—В–Њ –Є–љ–і–µ–Ї—Б –≤ –њ—А–µ–і–µ–ї–∞—Е
            columns = self.order_details_tree['columns']
            if column_index < 0 or column_index >= len(columns):
                return

            column_name = columns[column_index]

            # –†–∞–Ј—А–µ—И–∞–µ–Љ —А–µ–і–∞–Ї—В–Є—А–Њ–≤–∞—В—М —В–Њ–ї—М–Ї–Њ –Я–Њ—А–µ–Ј–∞–љ–Њ –Є –Я–Њ–≥–љ—Г—В–Њ
            if column_name not in ["–Я–Њ—А–µ–Ј–∞–љ–Њ", "–Я–Њ–≥–љ—Г—В–Њ"]:
                return

            # –Ю–њ—А–µ–і–µ–ї—П–µ–Љ —Б—В—А–Њ–Ї—Г
            item = self.order_details_tree.identify_row(event.y)
            if not item:
                return

            # –Я–Њ–ї—Г—З–∞–µ–Љ –і–∞–љ–љ—Л–µ —Б—В—А–Њ–Ї–Є
            values = self.order_details_tree.item(item, 'values')
            if not values or len(values) < 6:
                return

            try:
                detail_id = int(values[0])
            except (ValueError, TypeError):
                messagebox.showerror("–Ю—И–Є–±–Ї–∞", "–Э–µ —Г–і–∞–ї–Њ—Б—М –Њ–њ—А–µ–і–µ–ї–Є—В—М ID –і–µ—В–∞–ї–Є")
                return

            # –°–†–Р–Ч–£ –Я–†–Ю–Т–Х–†–ѓ–Х–Ь —Б—Г—Й–µ—Б—В–≤–Њ–≤–∞–љ–Є–µ –і–µ—В–∞–ї–Є –≤ –±–∞–Ј–µ
            df = load_data("OrderDetails")
            if df.empty:
                messagebox.showwarning("–Я—А–µ–і—Г–њ—А–µ–ґ–і–µ–љ–Є–µ", "–Ґ–∞–±–ї–Є—Ж–∞ –і–µ—В–∞–ї–µ–є –њ—Г—Б—В–∞")
                return

            detail_exists = df[df["ID"] == detail_id]
            if detail_exists.empty:
                messagebox.showerror("–Ю—И–Є–±–Ї–∞",
                                     f"–Ф–µ—В–∞–ї—М ID {detail_id} –љ–µ –љ–∞–є–і–µ–љ–∞ –≤ –±–∞–Ј–µ –і–∞–љ–љ—Л—Е!\n\n"
                                     f"–Т–Њ–Ј–Љ–Њ–ґ–љ–Њ –і–∞–љ–љ—Л–µ —Г—Б—В–∞—А–µ–ї–Є. –Э–∞–ґ–Љ–Є—В–µ '–Ю–±–љ–Њ–≤–Є—В—М'.")
                self.refresh_order_details()
                return

            detail_name = values[2]

            try:
                total_qty = int(values[3])
                current_cut = int(values[4]) if values[4] and str(values[4]).strip() != '' else 0
                current_bent = int(values[5]) if values[5] and str(values[5]).strip() != '' else 0
            except (ValueError, IndexError):
                messagebox.showerror("–Ю—И–Є–±–Ї–∞", "–Э–µ —Г–і–∞–ї–Њ—Б—М –њ—А–Њ—З–Є—В–∞—В—М –Ј–љ–∞—З–µ–љ–Є—П –і–µ—В–∞–ї–Є")
                return

            # –Я–Њ–ї—Г—З–∞–µ–Љ –Ї–Њ–Њ—А–і–Є–љ–∞—В—Л —П—З–µ–є–Ї–Є
            x, y, width, height = self.order_details_tree.bbox(item, column)

            # –°–Њ–Ј–і–∞–µ–Љ Entry –і–ї—П —А–µ–і–∞–Ї—В–Є—А–Њ–≤–∞–љ–Є—П
            edit_entry = tk.Entry(self.order_details_tree, font=("Arial", 10))
            edit_entry.place(x=x, y=y, width=width, height=height)

            # –Т—Б—В–∞–≤–ї—П–µ–Љ —В–µ–Ї—Г—Й–µ–µ –Ј–љ–∞—З–µ–љ–Є–µ
            current_value = values[column_index]
            edit_entry.insert(0, str(current_value))
            edit_entry.select_range(0, tk.END)
            edit_entry.focus()

            def save_cell_edit(event=None):
                try:
                    new_value_str = edit_entry.get().strip()
                    if not new_value_str:
                        new_value = 0
                    else:
                        new_value = int(new_value_str)

                    if new_value < 0:
                        messagebox.showerror("–Ю—И–Є–±–Ї–∞", "–Ч–љ–∞—З–µ–љ–Є–µ –љ–µ –Љ–Њ–ґ–µ—В –±—Л—В—М –Њ—В—А–Є—Ж–∞—В–µ–ї—М–љ—Л–Љ!")
                        edit_entry.destroy()
                        return

                    # –Я–Х–†–Х–Ч–Р–У–†–£–Ц–Р–Х–Ь –і–∞–љ–љ—Л–µ –і–ї—П –∞–Ї—В—Г–∞–ї—М–љ–Њ—Б—В–Є
                    df = load_data("OrderDetails")
                    if df.empty:
                        messagebox.showerror("–Ю—И–Є–±–Ї–∞", "–Э–µ —Г–і–∞–ї–Њ—Б—М –Ј–∞–≥—А—Г–Ј–Є—В—М –і–µ—В–∞–ї–Є")
                        edit_entry.destroy()
                        return

                    # –Я–†–Ю–Т–Х–†–ѓ–Х–Ь —Б—Г—Й–µ—Б—В–≤–Њ–≤–∞–љ–Є–µ –і–µ—В–∞–ї–Є –Х–©–Х –†–Р–Ч
                    detail_row = df[df["ID"] == detail_id]
                    if detail_row.empty:
                        messagebox.showerror("–Ю—И–Є–±–Ї–∞",
                                             f"–Ф–µ—В–∞–ї—М ID {detail_id} –±—Л–ї–∞ —Г–і–∞–ї–µ–љ–∞!\n\n"
                                             f"–Ю–±–љ–Њ–≤–Є—В–µ —Б–њ–Є—Б–Њ–Ї –і–µ—В–∞–ї–µ–є.")
                        edit_entry.destroy()
                        self.refresh_order_details()
                        return

                    # –Я–Њ–ї—Г—З–∞–µ–Љ –∞–Ї—В—Г–∞–ї—М–љ—Л–µ –і–∞–љ–љ—Л–µ –Є–Ј –±–∞–Ј—Л
                    actual_row = detail_row.iloc[0]
                    actual_cut = int(actual_row.get("–Я–Њ—А–µ–Ј–∞–љ–Њ", 0)) if pd.notna(actual_row.get("–Я–Њ—А–µ–Ј–∞–љ–Њ")) else 0
                    actual_bent = int(actual_row.get("–Я–Њ–≥–љ—Г—В–Њ", 0)) if pd.notna(actual_row.get("–Я–Њ–≥–љ—Г—В–Њ")) else 0
                    actual_qty = int(actual_row["–Ъ–Њ–ї–Є—З–µ—Б—В–≤–Њ"])

                    # –Ю–њ—А–µ–і–µ–ї—П–µ–Љ —З—В–Њ —А–µ–і–∞–Ї—В–Є—А—Г–µ–Љ
                    if column_name == "–Я–Њ—А–µ–Ј–∞–љ–Њ":
                        new_cut = new_value
                        new_bent = actual_bent

                        if new_cut < new_bent:
                            if not messagebox.askyesno("–Я—А–µ–і—Г–њ—А–µ–ґ–і–µ–љ–Є–µ",
                                                       f"–Я–Њ—А–µ–Ј–∞–љ–Њ ({new_cut}) –Љ–µ–љ—М—И–µ –њ–Њ–≥–љ—Г—В–Њ–≥–Њ ({new_bent}).\n"
                                                       f"–≠—В–Њ –Њ–Ј–љ–∞—З–∞–µ—В, —З—В–Њ –њ–Њ–≥–љ—Г—В–Њ –±–Њ–ї—М—И–µ –Ј–∞–≥–Њ—В–Њ–≤–Њ–Ї —З–µ–Љ –µ—Б—В—М.\n\n"
                                                       f"–Я—А–Њ–і–Њ–ї–ґ–Є—В—М?"):
                                edit_entry.destroy()
                                return

                        if new_cut > actual_qty:
                            if not messagebox.askyesno("–Я—А–µ–і—Г–њ—А–µ–ґ–і–µ–љ–Є–µ",
                                                       f"–Я–Њ—А–µ–Ј–∞–љ–Њ ({new_cut}) –±–Њ–ї—М—И–µ –Њ–±—Й–µ–≥–Њ –Ї–Њ–ї–Є—З–µ—Б—В–≤–∞ ({actual_qty}).\n"
                                                       f"–Т–Њ–Ј–Љ–Њ–ґ–љ–Њ –µ—Б—В—М –Є–Ј–ї–Є—И–Ї–Є –Ј–∞–≥–Њ—В–Њ–≤–Њ–Ї.\n\n"
                                                       f"–Я—А–Њ–і–Њ–ї–ґ–Є—В—М?"):
                                edit_entry.destroy()
                                return

                        df.loc[df["ID"] == detail_id, "–Я–Њ—А–µ–Ј–∞–љ–Њ"] = new_cut

                    elif column_name == "–Я–Њ–≥–љ—Г—В–Њ":
                        new_cut = actual_cut
                        new_bent = new_value

                        if new_bent > new_cut:
                            if not messagebox.askyesno("–Я—А–µ–і—Г–њ—А–µ–ґ–і–µ–љ–Є–µ",
                                                       f"–Я–Њ–≥–љ—Г—В–Њ ({new_bent}) –±–Њ–ї—М—И–µ –њ–Њ—А–µ–Ј–∞–љ–љ–Њ–≥–Њ ({new_cut}).\n"
                                                       f"–Э—Г–ґ–љ–Њ —Б–љ–∞—З–∞–ї–∞ –њ–Њ—А–µ–Ј–∞—В—М –Ј–∞–≥–Њ—В–Њ–≤–Ї–Є.\n\n"
                                                       f"–Я—А–Њ–і–Њ–ї–ґ–Є—В—М?"):
                                edit_entry.destroy()
                                return

                        df.loc[df["ID"] == detail_id, "–Я–Њ–≥–љ—Г—В–Њ"] = new_bent

                    # –°–Њ—Е—А–∞–љ—П–µ–Љ
                    save_data("OrderDetails", df)
                    self.refresh_order_details()
                    edit_entry.destroy()

                    # –Я–Њ–Ї–∞–Ј—Л–≤–∞–µ–Љ –Ї—А–∞—В–Ї–Њ–µ —Г–≤–µ–і–Њ–Љ–ї–µ–љ–Є–µ
                    to_cut = actual_qty - new_cut
                    to_bend = new_cut - new_bent

                    status_msg = f"вЬЕ {detail_name}\n"
                    status_msg += f"–Я–Њ—А–µ–Ј–∞–љ–Њ: {new_cut}/{actual_qty} (–Њ—Б—В–∞–ї–Њ—Б—М: {to_cut})\n"
                    status_msg += f"–Я–Њ–≥–љ—Г—В–Њ: {new_bent}/{new_cut} (–Њ—Б—В–∞–ї–Њ—Б—М: {to_bend})"

                    self.show_status_tooltip(status_msg)

                except ValueError:
                    messagebox.showerror("–Ю—И–Є–±–Ї–∞", "–Т–≤–µ–і–Є—В–µ –Ї–Њ—А—А–µ–Ї—В–љ–Њ–µ —З–Є—Б–ї–Њ!")
                    edit_entry.destroy()
                except Exception as e:
                    messagebox.showerror("–Ю—И–Є–±–Ї–∞", f"–Э–µ —Г–і–∞–ї–Њ—Б—М –Њ–±–љ–Њ–≤–Є—В—М: {e}")
                    edit_entry.destroy()
                    import traceback
                    traceback.print_exc()

            # –Я—А–Є–≤—П–Ј—Л–≤–∞–µ–Љ —Б–Њ–±—Л—В–Є—П
            edit_entry.bind('<Return>', save_cell_edit)
            edit_entry.bind('<FocusOut>', save_cell_edit)
            edit_entry.bind('<Escape>', lambda e: edit_entry.destroy())

        except Exception as e:
            print(f"–Ю—И–Є–±–Ї–∞ –≤ on_detail_double_click: {e}")
            import traceback
            traceback.print_exc()

    def on_detail_right_click(self, event):
        """–Ъ–Њ–љ—В–µ–Ї—Б—В–љ–Њ–µ –Љ–µ–љ—О –њ—А–Є –њ—А–∞–≤–Њ–Љ –Ї–ї–Є–Ї–µ –љ–∞ –і–µ—В–∞–ї—М"""
        # –Ю–њ—А–µ–і–µ–ї—П–µ–Љ —Б—В—А–Њ–Ї—Г –њ–Њ–і –Ї—Г—А—Б–Њ—А–Њ–Љ
        item = self.order_details_tree.identify_row(event.y)
        if not item:
            return

        # –Т—Л–і–µ–ї—П–µ–Љ —Б—В—А–Њ–Ї—Г
        self.order_details_tree.selection_set(item)

        # –°–Њ–Ј–і–∞—С–Љ –Ї–Њ–љ—В–µ–Ї—Б—В–љ–Њ–µ –Љ–µ–љ—О
        context_menu = tk.Menu(self.root, tearoff=0)
        context_menu.add_command(
            label="рЯУЛ –Ъ–Њ–њ–Є—А–Њ–≤–∞—В—М –Є–љ—Д–Њ—А–Љ–∞—Ж–Є—О –Њ –і–µ—В–∞–ї–Є",
            command=lambda: self.copy_detail_info(item)
        )

        # –Я–Њ–Ї–∞–Ј—Л–≤–∞–µ–Љ –Љ–µ–љ—О
        try:
            context_menu.tk_popup(event.x_root, event.y_root)
        finally:
            context_menu.grab_release()

    def copy_detail_info(self, item_id):
        """–Ъ–Њ–њ–Є—А–Њ–≤–∞–љ–Є–µ –Є–љ—Д–Њ—А–Љ–∞—Ж–Є–Є –Њ –і–µ—В–∞–ї–Є –≤ –±—Г—Д–µ—А –Њ–±–Љ–µ–љ–∞"""
        try:
            # –Я–Њ–ї—Г—З–∞–µ–Љ –і–∞–љ–љ—Л–µ –Є–Ј —Б—В—А–Њ–Ї–Є —В–∞–±–ї–Є—Ж—Л
            values = self.order_details_tree.item(item_id)["values"]
            if not values or len(values) < 6:
                messagebox.showwarning("–Я—А–µ–і—Г–њ—А–µ–ґ–і–µ–љ–Є–µ", "–Э–µ —Г–і–∞–ї–Њ—Б—М –њ–Њ–ї—Г—З–Є—В—М –і–∞–љ–љ—Л–µ –і–µ—В–∞–ї–Є")
                return

            detail_id = values[0]
            order_id = values[1]
            detail_name = values[2]
            quantity = int(values[3])
            cut = int(values[4])
            bent = int(values[5])

            # –†–∞—Б—Б—З–Є—В—Л–≤–∞–µ–Љ –Њ—Б—В–∞—В–Њ–Ї –і–ї—П –љ–∞—А–µ–Ј–Ї–Є
            remaining_to_cut = quantity - cut

            # –Ч–∞–≥—А—Г–ґ–∞–µ–Љ –і–∞–љ–љ—Л–µ –Ј–∞–Ї–∞–Ј–∞
            orders_df = load_data("Orders")
            order_row = orders_df[orders_df["ID –Ј–∞–Ї–∞–Ј–∞"] == order_id]

            if order_row.empty:
                customer = "–Э–µ–Є–Ј–≤–µ—Б—В–љ–Њ"
                order_name = "–Э–µ–Є–Ј–≤–µ—Б—В–љ–Њ"
            else:
                customer = order_row.iloc[0]["–Ч–∞–Ї–∞–Ј—З–Є–Ї"]
                order_name = order_row.iloc[0]["–Э–∞–Ј–≤–∞–љ–Є–µ –Ј–∞–Ї–∞–Ј–∞"]

            # –Ч–∞–≥—А—Г–ґ–∞–µ–Љ –і–∞–љ–љ—Л–µ —А–µ–Ј–µ—А–≤–Є—А–Њ–≤–∞–љ–Є—П
            reservations_df = load_data("Reservations")

            # –Ш—Й–µ–Љ —А–µ–Ј–µ—А–≤—Л –і–ї—П —Н—В–Њ–є –і–µ—В–∞–ї–Є –Є –Ј–∞–Ї–∞–Ј–∞
            detail_reserves = reservations_df[
                (reservations_df["ID –Ј–∞–Ї–∞–Ј–∞"] == order_id) &
                (reservations_df["ID –і–µ—В–∞–ї–Є"] == detail_id)
                ]

            # –Я–µ—А–µ–Љ–µ–љ–љ—Л–µ –і–ї—П –≤—Л–≤–Њ–і–∞
            material_info = ""
            material_stock = ""
            remaining_reserved_count = ""

            if not detail_reserves.empty:
                # –Ч–∞–≥—А—Г–ґ–∞–µ–Љ –і–∞–љ–љ—Л–µ –Љ–∞—В–µ—А–Є–∞–ї–Њ–≤ –і–ї—П –њ—А–Њ–≤–µ—А–Ї–Є –Њ—Б—В–∞—В–Ї–∞ –љ–∞ —Б–Ї–ї–∞–і–µ
                materials_df = load_data("Materials")

                material_parts = []
                stock_parts = []
                remaining_count_list = []

                for _, reserve in detail_reserves.iterrows():
                    material_id = reserve["ID –Љ–∞—В–µ—А–Є–∞–ї–∞"]
                    marka = reserve["–Ь–∞—А–Ї–∞"]
                    thickness = reserve["–Ґ–Њ–ї—Й–Є–љ–∞"]
                    width = reserve["–®–Є—А–Є–љ–∞"]
                    length = reserve["–Ф–ї–Є–љ–∞"]
                    remaining_qty = int(reserve["–Ю—Б—В–∞—В–Њ–Ї –Ї —Б–њ–Є—Б–∞–љ–Є—О"])

                    # –Ю–њ–Є—Б–∞–љ–Є–µ –Љ–∞—В–µ—А–Є–∞–ї–∞ (–±–µ–Ј –Ї–Њ–ї–Є—З–µ—Б—В–≤–∞)
                    material_desc = f"{marka} {thickness}–Љ–Љ {width}x{length}"

                    # –Ф–Њ–±–∞–≤–ї—П–µ–Љ —В–Њ–ї—М–Ї–Њ –Ї–Њ–ї–Є—З–µ—Б—В–≤–Њ –ї–Є—Б—В–Њ–≤ –Ї —Б–њ–Є—Б–∞–љ–Є—О
                    if remaining_qty > 0:
                        remaining_count_list.append(str(remaining_qty))

                    # –Ш—Й–µ–Љ –Ю–С–©–Ш–Щ —Д–∞–Ї—В–Є—З–µ—Б–Ї–Є–є –Њ—Б—В–∞—В–Њ–Ї –љ–∞ —Б–Ї–ї–∞–і–µ (–Ї–Њ–ї–Њ–љ–Ї–∞ "–Ъ–Њ–ї–Є—З–µ—Б—В–≤–Њ —И—В—Г–Ї")
                    if material_id != -1 and not materials_df.empty:
                        material_row = materials_df[materials_df["ID"] == material_id]
                        if not material_row.empty:
                            total_quantity = int(material_row.iloc[0]["–Ъ–Њ–ї–Є—З–µ—Б—В–≤–Њ —И—В—Г–Ї"])
                            material_parts.append(material_desc)
                            stock_parts.append(str(total_quantity))

                material_info = "; ".join(material_parts) if material_parts else ""
                material_stock = "; ".join(stock_parts) if stock_parts else ""
                remaining_reserved_count = "; ".join(remaining_count_list) if remaining_count_list else ""

            # –§–Њ—А–Љ–Є—А—Г–µ–Љ —В–µ–Ї—Б—В –і–ї—П –Ї–Њ–њ–Є—А–Њ–≤–∞–љ–Є—П
            parts = [
                f"{customer} | {order_name}",
                f"{detail_name}",
                f"–Ю—Б—В–∞–ї–Њ—Б—М: {remaining_to_cut} —И—В"
            ]

            # –Ф–Њ–±–∞–≤–ї—П–µ–Љ –Љ–∞—В–µ—А–Є–∞–ї (–µ—Б–ї–Є –µ—Б—В—М)
            if material_info:
                parts.append(f"–Ь–∞—В–µ—А–Є–∞–ї: {material_info}")
            else:
                parts.append("–Ь–∞—В–µ—А–Є–∞–ї: ")

            # –Ф–Њ–±–∞–≤–ї—П–µ–Љ –Њ—Б—В–∞—В–Њ–Ї –љ–∞ —Б–Ї–ї–∞–і–µ (–µ—Б–ї–Є –µ—Б—В—М)
            if material_stock:
                parts.append(f"–Ю—Б—В–∞—В–Њ–Ї –љ–∞ —Б–Ї–ї–∞–і–µ: {material_stock} —И—В")
            else:
                parts.append("–Ю—Б—В–∞—В–Њ–Ї –љ–∞ —Б–Ї–ї–∞–і–µ: ")

            # –Ф–Њ–±–∞–≤–ї—П–µ–Љ –Ї–Њ–ї–Є—З–µ—Б—В–≤–Њ –Ї —Б–њ–Є—Б–∞–љ–Є—О (–µ—Б–ї–Є –µ—Б—В—М)
            if remaining_reserved_count:
                parts.append(f"–Ю—Б—В–∞—В–Њ–Ї –њ–Њ—А–µ–Ј–∞—В—М: {remaining_reserved_count} —И—В")
            else:
                parts.append("–Ю—Б—В–∞—В–Њ–Ї –њ–Њ—А–µ–Ј–∞—В—М: ")

            copy_text = " | ".join(parts)

            # –Ъ–Њ–њ–Є—А—Г–µ–Љ –≤ –±—Г—Д–µ—А –Њ–±–Љ–µ–љ–∞
            self.root.clipboard_clear()
            self.root.clipboard_append(copy_text)
            self.root.update()  # –Ю–±–љ–Њ–≤–ї—П–µ–Љ –±—Г—Д–µ—А –Њ–±–Љ–µ–љ–∞

            # –Я–Њ–Ї–∞–Ј—Л–≤–∞–µ–Љ —Г–≤–µ–і–Њ–Љ–ї–µ–љ–Є–µ
            messagebox.showinfo(
                "–°–Ї–Њ–њ–Є—А–Њ–≤–∞–љ–Њ",
                f"–Ш–љ—Д–Њ—А–Љ–∞—Ж–Є—П –Њ –і–µ—В–∞–ї–Є —Б–Ї–Њ–њ–Є—А–Њ–≤–∞–љ–∞ –≤ –±—Г—Д–µ—А –Њ–±–Љ–µ–љ–∞:\n\n{copy_text}"
            )

        except Exception as e:
            messagebox.showerror("–Ю—И–Є–±–Ї–∞", f"–Э–µ —Г–і–∞–ї–Њ—Б—М —Б–Ї–Њ–њ–Є—А–Њ–≤–∞—В—М –Є–љ—Д–Њ—А–Љ–∞—Ж–Є—О: {e}")
            import traceback
            traceback.print_exc()

    def on_details_tab_right_click(self, event):
        """–Ъ–Њ–љ—В–µ–Ї—Б—В–љ–Њ–µ –Љ–µ–љ—О –њ—А–Є –њ—А–∞–≤–Њ–Љ –Ї–ї–Є–Ї–µ –љ–∞ –і–µ—В–∞–ї—М –≤–Њ –≤–Ї–ї–∞–і–Ї–µ '–£—З—С—В –і–µ—В–∞–ї–µ–є'"""
        # –Ю–њ—А–µ–і–µ–ї—П–µ–Љ —Б—В—А–Њ–Ї—Г –њ–Њ–і –Ї—Г—А—Б–Њ—А–Њ–Љ
        item = self.details_tree.identify_row(event.y)
        if not item:
            return

        # –Т—Л–і–µ–ї—П–µ–Љ —Б—В—А–Њ–Ї—Г
        self.details_tree.selection_set(item)

        # –°–Њ–Ј–і–∞—С–Љ –Ї–Њ–љ—В–µ–Ї—Б—В–љ–Њ–µ –Љ–µ–љ—О
        context_menu = tk.Menu(self.root, tearoff=0)
        context_menu.add_command(
            label="рЯУЛ –Ъ–Њ–њ–Є—А–Њ–≤–∞—В—М –Є–љ—Д–Њ—А–Љ–∞—Ж–Є—О –Њ –і–µ—В–∞–ї–Є",
            command=lambda: self.copy_details_tab_info(item)
        )

        # –Я–Њ–Ї–∞–Ј—Л–≤–∞–µ–Љ –Љ–µ–љ—О
        try:
            context_menu.tk_popup(event.x_root, event.y_root)
        finally:
            context_menu.grab_release()

    def copy_details_tab_info(self, item_id):
        """–Ъ–Њ–њ–Є—А–Њ–≤–∞–љ–Є–µ –Є–љ—Д–Њ—А–Љ–∞—Ж–Є–Є –Њ –і–µ—В–∞–ї–Є –Є–Ј –≤–Ї–ї–∞–і–Ї–Є '–£—З—С—В –і–µ—В–∞–ї–µ–є' –≤ –±—Г—Д–µ—А –Њ–±–Љ–µ–љ–∞"""
        try:
            # –Я–Њ–ї—Г—З–∞–µ–Љ –і–∞–љ–љ—Л–µ –Є–Ј —Б—В—А–Њ–Ї–Є —В–∞–±–ї–Є—Ж—Л
            values = self.details_tree.item(item_id)["values"]
            if not values or len(values) < 9:
                messagebox.showwarning("–Я—А–µ–і—Г–њ—А–µ–ґ–і–µ–љ–Є–µ", "–Э–µ —Г–і–∞–ї–Њ—Б—М –њ–Њ–ї—Г—З–Є—В—М –і–∞–љ–љ—Л–µ –і–µ—В–∞–ї–Є")
                return

            detail_id = values[0]
            customer = values[1]
            detail_name = values[2]
            order_name = values[3]
            quantity = int(values[4])
            cut = int(values[5])
            bent = int(values[6])
            remaining = int(values[7])

            # –Ч–∞–≥—А—Г–ґ–∞–µ–Љ –і–∞–љ–љ—Л–µ —А–µ–Ј–µ—А–≤–Є—А–Њ–≤–∞–љ–Є—П
            reservations_df = load_data("Reservations")

            # –Ш—Й–µ–Љ —А–µ–Ј–µ—А–≤—Л –і–ї—П —Н—В–Њ–є –і–µ—В–∞–ї–Є
            detail_reserves = reservations_df[
                reservations_df["ID –і–µ—В–∞–ї–Є"] == detail_id
                ]

            # –Я–µ—А–µ–Љ–µ–љ–љ—Л–µ –і–ї—П –≤—Л–≤–Њ–і–∞
            material_info = ""
            material_stock = ""
            remaining_reserved_count = ""

            if not detail_reserves.empty:
                # –Ч–∞–≥—А—Г–ґ–∞–µ–Љ –і–∞–љ–љ—Л–µ –Љ–∞—В–µ—А–Є–∞–ї–Њ–≤ –і–ї—П –њ—А–Њ–≤–µ—А–Ї–Є –Њ—Б—В–∞—В–Ї–∞ –љ–∞ —Б–Ї–ї–∞–і–µ
                materials_df = load_data("Materials")

                material_parts = []
                stock_parts = []
                remaining_count_list = []

                for _, reserve in detail_reserves.iterrows():
                    material_id = reserve["ID –Љ–∞—В–µ—А–Є–∞–ї–∞"]
                    marka = reserve["–Ь–∞—А–Ї–∞"]
                    thickness = reserve["–Ґ–Њ–ї—Й–Є–љ–∞"]
                    width = reserve["–®–Є—А–Є–љ–∞"]
                    length = reserve["–Ф–ї–Є–љ–∞"]
                    remaining_qty = int(reserve["–Ю—Б—В–∞—В–Њ–Ї –Ї —Б–њ–Є—Б–∞–љ–Є—О"])

                    # –Ю–њ–Є—Б–∞–љ–Є–µ –Љ–∞—В–µ—А–Є–∞–ї–∞ (–±–µ–Ј –Ї–Њ–ї–Є—З–µ—Б—В–≤–∞)
                    material_desc = f"{marka} {thickness}–Љ–Љ {width}x{length}"

                    # –Ф–Њ–±–∞–≤–ї—П–µ–Љ —В–Њ–ї—М–Ї–Њ –Ї–Њ–ї–Є—З–µ—Б—В–≤–Њ –ї–Є—Б—В–Њ–≤ –Ї —Б–њ–Є—Б–∞–љ–Є—О
                    if remaining_qty > 0:
                        remaining_count_list.append(str(remaining_qty))

                    # –Ш—Й–µ–Љ –Ю–С–©–Ш–Щ —Д–∞–Ї—В–Є—З–µ—Б–Ї–Є–є –Њ—Б—В–∞—В–Њ–Ї –љ–∞ —Б–Ї–ї–∞–і–µ (–Ї–Њ–ї–Њ–љ–Ї–∞ "–Ъ–Њ–ї–Є—З–µ—Б—В–≤–Њ —И—В—Г–Ї")
                    if material_id != -1 and not materials_df.empty:
                        material_row = materials_df[materials_df["ID"] == material_id]
                        if not material_row.empty:
                            total_quantity = int(material_row.iloc[0]["–Ъ–Њ–ї–Є—З–µ—Б—В–≤–Њ —И—В—Г–Ї"])
                            material_parts.append(material_desc)
                            stock_parts.append(str(total_quantity))

                material_info = "; ".join(material_parts) if material_parts else ""
                material_stock = "; ".join(stock_parts) if stock_parts else ""
                remaining_reserved_count = "; ".join(remaining_count_list) if remaining_count_list else ""

            # –§–Њ—А–Љ–Є—А—Г–µ–Љ —В–µ–Ї—Б—В –і–ї—П –Ї–Њ–њ–Є—А–Њ–≤–∞–љ–Є—П
            parts = [
                f"{customer} | {order_name}",
                f"{detail_name}",
                f"–Ю—Б—В–∞–ї–Њ—Б—М: {remaining} —И—В"
            ]

            # –Ф–Њ–±–∞–≤–ї—П–µ–Љ –Љ–∞—В–µ—А–Є–∞–ї (–µ—Б–ї–Є –µ—Б—В—М)
            if material_info:
                parts.append(f"–Ь–∞—В–µ—А–Є–∞–ї: {material_info}")
            else:
                parts.append("–Ь–∞—В–µ—А–Є–∞–ї: ")

            # –Ф–Њ–±–∞–≤–ї—П–µ–Љ –Њ—Б—В–∞—В–Њ–Ї –љ–∞ —Б–Ї–ї–∞–і–µ (–µ—Б–ї–Є –µ—Б—В—М)
            if material_stock:
                parts.append(f"–Ю—Б—В–∞—В–Њ–Ї –љ–∞ —Б–Ї–ї–∞–і–µ: {material_stock} —И—В")
            else:
                parts.append("–Ю—Б—В–∞—В–Њ–Ї –љ–∞ —Б–Ї–ї–∞–і–µ: ")

            # –Ф–Њ–±–∞–≤–ї—П–µ–Љ –Ї–Њ–ї–Є—З–µ—Б—В–≤–Њ –Ї —Б–њ–Є—Б–∞–љ–Є—О (–µ—Б–ї–Є –µ—Б—В—М)
            if remaining_reserved_count:
                parts.append(f"–Ю—Б—В–∞—В–Њ–Ї –њ–Њ—А–µ–Ј–∞—В—М: {remaining_reserved_count} —И—В")
            else:
                parts.append("–Ю—Б—В–∞—В–Њ–Ї –њ–Њ—А–µ–Ј–∞—В—М: ")

            copy_text = " | ".join(parts)

            # –Ъ–Њ–њ–Є—А—Г–µ–Љ –≤ –±—Г—Д–µ—А –Њ–±–Љ–µ–љ–∞
            self.root.clipboard_clear()
            self.root.clipboard_append(copy_text)
            self.root.update()  # –Ю–±–љ–Њ–≤–ї—П–µ–Љ –±—Г—Д–µ—А –Њ–±–Љ–µ–љ–∞

            # –Я–Њ–Ї–∞–Ј—Л–≤–∞–µ–Љ —Г–≤–µ–і–Њ–Љ–ї–µ–љ–Є–µ
            messagebox.showinfo(
                "–°–Ї–Њ–њ–Є—А–Њ–≤–∞–љ–Њ",
                f"–Ш–љ—Д–Њ—А–Љ–∞—Ж–Є—П –Њ –і–µ—В–∞–ї–Є —Б–Ї–Њ–њ–Є—А–Њ–≤–∞–љ–∞ –≤ –±—Г—Д–µ—А –Њ–±–Љ–µ–љ–∞:\n\n{copy_text}"
            )

        except Exception as e:
            messagebox.showerror("–Ю—И–Є–±–Ї–∞", f"–Э–µ —Г–і–∞–ї–Њ—Б—М —Б–Ї–Њ–њ–Є—А–Њ–≤–∞—В—М –Є–љ—Д–Њ—А–Љ–∞—Ж–Є—О: {e}")
            import traceback
            traceback.print_exc()

    def show_status_tooltip(self, message):
        """–Я–Њ–Ї–∞–Ј—Л–≤–∞–µ—В –≤—А–µ–Љ–µ–љ–љ–Њ–µ –≤—Б–њ–ї—Л–≤–∞—О—Й–µ–µ –Њ–Ї–љ–Њ —Б–Њ —Б—В–∞—В—Г—Б–Њ–Љ"""
        try:
            tooltip = tk.Toplevel(self.root)
            tooltip.wm_overrideredirect(True)

            # –Я–Њ–Ј–Є—Ж–Є–Њ–љ–Є—А—Г–µ–Љ –Њ–Ї–љ–Њ —А—П–і–Њ–Љ —Б –Ї—Г—А—Б–Њ—А–Њ–Љ
            x = self.root.winfo_pointerx() + 10
            y = self.root.winfo_pointery() + 10
            tooltip.wm_geometry(f"+{x}+{y}")

            label = tk.Label(tooltip, text=message, background="#d4edda",
                             foreground="#155724", relief=tk.SOLID, borderwidth=1,
                             font=("Arial", 9), padx=10, pady=5, justify=tk.LEFT)
            label.pack()

            # –Р–≤—В–Њ–Љ–∞—В–Є—З–µ—Б–Ї–Є –Ј–∞–Ї—А—Л–≤–∞–µ–Љ —З–µ—А–µ–Ј 2 —Б–µ–Ї—Г–љ–і—Л
            tooltip.after(2000, tooltip.destroy)
        except Exception as e:
            print(f"–Ю—И–Є–±–Ї–∞ –≤ show_status_tooltip: {e}")

    def download_orders_template(self):
        file_path = filedialog.asksaveasfilename(title="–°–Њ—Е—А–∞–љ–Є—В—М —И–∞–±–ї–Њ–љ", defaultextension=".xlsx",
                                                 filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                                                 initialfile="template_orders.xlsx")
        if not file_path:
            return
        try:
            wb = Workbook()
            ws_orders = wb.active
            ws_orders.title = "–Ч–∞–Ї–∞–Ј—Л"
            headers_orders = ["–Э–∞–Ј–≤–∞–љ–Є–µ –Ј–∞–Ї–∞–Ј–∞", "–Ч–∞–Ї–∞–Ј—З–Є–Ї", "–°—В–∞—В—Г—Б", "–Я—А–Є–Љ–µ—З–∞–љ–Є—П"]
            ws_orders.append(headers_orders)
            examples_orders = [
                ["–Ч–∞–Ї–∞–Ј вДЦ1 - –Ь–µ—В–∞–ї–ї–Њ–Ї–Њ–љ—Б—В—А—Г–Ї—Ж–Є–Є", "–Ю–Ю–Ю –°—В—А–Њ–є—В–µ—Е", "–Э–Њ–≤—Л–є", "–°—А–Њ—З–љ—Л–є –Ј–∞–Ї–∞–Ј"],
                ["–Ч–∞–Ї–∞–Ј вДЦ2 - –Ы–µ—Б—В–љ–Є—Ж–∞", "–Ш–Я –Ш–≤–∞–љ–Њ–≤", "–Т —А–∞–±–Њ—В–µ", ""],
                ["–Ч–∞–Ї–∞–Ј вДЦ3 - –Ю–≥—А–∞–ґ–і–µ–љ–Є–µ", "–Ю–Ю–Ю –Ь–µ—В–њ—А–Њ–Љ", "–Э–Њ–≤—Л–є", "–Ґ—А–µ–±—Г–µ—В—Б—П –њ—А–µ–і–Њ–њ–ї–∞—В–∞"]
            ]
            for example in examples_orders:
                ws_orders.append(example)
            for col in ws_orders.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                ws_orders.column_dimensions[column].width = max_length + 2
            ws_details = wb.create_sheet("–Ф–µ—В–∞–ї–Є")
            headers_details = ["–Э–∞–Ј–≤–∞–љ–Є–µ –Ј–∞–Ї–∞–Ј–∞", "–Э–∞–Ј–≤–∞–љ–Є–µ –і–µ—В–∞–ї–Є", "–Ъ–Њ–ї–Є—З–µ—Б—В–≤–Њ"]
            ws_details.append(headers_details)
            examples_details = [
                ["–Ч–∞–Ї–∞–Ј вДЦ1 - –Ь–µ—В–∞–ї–ї–Њ–Ї–Њ–љ—Б—В—А—Г–Ї—Ж–Є–Є", "–С–∞–ї–Ї–∞ –і–≤—Г—В–∞–≤—А–Њ–≤–∞—П 20", 15],
                ["–Ч–∞–Ї–∞–Ј вДЦ1 - –Ь–µ—В–∞–ї–ї–Њ–Ї–Њ–љ—Б—В—А—Г–Ї—Ж–Є–Є", "–®–≤–µ–ї–ї–µ—А 16", 8],
                ["–Ч–∞–Ї–∞–Ј вДЦ2 - –Ы–µ—Б—В–љ–Є—Ж–∞", "–°—В—Г–њ–µ–љ—М 300x250", 12],
                ["–Ч–∞–Ї–∞–Ј вДЦ2 - –Ы–µ—Б—В–љ–Є—Ж–∞", "–Я–Њ—А—Г—З–µ–љ—М", 2],
                ["–Ч–∞–Ї–∞–Ј вДЦ3 - –Ю–≥—А–∞–ґ–і–µ–љ–Є–µ", "–°—В–Њ–є–Ї–∞ 50x50", 20]
            ]
            for example in examples_details:
                ws_details.append(example)
            for col in ws_details.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                ws_details.column_dimensions[column].width = max_length + 2
            wb.save(file_path)
            messagebox.showinfo("–£—Б–њ–µ—Е",
                                f"–®–∞–±–ї–Њ–љ —Б–Њ—Е—А–∞–љ–µ–љ –≤:\n{file_path}\n\nрЯУЛ –Ш–Э–°–Ґ–†–£–Ъ–¶–Ш–ѓ:\n\n–Ы–Є—Б—В '–Ч–∞–Ї–∞–Ј—Л':\nвАҐ –Э–∞–Ј–≤–∞–љ–Є–µ –Ј–∞–Ї–∞–Ј–∞ - —Г–љ–Є–Ї–∞–ї—М–љ–Њ–µ –Є–Љ—П\nвАҐ –Ч–∞–Ї–∞–Ј—З–Є–Ї - –Њ–±—П–Ј–∞—В–µ–ї—М–љ–Њ\nвАҐ –°—В–∞—В—Г—Б: –Э–Њ–≤—Л–є, –Т —А–∞–±–Њ—В–µ, –Ч–∞–≤–µ—А—И–µ–љ, –Ю—В–Љ–µ–љ–µ–љ\nвАҐ –Я—А–Є–Љ–µ—З–∞–љ–Є—П - –Њ–њ—Ж–Є–Њ–љ–∞–ї—М–љ–Њ\n\n–Ы–Є—Б—В '–Ф–µ—В–∞–ї–Є':\nвАҐ –Э–∞–Ј–≤–∞–љ–Є–µ –Ј–∞–Ї–∞–Ј–∞ - –і–Њ–ї–ґ–љ–Њ —Б–Њ–≤–њ–∞–і–∞—В—М —Б –ї–Є—Б—В–Њ–Љ '–Ч–∞–Ї–∞–Ј—Л'\nвАҐ –Э–∞–Ј–≤–∞–љ–Є–µ –і–µ—В–∞–ї–Є - –Њ–±—П–Ј–∞—В–µ–ї—М–љ–Њ\nвАҐ –Ъ–Њ–ї–Є—З–µ—Б—В–≤–Њ - —З–Є—Б–ї–Њ")
        except Exception as e:
            messagebox.showerror("–Ю—И–Є–±–Ї–∞", f"–Э–µ —Г–і–∞–ї–Њ—Б—М —Б–Њ–Ј–і–∞—В—М —И–∞–±–ї–Њ–љ: {e}")

    def import_orders(self):
        file_path = filedialog.askopenfilename(title="–Т—Л–±–µ—А–Є—В–µ —Д–∞–є–ї Excel —Б –Ј–∞–Ї–∞–Ј–∞–Љ–Є",
                                               filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")])
        if not file_path:
            return
        try:
            try:
                orders_import_df = pd.read_excel(file_path, sheet_name="–Ч–∞–Ї–∞–Ј—Л", engine='openpyxl')
            except:
                messagebox.showerror("–Ю—И–Є–±–Ї–∞", "–Т —Д–∞–є–ї–µ –Њ—В—Б—Г—В—Б—В–≤—Г–µ—В –ї–Є—Б—В '–Ч–∞–Ї–∞–Ј—Л'!\n\n–Ш—Б–њ–Њ–ї—М–Ј—Г–є—В–µ —И–∞–±–ї–Њ–љ.")
                return
            try:
                details_import_df = pd.read_excel(file_path, sheet_name="–Ф–µ—В–∞–ї–Є", engine='openpyxl')
                has_details = True
            except:
                details_import_df = pd.DataFrame()
                has_details = False
            required_columns_orders = ["–Э–∞–Ј–≤–∞–љ–Є–µ –Ј–∞–Ї–∞–Ј–∞", "–Ч–∞–Ї–∞–Ј—З–Є–Ї"]
            missing_columns = [col for col in required_columns_orders if col not in orders_import_df.columns]
            if missing_columns:
                messagebox.showerror("–Ю—И–Є–±–Ї–∞",
                                     f"–Т –ї–Є—Б—В–µ '–Ч–∞–Ї–∞–Ј—Л' –Њ—В—Б—Г—В—Б—В–≤—Г—О—В –Ї–Њ–ї–Њ–љ–Ї–Є:\n{', '.join(missing_columns)}\n\n–Ш—Б–њ–Њ–ї—М–Ј—Г–є—В–µ –Ї–љ–Њ–њ–Ї—Г '–°–Ї–∞—З–∞—В—М —И–∞–±–ї–Њ–љ'.")
                return
            if has_details and not details_import_df.empty:
                required_columns_details = ["–Э–∞–Ј–≤–∞–љ–Є–µ –Ј–∞–Ї–∞–Ј–∞", "–Э–∞–Ј–≤–∞–љ–Є–µ –і–µ—В–∞–ї–Є", "–Ъ–Њ–ї–Є—З–µ—Б—В–≤–Њ"]
                missing_details = [col for col in required_columns_details if col not in details_import_df.columns]
                if missing_details:
                    messagebox.showwarning("–Я—А–µ–і—Г–њ—А–µ–ґ–і–µ–љ–Є–µ",
                                           f"–Т –ї–Є—Б—В–µ '–Ф–µ—В–∞–ї–Є' –Њ—В—Б—Г—В—Б—В–≤—Г—О—В –Ї–Њ–ї–Њ–љ–Ї–Є:\n{', '.join(missing_details)}\n\n–Ф–µ—В–∞–ї–Є –љ–µ –±—Г–і—Г—В –Є–Љ–њ–Њ—А—В–Є—А–Њ–≤–∞–љ—Л.")
                    has_details = False
            orders_df = load_data("Orders")
            current_max_order_id = 1000 if orders_df.empty else int(orders_df["ID –Ј–∞–Ї–∞–Ј–∞"].max())
            order_details_df = load_data("OrderDetails")
            current_max_detail_id = 0 if order_details_df.empty else int(order_details_df["ID"].max())
            imported_orders = 0
            imported_details = 0
            errors = []
            valid_statuses = ["–Э–Њ–≤—Л–є", "–Т —А–∞–±–Њ—В–µ", "–Ч–∞–≤–µ—А—И–µ–љ", "–Ю—В–Љ–µ–љ–µ–љ"]
            order_name_to_id = {}
            for idx, row in orders_import_df.iterrows():
                try:
                    if pd.isna(row["–Э–∞–Ј–≤–∞–љ–Є–µ –Ј–∞–Ї–∞–Ј–∞"]) or str(row["–Э–∞–Ј–≤–∞–љ–Є–µ –Ј–∞–Ї–∞–Ј–∞"]).strip() == "":
                        continue
                    if pd.isna(row["–Ч–∞–Ї–∞–Ј—З–Є–Ї"]) or str(row["–Ч–∞–Ї–∞–Ј—З–Є–Ї"]).strip() == "":
                        errors.append(f"–Ч–∞–Ї–∞–Ј—Л, —Б—В—А–Њ–Ї–∞ {idx + 2}: –Ю—В—Б—Г—В—Б—В–≤—Г–µ—В –Ј–∞–Ї–∞–Ј—З–Є–Ї")
                        continue
                    order_name = str(row["–Э–∞–Ј–≤–∞–љ–Є–µ –Ј–∞–Ї–∞–Ј–∞"]).strip()
                    customer = str(row["–Ч–∞–Ї–∞–Ј—З–Є–Ї"]).strip()
                    status = "–Э–Њ–≤—Л–є"
                    if "–°—В–∞—В—Г—Б" in orders_import_df.columns and not pd.isna(row["–°—В–∞—В—Г—Б"]):
                        status_input = str(row["–°—В–∞—В—Г—Б"]).strip()
                        if status_input in valid_statuses:
                            status = status_input
                        else:
                            errors.append(
                                f"–Ч–∞–Ї–∞–Ј—Л, —Б—В—А–Њ–Ї–∞ {idx + 2}: –Э–µ–≤–µ—А–љ—Л–є —Б—В–∞—В—Г—Б '{status_input}', —Г—Б—В–∞–љ–Њ–≤–ї–µ–љ '–Э–Њ–≤—Л–є'")
                    notes = ""
                    if "–Я—А–Є–Љ–µ—З–∞–љ–Є—П" in orders_import_df.columns and not pd.isna(row["–Я—А–Є–Љ–µ—З–∞–љ–Є—П"]):
                        notes = str(row["–Я—А–Є–Љ–µ—З–∞–љ–Є—П"]).strip()
                    current_max_order_id += 1
                    new_order_id = current_max_order_id
                    new_row = pd.DataFrame([{
                        "ID –Ј–∞–Ї–∞–Ј–∞": new_order_id,
                        "–Э–∞–Ј–≤–∞–љ–Є–µ –Ј–∞–Ї–∞–Ј–∞": order_name,
                        "–Ч–∞–Ї–∞–Ј—З–Є–Ї": customer,
                        "–Ф–∞—В–∞ —Б–Њ–Ј–і–∞–љ–Є—П": datetime.now().strftime("%Y-%m-%d"),
                        "–°—В–∞—В—Г—Б": status,
                        "–Я—А–Є–Љ–µ—З–∞–љ–Є—П": notes
                    }])
                    orders_df = pd.concat([orders_df, new_row], ignore_index=True)
                    imported_orders += 1
                    order_name_to_id[order_name] = new_order_id
                except Exception as e:
                    errors.append(f"–Ч–∞–Ї–∞–Ј—Л, —Б—В—А–Њ–Ї–∞ {idx + 2}: {str(e)}")
            if has_details and not details_import_df.empty:
                for idx, row in details_import_df.iterrows():
                    try:
                        if pd.isna(row["–Э–∞–Ј–≤–∞–љ–Є–µ –Ј–∞–Ї–∞–Ј–∞"]) or str(row["–Э–∞–Ј–≤–∞–љ–Є–µ –Ј–∞–Ї–∞–Ј–∞"]).strip() == "":
                            continue
                        order_name = str(row["–Э–∞–Ј–≤–∞–љ–Є–µ –Ј–∞–Ї–∞–Ј–∞"]).strip()
                        if order_name not in order_name_to_id:
                            errors.append(f"–Ф–µ—В–∞–ї–Є, —Б—В—А–Њ–Ї–∞ {idx + 2}: –Ч–∞–Ї–∞–Ј '{order_name}' –љ–µ –љ–∞–є–і–µ–љ –≤ –ї–Є—Б—В–µ '–Ч–∞–Ї–∞–Ј—Л'")
                            continue
                        if pd.isna(row["–Э–∞–Ј–≤–∞–љ–Є–µ –і–µ—В–∞–ї–Є"]) or str(row["–Э–∞–Ј–≤–∞–љ–Є–µ –і–µ—В–∞–ї–Є"]).strip() == "":
                            errors.append(f"–Ф–µ—В–∞–ї–Є, —Б—В—А–Њ–Ї–∞ {idx + 2}: –Ю—В—Б—Г—В—Б—В–≤—Г–µ—В –љ–∞–Ј–≤–∞–љ–Є–µ –і–µ—В–∞–ї–Є")
                            continue
                        detail_name = str(row["–Э–∞–Ј–≤–∞–љ–Є–µ –і–µ—В–∞–ї–Є"]).strip()
                        if pd.isna(row["–Ъ–Њ–ї–Є—З–µ—Б—В–≤–Њ"]):
                            errors.append(
                                f"–Ф–µ—В–∞–ї–Є, —Б—В—А–Њ–Ї–∞ {idx + 2}: –Ю—В—Б—Г—В—Б—В–≤—Г–µ—В –Ї–Њ–ї–Є—З–µ—Б—В–≤–Њ –і–ї—П –і–µ—В–∞–ї–Є '{detail_name}'")
                            continue
                        try:
                            quantity = float(row["–Ъ–Њ–ї–Є—З–µ—Б—В–≤–Њ"])
                            quantity = int(quantity)
                            if quantity <= 0:
                                errors.append(
                                    f"–Ф–µ—В–∞–ї–Є, —Б—В—А–Њ–Ї–∞ {idx + 2}: –Ъ–Њ–ї–Є—З–µ—Б—В–≤–Њ –і–Њ–ї–ґ–љ–Њ –±—Л—В—М –±–Њ–ї—М—И–µ –љ—Г–ї—П –і–ї—П –і–µ—В–∞–ї–Є '{detail_name}'")
                                continue
                        except (ValueError, TypeError):
                            errors.append(
                                f"–Ф–µ—В–∞–ї–Є, —Б—В—А–Њ–Ї–∞ {idx + 2}: –Э–µ–≤–µ—А–љ–Њ–µ –Ї–Њ–ї–Є—З–µ—Б—В–≤–Њ '{row['–Ъ–Њ–ї–Є—З–µ—Б—В–≤–Њ']}' –і–ї—П –і–µ—В–∞–ї–Є '{detail_name}'")
                            continue
                        current_max_detail_id += 1
                        order_id = order_name_to_id[order_name]
                        new_detail = pd.DataFrame([{
                            "ID": current_max_detail_id,
                            "ID –Ј–∞–Ї–∞–Ј–∞": order_id,
                            "–Э–∞–Ј–≤–∞–љ–Є–µ –і–µ—В–∞–ї–Є": detail_name,
                            "–Ъ–Њ–ї–Є—З–µ—Б—В–≤–Њ": quantity
                        }])
                        order_details_df = pd.concat([order_details_df, new_detail], ignore_index=True)
                        imported_details += 1
                    except Exception as e:
                        errors.append(f"–Ф–µ—В–∞–ї–Є, —Б—В—А–Њ–Ї–∞ {idx + 2}: {str(e)}")
            save_data("Orders", orders_df)
            if imported_details > 0:
                save_data("OrderDetails", order_details_df)
            self.refresh_orders()
            result_msg = f"вЬЕ –£—Б–њ–µ—И–љ–Њ –Є–Љ–њ–Њ—А—В–Є—А–Њ–≤–∞–љ–Њ:\nвАҐ –Ч–∞–Ї–∞–Ј–Њ–≤: {imported_orders}\nвАҐ –Ф–µ—В–∞–ї–µ–є: {imported_details}"
            if errors:
                result_msg += f"\n\nвЪ† –Ю—И–Є–±–Ї–Є ({len(errors)}):\n" + "\n".join(errors[:15])
                if len(errors) > 15:
                    result_msg += f"\n... –Є –µ—Й–µ {len(errors) - 15} –Њ—И–Є–±–Њ–Ї"
            messagebox.showinfo("–†–µ–Ј—Г–ї—М—В–∞—В –Є–Љ–њ–Њ—А—В–∞", result_msg)
        except Exception as e:
            messagebox.showerror("–Ю—И–Є–±–Ї–∞", f"–Э–µ —Г–і–∞–ї–Њ—Б—М –Є–Љ–њ–Њ—А—В–Є—А–Њ–≤–∞—В—М –і–∞–љ–љ—Л–µ:\n{e}")

    def add_order(self):
        add_window = tk.Toplevel(self.root)
        add_window.title("–Ф–Њ–±–∞–≤–Є—В—М –Ј–∞–Ї–∞–Ј")
        add_window.geometry("450x450")
        add_window.configure(bg='#ecf0f1')
        tk.Label(add_window, text="–°–Њ–Ј–і–∞–љ–Є–µ –љ–Њ–≤–Њ–≥–Њ –Ј–∞–Ї–∞–Ј–∞", font=("Arial", 12, "bold"), bg='#ecf0f1').pack(pady=10)
        fields = [("–Э–∞–Ј–≤–∞–љ–Є–µ –Ј–∞–Ї–∞–Ј–∞:", "name"), ("–Ч–∞–Ї–∞–Ј—З–Є–Ї:", "customer"), ("–Я—А–Є–Љ–µ—З–∞–љ–Є—П:", "notes")]
        entries = {}
        for label_text, key in fields:
            frame = tk.Frame(add_window, bg='#ecf0f1')
            frame.pack(fill=tk.X, padx=20, pady=5)
            tk.Label(frame, text=label_text, width=20, anchor='w', bg='#ecf0f1', font=("Arial", 10)).pack(side=tk.LEFT)
            entry = tk.Entry(frame, font=("Arial", 10))
            entry.pack(side=tk.RIGHT, expand=True, fill=tk.X)
            entries[key] = entry
        status_frame = tk.Frame(add_window, bg='#ecf0f1')
        status_frame.pack(fill=tk.X, padx=20, pady=5)
        tk.Label(status_frame, text="–°—В–∞—В—Г—Б:", width=20, anchor='w', bg='#ecf0f1', font=("Arial", 10)).pack(
            side=tk.LEFT)
        status_var = tk.StringVar(value="–Э–Њ–≤—Л–є")
        status_combo = ttk.Combobox(status_frame, textvariable=status_var,
                                    values=["–Э–Њ–≤—Л–є", "–Т —А–∞–±–Њ—В–µ", "–Ч–∞–≤–µ—А—И–µ–љ", "–Ю—В–Љ–µ–љ–µ–љ"],
                                    font=("Arial", 10), state="readonly")
        status_combo.pack(side=tk.RIGHT, expand=True, fill=tk.X)

        def save_order():
            try:
                name = entries["name"].get().strip()
                customer = entries["customer"].get().strip()
                if not name or not customer:
                    messagebox.showwarning("–Я—А–µ–і—Г–њ—А–µ–ґ–і–µ–љ–Є–µ", "–Ч–∞–њ–Њ–ї–љ–Є—В–µ –љ–∞–Ј–≤–∞–љ–Є–µ –Є –Ј–∞–Ї–∞–Ј—З–Є–Ї–∞!")
                    return
                df = load_data("Orders")
                new_id = 1001 if df.empty else int(df["ID –Ј–∞–Ї–∞–Ј–∞"].max()) + 1
                new_row = pd.DataFrame([{"ID –Ј–∞–Ї–∞–Ј–∞": new_id, "–Э–∞–Ј–≤–∞–љ–Є–µ –Ј–∞–Ї–∞–Ј–∞": name, "–Ч–∞–Ї–∞–Ј—З–Є–Ї": customer,
                                         "–Ф–∞—В–∞ —Б–Њ–Ј–і–∞–љ–Є—П": datetime.now().strftime("%Y-%m-%d"),
                                         "–°—В–∞—В—Г—Б": status_var.get(), "–Я—А–Є–Љ–µ—З–∞–љ–Є—П": entries["notes"].get()}])
                df = pd.concat([df, new_row], ignore_index=True)
                save_data("Orders", df)
                self.refresh_orders()
                add_window.destroy()
                messagebox.showinfo("–£—Б–њ–µ—Е", f"–Ч–∞–Ї–∞–Ј #{new_id} —Г—Б–њ–µ—И–љ–Њ —Б–Њ–Ј–і–∞–љ!")
            except Exception as e:
                messagebox.showerror("–Ю—И–Є–±–Ї–∞", f"–Э–µ —Г–і–∞–ї–Њ—Б—М —Б–Њ–Ј–і–∞—В—М –Ј–∞–Ї–∞–Ј: {e}")

        tk.Button(add_window, text="–°–Њ–Ј–і–∞—В—М –Ј–∞–Ї–∞–Ј", bg='#27ae60', fg='white', font=("Arial", 12, "bold"),
                  command=save_order).pack(pady=20)

    def edit_order(self):
        selected = self.orders_tree.selection()
        if not selected:
            messagebox.showwarning("–Я—А–µ–і—Г–њ—А–µ–ґ–і–µ–љ–Є–µ", "–Т—Л–±–µ—А–Є—В–µ –Ј–∞–Ї–∞–Ј –і–ї—П —А–µ–і–∞–Ї—В–Є—А–Њ–≤–∞–љ–Є—П")
            return
        item_id = self.orders_tree.item(selected)["values"][0]
        df = load_data("Orders")
        row = df[df["ID –Ј–∞–Ї–∞–Ј–∞"] == item_id].iloc[0]
        edit_window = tk.Toplevel(self.root)
        edit_window.title("–†–µ–і–∞–Ї—В–Є—А–Њ–≤–∞—В—М –Ј–∞–Ї–∞–Ј")
        edit_window.geometry("450x450")
        edit_window.configure(bg='#ecf0f1')
        tk.Label(edit_window, text=f"–†–µ–і–∞–Ї—В–Є—А–Њ–≤–∞–љ–Є–µ –Ј–∞–Ї–∞–Ј–∞ #{item_id}", font=("Arial", 12, "bold"), bg='#ecf0f1').pack(
            pady=10)
        fields = [("–Э–∞–Ј–≤–∞–љ–Є–µ –Ј–∞–Ї–∞–Ј–∞:", "–Э–∞–Ј–≤–∞–љ–Є–µ –Ј–∞–Ї–∞–Ј–∞"), ("–Ч–∞–Ї–∞–Ј—З–Є–Ї:", "–Ч–∞–Ї–∞–Ј—З–Є–Ї"), ("–Я—А–Є–Љ–µ—З–∞–љ–Є—П:", "–Я—А–Є–Љ–µ—З–∞–љ–Є—П")]
        entries = {}
        for label_text, key in fields:
            frame = tk.Frame(edit_window, bg='#ecf0f1')
            frame.pack(fill=tk.X, padx=20, pady=5)
            tk.Label(frame, text=label_text, width=20, anchor='w', bg='#ecf0f1', font=("Arial", 10)).pack(side=tk.LEFT)
            entry = tk.Entry(frame, font=("Arial", 10))
            entry.insert(0, str(row[key]))
            entry.pack(side=tk.RIGHT, expand=True, fill=tk.X)
            entries[key] = entry
        status_frame = tk.Frame(edit_window, bg='#ecf0f1')
        status_frame.pack(fill=tk.X, padx=20, pady=5)
        tk.Label(status_frame, text="–°—В–∞—В—Г—Б:", width=20, anchor='w', bg='#ecf0f1', font=("Arial", 10)).pack(
            side=tk.LEFT)
        status_var = tk.StringVar(value=row["–°—В–∞—В—Г—Б"])
        status_combo = ttk.Combobox(status_frame, textvariable=status_var,
                                    values=["–Э–Њ–≤—Л–є", "–Т —А–∞–±–Њ—В–µ", "–Ч–∞–≤–µ—А—И–µ–љ", "–Ю—В–Љ–µ–љ–µ–љ"],
                                    font=("Arial", 10), state="readonly")
        status_combo.pack(side=tk.RIGHT, expand=True, fill=tk.X)

        def save_changes():
            try:
                df.loc[df["ID –Ј–∞–Ї–∞–Ј–∞"] == item_id, "–Э–∞–Ј–≤–∞–љ–Є–µ –Ј–∞–Ї–∞–Ј–∞"] = entries["–Э–∞–Ј–≤–∞–љ–Є–µ –Ј–∞–Ї–∞–Ј–∞"].get()
                df.loc[df["ID –Ј–∞–Ї–∞–Ј–∞"] == item_id, "–Ч–∞–Ї–∞–Ј—З–Є–Ї"] = entries["–Ч–∞–Ї–∞–Ј—З–Є–Ї"].get()
                df.loc[df["ID –Ј–∞–Ї–∞–Ј–∞"] == item_id, "–°—В–∞—В—Г—Б"] = status_var.get()
                df.loc[df["ID –Ј–∞–Ї–∞–Ј–∞"] == item_id, "–Я—А–Є–Љ–µ—З–∞–љ–Є—П"] = entries["–Я—А–Є–Љ–µ—З–∞–љ–Є—П"].get()
                save_data("Orders", df)
                self.refresh_orders()
                edit_window.destroy()
                messagebox.showinfo("–£—Б–њ–µ—Е", "–Ч–∞–Ї–∞–Ј —Г—Б–њ–µ—И–љ–Њ –Њ–±–љ–Њ–≤–ї–µ–љ!")
            except Exception as e:
                messagebox.showerror("–Ю—И–Є–±–Ї–∞", f"–Э–µ —Г–і–∞–ї–Њ—Б—М –Њ–±–љ–Њ–≤–Є—В—М –Ј–∞–Ї–∞–Ј: {e}")

        tk.Button(edit_window, text="–°–Њ—Е—А–∞–љ–Є—В—М", bg='#3498db', fg='white', font=("Arial", 12, "bold"),
                  command=save_changes).pack(pady=20)

    def delete_order(self):
        selected = self.orders_tree.selection()
        if not selected:
            messagebox.showwarning("–Я—А–µ–і—Г–њ—А–µ–ґ–і–µ–љ–Є–µ", "–Т—Л–±–µ—А–Є—В–µ –Ј–∞–Ї–∞–Ј—Л –і–ї—П —Г–і–∞–ї–µ–љ–Є—П")
            return
        count = len(selected)
        if messagebox.askyesno("–Я–Њ–і—В–≤–µ—А–ґ–і–µ–љ–Є–µ", f"–£–і–∞–ї–Є—В—М –≤—Л–±—А–∞–љ–љ—Л–µ –Ј–∞–Ї–∞–Ј—Л ({count} —И—В)?"):
            df = load_data("Orders")
            details_df = load_data("OrderDetails")
            for item in selected:
                item_id = self.orders_tree.item(item)["values"][0]
                df = df[df["ID –Ј–∞–Ї–∞–Ј–∞"] != item_id]
                if not details_df.empty:
                    details_df = details_df[details_df["ID –Ј–∞–Ї–∞–Ј–∞"] != item_id]
            save_data("Orders", df)
            if not details_df.empty or len(selected) > 0:
                save_data("OrderDetails", details_df)
            self.refresh_orders()
            self.refresh_order_details()
            messagebox.showinfo("–£—Б–њ–µ—Е", f"–£–і–∞–ї–µ–љ–Њ –Ј–∞–Ї–∞–Ј–Њ–≤: {count}")

    def add_order_detail(self):
        selected = self.orders_tree.selection()
        if not selected:
            messagebox.showwarning("–Я—А–µ–і—Г–њ—А–µ–ґ–і–µ–љ–Є–µ", "–°–љ–∞—З–∞–ї–∞ –≤—Л–±–µ—А–Є—В–µ –Ј–∞–Ї–∞–Ј!")
            return
        order_id = self.orders_tree.item(selected)["values"][0]
        add_window = tk.Toplevel(self.root)
        add_window.title("–Ф–Њ–±–∞–≤–Є—В—М –і–µ—В–∞–ї—М")
        add_window.geometry("400x300")
        add_window.configure(bg='#ecf0f1')
        tk.Label(add_window, text=f"–Ф–Њ–±–∞–≤–ї–µ–љ–Є–µ –і–µ—В–∞–ї–Є –Ї –Ј–∞–Ї–∞–Ј—Г #{order_id}", font=("Arial", 12, "bold"),
                 bg='#ecf0f1').pack(pady=10)
        name_frame = tk.Frame(add_window, bg='#ecf0f1')
        name_frame.pack(fill=tk.X, padx=20, pady=5)
        tk.Label(name_frame, text="–Э–∞–Ј–≤–∞–љ–Є–µ –і–µ—В–∞–ї–Є:", width=20, anchor='w', bg='#ecf0f1', font=("Arial", 10)).pack(
            side=tk.LEFT)
        name_entry = tk.Entry(name_frame, font=("Arial", 10))
        name_entry.pack(side=tk.RIGHT, expand=True, fill=tk.X)
        qty_frame = tk.Frame(add_window, bg='#ecf0f1')
        qty_frame.pack(fill=tk.X, padx=20, pady=5)
        tk.Label(qty_frame, text="–Ъ–Њ–ї–Є—З–µ—Б—В–≤–Њ:", width=20, anchor='w', bg='#ecf0f1', font=("Arial", 10)).pack(
            side=tk.LEFT)
        qty_entry = tk.Entry(qty_frame, font=("Arial", 10))
        qty_entry.pack(side=tk.RIGHT, expand=True, fill=tk.X)

        def save_detail():
            try:
                detail_name = name_entry.get().strip()
                quantity = int(qty_entry.get().strip())
                if not detail_name:
                    messagebox.showwarning("–Я—А–µ–і—Г–њ—А–µ–ґ–і–µ–љ–Є–µ", "–Т–≤–µ–і–Є—В–µ –љ–∞–Ј–≤–∞–љ–Є–µ –і–µ—В–∞–ї–Є!")
                    return
                df = load_data("OrderDetails")
                new_id = 1 if df.empty else int(df["ID"].max()) + 1
                new_row = pd.DataFrame(
                    [{"ID": new_id, "ID –Ј–∞–Ї–∞–Ј–∞": order_id, "–Э–∞–Ј–≤–∞–љ–Є–µ –і–µ—В–∞–ї–Є": detail_name,
                      "–Ъ–Њ–ї–Є—З–µ—Б—В–≤–Њ": quantity, "–Я–Њ—А–µ–Ј–∞–љ–Њ": 0, "–Я–Њ–≥–љ—Г—В–Њ": 0}])
                df = pd.concat([df, new_row], ignore_index=True)
                save_data("OrderDetails", df)
                self.refresh_order_details()
                add_window.destroy()
                messagebox.showinfo("–£—Б–њ–µ—Е", "–Ф–µ—В–∞–ї—М –і–Њ–±–∞–≤–ї–µ–љ–∞!")
            except ValueError:
                messagebox.showerror("–Ю—И–Є–±–Ї–∞", "–Ъ–Њ–ї–Є—З–µ—Б—В–≤–Њ –і–Њ–ї–ґ–љ–Њ –±—Л—В—М —З–Є—Б–ї–Њ–Љ!")
            except Exception as e:
                messagebox.showerror("–Ю—И–Є–±–Ї–∞", f"–Э–µ —Г–і–∞–ї–Њ—Б—М –і–Њ–±–∞–≤–Є—В—М –і–µ—В–∞–ї—М: {e}")

        tk.Button(add_window, text="–Ф–Њ–±–∞–≤–Є—В—М", bg='#27ae60', fg='white', font=("Arial", 12, "bold"),
                  command=save_detail).pack(pady=20)

    def delete_order_detail(self):
        selected = self.order_details_tree.selection()
        if not selected:
            messagebox.showwarning("–Я—А–µ–і—Г–њ—А–µ–ґ–і–µ–љ–Є–µ", "–Т—Л–±–µ—А–Є—В–µ –і–µ—В–∞–ї–Є –і–ї—П —Г–і–∞–ї–µ–љ–Є—П")
            return
        count = len(selected)
        if messagebox.askyesno("–Я–Њ–і—В–≤–µ—А–ґ–і–µ–љ–Є–µ", f"–£–і–∞–ї–Є—В—М –≤—Л–±—А–∞–љ–љ—Л–µ –і–µ—В–∞–ї–Є ({count} —И—В)?"):
            df = load_data("OrderDetails")
            for item in selected:
                detail_id = self.order_details_tree.item(item)["values"][0]
                df = df[df["ID"] != detail_id]
            save_data("OrderDetails", df)
            self.refresh_order_details()
            messagebox.showinfo("–£—Б–њ–µ—Е", f"–£–і–∞–ї–µ–љ–Њ –і–µ—В–∞–ї–µ–є: {count}")

    def edit_order_detail(self):
        """–†–µ–і–∞–Ї—В–Є—А–Њ–≤–∞–љ–Є–µ –і–µ—В–∞–ї–Є –Ј–∞–Ї–∞–Ј–∞ —Б —Г—З–µ—В–Њ–Љ —Н—В–∞–њ–Њ–≤ –њ—А–Њ–Є–Ј–≤–Њ–і—Б—В–≤–∞"""
        selected = self.order_details_tree.selection()
        if not selected:
            messagebox.showwarning("–Я—А–µ–і—Г–њ—А–µ–ґ–і–µ–љ–Є–µ", "–Т—Л–±–µ—А–Є—В–µ –і–µ—В–∞–ї—М –і–ї—П —А–µ–і–∞–Ї—В–Є—А–Њ–≤–∞–љ–Є—П")
            return

        detail_id = self.order_details_tree.item(selected)["values"][0]
        df = load_data("OrderDetails")
        row = df[df["ID"] == detail_id].iloc[0]

        edit_window = tk.Toplevel(self.root)
        edit_window.title("–†–µ–і–∞–Ї—В–Є—А–Њ–≤–∞—В—М –і–µ—В–∞–ї—М")
        edit_window.geometry("450x550")
        edit_window.configure(bg='#ecf0f1')

        tk.Label(edit_window, text=f"–†–µ–і–∞–Ї—В–Є—А–Њ–≤–∞–љ–Є–µ –і–µ—В–∞–ї–Є #{detail_id}",
                 font=("Arial", 12, "bold"), bg='#ecf0f1').pack(pady=10)

        # –Э–∞–Ј–≤–∞–љ–Є–µ –і–µ—В–∞–ї–Є
        name_frame = tk.Frame(edit_window, bg='#ecf0f1')
        name_frame.pack(fill=tk.X, padx=20, pady=5)
        tk.Label(name_frame, text="–Э–∞–Ј–≤–∞–љ–Є–µ –і–µ—В–∞–ї–Є:", width=20, anchor='w',
                 bg='#ecf0f1', font=("Arial", 10)).pack(side=tk.LEFT)
        name_entry = tk.Entry(name_frame, font=("Arial", 10))
        name_entry.insert(0, str(row["–Э–∞–Ј–≤–∞–љ–Є–µ –і–µ—В–∞–ї–Є"]))
        name_entry.pack(side=tk.RIGHT, expand=True, fill=tk.X)

        # –Ю–±—Й–µ–µ –Ї–Њ–ї–Є—З–µ—Б—В–≤–Њ
        qty_frame = tk.Frame(edit_window, bg='#ecf0f1')
        qty_frame.pack(fill=tk.X, padx=20, pady=5)
        tk.Label(qty_frame, text="рЯУЛ –Ю–±—Й–µ–µ –Ї–Њ–ї–Є—З–µ—Б—В–≤–Њ:", width=20, anchor='w',
                 bg='#ecf0f1', font=("Arial", 10, "bold")).pack(side=tk.LEFT)
        qty_entry = tk.Entry(qty_frame, font=("Arial", 10))
        qty_entry.insert(0, str(int(row["–Ъ–Њ–ї–Є—З–µ—Б—В–≤–Њ"])))
        qty_entry.pack(side=tk.RIGHT, expand=True, fill=tk.X)

        # –†–∞–Ј–і–µ–ї–Є—В–µ–ї—М –і–ї—П —Н—В–∞–њ–Њ–≤ –њ—А–Њ–Є–Ј–≤–Њ–і—Б—В–≤–∞
        tk.Label(edit_window, text="вФБ" * 50, bg='#ecf0f1', fg='#95a5a6').pack(pady=10)
        tk.Label(edit_window, text="–≠—В–∞–њ—Л –њ—А–Њ–Є–Ј–≤–Њ–і—Б—В–≤–∞", font=("Arial", 11, "bold"),
                 bg='#ecf0f1', fg='#2980b9').pack(pady=5)

        # –Я–Њ—А–µ–Ј–∞–љ–Њ (—Н—В–∞–њ 1)
        cut_frame = tk.Frame(edit_window, bg='#ecf0f1')
        cut_frame.pack(fill=tk.X, padx=20, pady=5)
        tk.Label(cut_frame, text="вЬВпЄП –Я–Њ—А–µ–Ј–∞–љ–Њ:", width=20, anchor='w',
                 bg='#ecf0f1', font=("Arial", 10, "bold"), fg='#27ae60').pack(side=tk.LEFT)
        cut_entry = tk.Entry(cut_frame, font=("Arial", 10))
        cut_value = row.get("–Я–Њ—А–µ–Ј–∞–љ–Њ", 0) if "–Я–Њ—А–µ–Ј–∞–љ–Њ" in row and pd.notna(row["–Я–Њ—А–µ–Ј–∞–љ–Њ"]) else 0
        cut_entry.insert(0, str(int(cut_value)))
        cut_entry.pack(side=tk.RIGHT, expand=True, fill=tk.X)

        # –Я–Њ–≥–љ—Г—В–Њ (—Н—В–∞–њ 2)
        bent_frame = tk.Frame(edit_window, bg='#ecf0f1')
        bent_frame.pack(fill=tk.X, padx=20, pady=5)
        tk.Label(bent_frame, text="рЯФІ –Я–Њ–≥–љ—Г—В–Њ:", width=20, anchor='w',
                 bg='#ecf0f1', font=("Arial", 10, "bold"), fg='#f39c12').pack(side=tk.LEFT)
        bent_entry = tk.Entry(bent_frame, font=("Arial", 10))
        bent_value = row.get("–Я–Њ–≥–љ—Г—В–Њ", 0) if "–Я–Њ–≥–љ—Г—В–Њ" in row and pd.notna(row["–Я–Њ–≥–љ—Г—В–Њ"]) else 0
        bent_entry.insert(0, str(int(bent_value)))
        bent_entry.pack(side=tk.RIGHT, expand=True, fill=tk.X)

        # –Ш–љ—Д–Њ—А–Љ–∞—Ж–Є—П
        info_frame = tk.Frame(edit_window, bg='#d1ecf1', relief=tk.RIDGE, borderwidth=2)
        info_frame.pack(fill=tk.X, padx=20, pady=10)
        tk.Label(info_frame, text="вДєпЄП –Ш–љ—Д–Њ—А–Љ–∞—Ж–Є—П –Њ –њ—А–Њ–Є–Ј–≤–Њ–і—Б—В–≤–µ:", font=("Arial", 9, "bold"),
                 bg='#d1ecf1', fg='#0c5460').pack(anchor='w', padx=5, pady=2)
        tk.Label(info_frame, text="вАҐ –Ю–±—Й–µ–µ –Ї–Њ–ї–Є—З–µ—Б—В–≤–Њ - –≤—Б–µ–≥–Њ –і–µ—В–∞–ї–µ–є –≤ –Ј–∞–Ї–∞–Ј–µ",
                 font=("Arial", 8), bg='#d1ecf1', fg='#0c5460').pack(anchor='w', padx=10)
        tk.Label(info_frame, text="вАҐ –Я–Њ—А–µ–Ј–∞–љ–Њ - –Ї–Њ–ї–Є—З–µ—Б—В–≤–Њ –Ј–∞–≥–Њ—В–Њ–≤–Њ–Ї –њ–Њ—Б–ї–µ —А–µ–Ј–Ї–Є –Љ–µ—В–∞–ї–ї–∞",
                 font=("Arial", 8), bg='#d1ecf1', fg='#0c5460').pack(anchor='w', padx=10)
        tk.Label(info_frame, text="вАҐ –Я–Њ–≥–љ—Г—В–Њ - –Ї–Њ–ї–Є—З–µ—Б—В–≤–Њ –і–µ—В–∞–ї–µ–є –њ–Њ—Б–ї–µ –≥–Є–±–Ї–Є",
                 font=("Arial", 8), bg='#d1ecf1', fg='#0c5460').pack(anchor='w', padx=10)
        tk.Label(info_frame, text="вАҐ –Ъ–Њ—А—А–µ–Ї—В–Є—А–Њ–≤–Ї–∞ –Ј–љ–∞—З–µ–љ–Є–є –њ—А–Њ–Є–Ј–≤–Њ–і–Є—В—Б—П –≤—А—Г—З–љ—Г—О",
                 font=("Arial", 8), bg='#d1ecf1', fg='#0c5460').pack(anchor='w', padx=10)

        def save_changes():
            try:
                new_name = name_entry.get().strip()
                new_qty = int(qty_entry.get().strip())
                new_cut = int(cut_entry.get().strip())
                new_bent = int(bent_entry.get().strip())

                if not new_name:
                    messagebox.showwarning("–Я—А–µ–і—Г–њ—А–µ–ґ–і–µ–љ–Є–µ", "–Т–≤–µ–і–Є—В–µ –љ–∞–Ј–≤–∞–љ–Є–µ –і–µ—В–∞–ї–Є!")
                    return

                if new_qty < 0 or new_cut < 0 or new_bent < 0:
                    messagebox.showerror("–Ю—И–Є–±–Ї–∞", "–Ч–љ–∞—З–µ–љ–Є—П –љ–µ –Љ–Њ–≥—Г—В –±—Л—В—М –Њ—В—А–Є—Ж–∞—В–µ–ї—М–љ—Л–Љ–Є!")
                    return

                if new_cut > new_qty:
                    if not messagebox.askyesno("–Я—А–µ–і—Г–њ—А–µ–ґ–і–µ–љ–Є–µ",
                                               f"–Я–Њ—А–µ–Ј–∞–љ–Њ ({new_cut}) –±–Њ–ї—М—И–µ –Њ–±—Й–µ–≥–Њ –Ї–Њ–ї–Є—З–µ—Б—В–≤–∞ ({new_qty}).\n"
                                               "–Т–Њ–Ј–Љ–Њ–ґ–љ–Њ, –µ—Б—В—М –Є–Ј–ї–Є—И–Ї–Є –Ј–∞–≥–Њ—В–Њ–≤–Њ–Ї.\n\n–Я—А–Њ–і–Њ–ї–ґ–Є—В—М?"):
                        return

                if new_bent > new_cut:
                    if not messagebox.askyesno("–Я—А–µ–і—Г–њ—А–µ–ґ–і–µ–љ–Є–µ",
                                               f"–Я–Њ–≥–љ—Г—В–Њ ({new_bent}) –±–Њ–ї—М—И–µ –њ–Њ—А–µ–Ј–∞–љ–љ—Л—Е ({new_cut}).\n"
                                               "–Я—А–Њ–≤–µ—А—М—В–µ –њ—А–∞–≤–Є–ї—М–љ–Њ—Б—В—М –і–∞–љ–љ—Л—Е.\n\n–Я—А–Њ–і–Њ–ї–ґ–Є—В—М?"):
                        return

                # –Ю–±–љ–Њ–≤–ї—П–µ–Љ –і–∞–љ–љ—Л–µ
                df.loc[df["ID"] == detail_id, "–Э–∞–Ј–≤–∞–љ–Є–µ –і–µ—В–∞–ї–Є"] = new_name
                df.loc[df["ID"] == detail_id, "–Ъ–Њ–ї–Є—З–µ—Б—В–≤–Њ"] = new_qty
                df.loc[df["ID"] == detail_id, "–Я–Њ—А–µ–Ј–∞–љ–Њ"] = new_cut
                df.loc[df["ID"] == detail_id, "–Я–Њ–≥–љ—Г—В–Њ"] = new_bent

                save_data("OrderDetails", df)
                self.refresh_order_details()
                edit_window.destroy()

                # –†–∞—Б—З–µ—В –Њ—Б—В–∞—В–Ї–Њ–≤
                to_cut = new_qty - new_cut
                to_bend = new_cut - new_bent

                messagebox.showinfo("–£—Б–њ–µ—Е",
                                    f"вЬЕ –Ф–µ—В–∞–ї—М –Њ–±–љ–Њ–≤–ї–µ–љ–∞!\n\n"
                                    f"рЯУЛ –Ю–±—Й–µ–µ –Ї–Њ–ї–Є—З–µ—Б—В–≤–Њ: {new_qty}\n"
                                    f"вЬВпЄП –Я–Њ—А–µ–Ј–∞–љ–Њ: {new_cut} (–Њ—Б—В–∞–ї–Њ—Б—М –њ–Њ—А–µ–Ј–∞—В—М: {to_cut})\n"
                                    f"рЯФІ –Я–Њ–≥–љ—Г—В–Њ: {new_bent} (–Њ—Б—В–∞–ї–Њ—Б—М –њ–Њ–≥–љ—Г—В—М: {to_bend})")

            except ValueError:
                messagebox.showerror("–Ю—И–Є–±–Ї–∞", "–Я—А–Њ–≤–µ—А—М—В–µ –њ—А–∞–≤–Є–ї—М–љ–Њ—Б—В—М –≤–≤–Њ–і–∞ —З–Є—Б–ї–Њ–≤—Л—Е –Ј–љ–∞—З–µ–љ–Є–є!")
            except Exception as e:
                messagebox.showerror("–Ю—И–Є–±–Ї–∞", f"–Э–µ —Г–і–∞–ї–Њ—Б—М –Њ–±–љ–Њ–≤–Є—В—М –і–µ—В–∞–ї—М: {e}")

        tk.Button(edit_window, text="рЯТЊ –°–Њ—Е—А–∞–љ–Є—В—М –Є–Ј–Љ–µ–љ–µ–љ–Є—П", bg='#3498db', fg='white',
                  font=("Arial", 12, "bold"), command=save_changes).pack(pady=15)

    def setup_reservations_tab(self):
        header = tk.Label(self.reservations_frame, text="–†–µ–Ј–µ—А–≤–Є—А–Њ–≤–∞–љ–Є–µ –Љ–∞—В–µ—А–Є–∞–ї–Њ–≤", font=("Arial", 16, "bold"),
                          bg='white', fg='#2c3e50')
        header.pack(pady=10)

        tree_frame = tk.Frame(self.reservations_frame, bg='white')
        tree_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)

        scroll_y = tk.Scrollbar(tree_frame, orient=tk.VERTICAL)
        scroll_x = tk.Scrollbar(tree_frame, orient=tk.HORIZONTAL)

        self.reservations_tree = ttk.Treeview(tree_frame,
                                              columns=("ID", "–Ч–∞–Ї–∞–Ј—З–Є–Ї | –Ч–∞–Ї–∞–Ј", "–Ф–µ—В–∞–ї—М", "–Ь–∞—В–µ—А–Є–∞–ї", "–Ь–∞—А–Ї–∞",
                                                       "–Ґ–Њ–ї—Й–Є–љ–∞", "–†–∞–Ј–Љ–µ—А", "–†–µ–Ј–µ—А–≤", "–°–њ–Є—Б–∞–љ–Њ", "–Ю—Б—В–∞—В–Њ–Ї", "–Ф–∞—В–∞"),
                                              show="headings", yscrollcommand=scroll_y.set, xscrollcommand=scroll_x.set)
        scroll_y.config(command=self.reservations_tree.yview)
        scroll_x.config(command=self.reservations_tree.xview)
        scroll_y.pack(side=tk.RIGHT, fill=tk.Y)
        scroll_x.pack(side=tk.BOTTOM, fill=tk.X)

        # –Э–∞—Б—В—А–Њ–є–Ї–∞ –Ї–Њ–ї–Њ–љ–Њ–Ї –С–Х–Ч —А–∞—Б—В—П–≥–Є–≤–∞–љ–Є—П
        for col in self.reservations_tree["columns"]:
            self.reservations_tree.heading(col, text=col)
            self.reservations_tree.column(col, anchor=tk.CENTER, width=100, minwidth=80, stretch=False)

        self.reservations_tree.pack(fill=tk.BOTH, expand=True)

        # рЯЖХ –Ш–Э–Ш–¶–Ш–Р–Ы–Ш–Ч–Р–¶–Ш–ѓ EXCEL-–§–Ш–Ы–ђ–Ґ–†–Р –Ф–Ы–ѓ –†–Х–Ч–Х–†–Т–Ш–†–Ю–Т–Р–Э–Ш–ѓ
        self.reservations_excel_filter = ExcelStyleFilter(
            tree=self.reservations_tree,
            refresh_callback=self.refresh_reservations
        )

        # рЯЖХ –Ш–Э–Ф–Ш–Ъ–Р–Ґ–Ю–† –Р–Ъ–Ґ–Ш–Т–Э–Ђ–• –§–Ш–Ы–ђ–Ґ–†–Ю–Т
        self.reservations_filter_status = tk.Label(
            self.reservations_frame,
            text="",
            font=("Arial", 9),
            bg='#d1ecf1',
            fg='#0c5460'
        )
        self.reservations_filter_status.pack(pady=5)

        # –Я–µ—А–µ–Ї–ї—О—З–∞—В–µ–ї–Є –≤–Є–і–Є–Љ–Њ—Б—В–Є
        self.reservations_toggles = self.create_visibility_toggles(
            self.reservations_frame,
            self.reservations_tree,
            {
                'show_fully_written_off': 'рЯУЭ –Я–Њ–Ї–∞–Ј–∞—В—М –њ–Њ–ї–љ–Њ—Б—В—М—О —Б–њ–Є—Б–∞–љ–љ—Л–µ'
            },
            self.refresh_reservations
        )

        # –Ъ–љ–Њ–њ–Ї–Є —Г–њ—А–∞–≤–ї–µ–љ–Є—П
        buttons_frame = tk.Frame(self.reservations_frame, bg='white')
        buttons_frame.pack(fill=tk.X, padx=10, pady=10)
        btn_style = {"font": ("Arial", 10), "width": 18, "height": 2}

        tk.Button(buttons_frame, text="–Ч–∞—А–µ–Ј–µ—А–≤–Є—А–Њ–≤–∞—В—М", bg='#27ae60', fg='white', command=self.add_reservation,
                  **btn_style).pack(side=tk.LEFT, padx=5)
        tk.Button(buttons_frame, text="–£–і–∞–ї–Є—В—М —А–µ–Ј–µ—А–≤", bg='#e74c3c', fg='white', command=self.delete_reservation,
                  **btn_style).pack(side=tk.LEFT, padx=5)
        tk.Button(buttons_frame, text="–†–µ–і–∞–Ї—В–Є—А–Њ–≤–∞—В—М", bg='#f39c12', fg='white', command=self.edit_reservation,
                  **btn_style).pack(side=tk.LEFT, padx=5)
        tk.Button(buttons_frame, text="–Ю–±–љ–Њ–≤–Є—В—М", bg='#95a5a6', fg='white', command=self.refresh_reservations,
                  **btn_style).pack(side=tk.LEFT, padx=5)
        tk.Button(buttons_frame, text="–Ч–∞–і–∞–љ–Є–µ –љ–∞ –ї–∞–Ј–µ—А", bg='#e67e22', fg='white', command=self.export_laser_task,
                  **btn_style).pack(side=tk.LEFT, padx=5)
        tk.Button(buttons_frame, text="вЬЦ –°–±—А–Њ—Б–Є—В—М —Д–Є–ї—М—В—А—Л", bg='#e67e22', fg='white',
                  command=self.clear_reservations_filters, **btn_style).pack(side=tk.LEFT, padx=5)

        self.refresh_reservations()

    def clear_reservations_filters(self):
        """–°–±—А–Њ—Б–Є—В—М –≤—Б–µ —Д–Є–ї—М—В—А—Л —А–µ–Ј–µ—А–≤–Є—А–Њ–≤–∞–љ–Є—П"""
        if hasattr(self, 'reservations_excel_filter'):
            self.reservations_excel_filter.clear_all_filters()

    def clear_writeoffs_filters(self):
        """–°–±—А–Њ—Б–Є—В—М –≤—Б–µ —Д–Є–ї—М—В—А—Л —Б–њ–Є—Б–∞–љ–Є–є"""
        if hasattr(self, 'writeoffs_excel_filter'):
            self.writeoffs_excel_filter.clear_all_filters()

    def clear_details_filters(self):
        """–°–±—А–Њпњљпњљ–Є—В—М –≤—Б–µ —Д–Є–ї—М—В—А—Л –і–µ—В–∞–ї–µ–є"""
        if hasattr(self, 'details_excel_filter'):
            self.details_excel_filter.clear_all_filters()

    def clear_laser_import_filters(self):
        """–°–±—А–Њ—Б–Є—В—М –≤—Б–µ —Д–Є–ї—М—В—А—Л –Є–Љ–њ–Њ—А—В–∞ –Њ—В –ї–∞–Ј–µ—А—Й–Є–Ї–Њ–≤"""
        if hasattr(self, 'laser_import_excel_filter'):
            self.laser_import_excel_filter.clear_all_filters()

    def refresh_reservations(self):
        """–Ю–±–љ–Њ–≤–ї–µ–љ–Є–µ —Б–њ–Є—Б–Ї–∞ —А–µ–Ј–µ—А–≤–Њ–≤"""

        # –°–Ю–•–†–Р–Э–ѓ–Х–Ь –Р–Ъ–Ґ–Ш–Т–Э–Ђ–Х –§–Ш–Ы–ђ–Ґ–†–Ђ –Я–Х–†–Х–Ф –Ю–І–Ш–°–Ґ–Ъ–Ю–Щ
        active_filters_backup = {}
        if hasattr(self, 'reservations_excel_filter') and self.reservations_excel_filter.active_filters:
            active_filters_backup = self.reservations_excel_filter.active_filters.copy()

        # –Я–Ю–Ы–Э–Ю–°–Ґ–ђ–Ѓ –Ю–І–Ш–©–Р–Х–Ь –Ф–Х–†–Х–Т–Ю
        for i in self.reservations_tree.get_children():
            self.reservations_tree.delete(i)

        # –Ю–І–Ш–©–Р–Х–Ь –Ъ–≠–® –≠–Ы–Х–Ь–Х–Э–Ґ–Ю–Т
        if hasattr(self, 'reservations_excel_filter'):
            self.reservations_excel_filter._all_item_cache = set()

        reservations_df = load_data("Reservations")
        orders_df = load_data("Orders")

        if not reservations_df.empty:
            show_fully_written_off = True

            if hasattr(self, 'reservations_toggles') and self.reservations_toggles:
                show_fully_written_off = self.reservations_toggles.get('show_fully_written_off',
                                                                       tk.BooleanVar(value=True)).get()

            for index, row in reservations_df.iterrows():
                remainder = int(row["–Ю—Б—В–∞—В–Њ–Ї –Ї —Б–њ–Є—Б–∞–љ–Є—О"])
                if not show_fully_written_off and remainder == 0:
                    continue

                # –Я–Њ–ї—Г—З–∞–µ–Љ –Є–љ—Д–Њ—А–Љ–∞—Ж–Є—О –Њ –Ј–∞–Ї–∞–Ј–µ
                order_id = int(row["ID –Ј–∞–Ї–∞–Ј–∞"])
                order_display = f"#{order_id}"

                if not orders_df.empty:
                    order_row = orders_df[orders_df["ID –Ј–∞–Ї–∞–Ј–∞"] == order_id]
                    if not order_row.empty:
                        customer = order_row.iloc[0]["–Ч–∞–Ї–∞–Ј—З–Є–Ї"]
                        order_name = order_row.iloc[0]["–Э–∞–Ј–≤–∞–љ–Є–µ –Ј–∞–Ї–∞–Ј–∞"]
                        order_display = f"{customer} | {order_name}"

                size_str = f"{row['–®–Є—А–Є–љ–∞']}x{row['–Ф–ї–Є–љ–∞']}"
                detail_name = row.get("–Э–∞–Ј–≤–∞–љ–Є–µ –і–µ—В–∞–ї–Є", "–Э–µ —Г–Ї–∞–Ј–∞–љ–∞") if "–Э–∞–Ј–≤–∞–љ–Є–µ –і–µ—В–∞–ї–Є" in row else "–Э–µ —Г–Ї–∞–Ј–∞–љ–∞"

                values = [
                    row["ID —А–µ–Ј–µ—А–≤–∞"],
                    order_display,
                    detail_name,
                    row["ID –Љ–∞—В–µ—А–Є–∞–ї–∞"],
                    row["–Ь–∞—А–Ї–∞"],
                    row["–Ґ–Њ–ї—Й–Є–љ–∞"],
                    size_str,
                    row["–Ч–∞—А–µ–Ј–µ—А–≤–Є—А–Њ–≤–∞–љ–Њ —И—В—Г–Ї"],
                    row["–°–њ–Є—Б–∞–љ–Њ"],
                    row["–Ю—Б—В–∞—В–Њ–Ї –Ї —Б–њ–Є—Б–∞–љ–Є—О"],
                    row["–Ф–∞—В–∞ —А–µ–Ј–µ—А–≤–∞"]
                ]

                item_id = self.reservations_tree.insert("", "end", values=values)

                # –°–Ю–•–†–Р–Э–ѓ–Х–Ь item_id –Т –Ъ–≠–®
                if hasattr(self, 'reservations_excel_filter'):
                    if not hasattr(self.reservations_excel_filter, '_all_item_cache'):
                        self.reservations_excel_filter._all_item_cache = set()
                    self.reservations_excel_filter._all_item_cache.add(item_id)

        # вЬЕ –Р–Т–Ґ–Ю–Я–Ю–Ф–С–Ю–† –®–Ш–†–Ш–Э–Ђ –Ъ–Ю–Ы–Ю–Э–Ю–Ъ (–Ф–Ю–Ы–Ц–Х–Э –С–Ђ–Ґ–ђ –Ч–Ф–Х–°–ђ!)
        self.auto_resize_columns(self.reservations_tree, min_width=80, max_width=400)

        # –Я–Х–†–Х–Я–†–Ш–Ь–Х–Э–ѓ–Х–Ь –§–Ш–Ы–ђ–Ґ–†–Ђ –Я–Ю–°–Ы–Х –Ч–Р–У–†–£–Ч–Ъ–Ш –Ф–Р–Э–Э–Ђ–•
        if active_filters_backup and hasattr(self, 'reservations_excel_filter'):
            self.reservations_excel_filter.active_filters = active_filters_backup
            self.reservations_excel_filter.reapply_all_filters()

    def add_reservation(self):
        orders_df = load_data("Orders")
        if orders_df.empty:
            messagebox.showwarning("–Я—А–µ–і—Г–њ—А–µ–ґ–і–µ–љ–Є–µ", "–°–љ–∞—З–∞–ї–∞ —Б–Њ–Ј–і–∞–є—В–µ –Ј–∞–Ї–∞–Ј—Л!")
            return

        add_window = tk.Toplevel(self.root)
        add_window.title("–°–Њ–Ј–і–∞—В—М —А–µ–Ј–µ—А–≤")
        add_window.geometry("550x850")
        add_window.configure(bg='#ecf0f1')
        tk.Label(add_window, text="–†–µ–Ј–µ—А–≤–Є—А–Њ–≤–∞–љ–Є–µ –Љ–∞—В–µ—А–Є–∞–ї–∞ –њ–Њ–і –Ј–∞–Ї–∞–Ј", font=("Arial", 12, "bold"), bg='#ecf0f1').pack(
            pady=10)

        # –Ч–Р–Ъ–Р–Ч –° –Я–Ю–Ш–°–Ъ–Ю–Ь
        order_frame = tk.Frame(add_window, bg='#ecf0f1')
        order_frame.pack(fill=tk.X, padx=20, pady=5)
        tk.Label(order_frame, text="–Ч–∞–Ї–∞–Ј:", width=20, anchor='w', bg='#ecf0f1', font=("Arial", 10)).pack(side=tk.LEFT)

        all_order_options = [
            f"ID:{int(row['ID –Ј–∞–Ї–∞–Ј–∞'])} | {row['–Ч–∞–Ї–∞–Ј—З–Є–Ї']} | {row['–Э–∞–Ј–≤–∞–љ–Є–µ –Ј–∞–Ї–∞–Ј–∞']}"
            for _, row in orders_df.iterrows()
        ]

        order_search_var = tk.StringVar()
        order_search_entry = tk.Entry(order_frame, textvariable=order_search_var, font=("Arial", 10), width=35)
        order_search_entry.pack(side=tk.RIGHT, expand=True, fill=tk.X)

        order_results_frame = tk.Frame(add_window, bg='#ecf0f1')
        order_results_frame.pack(fill=tk.X, padx=20, pady=5)

        order_scroll = tk.Scrollbar(order_results_frame, orient=tk.VERTICAL)
        order_listbox = tk.Listbox(order_results_frame, height=3, font=("Arial", 9),
                                   yscrollcommand=order_scroll.set)
        order_scroll.config(command=order_listbox.yview)
        order_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        order_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        for option in all_order_options:
            order_listbox.insert(tk.END, option)

        selected_order = {"value": None}

        def on_order_search(*args):
            search_text = order_search_var.get().lower()
            order_listbox.delete(0, tk.END)
            for option in all_order_options:
                if search_text in option.lower():
                    order_listbox.insert(tk.END, option)

        def on_select_order(event):
            try:
                selection = order_listbox.get(order_listbox.curselection())
                selected_order["value"] = selection
                order_search_var.set(selection)
                update_details_list()
            except:
                pass

        order_search_var.trace('w', on_order_search)
        order_listbox.bind('<<ListboxSelect>>', on_select_order)
        order_listbox.bind('<Double-Button-1>', on_select_order)

        # –Ф–Х–Ґ–Р–Ы–ђ –Ч–Р–Ъ–Р–Ч–Р
        detail_frame = tk.Frame(add_window, bg='#ecf0f1')
        detail_frame.pack(fill=tk.X, padx=20, pady=5)
        tk.Label(detail_frame, text="–Ф–µ—В–∞–ї—М –Ј–∞–Ї–∞–Ј–∞:", width=20, anchor='w', bg='#ecf0f1',
                 font=("Arial", 10, "bold")).pack(side=tk.LEFT)

        detail_var = tk.StringVar()
        detail_combo = ttk.Combobox(detail_frame, textvariable=detail_var, font=("Arial", 10), state="readonly",
                                    width=35)
        detail_combo.pack(side=tk.RIGHT, expand=True, fill=tk.X)

        selected_detail = {"id": None, "name": None}

        def update_details_list():
            detail_combo['values'] = []
            detail_var.set("")
            selected_detail["id"] = None
            selected_detail["name"] = None

            if not selected_order["value"]:
                return

            try:
                # рЯЖХ –Я–†–Р–Т–Ш–Ы–ђ–Э–Ђ–Щ –Я–Р–†–°–Ш–Э–У: "ID:123 | –Ч–∞–Ї–∞–Ј—З–Є–Ї | –Э–∞–Ј–≤–∞–љ–Є–µ"
                order_str = selected_order["value"]

                if order_str.startswith("ID:"):
                    order_id = int(order_str.split("ID:")[1].split(" | ")[0].strip())
                else:
                    # –°—В–∞—А—Л–є —Д–Њ—А–Љ–∞—В –і–ї—П —Б–Њ–≤–Љ–µ—Б—В–Є–Љ–Њ—Б—В–Є
                    order_id = int(order_str.split(" - ")[0])

                print(f"рЯФН –Ч–∞–≥—А—Г–Ј–Ї–∞ –і–µ—В–∞–ї–µ–є –і–ї—П –Ј–∞–Ї–∞–Ј–∞ ID={order_id}")

                order_details_df = load_data("OrderDetails")

                if not order_details_df.empty:
                    details = order_details_df[order_details_df["ID –Ј–∞–Ї–∞–Ј–∞"] == order_id]

                    if not details.empty:
                        detail_options = ["[–С–µ–Ј –њ—А–Є–≤—П–Ј–Ї–Є –Ї –і–µ—В–∞–ї–Є]"]
                        detail_options.extend([f"ID:{int(row['ID'])} - {row['–Э–∞–Ј–≤–∞–љ–Є–µ –і–µ—В–∞–ї–Є']}"
                                               for _, row in details.iterrows()])
                        detail_combo['values'] = detail_options
                        detail_combo.current(0)
                        print(f"вЬЕ –Э–∞–є–і–µ–љ–Њ –і–µ—В–∞–ї–µ–є: {len(details)}")
                    else:
                        detail_combo['values'] = ["[–Э–µ—В –і–µ—В–∞–ї–µ–є —Г –Ј–∞–Ї–∞–Ј–∞]"]
                        detail_combo.current(0)
                        print(f"вЪ†пЄП –£ –Ј–∞–Ї–∞–Ј–∞ ID={order_id} –љ–µ—В –і–µ—В–∞–ї–µ–є")
                else:
                    detail_combo['values'] = ["[–Э–µ—В –і–µ—В–∞–ї–µ–є —Г –Ј–∞–Ї–∞–Ј–∞]"]
                    detail_combo.current(0)
                    print(f"вЪ†пЄП –Ґ–∞–±–ї–Є—Ж–∞ –і–µ—В–∞–ї–µ–є –њ—Г—Б—В–∞")
            except Exception as e:
                print(f"вЭМ –Ю—И–Є–±–Ї–∞ –Њ–±–љ–Њ–≤–ї–µ–љ–Є—П —Б–њ–Є—Б–Ї–∞ –і–µ—В–∞–ї–µ–є: {e}")
                import traceback
                traceback.print_exc()

        def on_detail_select(event):
            value = detail_var.get()
            if value and value.startswith("ID:"):
                try:
                    selected_detail["id"] = int(value.split("ID:")[1].split(" - ")[0])
                    selected_detail["name"] = value.split(" - ")[1]
                except:
                    selected_detail["id"] = None
                    selected_detail["name"] = None
            else:
                selected_detail["id"] = None
                selected_detail["name"] = None

        detail_combo.bind('<<ComboboxSelected>>', on_detail_select)

        # –Ь–Р–Ґ–Х–†–Ш–Р–Ы –° –Я–Ю–Ш–°–Ъ–Ю–Ь
        material_frame = tk.Frame(add_window, bg='#ecf0f1')
        material_frame.pack(fill=tk.X, padx=20, pady=5)
        tk.Label(material_frame, text="–Ь–∞—В–µ—А–Є–∞–ї:", width=20, anchor='w', bg='#ecf0f1', font=("Arial", 10)).pack(
            side=tk.LEFT)

        materials_df = load_data("Materials")
        all_material_options = ["[–Ф–Њ–±–∞–≤–Є—В—М –≤—А—Г—З–љ—Г—О]"]
        if not materials_df.empty:
            all_material_options.extend([
                                            f"{int(row['ID'])} - {row['–Ь–∞—А–Ї–∞']} {row['–Ґ–Њ–ї—Й–Є–љ–∞']}–Љ–Љ {row['–®–Є—А–Є–љ–∞']}x{row['–Ф–ї–Є–љ–∞']} (–і–Њ—Б—В—Г–њ–љ–Њ: {int(row['–Ф–Њ—Б—В—Г–њ–љ–Њ'])} —И—В)"
                                            for _, row in materials_df.iterrows()])

        search_container = tk.Frame(material_frame, bg='#ecf0f1')
        search_container.pack(side=tk.RIGHT, expand=True, fill=tk.X)

        material_search_var = tk.StringVar()
        material_search_entry = tk.Entry(search_container, textvariable=material_search_var, font=("Arial", 10))
        material_search_entry.pack(fill=tk.X)

        selected_reserve = {"value": None}

        search_results_frame = tk.Frame(add_window, bg='#ecf0f1')
        search_results_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=5)

        scroll_results = tk.Scrollbar(search_results_frame, orient=tk.VERTICAL)
        results_listbox = tk.Listbox(search_results_frame, height=5, font=("Arial", 9),
                                     yscrollcommand=scroll_results.set)
        scroll_results.config(command=results_listbox.yview)
        scroll_results.pack(side=tk.RIGHT, fill=tk.Y)
        results_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        for option in all_material_options:
            results_listbox.insert(tk.END, option)

        selected_material = {"value": None}

        def on_search_change(*args):
            search_text = material_search_var.get().lower()
            results_listbox.delete(0, tk.END)

            for option in all_material_options:
                if search_text in option.lower():
                    results_listbox.insert(tk.END, option)

        def on_select_material(event):
            try:
                selection = results_listbox.get(results_listbox.curselection())
                selected_material["value"] = selection
                material_search_var.set(selection)
            except:
                pass

        material_search_var.trace('w', on_search_change)
        results_listbox.bind('<<ListboxSelect>>', on_select_material)
        results_listbox.bind('<Double-Button-1>', on_select_material)

        # –Я–Р–†–Р–Ь–Х–Ґ–†–Ђ –Ь–Р–Ґ–Х–†–Ш–Р–Ы–Р (—А—Г—З–љ–Њ–є –≤–≤–Њ–і)
        manual_frame = tk.LabelFrame(add_window, text="–Я–∞—А–∞–Љ–µ—В—А—Л –Љ–∞—В–µ—А–Є–∞–ї–∞ (–і–ї—П —А—Г—З–љ–Њ–≥–Њ –≤–≤–Њ–і–∞)", bg='#ecf0f1',
                                     font=("Arial", 10, "bold"))
        manual_frame.pack(fill=tk.X, padx=20, pady=10)
        manual_entries = {}
        manual_fields = [("–Ь–∞—А–Ї–∞ —Б—В–∞–ї–Є:", "marka"), ("–Ґ–Њ–ї—Й–Є–љ–∞ (–Љ–Љ):", "thickness"), ("–Ф–ї–Є–љ–∞ (–Љ–Љ):", "length"),
                         ("–®–Є—А–Є–љ–∞ (–Љ–Љ):", "width")]
        for label_text, key in manual_fields:
            frame = tk.Frame(manual_frame, bg='#ecf0f1')
            frame.pack(fill=tk.X, padx=10, pady=3)
            tk.Label(frame, text=label_text, width=18, anchor='w', bg='#ecf0f1', font=("Arial", 9)).pack(side=tk.LEFT)
            entry = tk.Entry(frame, font=("Arial", 9))
            entry.pack(side=tk.RIGHT, expand=True, fill=tk.X)
            manual_entries[key] = entry

        # –Ъ–Ю–Ы–Ш–І–Х–°–Ґ–Т–Ю
        qty_frame = tk.Frame(add_window, bg='#ecf0f1')
        qty_frame.pack(fill=tk.X, padx=20, pady=5)
        tk.Label(qty_frame, text="–Ъ–Њ–ї–Є—З–µ—Б—В–≤–Њ (—И—В):", width=20, anchor='w', bg='#ecf0f1',
                 font=("Arial", 10, "bold")).pack(side=tk.LEFT)
        qty_entry = tk.Entry(qty_frame, font=("Arial", 10))
        qty_entry.pack(side=tk.RIGHT, expand=True, fill=tk.X)

        def save_reservation():
            try:
                order_value = selected_order["value"] or order_search_var.get()
                if not order_value:
                    messagebox.showwarning("–Я—А–µ–і—Г–њ—А–µ–ґ–і–µ–љ–Є–µ", "–Т—Л–±–µ—А–Є—В–µ –Ј–∞–Ї–∞–Ј!")
                    return

                material_value = selected_material["value"] or material_search_var.get()
                if not material_value:
                    messagebox.showwarning("–Я—А–µ–і—Г–њ—А–µ–ґ–і–µ–љ–Є–µ", "–Т—Л–±–µ—А–Є—В–µ –Љ–∞—В–µ—А–Є–∞–ї!")
                    return

                # –Я–∞—А—Б–Є–Љ ID –Є–Ј —Д–Њ—А–Љ–∞—В–∞ "ID:1001 | –Ч–∞–Ї–∞–Ј—З–Є–Ї | –Э–∞–Ј–≤–∞–љ–Є–µ"
                order_id = int(order_value.split("ID:")[1].split(" | ")[0])
                quantity = int(qty_entry.get())

                # –Я–Њ–ї—Г—З–∞–µ–Љ ID –Є –љ–∞–Ј–≤–∞–љ–Є–µ –і–µ—В–∞–ї–Є
                detail_id = selected_detail["id"] if selected_detail["id"] else -1
                detail_name = selected_detail["name"] if selected_detail["name"] else "–Э–µ —Г–Ї–∞–Ј–∞–љ–∞"

                if material_value == "[–Ф–Њ–±–∞–≤–Є—В—М –≤—А—Г—З–љ—Г—О]":
                    marka = manual_entries["marka"].get().strip()
                    thickness = float(manual_entries["thickness"].get().strip())
                    length = float(manual_entries["length"].get().strip())
                    width = float(manual_entries["width"].get().strip())
                    if not marka:
                        messagebox.showwarning("–Я—А–µ–і—Г–њ—А–µ–ґ–і–µ–љ–Є–µ", "–Ч–∞–њ–Њ–ї–љ–Є—В–µ –Љ–∞—А–Ї—Г —Б—В–∞–ї–Є!")
                        return
                    material_id = -1
                else:
                    material_id = int(material_value.split(" - ")[0])
                    material_row = materials_df[materials_df["ID"] == material_id].iloc[0]
                    marka = material_row["–Ь–∞—А–Ї–∞"]
                    thickness = material_row["–Ґ–Њ–ї—Й–Є–љ–∞"]
                    length = material_row["–Ф–ї–Є–љ–∞"]
                    width = material_row["–®–Є—А–Є–љ–∞"]

                reservations_df = load_data("Reservations")
                new_id = 1 if reservations_df.empty else int(reservations_df["ID —А–µ–Ј–µ—А–≤–∞"].max()) + 1

                new_row = pd.DataFrame([{
                    "ID —А–µ–Ј–µ—А–≤–∞": new_id,
                    "ID –Ј–∞–Ї–∞–Ј–∞": order_id,
                    "ID –і–µ—В–∞–ї–Є": detail_id,
                    "–Э–∞–Ј–≤–∞–љ–Є–µ –і–µ—В–∞–ї–Є": detail_name,
                    "ID –Љ–∞—В–µ—А–Є–∞–ї–∞": material_id,
                    "–Ь–∞—А–Ї–∞": marka,
                    "–Ґ–Њ–ї—Й–Є–љ–∞": thickness,
                    "–Ф–ї–Є–љ–∞": length,
                    "–®–Є—А–Є–љ–∞": width,
                    "–Ч–∞—А–µ–Ј–µ—А–≤–Є—А–Њ–≤–∞–љ–Њ —И—В—Г–Ї": quantity,
                    "–°–њ–Є—Б–∞–љ–Њ": 0,
                    "–Ю—Б—В–∞—В–Њ–Ї –Ї —Б–њ–Є—Б–∞–љ–Є—О": quantity,
                    "–Ф–∞—В–∞ —А–µ–Ј–µ—А–≤–∞": datetime.now().strftime("%Y-%m-%d")
                }])

                reservations_df = pd.concat([reservations_df, new_row], ignore_index=True)
                save_data("Reservations", reservations_df)

                if material_id != -1:
                    materials_df.loc[materials_df["ID"] == material_id, "–Ч–∞—А–µ–Ј–µ—А–≤–Є—А–Њ–≤–∞–љ–Њ"] = int(
                        material_row["–Ч–∞—А–µ–Ј–µ—А–≤–Є—А–Њ–≤–∞–љ–Њ"]) + quantity
                    materials_df.loc[materials_df["ID"] == material_id, "–Ф–Њ—Б—В—Г–њ–љ–Њ"] = int(
                        material_row["–Ф–Њ—Б—В—Г–њ–љ–Њ"]) - quantity
                    save_data("Materials", materials_df)
                    self.refresh_materials()

                self.refresh_reservations()
                self.refresh_balance()
                add_window.destroy()

                detail_info = f"\n–Ф–µ—В–∞–ї—М: {detail_name}" if detail_name != "–Э–µ —Г–Ї–∞–Ј–∞–љ–∞" else ""
                messagebox.showinfo("–£—Б–њ–µ—Е", f"–†–µ–Ј–µ—А–≤ #{new_id} —Г—Б–њ–µ—И–љ–Њ —Б–Њ–Ј–і–∞–љ!{detail_info}")

            except ValueError:
                messagebox.showerror("–Ю—И–Є–±–Ї–∞", "–Я—А–Њ–≤–µ—А—М—В–µ –њ—А–∞–≤–Є–ї—М–љ–Њ—Б—В—М –≤–≤–Њ–і–∞ —З–Є—Б–ї–Њ–≤—Л—Е –Ј–љ–∞—З–µ–љ–Є–є!")
            except Exception as e:
                messagebox.showerror("–Ю—И–Є–±–Ї–∞", f"–Э–µ —Г–і–∞–ї–Њ—Б—М —Б–Њ–Ј–і–∞—В—М —А–µ–Ј–µ—А–≤: {e}")

        tk.Button(add_window, text="–Ч–∞—А–µ–Ј–µ—А–≤–Є—А–Њ–≤–∞—В—М", bg='#27ae60', fg='white', font=("Arial", 12, "bold"),
                  command=save_reservation).pack(pady=15)

    def delete_reservation(self):
        selected = self.reservations_tree.selection()
        if not selected:
            messagebox.showwarning("–Я—А–µ–і—Г–њ—А–µ–ґ–і–µ–љ–Є–µ", "–Т—Л–±–µ—А–Є—В–µ —А–µ–Ј–µ—А–≤—Л –і–ї—П —Г–і–∞–ї–µ–љ–Є—П")
            return
        count = len(selected)
        if messagebox.askyesno("–Я–Њ–і—В–≤–µ—А–ґ–і–µ–љ–Є–µ",
                               f"–£–і–∞–ї–Є—В—М –≤—Л–±—А–∞–љ–љ—Л–µ —А–µ–Ј–µ—А–≤—Л ({count} —И—В)?\n\n–Ь–∞—В–µ—А–Є–∞–ї—Л –≤–µ—А–љ—Г—В—Б—П –љ–∞ —Б–Ї–ї–∞–і!"):
            reservations_df = load_data("Reservations")
            materials_df = load_data("Materials")
            for item in selected:
                reserve_id = self.reservations_tree.item(item)["values"][0]
                reserve_row = reservations_df[reservations_df["ID —А–µ–Ј–µ—А–≤–∞"] == reserve_id].iloc[0]
                material_id = reserve_row["ID –Љ–∞—В–µ—А–Є–∞–ї–∞"]
                if material_id != -1:
                    quantity_to_return = int(reserve_row["–Ю—Б—В–∞—В–Њ–Ї –Ї —Б–њ–Є—Б–∞–љ–Є—О"])
                    if not materials_df[materials_df["ID"] == material_id].empty:
                        mat_row = materials_df[materials_df["ID"] == material_id].iloc[0]
                        materials_df.loc[materials_df["ID"] == material_id, "–Ч–∞—А–µ–Ј–µ—А–≤–Є—А–Њ–≤–∞–љ–Њ"] = int(
                            mat_row["–Ч–∞—А–µ–Ј–µ—А–≤–Є—А–Њ–≤–∞–љ–Њ"]) - quantity_to_return
                        materials_df.loc[materials_df["ID"] == material_id, "–Ф–Њ—Б—В—Г–њ–љ–Њ"] = int(
                            mat_row["–Ф–Њ—Б—В—Г–њ–љ–Њ"]) + quantity_to_return
                reservations_df = reservations_df[reservations_df["ID —А–µ–Ј–µ—А–≤–∞"] != reserve_id]
            save_data("Reservations", reservations_df)
            save_data("Materials", materials_df)
            self.refresh_materials()
            self.refresh_reservations()
            self.refresh_balance()
            messagebox.showinfo("–£—Б–њ–µ—Е", f"–£–і–∞–ї–µ–љ–Њ —А–µ–Ј–µ—А–≤–Њ–≤: {count}")

    def edit_reservation(self):
        """–†–µ–і–∞–Ї—В–Є—А–Њ–≤–∞–љ–Є–µ —А–µ–Ј–µ—А–≤–Є—А–Њ–≤–∞–љ–Є—П —Б –≤–Њ–Ј–Љ–Њ–ґ–љ–Њ—Б—В—М—О –Є–Ј–Љ–µ–љ–µ–љ–Є—П –Ј–∞–Ї–∞–Ј–∞ –Є –і–µ—В–∞–ї–Є"""
        selected = self.reservations_tree.selection()
        if not selected:
            messagebox.showwarning("–Я—А–µ–і—Г–њ—А–µ–ґ–і–µ–љ–Є–µ", "–Т—Л–±–µ—А–Є—В–µ —А–µ–Ј–µ—А–≤ –і–ї—П —А–µ–і–∞–Ї—В–Є—А–Њ–≤–∞–љ–Є—П")
            return

        reserve_id = self.reservations_tree.item(selected)["values"][0]
        reservations_df = load_data("Reservations")
        reserve_row = reservations_df[reservations_df["ID —А–µ–Ј–µ—А–≤–∞"] == reserve_id].iloc[0]

        edit_window = tk.Toplevel(self.root)
        edit_window.title("–†–µ–і–∞–Ї—В–Є—А–Њ–≤–∞—В—М —А–µ–Ј–µ—А–≤")
        edit_window.geometry("650x800")
        edit_window.configure(bg='#ecf0f1')

        tk.Label(edit_window, text=f"–†–µ–і–∞–Ї—В–Є—А–Њ–≤–∞–љ–Є–µ —А–µ–Ј–µ—А–≤–∞ #{reserve_id}",
                 font=("Arial", 12, "bold"), bg='#ecf0f1', fg='#2c3e50').pack(pady=10)

        # –Ч–∞–≥—А—Г–ґ–∞–µ–Љ –і–∞–љ–љ—Л–µ
        orders_df = load_data("Orders")
        order_details_df = load_data("OrderDetails")

        # –Ґ–µ–Ї—Г—Й–Є–µ –і–∞–љ–љ—Л–µ —А–µ–Ј–µ—А–≤–∞
        current_order_id = int(reserve_row["ID –Ј–∞–Ї–∞–Ј–∞"])
        current_detail_id = reserve_row.get("ID –і–µ—В–∞–ї–Є", -1)
        if pd.isna(current_detail_id):
            current_detail_id = -1
        else:
            current_detail_id = int(current_detail_id)

        written_off = int(reserve_row["–°–њ–Є—Б–∞–љ–Њ"])

        # === –Ч–Р–Ъ–Р–Ч ===
        order_frame = tk.LabelFrame(edit_window, text="–Ч–∞–Ї–∞–Ј", bg='#ecf0f1', font=("Arial", 10, "bold"))
        order_frame.pack(fill=tk.X, padx=20, pady=10)

        tk.Label(order_frame, text="–Т—Л–±–µ—А–Є—В–µ –Ј–∞–Ї–∞–Ј:", bg='#ecf0f1', font=("Arial", 9)).pack(anchor='w', padx=10, pady=5)

        # –§–Њ—А–Љ–Є—А—Г–µ–Љ —Б–њ–Є—Б–Њ–Ї –Ј–∞–Ї–∞–Ј–Њ–≤
        all_order_options = []
        order_map = {}

        if not orders_df.empty:
            for _, row in orders_df.iterrows():
                order_id = int(row['ID –Ј–∞–Ї–∞–Ј–∞'])
                display_text = f"ID:{order_id} | {row['–Ч–∞–Ї–∞–Ј—З–Є–Ї']} | {row['–Э–∞–Ј–≤–∞–љ–Є–µ –Ј–∞–Ї–∞–Ј–∞']}"
                all_order_options.append(display_text)
                order_map[display_text] = order_id

        order_search_var = tk.StringVar()
        order_search_entry = tk.Entry(order_frame, textvariable=order_search_var, font=("Arial", 9))
        order_search_entry.pack(fill=tk.X, padx=10, pady=5)

        order_listbox = tk.Listbox(order_frame, height=4, font=("Arial", 9))
        order_listbox.pack(fill=tk.BOTH, padx=10, pady=5)

        for option in all_order_options:
            order_listbox.insert(tk.END, option)

        selected_order = {"value": None, "id": current_order_id}

        def on_order_search(*args):
            search_text = order_search_var.get().lower()
            order_listbox.delete(0, tk.END)
            for option in all_order_options:
                if search_text in option.lower():
                    order_listbox.insert(tk.END, option)

        def on_select_order(event):
            try:
                selection = order_listbox.get(order_listbox.curselection())
                selected_order["value"] = selection
                selected_order["id"] = order_map[selection]
                order_search_var.set(selection)
                update_details_list()
            except:
                pass

        order_search_var.trace('w', on_order_search)
        order_listbox.bind('<<ListboxSelect>>', on_select_order)
        order_listbox.bind('<Double-Button-1>', on_select_order)

        # –£—Б—В–∞–љ–∞–≤–ї–Є–≤–∞–µ–Љ —В–µ–Ї—Г—Й–Є–є –Ј–∞–Ї–∞–Ј
        for i, option in enumerate(all_order_options):
            if order_map[option] == current_order_id:
                order_listbox.selection_set(i)
                order_listbox.see(i)
                order_search_var.set(option)
                selected_order["value"] = option
                break

        # === –Ф–Х–Ґ–Р–Ы–ђ ===
        detail_frame = tk.LabelFrame(edit_window, text="–Ф–µ—В–∞–ї—М", bg='#ecf0f1', font=("Arial", 10, "bold"))
        detail_frame.pack(fill=tk.X, padx=20, pady=10)

        tk.Label(detail_frame, text="–Т—Л–±–µ—А–Є—В–µ –і–µ—В–∞–ї—М:", bg='#ecf0f1', font=("Arial", 9)).pack(anchor='w', padx=10,
                                                                                              pady=5)

        detail_var = tk.StringVar()
        detail_combo = ttk.Combobox(detail_frame, textvariable=detail_var, font=("Arial", 9),
                                    state="readonly", width=50)
        detail_combo.pack(fill=tk.X, padx=10, pady=5)

        selected_detail = {"id": current_detail_id, "name": None}

        def update_details_list():
            detail_combo['values'] = []
            detail_var.set("")
            selected_detail["id"] = -1
            selected_detail["name"] = None

            order_id = selected_order["id"]
            if not order_id:
                return

            try:
                if not order_details_df.empty:
                    details = order_details_df[order_details_df["ID –Ј–∞–Ї–∞–Ј–∞"] == order_id]

                    if not details.empty:
                        detail_options = ["[–С–µ–Ј –њ—А–Є–≤—П–Ј–Ї–Є –Ї –і–µ—В–∞–ї–Є]"]
                        detail_options.extend([f"ID:{int(row['ID'])} - {row['–Э–∞–Ј–≤–∞–љ–Є–µ –і–µ—В–∞–ї–Є']}"
                                               for _, row in details.iterrows()])
                        detail_combo['values'] = detail_options

                        # –Я—Л—В–∞–µ–Љ—Б—П —Г—Б—В–∞–љ–Њ–≤–Є—В—М —В–µ–Ї—Г—Й—Г—О –і–µ—В–∞–ї—М
                        if current_detail_id != -1:
                            for opt in detail_options:
                                if opt.startswith(f"ID:{current_detail_id} -"):
                                    detail_combo.set(opt)
                                    selected_detail["id"] = current_detail_id
                                    selected_detail["name"] = opt.split(" - ")[1]
                                    break
                        else:
                            detail_combo.current(0)
                    else:
                        detail_combo['values'] = ["[–Э–µ—В –і–µ—В–∞–ї–µ–є —Г –Ј–∞–Ї–∞–Ј–∞]"]
                        detail_combo.current(0)
                else:
                    detail_combo['values'] = ["[–Э–µ—В –і–µ—В–∞–ї–µ–є —Г –Ј–∞–Ї–∞–Ј–∞]"]
                    detail_combo.current(0)
            except Exception as e:
                print(f"–Ю—И–Є–±–Ї–∞ –Њ–±–љ–Њ–≤–ї–µ–љ–Є—П —Б–њ–Є—Б–Ї–∞ –і–µ—В–∞–ї–µ–є: {e}")

        def on_detail_select(event):
            value = detail_var.get()
            if value and value.startswith("ID:"):
                try:
                    selected_detail["id"] = int(value.split("ID:")[1].split(" - ")[0])
                    selected_detail["name"] = value.split(" - ")[1]
                except:
                    selected_detail["id"] = -1
                    selected_detail["name"] = None
            else:
                selected_detail["id"] = -1
                selected_detail["name"] = None

        detail_combo.bind('<<ComboboxSelected>>', on_detail_select)

        # –Ш–љ–Є—Ж–Є–∞–ї–Є–Ј–Є—А—Г–µ–Љ —Б–њ–Є—Б–Њ–Ї –і–µ—В–∞–ї–µ–є
        update_details_list()

        # === –Ь–Р–Ґ–Х–†–Ш–Р–Ы (—В–Њ–ї—М–Ї–Њ –і–ї—П —З—В–µ–љ–Є—П) ===
        material_frame = tk.LabelFrame(edit_window, text="–Ь–∞—В–µ—А–Є–∞–ї (–љ–µ —А–µ–і–∞–Ї—В–Є—А—Г–µ—В—Б—П)",
                                       bg='#e8f4f8', font=("Arial", 9, "bold"))
        material_frame.pack(fill=tk.X, padx=20, pady=10)

        material_info = f"{reserve_row['–Ь–∞—А–Ї–∞']} {reserve_row['–Ґ–Њ–ї—Й–Є–љ–∞']}–Љ–Љ {reserve_row['–®–Є—А–Є–љ–∞']}x{reserve_row['–Ф–ї–Є–љ–∞']}"
        tk.Label(material_frame, text=material_info, bg='#e8f4f8', font=("Arial", 10)).pack(padx=10, pady=5)

        # === –Ъ–Ю–Ы–Ш–І–Х–°–Ґ–Т–Ю ===
        qty_frame = tk.Frame(edit_window, bg='#ecf0f1')
        qty_frame.pack(fill=tk.X, padx=20, pady=10)
        tk.Label(qty_frame, text="–Ч–∞—А–µ–Ј–µ—А–≤–Є—А–Њ–≤–∞–љ–Њ (—И—В):", width=25, anchor='w',
                 bg='#ecf0f1', font=("Arial", 10, "bold")).pack(side=tk.LEFT)
        qty_entry = tk.Entry(qty_frame, font=("Arial", 10))
        qty_entry.insert(0, str(int(reserve_row["–Ч–∞—А–µ–Ј–µ—А–≤–Є—А–Њ–≤–∞–љ–Њ —И—В—Г–Ї"])))
        qty_entry.pack(side=tk.RIGHT, expand=True, fill=tk.X)

        # === –°–Ґ–Р–Ґ–Ш–°–Ґ–Ш–Ъ–Р ===
        remainder = int(reserve_row["–Ю—Б—В–∞—В–Њ–Ї –Ї —Б–њ–Є—Б–∞–љ–Є—О"])

        stats_frame = tk.LabelFrame(edit_window, text="–°—В–∞—В–Є—Б—В–Є–Ї–∞", bg='#fff3cd', font=("Arial", 9, "bold"))
        stats_frame.pack(fill=tk.X, padx=20, pady=10)
        tk.Label(stats_frame, text=f"–£–ґ–µ —Б–њ–Є—Б–∞–љ–Њ: {written_off} —И—В",
                 bg='#fff3cd', font=("Arial", 9)).pack(anchor='w', padx=10, pady=2)
        tk.Label(stats_frame, text=f"–Ю—Б—В–∞—В–Њ–Ї –Ї —Б–њ–Є—Б–∞–љ–Є—О: {remainder} —И—В",
                 bg='#fff3cd', font=("Arial", 9)).pack(anchor='w', padx=10, pady=2)

        # === –Я–†–Х–Ф–£–Я–†–Х–Ц–Ф–Х–Э–Ш–Х ===
        warning_frame = tk.Frame(edit_window, bg='#ffcccc', relief=tk.RIDGE, borderwidth=2)
        warning_frame.pack(fill=tk.X, padx=20, pady=10)
        tk.Label(warning_frame, text="вЪ† –Т–Р–Ц–Э–Ю!", font=("Arial", 9, "bold"),
                 bg='#ffcccc', fg='#c0392b').pack(anchor='w', padx=5, pady=2)
        tk.Label(warning_frame, text="вАҐ –Э–µ–ї—М–Ј—П —Г–Љ–µ–љ—М—И–Є—В—М –Ї–Њ–ї–Є—З–µ—Б—В–≤–Њ –љ–Є–ґ–µ —Г–ґ–µ —Б–њ–Є—Б–∞–љ–љ–Њ–≥–Њ",
                 font=("Arial", 8), bg='#ffcccc', fg='#c0392b').pack(anchor='w', padx=10)
        tk.Label(warning_frame, text="вАҐ –Ь–Њ–ґ–љ–Њ –Є–Ј–Љ–µ–љ–Є—В—М –Ј–∞–Ї–∞–Ј –Є –і–µ—В–∞–ї—М",
                 font=("Arial", 8), bg='#ffcccc', fg='#c0392b').pack(anchor='w', padx=10)
        tk.Label(warning_frame, text="вАҐ –Ш–Ј–Љ–µ–љ–µ–љ–Є–µ –Ї–Њ–ї–Є—З–µ—Б—В–≤–∞ –≤–ї–Є—П–µ—В –љ–∞ –±–∞–ї–∞–љ—Б –Љ–∞—В–µ—А–Є–∞–ї–Њ–≤",
                 font=("Arial", 8), bg='#ffcccc', fg='#c0392b').pack(anchor='w', padx=10)

        def save_changes():
            try:
                new_qty = int(qty_entry.get().strip())
                new_order_id = selected_order["id"]
                new_detail_id = selected_detail["id"]
                new_detail_name = selected_detail["name"] if selected_detail["name"] else "–Э–µ —Г–Ї–∞–Ј–∞–љ–∞"

                if not new_order_id:
                    messagebox.showwarning("–Я—А–µ–і—Г–њ—А–µ–ґ–і–µ–љ–Є–µ", "–Т—Л–±–µ—А–Є—В–µ –Ј–∞–Ї–∞–Ј!")
                    return

                if new_qty < written_off:
                    messagebox.showerror("–Ю—И–Є–±–Ї–∞",
                                         f"–Э–µ–ї—М–Ј—П —Г—Б—В–∞–љ–Њ–≤–Є—В—М –Ї–Њ–ї–Є—З–µ—Б—В–≤–Њ ({new_qty}) –Љ–µ–љ—М—И–µ —Г–ґ–µ —Б–њ–Є—Б–∞–љ–љ–Њ–≥–Њ ({written_off})!")
                    return

                if new_qty <= 0:
                    messagebox.showerror("–Ю—И–Є–±–Ї–∞", "–Ъ–Њ–ї–Є—З–µ—Б—В–≤–Њ –і–Њ–ї–ґ–љ–Њ –±—Л—В—М –±–Њ–ї—М—И–µ –љ—Г–ї—П!")
                    return

                old_qty = int(reserve_row["–Ч–∞—А–µ–Ј–µ—А–≤–Є—А–Њ–≤–∞–љ–Њ —И—В—Г–Ї"])
                qty_difference = new_qty - old_qty

                # –Я—А–Њ–≤–µ—А—П–µ–Љ –Є–Ј–Љ–µ–љ–µ–љ–Є—П
                order_changed = new_order_id != current_order_id
                detail_changed = new_detail_id != current_detail_id
                qty_changed = qty_difference != 0

                if not order_changed and not detail_changed and not qty_changed:
                    messagebox.showinfo("–Ш–љ—Д–Њ—А–Љ–∞—Ж–Є—П", "–Ш–Ј–Љ–µ–љ–µ–љ–Є–є –љ–µ –±—Л–ї–Њ")
                    edit_window.destroy()
                    return

                # –§–Њ—А–Љ–Є—А—Г–µ–Љ —Б–Њ–Њ–±—Й–µ–љ–Є–µ —Б –Є–Ј–Љ–µ–љ–µ–љ–Є—П–Љ–Є
                changes_msg = "–С—Г–і—Г—В –≤–љ–µ—Б–µ–љ—Л —Б–ї–µ–і—Г—О—Й–Є–µ –Є–Ј–Љ–µ–љ–µ–љ–Є—П:\n\n"

                if order_changed:
                    old_order = orders_df[orders_df["ID –Ј–∞–Ї–∞–Ј–∞"] == current_order_id].iloc[0]
                    new_order = orders_df[orders_df["ID –Ј–∞–Ї–∞–Ј–∞"] == new_order_id].iloc[0]
                    changes_msg += f"рЯУЛ –Ч–∞–Ї–∞–Ј:\n"
                    changes_msg += f"  –°—В–∞—А—Л–є: {old_order['–Ч–∞–Ї–∞–Ј—З–Є–Ї']} | {old_order['–Э–∞–Ј–≤–∞–љ–Є–µ –Ј–∞–Ї–∞–Ј–∞']}\n"
                    changes_msg += f"  –Э–Њ–≤—Л–є: {new_order['–Ч–∞–Ї–∞–Ј—З–Є–Ї']} | {new_order['–Э–∞–Ј–≤–∞–љ–Є–µ –Ј–∞–Ї–∞–Ј–∞']}\n\n"

                if detail_changed:
                    old_detail_name = reserve_row.get("–Э–∞–Ј–≤–∞–љ–Є–µ –і–µ—В–∞–ї–Є", "–Э–µ —Г–Ї–∞–Ј–∞–љ–∞")
                    if pd.isna(old_detail_name) or old_detail_name == "":
                        old_detail_name = "–Э–µ —Г–Ї–∞–Ј–∞–љ–∞"
                    changes_msg += f"рЯФІ –Ф–µ—В–∞–ї—М:\n"
                    changes_msg += f"  –°—В–∞—А–∞—П: {old_detail_name}\n"
                    changes_msg += f"  –Э–Њ–≤–∞—П: {new_detail_name}\n\n"

                if qty_changed:
                    changes_msg += f"рЯУ¶ –Ъ–Њ–ї–Є—З–µ—Б—В–≤–Њ:\n"
                    changes_msg += f"  –°—В–∞—А–Њ–µ: {old_qty} —И—В\n"
                    changes_msg += f"  –Э–Њ–≤–Њ–µ: {new_qty} —И—В\n"
                    changes_msg += f"  –†–∞–Ј–љ–Є—Ж–∞: {'+' if qty_difference > 0 else ''}{qty_difference} —И—В\n"
                    changes_msg += f"  –Э–Њ–≤—Л–є –Њ—Б—В–∞—В–Њ–Ї –Ї —Б–њ–Є—Б–∞–љ–Є—О: {new_qty - written_off} —И—В\n\n"

                changes_msg += "–Я—А–Њ–і–Њ–ї–ґ–Є—В—М?"

                if not messagebox.askyesno("–Я–Њ–і—В–≤–µ—А–ґ–і–µ–љ–Є–µ –Є–Ј–Љ–µ–љ–µ–љ–Є–є", changes_msg):
                    return

                # –Ю–±–љ–Њ–≤–ї—П–µ–Љ —А–µ–Ј–µ—А–≤
                new_remainder = new_qty - written_off
                reservations_df.loc[reservations_df["ID —А–µ–Ј–µ—А–≤–∞"] == reserve_id, "ID –Ј–∞–Ї–∞–Ј–∞"] = new_order_id
                reservations_df.loc[reservations_df["ID —А–µ–Ј–µ—А–≤–∞"] == reserve_id, "ID –і–µ—В–∞–ї–Є"] = new_detail_id
                reservations_df.loc[reservations_df["ID —А–µ–Ј–µ—А–≤–∞"] == reserve_id, "–Э–∞–Ј–≤–∞–љ–Є–µ –і–µ—В–∞–ї–Є"] = new_detail_name
                reservations_df.loc[reservations_df["ID —А–µ–Ј–µ—А–≤–∞"] == reserve_id, "–Ч–∞—А–µ–Ј–µ—А–≤–Є—А–Њ–≤–∞–љ–Њ —И—В—Г–Ї"] = new_qty
                reservations_df.loc[reservations_df["ID —А–µ–Ј–µ—А–≤–∞"] == reserve_id, "–Ю—Б—В–∞—В–Њ–Ї –Ї —Б–њ–Є—Б–∞–љ–Є—О"] = new_remainder
                save_data("Reservations", reservations_df)

                # –Ю–±–љ–Њ–≤–ї—П–µ–Љ –Љ–∞—В–µ—А–Є–∞–ї –љ–∞ —Б–Ї–ї–∞–і–µ (–µ—Б–ї–Є –Ї–Њ–ї–Є—З–µ—Б—В–≤–Њ –Є–Ј–Љ–µ–љ–Є–ї–Њ—Б—М –Є –љ–µ –≤—А—Г—З–љ—Г—О –і–Њ–±–∞–≤–ї–µ–љ–љ—Л–є)
                if qty_changed:
                    material_id = int(reserve_row["ID –Љ–∞—В–µ—А–Є–∞–ї–∞"])
                    if material_id != -1:
                        materials_df = load_data("Materials")
                        if not materials_df[materials_df["ID"] == material_id].empty:
                            mat_row = materials_df[materials_df["ID"] == material_id].iloc[0]
                            current_reserved = int(mat_row["–Ч–∞—А–µ–Ј–µ—А–≤–Є—А–Њ–≤–∞–љ–Њ"])
                            current_available = int(mat_row["–Ф–Њ—Б—В—Г–њ–љ–Њ"])

                            new_reserved = current_reserved + qty_difference
                            new_available = current_available - qty_difference

                            materials_df.loc[materials_df["ID"] == material_id, "–Ч–∞—А–µ–Ј–µ—А–≤–Є—А–Њ–≤–∞–љ–Њ"] = new_reserved
                            materials_df.loc[materials_df["ID"] == material_id, "–Ф–Њ—Б—В—Г–њ–љ–Њ"] = new_available
                            save_data("Materials", materials_df)
                            self.refresh_materials()

                self.refresh_reservations()
                self.refresh_balance()
                edit_window.destroy()

                result_msg = f"вЬЕ –†–µ–Ј–µ—А–≤ #{reserve_id} –Њ–±–љ–Њ–≤–ї–µ–љ!\n\n"
                if order_changed:
                    result_msg += "рЯУЛ –Ч–∞–Ї–∞–Ј –Є–Ј–Љ–µ–љ–µ–љ\n"
                if detail_changed:
                    result_msg += f"рЯФІ –Ф–µ—В–∞–ї—М –Є–Ј–Љ–µ–љ–µ–љ–∞ –љ–∞: {new_detail_name}\n"
                if qty_changed:
                    result_msg += f"рЯУ¶ –Ъ–Њ–ї–Є—З–µ—Б—В–≤–Њ: {new_qty} —И—В (–Њ—Б—В–∞—В–Њ–Ї: {new_remainder} —И—В)\n"

                messagebox.showinfo("–£—Б–њ–µ—Е", result_msg)

            except ValueError:
                messagebox.showerror("–Ю—И–Є–±–Ї–∞", "–Я—А–Њ–≤–µ—А—М—В–µ –њ—А–∞–≤–Є–ї—М–љ–Њ—Б—В—М –≤–≤–Њ–і–∞ —З–Є—Б–ї–Њ–≤—Л—Е –Ј–љ–∞—З–µ–љ–Є–є!")
            except Exception as e:
                messagebox.showerror("–Ю—И–Є–±–Ї–∞", f"–Э–µ —Г–і–∞–ї–Њ—Б—М –Њ–±–љ–Њ–≤–Є—В—М —А–µ–Ј–µ—А–≤: {e}")
                import traceback
                traceback.print_exc()

        tk.Button(edit_window, text="рЯТЊ –°–Њ—Е—А–∞–љ–Є—В—М –Є–Ј–Љ–µ–љ–µ–љ–Є—П", bg='#f39c12', fg='white',
                  font=("Arial", 12, "bold"), command=save_changes).pack(pady=15)

    def export_laser_task(self):
        """–§–Њ—А–Љ–Є—А–Њ–≤–∞–љ–Є–µ –Ј–∞–і–∞–љ–Є—П –љ–∞ –ї–∞–Ј–µ—А –Є–Ј —А–µ–Ј–µ—А–≤–Њ–≤"""
        try:
            # –Ч–∞–≥—А—Г–ґ–∞–µ–Љ –і–∞–љ–љ—Л–µ
            orders_df = load_data("Orders")
            reservations_df = load_data("Reservations")
            order_details_df = load_data("OrderDetails")

            if orders_df.empty:
                messagebox.showwarning("–Я—А–µ–і—Г–њ—А–µ–ґ–і–µ–љ–Є–µ", "–Э–µ—В –Ј–∞–Ї–∞–Ј–Њ–≤ –≤ –±–∞–Ј–µ!")
                return

            # –§–Є–ї—М—В—А—Г–µ–Љ –Ј–∞–Ї–∞–Ј—Л "–Т —А–∞–±–Њ—В–µ"
            active_orders = orders_df[orders_df["–°—В–∞—В—Г—Б"] == "–Т —А–∞–±–Њ—В–µ"]

            if active_orders.empty:
                messagebox.showwarning("–Я—А–µ–і—Г–њ—А–µ–ґ–і–µ–љ–Є–µ", "–Э–µ—В –Ј–∞–Ї–∞–Ј–Њ–≤ —Б–Њ —Б—В–∞—В—Г—Б–Њ–Љ '–Т —А–∞–±–Њ—В–µ'!")
                return

            # –Я—А–Њ–≤–µ—А—П–µ–Љ –љ–∞–ї–Є—З–Є–µ —А–µ–Ј–µ—А–≤–Њ–≤
            if reservations_df.empty:
                messagebox.showwarning("–Я—А–µ–і—Г–њ—А–µ–ґ–і–µ–љ–Є–µ", "–Э–µ—В –Ј–∞—А–µ–Ј–µ—А–≤–Є—А–Њ–≤–∞–љ–љ—Л—Е –Љ–∞—В–µ—А–Є–∞–ї–Њ–≤!")
                return

            # –Ю–Ї–љ–Њ –≤—Л–±–Њ—А–∞ –Ј–∞–Ї–∞–Ј–Њ–≤
            select_window = tk.Toplevel(self.root)
            select_window.title("–Т—Л–±–Њ—А –Ј–∞–Ї–∞–Ј–Њ–≤ –і–ї—П –Ј–∞–і–∞–љ–Є—П –љ–∞ –ї–∞–Ј–µ—А")
            select_window.geometry("700x600")
            select_window.configure(bg='#ecf0f1')

            tk.Label(select_window, text="–§–Њ—А–Љ–Є—А–Њ–≤–∞–љ–Є–µ –Ј–∞–і–∞–љ–Є—П –љ–∞ –ї–∞–Ј–µ—А",
                     font=("Arial", 14, "bold"), bg='#ecf0f1', fg='#e67e22').pack(pady=10)

            tk.Label(select_window, text="–Т—Л–±–µ—А–Є—В–µ –Ј–∞–Ї–∞–Ј—Л (—Б—В–∞—В—Г—Б: –Т —А–∞–±–Њ—В–µ)",
                     font=("Arial", 10), bg='#ecf0f1').pack(pady=5)

            # –§—А–µ–є–Љ —Б–Њ —Б–њ–Є—Б–Ї–Њ–Љ –Ј–∞–Ї–∞–Ј–Њ–≤
            list_frame = tk.Frame(select_window, bg='#ecf0f1')
            list_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)

            scroll_y = tk.Scrollbar(list_frame, orient=tk.VERTICAL)

            # –°–Њ–Ј–і–∞–µ–Љ Listbox —Б –Љ–љ–Њ–ґ–µ—Б—В–≤–µ–љ–љ—Л–Љ –≤—Л–±–Њ—А–Њ–Љ
            orders_listbox = tk.Listbox(list_frame, selectmode=tk.MULTIPLE,
                                        font=("Arial", 10), yscrollcommand=scroll_y.set)
            scroll_y.config(command=orders_listbox.yview)
            scroll_y.pack(side=tk.RIGHT, fill=tk.Y)
            orders_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

            # –Ч–∞–њ–Њ–ї–љ—П–µ–Љ —Б–њ–Є—Б–Њ–Ї –Ј–∞–Ї–∞–Ј–Њ–≤ "–Т —А–∞–±–Њ—В–µ"
            order_map = {}
            orders_without_reserves = []

            for _, order in active_orders.iterrows():
                order_id = order["ID –Ј–∞–Ї–∞–Ј–∞"]
                order_name = order["–Э–∞–Ј–≤–∞–љ–Є–µ –Ј–∞–Ї–∞–Ј–∞"]
                customer = order["–Ч–∞–Ї–∞–Ј—З–Є–Ї"]

                # –Я—А–Њ–≤–µ—А—П–µ–Љ –љ–∞–ї–Є—З–Є–µ —А–µ–Ј–µ—А–≤–Њ–≤
                has_reserves = not reservations_df[reservations_df["ID –Ј–∞–Ї–∞–Ј–∞"] == order_id].empty

                if has_reserves:
                    display_text = f"ID:{int(order_id)} | {customer} | {order_name}"
                    orders_listbox.insert(tk.END, display_text)
                    order_map[display_text] = order_id
                else:
                    orders_without_reserves.append(f"{customer} - {order_name}")

            if orders_listbox.size() == 0:
                messagebox.showwarning("–Я—А–µ–і—Г–њ—А–µ–ґ–і–µ–љ–Є–µ",
                                       "–Э–µ—В –Ј–∞–Ї–∞–Ј–Њ–≤ '–Т —А–∞–±–Њ—В–µ' —Б –Ј–∞—А–µ–Ј–µ—А–≤–Є—А–Њ–≤–∞–љ–љ—Л–Љ–Є –Љ–∞—В–µ—А–Є–∞–ї–∞–Љ–Є!")
                select_window.destroy()
                return

            # –Ъ–љ–Њ–њ–Ї–Є –≤—Л–±–Њ—А–∞
            btn_frame = tk.Frame(select_window, bg='#ecf0f1')
            btn_frame.pack(fill=tk.X, padx=20, pady=5)

            def select_all():
                orders_listbox.select_set(0, tk.END)

            def deselect_all():
                orders_listbox.select_clear(0, tk.END)

            tk.Button(btn_frame, text="–Т—Л–±—А–∞—В—М –≤—Б–µ", bg='#3498db', fg='white',
                      font=("Arial", 9), command=select_all).pack(side=tk.LEFT, padx=5)
            tk.Button(btn_frame, text="–°–љ—П—В—М –≤—Л–±–Њ—А", bg='#95a5a6', fg='white',
                      font=("Arial", 9), command=deselect_all).pack(side=tk.LEFT, padx=5)

            # –Ш–љ—Д–Њ—А–Љ–∞—Ж–Є—П
            info_frame = tk.Frame(select_window, bg='#d1ecf1', relief=tk.RIDGE, borderwidth=2)
            info_frame.pack(fill=tk.X, padx=20, pady=10)
            tk.Label(info_frame, text="–Ш–љ—Д–Њ—А–Љ–∞—Ж–Є—П:", font=("Arial", 9, "bold"),
                     bg='#d1ecf1', fg='#0c5460').pack(anchor='w', padx=5, pady=2)
            tk.Label(info_frame, text="- –Ю—В–Њ–±—А–∞–ґ–∞—О—В—Б—П —В–Њ–ї—М–Ї–Њ –Ј–∞–Ї–∞–Ј—Л —Б–Њ —Б—В–∞—В—Г—Б–Њ–Љ '–Т —А–∞–±–Њ—В–µ'",
                     font=("Arial", 8), bg='#d1ecf1', fg='#0c5460').pack(anchor='w', padx=10)
            tk.Label(info_frame, text="- –Ф–ї—П –Ї–∞–ґ–і–Њ–≥–Њ —А–µ–Ј–µ—А–≤–∞ —Б–Њ–Ј–і–∞–µ—В—Б—П –Њ—В–і–µ–ї—М–љ–∞—П —Б—В—А–Њ–Ї–∞",
                     font=("Arial", 8), bg='#d1ecf1', fg='#0c5460').pack(anchor='w', padx=10)
            tk.Label(info_frame, text="- –§–Њ—А–Љ–∞—В: –Ч–∞–Ї–∞–Ј—З–Є–Ї | –Э–∞–Ј–≤–∞–љ–Є–µ –Ј–∞—П–≤–Ї–Є | –Ф–µ—В–∞–ї—М | –Ь–µ—В–∞–ї–ї",
                     font=("Arial", 8), bg='#d1ecf1', fg='#0c5460').pack(anchor='w', padx=10)
            tk.Label(info_frame, text="- –Х—Б–ї–Є –і–µ—В–∞–ї—М –љ–µ –њ—А–Є–≤—П–Ј–∞–љ–∞ - '–С–µ–Ј —Г—З–µ—В–∞ –і–µ—В–∞–ї–µ–є'",
                     font=("Arial", 8), bg='#d1ecf1', fg='#0c5460').pack(anchor='w', padx=10)

            # –Я—А–µ–і—Г–њ—А–µ–ґ–і–µ–љ–Є–µ –Њ –Ј–∞–Ї–∞–Ј–∞—Е –±–µ–Ј —А–µ–Ј–µ—А–≤–Њ–≤
            if orders_without_reserves:
                warning_frame = tk.Frame(select_window, bg='#fff3cd', relief=tk.RIDGE, borderwidth=2)
                warning_frame.pack(fill=tk.X, padx=20, pady=5)
                tk.Label(warning_frame, text="–Т–љ–Є–Љ–∞–љ–Є–µ! –Ч–∞–Ї–∞–Ј—Л '–Т —А–∞–±–Њ—В–µ' –±–µ–Ј —А–µ–Ј–µ—А–≤–Њ–≤:",
                         font=("Arial", 8, "bold"), bg='#fff3cd', fg='#856404').pack(anchor='w', padx=5, pady=2)
                for order_name in orders_without_reserves[:3]:
                    tk.Label(warning_frame, text=f"  - {order_name}",
                             font=("Arial", 7), bg='#fff3cd', fg='#856404').pack(anchor='w', padx=10)
                if len(orders_without_reserves) > 3:
                    tk.Label(warning_frame, text=f"  ... –Є –µ—Й—С {len(orders_without_reserves) - 3}",
                             font=("Arial", 7), bg='#fff3cd', fg='#856404').pack(anchor='w', padx=10)

            def generate_file():
                selected_indices = orders_listbox.curselection()
                if not selected_indices:
                    messagebox.showwarning("–Я—А–µ–і—Г–њ—А–µ–ґ–і–µ–љ–Є–µ", "–Т—Л–±–µ—А–Є—В–µ —Е–Њ—В—П –±—Л –Њ–і–Є–љ –Ј–∞–Ї–∞–Ј!")
                    return

                # –Я–Њ–ї—Г—З–∞–µ–Љ –≤—Л–±—А–∞–љ–љ—Л–µ ID –Ј–∞–Ї–∞–Ј–Њ–≤
                selected_order_ids = []
                for index in selected_indices:
                    display_text = orders_listbox.get(index)
                    selected_order_ids.append(order_map[display_text])

                # –§–Њ—А–Љ–Є—А—Г–µ–Љ –і–∞–љ–љ—Л–µ –і–ї—П —Н–Ї—Б–њ–Њ—А—В–∞
                export_data = []
                warnings = []

                for order_id in selected_order_ids:
                    # –Я–Њ–ї—Г—З–∞–µ–Љ –Є–љ—Д–Њ—А–Љ–∞—Ж–Є—О –Њ –Ј–∞–Ї–∞–Ј–µ
                    order_row = orders_df[orders_df["ID –Ј–∞–Ї–∞–Ј–∞"] == order_id]
                    if order_row.empty:
                        continue

                    customer = order_row.iloc[0]["–Ч–∞–Ї–∞–Ј—З–Є–Ї"]
                    order_name = order_row.iloc[0]["–Э–∞–Ј–≤–∞–љ–Є–µ –Ј–∞–Ї–∞–Ј–∞"]

                    # –Я–Њ–ї—Г—З–∞–µ–Љ —А–µ–Ј–µ—А–≤—Л —Н—В–Њ–≥–Њ –Ј–∞–Ї–∞–Ј–∞
                    order_reserves = reservations_df[reservations_df["ID –Ј–∞–Ї–∞–Ј–∞"] == order_id]

                    if order_reserves.empty:
                        warnings.append(f"{customer} - {order_name}: –љ–µ—В —А–µ–Ј–µ—А–≤–Њ–≤")
                        continue

                    for _, reserve in order_reserves.iterrows():
                        # –§–Њ—А–Љ–Є—А—Г–µ–Љ –љ–∞–Ј–≤–∞–љ–Є–µ –і–µ—В–∞–ї–Є
                        detail_id = reserve.get("ID –і–µ—В–∞–ї–Є", -1)
                        detail_name = reserve.get("–Э–∞–Ј–≤–∞–љ–Є–µ –і–µ—В–∞–ї–Є", "–С–µ–Ј —Г—З–µ—В–∞ –і–µ—В–∞–ї–µ–є")

                        # –Я—А–Њ–≤–µ—А—П–µ–Љ –Ї–Њ—А—А–µ–Ї—В–љ–Њ—Б—В—М –њ—А–Є–≤—П–Ј–Ї–Є –і–µ—В–∞–ї–Є
                        if pd.isna(detail_name) or detail_name == "" or detail_name == "–Э–µ —Г–Ї–∞–Ј–∞–љ–∞" or detail_id == -1:
                            detail_name = "–С–µ–Ј —Г—З–µ—В–∞ –і–µ—В–∞–ї–µ–є"

                        # –§–Њ—А–Љ–Є—А—Г–µ–Љ –Њ–њ–Є—Б–∞–љ–Є–µ –Љ–µ—В–∞–ї–ї–∞
                        metal_str = f"{reserve['–Ь–∞—А–Ї–∞']} {reserve['–Ґ–Њ–ї—Й–Є–љ–∞']}–Љ–Љ {reserve['–®–Є—А–Є–љ–∞']}x{reserve['–Ф–ї–Є–љ–∞']}"

                        # рЯЖХ –Ю–С–™–Х–Ф–Ш–Э–ѓ–Х–Ь –Ч–Р–Ъ–Р–Ч–І–Ш–Ъ–Р –Ш –Э–Р–Ч–Т–Р–Э–Ш–Х –Ч–Р–ѓ–Т–Ъ–Ш –Т –Ю–Ф–Ш–Э –°–Ґ–Ю–Ы–С–Х–¶
                        combined_order = f"{customer} | {order_name}"

                        # –Ф–Њ–±–∞–≤–ї—П–µ–Љ —Б—В—А–Њ–Ї—Г
                        export_data.append({
                            "–Ч–∞–Ї–∞–Ј": combined_order,  # вЖР –Ю–С–™–Х–Ф–Ш–Э–Б–Э–Э–Ђ–Щ –°–Ґ–Ю–Ы–С–Х–¶
                            "–Э–∞–Ј–≤–∞–љ–Є–µ –і–µ—В–∞–ї–Є": detail_name,
                            "–Ь–µ—В–∞–ї–ї": metal_str
                        })

                if not export_data:
                    messagebox.showwarning("–Я—А–µ–і—Г–њ—А–µ–ґ–і–µ–љ–Є–µ", "–Э–µ—В –і–∞–љ–љ—Л—Е –і–ї—П —Н–Ї—Б–њ–Њ—А—В–∞!")
                    return

                # –Я—А–Њ–≤–µ—А—П–µ–Љ –љ–∞–ї–Є—З–Є–µ —Б—В—А–Њ–Ї "–С–µ–Ј —Г—З–µ—В–∞ –і–µ—В–∞–ї–µ–є"
                rows_without_details = sum(1 for row in export_data if row["–Э–∞–Ј–≤–∞–љ–Є–µ –і–µ—В–∞–ї–Є"] == "–С–µ–Ј —Г—З–µ—В–∞ –і–µ—В–∞–ї–µ–є")

                if rows_without_details > 0:
                    if not messagebox.askyesno("–Я—А–µ–і—Г–њ—А–µ–ґ–і–µ–љ–Є–µ",
                                               f"–Т —В–∞–±–ї–Є—Ж–µ –±—Г–і–µ—В {rows_without_details} —Б—В—А–Њ–Ї(–Є) –±–µ–Ј –њ—А–Є–≤—П–Ј–Ї–Є –Ї –і–µ—В–∞–ї—П–Љ!\n\n"
                                               "–≠—В–Њ –Љ–∞—В–µ—А–Є–∞–ї—Л, –Ј–∞—А–µ–Ј–µ—А–≤–Є—А–Њ–≤–∞–љ–љ—Л–µ –±–µ–Ј —Г–Ї–∞–Ј–∞–љ–Є—П –Ї–Њ–љ–Ї—А–µ—В–љ–Њ–є –і–µ—В–∞–ї–Є.\n\n"
                                               "–Я—А–Њ–і–Њ–ї–ґ–Є—В—М —Д–Њ—А–Љ–Є—А–Њ–≤–∞–љ–Є–µ?"):
                        return

                # –Ф–Є–∞–ї–Њ–≥ —Б–Њ—Е—А–∞–љ–µ–љ–Є—П —Д–∞–є–ї–∞
                file_path = filedialog.asksaveasfilename(
                    title="–°–Њ—Е—А–∞–љ–Є—В—М –Ј–∞–і–∞–љ–Є–µ –љ–∞ –ї–∞–Ј–µ—А",
                    defaultextension=".xlsx",
                    filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                    initialfile=f"zadanie_na_laser_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                )

                if not file_path:
                    return

                # –°–Њ–Ј–і–∞—С–Љ DataFrame –Є —Б–Њ—Е—А–∞–љ—П–µ–Љ
                export_df = pd.DataFrame(export_data)

                # –°–Њ—Е—А–∞–љ—П–µ–Љ —Б –∞–≤—В–Њ–њ–Њ–і–±–Њ—А–Њ–Љ —И–Є—А–Є–љ—Л
                with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                    export_df.to_excel(writer, index=False, sheet_name='–Ч–∞–і–∞–љ–Є–µ –љ–∞ –ї–∞–Ј–µ—А')
                    worksheet = writer.sheets['–Ч–∞–і–∞–љ–Є–µ –љ–∞ –ї–∞–Ј–µ—А']

                    # –Р–≤—В–Њ–њ–Њ–і–±–Њ—А —И–Є—А–Є–љ—Л –Ї–Њ–ї–Њ–љ–Њ–Ї
                    for column in worksheet.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 60)
                        worksheet.column_dimensions[column_letter].width = adjusted_width

                select_window.destroy()

                result_msg = f"–Ч–∞–і–∞–љ–Є–µ –љ–∞ –ї–∞–Ј–µ—А —Г—Б–њ–µ—И–љ–Њ —Б–Њ–Ј–і–∞–љ–Њ!\n\n"
                result_msg += f"–Ч–∞–Ї–∞–Ј–Њ–≤ –Њ–±—А–∞–±–Њ—В–∞–љ–Њ: {len(selected_order_ids)}\n"
                result_msg += f"–°—В—А–Њ–Ї –≤ —В–∞–±–ї–Є—Ж–µ: {len(export_data)}\n"
                result_msg += f"–°—В—А–Њ–Ї –±–µ–Ј –і–µ—В–∞–ї–µ–є: {rows_without_details}\n\n"
                result_msg += f"–§–∞–є–ї —Б–Њ—Е—А–∞–љ–µ–љ:\n{file_path}"

                messagebox.showinfo("–£—Б–њ–µ—Е", result_msg)

            # –Ъ–љ–Њ–њ–Ї–∞ —Д–Њ—А–Љ–Є—А–Њ–≤–∞–љ–Є—П
            tk.Button(select_window, text="–°—Д–Њ—А–Љ–Є—А–Њ–≤–∞—В—М —Д–∞–є–ї", bg='#e67e22', fg='white',
                      font=("Arial", 12, "bold"), command=generate_file).pack(pady=15)

        except Exception as e:
            messagebox.showerror("–Ю—И–Є–±–Ї–∞", f"–Э–µ —Г–і–∞–ї–Њ—Б—М —Б–Њ–Ј–і–∞—В—М –Ј–∞–і–∞–љ–Є–µ –љ–∞ –ї–∞–Ј–µ—А:\n{e}")
            import traceback
            traceback.print_exc()

    def setup_writeoffs_tab(self):
        """–Т–Ї–ї–∞–і–Ї–∞ —Б–њ–Є—Б–∞–љ–Є—П –Љ–∞—В–µ—А–Є–∞–ї–Њ–≤ - –†–£–І–Э–Ю–Х —Б–њ–Є—Б–∞–љ–Є–µ (—Б–Њ–≤–Љ–µ—Б—В–Є–Љ–∞ —Б –Є–Љ–њ–Њ—А—В–Њ–Љ –Њ—В –ї–∞–Ј–µ—А—Й–Є–Ї–Њ–≤)"""
        header = tk.Label(self.writeoffs_frame, text="–°–њ–Є—Б–∞–љ–Є–µ –Ј–∞—А–µ–Ј–µ—А–≤–Є—А–Њ–≤–∞–љ–љ—Л—Е –Љ–∞—В–µ—А–Є–∞–ї–Њ–≤",
                          font=("Arial", 16, "bold"), bg='white', fg='#2c3e50')
        header.pack(pady=10)

        tree_frame = tk.Frame(self.writeoffs_frame, bg='white')
        tree_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)

        scroll_y = tk.Scrollbar(tree_frame, orient=tk.VERTICAL)
        scroll_x = tk.Scrollbar(tree_frame, orient=tk.HORIZONTAL)

        self.writeoffs_tree = ttk.Treeview(tree_frame,
                                           columns=("ID", "ID —А–µ–Ј–µ—А–≤–∞", "–Ч–∞–Ї–∞–Ј", "–Ф–µ—В–∞–ї—М", "–Ь–∞—В–µ—А–Є–∞–ї", "–Ь–∞—А–Ї–∞",
                                                    "–Ґ–Њ–ї—Й–Є–љ–∞", "–†–∞–Ј–Љ–µ—А", "–Ъ–Њ–ї–Є—З–µ—Б—В–≤–Њ", "–Ф–∞—В–∞", "–Ъ–Њ–Љ–Љ–µ–љ—В–∞—А–Є–є"),
                                           show="headings", yscrollcommand=scroll_y.set, xscrollcommand=scroll_x.set)

        scroll_y.config(command=self.writeoffs_tree.yview)
        scroll_x.config(command=self.writeoffs_tree.xview)
        scroll_y.pack(side=tk.RIGHT, fill=tk.Y)
        scroll_x.pack(side=tk.BOTTOM, fill=tk.X)

        columns_config = {
            "ID": 50, "ID —А–µ–Ј–µ—А–≤–∞": 80, "–Ч–∞–Ї–∞–Ј": 200, "–Ф–µ—В–∞–ї—М": 150,
            "–Ь–∞—В–µ—А–Є–∞–ї": 80, "–Ь–∞—А–Ї–∞": 90, "–Ґ–Њ–ї—Й–Є–љ–∞": 70, "–†–∞–Ј–Љ–µ—А": 110,
            "–Ъ–Њ–ї–Є—З–µ—Б—В–≤–Њ": 90, "–Ф–∞—В–∞": 140, "–Ъ–Њ–Љ–Љ–µ–љ—В–∞—А–Є–є": 180
        }

        # –Э–∞—Б—В—А–Њ–є–Ї–∞ –Ї–Њ–ї–Њ–љ–Њ–Ї –С–Х–Ч —А–∞—Б—В—П–≥–Є–≤–∞–љ–Є—П (–Ї–∞–Ї –≤ –і—А—Г–≥–Є—Е –≤–Ї–ї–∞–і–Ї–∞—Е)
        for col, width in columns_config.items():
            self.writeoffs_tree.heading(col, text=col)
            self.writeoffs_tree.column(col, width=width, anchor=tk.CENTER, minwidth=80, stretch=False)

        self.writeoffs_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # рЯЖХ –Ш–Э–Ш–¶–Ш–Р–Ы–Ш–Ч–Р–¶–Ш–ѓ EXCEL-–§–Ш–Ы–ђ–Ґ–†–Р –Ф–Ы–ѓ –°–Я–Ш–°–Р–Э–Ш–Щ
        self.writeoffs_excel_filter = ExcelStyleFilter(
            tree=self.writeoffs_tree,
            refresh_callback=self.refresh_writeoffs
        )

        # рЯЖХ –Ш–Э–Ф–Ш–Ъ–Р–Ґ–Ю–† –Р–Ъ–Ґ–Ш–Т–Э–Ђ–• –§–Ш–Ы–ђ–Ґ–†–Ю–Т
        self.writeoffs_filter_status = tk.Label(
            self.writeoffs_frame,
            text="",
            font=("Arial", 9),
            bg='#d1ecf1',
            fg='#0c5460'
        )
        self.writeoffs_filter_status.pack(pady=5)

        # –Ъ–љ–Њ–њ–Ї–Є —Г–њ—А–∞–≤–ї–µ–љ–Є—П
        buttons_frame = tk.Frame(self.writeoffs_frame, bg='white')
        buttons_frame.pack(fill=tk.X, padx=10, pady=10)

        btn_style = {"font": ("Arial", 10), "width": 18, "height": 2}

        tk.Button(buttons_frame, text="–°–њ–Є—Б–∞—В—М –Љ–∞—В–µ—А–Є–∞–ї", bg='#e67e22', fg='white',
                  command=self.add_writeoff, **btn_style).pack(side=tk.LEFT, padx=5)

        tk.Button(buttons_frame, text="–£–і–∞–ї–Є—В—М —Б–њ–Є—Б–∞–љ–Є–µ", bg='#e74c3c', fg='white',
                  command=self.delete_writeoff, **btn_style).pack(side=tk.LEFT, padx=5)

        tk.Button(buttons_frame, text="–†–µ–і–∞–Ї—В–Є—А–Њ–≤–∞—В—М", bg='#f39c12', fg='white',
                  command=self.edit_writeoff, **btn_style).pack(side=tk.LEFT, padx=5)

        tk.Button(buttons_frame, text="–Ю–±–љ–Њ–≤–Є—В—М", bg='#95a5a6', fg='white',
                  command=self.refresh_writeoffs, **btn_style).pack(side=tk.LEFT, padx=5)

        tk.Button(buttons_frame, text="вЬЦ –°–±—А–Њ—Б–Є—В—М —Д–Є–ї—М—В—А—Л", bg='#e67e22', fg='white',
                  command=self.clear_writeoffs_filters, **btn_style).pack(side=tk.LEFT, padx=5)

        self.refresh_writeoffs()

    def refresh_writeoffs(self):
        """–Ю–±–љ–Њ–≤–ї–µ–љ–Є–µ —В–∞–±–ї–Є—Ж—Л —Б–њ–Є—Б–∞–љ–Є–є"""

        # –°–Ю–•–†–Р–Э–ѓ–Х–Ь –Р–Ъ–Ґ–Ш–Т–Э–Ђ–Х –§–Ш–Ы–ђ–Ґ–†–Ђ –Я–Х–†–Х–Ф –Ю–І–Ш–°–Ґ–Ъ–Ю–Щ
        active_filters_backup = {}
        if hasattr(self, 'writeoffs_excel_filter') and self.writeoffs_excel_filter.active_filters:
            active_filters_backup = self.writeoffs_excel_filter.active_filters.copy()

        # –Я–Ю–Ы–Э–Ю–°–Ґ–ђ–Ѓ –Ю–І–Ш–©–Р–Х–Ь –Ф–Х–†–Х–Т–Ю
        for i in self.writeoffs_tree.get_children():
            self.writeoffs_tree.delete(i)

        # –Ю–І–Ш–©–Р–Х–Ь –Ъ–≠–® –≠–Ы–Х–Ь–Х–Э–Ґ–Ю–Т
        if hasattr(self, 'writeoffs_excel_filter'):
            self.writeoffs_excel_filter._all_item_cache = set()

        writeoffs_df = load_data("WriteOffs")
        orders_df = load_data("Orders")
        reservations_df = load_data("Reservations")

        if not writeoffs_df.empty:
            for index, row in writeoffs_df.iterrows():
                # –Я–Њ–ї—Г—З–∞–µ–Љ –Є–љ—Д–Њ—А–Љ–∞—Ж–Є—О –Њ –Ј–∞–Ї–∞–Ј–µ
                order_id = int(row["ID –Ј–∞–Ї–∞–Ј–∞"])
                order_display = f"#{order_id}"

                if not orders_df.empty:
                    order_row = orders_df[orders_df["ID –Ј–∞–Ї–∞–Ј–∞"] == order_id]
                    if not order_row.empty:
                        customer = order_row.iloc[0]["–Ч–∞–Ї–∞–Ј—З–Є–Ї"]
                        order_name = order_row.iloc[0]["–Э–∞–Ј–≤–∞–љ–Є–µ –Ј–∞–Ї–∞–Ј–∞"]
                        order_display = f"{customer} | {order_name}"

                # –Я–Њ–ї—Г—З–∞–µ–Љ –Є–љ—Д–Њ—А–Љ–∞—Ж–Є—О –Њ –і–µ—В–∞–ї–Є –Є–Ј —А–µ–Ј–µ—А–≤–∞
                reserve_id = int(row["ID —А–µ–Ј–µ—А–≤–∞"])
                detail_display = "–С–µ–Ј –і–µ—В–∞–ї–Є"

                if not reservations_df.empty:
                    reserve_row = reservations_df[reservations_df["ID —А–µ–Ј–µ—А–≤–∞"] == reserve_id]
                    if not reserve_row.empty:
                        detail_name = reserve_row.iloc[0].get("–Э–∞–Ј–≤–∞–љ–Є–µ –і–µ—В–∞–ї–Є", "–С–µ–Ј –і–µ—В–∞–ї–Є")
                        detail_id = reserve_row.iloc[0].get("ID –і–µ—В–∞–ї–Є", -1)

                        if pd.notna(
                                detail_name) and detail_name != "" and detail_name != "–Э–µ —Г–Ї–∞–Ј–∞–љ–∞" and detail_id != -1:
                            detail_display = detail_name

                size_str = f"{row['–®–Є—А–Є–љ–∞']}x{row['–Ф–ї–Є–љ–∞']}"

                values = [
                    row["ID —Б–њ–Є—Б–∞–љ–Є—П"],
                    row["ID —А–µ–Ј–µ—А–≤–∞"],
                    order_display,
                    detail_display,
                    row["ID –Љ–∞—В–µ—А–Є–∞–ї–∞"],
                    row["–Ь–∞—А–Ї–∞"],
                    row["–Ґ–Њ–ї—Й–Є–љ–∞"],
                    size_str,
                    row["–Ъ–Њ–ї–Є—З–µ—Б—В–≤–Њ"],
                    row["–Ф–∞—В–∞ —Б–њ–Є—Б–∞–љ–Є—П"],
                    row["–Ъ–Њ–Љ–Љ–µ–љ—В–∞—А–Є–є"]
                ]

                item_id = self.writeoffs_tree.insert("", "end", values=values)

                # –°–Ю–•–†–Р–Э–ѓ–Х–Ь item_id –Т –Ъ–≠–®
                if hasattr(self, 'writeoffs_excel_filter'):
                    if not hasattr(self.writeoffs_excel_filter, '_all_item_cache'):
                        self.writeoffs_excel_filter._all_item_cache = set()
                    self.writeoffs_excel_filter._all_item_cache.add(item_id)

            # –Р–Т–Ґ–Ю–Я–Ю–Ф–С–Ю–† –®–Ш–†–Ш–Э–Ђ –Ъ–Ю–Ы–Ю–Э–Ю–Ъ (–Ї–∞–Ї –≤ –і—А—Г–≥–Є—Е –≤–Ї–ї–∞–і–Ї–∞—Е)
            self.auto_resize_columns(self.writeoffs_tree, min_width=80, max_width=300)

            # –Я–Х–†–Х–Я–†–Ш–Ь–Х–Э–ѓ–Х–Ь –§–Ш–Ы–ђ–Ґ–†–Ђ –Я–Ю–°–Ы–Х –Ч–Р–У–†–£–Ч–Ъ–Ш –Ф–Р–Э–Э–Ђ–•
            if active_filters_backup and hasattr(self, 'writeoffs_excel_filter'):
                self.writeoffs_excel_filter.active_filters = active_filters_backup
                self.writeoffs_excel_filter.reapply_all_filters()

    def add_writeoff(self):
        reservations_df = load_data("Reservations")
        if reservations_df.empty:
            messagebox.showwarning("–Я—А–µ–і—Г–њ—А–µ–ґ–і–µ–љ–Є–µ", "–Э–µ—В —А–µ–Ј–µ—А–≤–Њ–≤ –і–ї—П —Б–њ–Є—Б–∞–љ–Є—П!")
            return

        active_reserves = reservations_df[reservations_df["–Ю—Б—В–∞—В–Њ–Ї –Ї —Б–њ–Є—Б–∞–љ–Є—О"] > 0]
        if active_reserves.empty:
            messagebox.showwarning("–Я—А–µ–і—Г–њ—А–µ–ґ–і–µ–љ–Є–µ", "–Э–µ—В —А–µ–Ј–µ—А–≤–Њ–≤ —Б –Њ—Б—В–∞—В–Ї–Њ–Љ –і–ї—П —Б–њ–Є—Б–∞–љ–Є—П!")
            return

        add_window = tk.Toplevel(self.root)
        add_window.title("–°–њ–Є—Б–∞–љ–Є–µ –Љ–∞—В–µ—А–Є–∞–ї–∞")
        add_window.geometry("550x500")
        add_window.configure(bg='#ecf0f1')

        tk.Label(add_window, text="–°–њ–Є—Б–∞–љ–Є–µ –Љ–∞—В–µ—А–Є–∞–ї–∞ —Б —А–µ–Ј–µ—А–≤–∞", font=("Arial", 12, "bold"), bg='#ecf0f1').pack(
            pady=10)

        # –†–Х–Ч–Х–†–Т –° –Я–Ю–Ш–°–Ъ–Ю–Ь
        reserve_frame = tk.Frame(add_window, bg='#ecf0f1')
        reserve_frame.pack(fill=tk.X, padx=20, pady=5)
        tk.Label(reserve_frame, text="–†–µ–Ј–µ—А–≤ (–њ–Њ–Є—Б–Ї):", width=20, anchor='w', bg='#ecf0f1', font=("Arial", 10)).pack(
            side=tk.LEFT)

        all_reserve_options = []

        # –Ч–∞–≥—А—Г–ґ–∞–µ–Љ –Ј–∞–Ї–∞–Ј—Л –і–ї—П –Њ—В–Њ–±—А–∞–ґ–µ–љ–Є—П –Ј–∞–Ї–∞–Ј—З–Є–Ї–∞ –Є –љ–∞–Ј–≤–∞–љ–Є—П
        orders_df = load_data("Orders")

        for _, row in active_reserves.iterrows():
            order_id = int(row['ID –Ј–∞–Ї–∞–Ј–∞'])

            # –Ш—Й–µ–Љ –Є–љ—Д–Њ—А–Љ–∞—Ж–Є—О –Њ –Ј–∞–Ї–∞–Ј–µ
            order_info = ""
            if not orders_df.empty:
                order_row = orders_df[orders_df["ID –Ј–∞–Ї–∞–Ј–∞"] == order_id]
                if not order_row.empty:
                    customer = order_row.iloc[0]["–Ч–∞–Ї–∞–Ј—З–Є–Ї"]
                    order_name = order_row.iloc[0]["–Э–∞–Ј–≤–∞–љ–Є–µ –Ј–∞–Ї–∞–Ј–∞"]
                    order_info = f"{customer} | {order_name}"
                else:
                    order_info = f"–Ч–∞–Ї–∞–Ј #{order_id}"
            else:
                order_info = f"–Ч–∞–Ї–∞–Ј #{order_id}"

            # –Я–Њ–ї—Г—З–∞–µ–Љ –љ–∞–Ј–≤–∞–љ–Є–µ –і–µ—В–∞–ї–Є
            detail_name = row.get("–Э–∞–Ј–≤–∞–љ–Є–µ –і–µ—В–∞–ї–Є", "–С–µ–Ј —Г—З–µ—В–∞ –і–µ—В–∞–ї–µ–є")
            detail_id = row.get("ID –і–µ—В–∞–ї–Є", -1)

            # –Я—А–Њ–≤–µ—А—П–µ–Љ, –њ—А–Є–≤—П–Ј–∞–љ–∞ –ї–Є –і–µ—В–∞–ї—М
            if pd.isna(detail_name) or detail_name == "" or detail_name == "–Э–µ —Г–Ї–∞–Ј–∞–љ–∞" or detail_id == -1:
                detail_info = "–С–µ–Ј –і–µ—В–∞–ї–Є"
            else:
                detail_info = f"–Ф–µ—В–∞–ї—М: {detail_name}"

            # –§–Њ—А–Љ–Є—А—Г–µ–Љ —Б—В—А–Њ–Ї—Г —Б –Є–љ—Д–Њ—А–Љ–∞—Ж–Є–µ–є –Њ –і–µ—В–∞–ї–Є
            reserve_str = f"–†–µ–Ј–µ—А–≤ #{int(row['ID —А–µ–Ј–µ—А–≤–∞'])} | {order_info} | {detail_info} | {row['–Ь–∞—А–Ї–∞']} {row['–Ґ–Њ–ї—Й–Є–љ–∞']}–Љ–Љ | –Ю—Б—В–∞–ї–Њ—Б—М: {int(row['–Ю—Б—В–∞—В–Њ–Ї –Ї —Б–њ–Є—Б–∞–љ–Є—О'])} —И—В"
            all_reserve_options.append(reserve_str)

        search_container = tk.Frame(reserve_frame, bg='#ecf0f1')
        search_container.pack(side=tk.RIGHT, expand=True, fill=tk.X)

        reserve_search_var = tk.StringVar()
        selected_reserve = {"value": None}

        reserve_search_entry = tk.Entry(search_container, textvariable=reserve_search_var, font=("Arial", 10))
        reserve_search_entry.pack(fill=tk.X)

        # Listbox –і–ї—П —А–µ–Ј—Г–ї—М—В–∞—В–Њ–≤ –њ–Њ–Є—Б–Ї–∞
        search_results_frame = tk.Frame(add_window, bg='#ecf0f1')
        search_results_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=5)

        scroll_results = tk.Scrollbar(search_results_frame, orient=tk.VERTICAL)
        results_listbox = tk.Listbox(search_results_frame, height=8, font=("Arial", 9),
                                     yscrollcommand=scroll_results.set)
        scroll_results.config(command=results_listbox.yview)
        scroll_results.pack(side=tk.RIGHT, fill=tk.Y)
        results_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        for option in all_reserve_options:
            results_listbox.insert(tk.END, option)

        def on_search_change(*args):
            search_text = reserve_search_var.get().lower()
            results_listbox.delete(0, tk.END)
            for option in all_reserve_options:
                if search_text in option.lower():
                    results_listbox.insert(tk.END, option)

        def on_select_reserve(event):
            try:
                selection = results_listbox.get(results_listbox.curselection())
                selected_reserve["value"] = selection
                reserve_search_var.set(selection)
            except:
                pass

        reserve_search_var.trace('w', on_search_change)
        results_listbox.bind('<<ListboxSelect>>', on_select_reserve)
        results_listbox.bind('<Double-Button-1>', on_select_reserve)

        # –Ъ–Њ–ї–Є—З–µ—Б—В–≤–Њ
        qty_frame = tk.Frame(add_window, bg='#ecf0f1')
        qty_frame.pack(fill=tk.X, padx=20, pady=5)
        tk.Label(qty_frame, text="–Ъ–Њ–ї–Є—З–µ—Б—В–≤–Њ (—И—В):", width=20, anchor='w', bg='#ecf0f1', font=("Arial", 10)).pack(
            side=tk.LEFT)
        qty_entry = tk.Entry(qty_frame, font=("Arial", 10))
        qty_entry.pack(side=tk.RIGHT, expand=True, fill=tk.X)

        # –Ъ–Њ–Љ–Љ–µ–љ—В–∞—А–Є–є
        comment_frame = tk.Frame(add_window, bg='#ecf0f1')
        comment_frame.pack(fill=tk.X, padx=20, pady=5)
        tk.Label(comment_frame, text="–Ъ–Њ–Љ–Љ–µ–љ—В–∞—А–Є–є:", width=20, anchor='w', bg='#ecf0f1', font=("Arial", 10)).pack(
            side=tk.LEFT)
        comment_entry = tk.Entry(comment_frame, font=("Arial", 10))
        comment_entry.pack(side=tk.RIGHT, expand=True, fill=tk.X)

        def save_writeoff():
            try:
                reserve_value = selected_reserve["value"] or reserve_search_var.get()
                if not reserve_value:
                    messagebox.showwarning("–Я—А–µ–і—Г–њ—А–µ–ґ–і–µ–љ–Є–µ", "–Т—Л–±–µ—А–Є—В–µ —А–µ–Ј–µ—А–≤!")
                    return

                # –Я–∞—А—Б–Є–Љ ID –Є–Ј —Д–Њ—А–Љ–∞—В–∞ "–†–µ–Ј–µ—А–≤ #123 | ..."
                reserve_id = int(reserve_value.split("–†–µ–Ј–µ—А–≤ #")[1].split(" | ")[0])
                quantity = int(qty_entry.get())
                comment = comment_entry.get().strip()

                # –Я—А–Њ–≤–µ—А—П–µ–Љ —А–µ–Ј–µ—А–≤
                reservations_df = load_data("Reservations")
                reservation = reservations_df[reservations_df["ID —А–µ–Ј–µ—А–≤–∞"] == reserve_id].iloc[0]
                remainder = int(reservation["–Ю—Б—В–∞—В–Њ–Ї –Ї —Б–њ–Є—Б–∞–љ–Є—О"])

                if quantity > remainder:
                    messagebox.showerror("–Ю—И–Є–±–Ї–∞", f"–Э–µ–ї—М–Ј—П —Б–њ–Є—Б–∞—В—М –±–Њ–ї—М—И–µ —З–µ–Љ –Њ—Б—В–∞–ї–Њ—Б—М!\n–Ю—Б—В–∞–ї–Њ—Б—М: {remainder} —И—В")
                    return

                # –Ф–Њ–±–∞–≤–ї—П–µ–Љ —Б–њ–Є—Б–∞–љ–Є–µ
                writeoffs_df = load_data("WriteOffs")
                new_id = 1 if writeoffs_df.empty else int(writeoffs_df["ID —Б–њ–Є—Б–∞–љ–Є—П"].max()) + 1

                new_row = pd.DataFrame([{
                    "ID —Б–њ–Є—Б–∞–љ–Є—П": new_id,
                    "ID —А–µ–Ј–µ—А–≤–∞": reserve_id,
                    "ID –Ј–∞–Ї–∞–Ј–∞": reservation["ID –Ј–∞–Ї–∞–Ј–∞"],
                    "ID –Љ–∞—В–µ—А–Є–∞–ї–∞": reservation["ID –Љ–∞—В–µ—А–Є–∞–ї–∞"],
                    "–Ь–∞—А–Ї–∞": reservation["–Ь–∞—А–Ї–∞"],
                    "–Ґ–Њ–ї—Й–Є–љ–∞": reservation["–Ґ–Њ–ї—Й–Є–љ–∞"],
                    "–Ф–ї–Є–љ–∞": reservation["–Ф–ї–Є–љ–∞"],
                    "–®–Є—А–Є–љ–∞": reservation["–®–Є—А–Є–љ–∞"],
                    "–Ъ–Њ–ї–Є—З–µ—Б—В–≤–Њ": quantity,
                    "–Ф–∞—В–∞ —Б–њ–Є—Б–∞–љ–Є—П": datetime.now().strftime("%Y-%m-%d"),
                    "–Ъ–Њ–Љ–Љ–µ–љ—В–∞—А–Є–є": comment
                }])

                writeoffs_df = pd.concat([writeoffs_df, new_row], ignore_index=True)
                save_data("WriteOffs", writeoffs_df)

                # –Ю–±–љ–Њ–≤–ї—П–µ–Љ —А–µ–Ј–µ—А–≤–Є—А–Њ–≤–∞–љ–Є–µ
                reservations_df = load_data("Reservations")
                reservation = reservations_df[reservations_df["ID —А–µ–Ј–µ—А–≤–∞"] == reserve_id].iloc[0]

                new_written_off = int(reservation["–°–њ–Є—Б–∞–љ–Њ"]) + quantity
                new_remainder = int(reservation["–Ч–∞—А–µ–Ј–µ—А–≤–Є—А–Њ–≤–∞–љ–Њ —И—В—Г–Ї"]) - new_written_off

                reservations_df.loc[reservations_df["ID —А–µ–Ј–µ—А–≤–∞"] == reserve_id, "–°–њ–Є—Б–∞–љ–Њ"] = new_written_off
                reservations_df.loc[reservations_df["ID —А–µ–Ј–µ—А–≤–∞"] == reserve_id, "–Ю—Б—В–∞—В–Њ–Ї –Ї —Б–њ–Є—Б–∞–љ–Є—О"] = new_remainder
                save_data("Reservations", reservations_df)

                # –Ю–±–љ–Њ–≤–ї—П–µ–Љ –Љ–∞—В–µ—А–Є–∞–ї (–Ш–°–Я–†–Р–Т–Ы–Х–Э–Ю: —Г–Љ–µ–љ—М—И–∞–µ–Љ –Ш –љ–∞–ї–Є—З–Є–µ –Ш —А–µ–Ј–µ—А–≤)
                material_id = int(reservation["ID –Љ–∞—В–µ—А–Є–∞–ї–∞"])
                if material_id != -1:
                    materials_df = load_data("Materials")
                    material = materials_df[materials_df["ID"] == material_id].iloc[0]

                    # –£–Љ–µ–љ—М—И–∞–µ–Љ –Ї–Њ–ї–Є—З–µ—Б—В–≤–Њ –≤ –љ–∞–ї–Є—З–Є–Є
                    new_qty = int(material["–Ъ–Њ–ї–Є—З–µ—Б—В–≤–Њ —И—В—Г–Ї"]) - quantity

                    # –£–Љ–µ–љ—М—И–∞–µ–Љ –Ј–∞—А–µ–Ј–µ—А–≤–Є—А–Њ–≤–∞–љ–Њ
                    new_reserved = int(material["–Ч–∞—А–µ–Ј–µ—А–≤–Є—А–Њ–≤–∞–љ–Њ"]) - quantity

                    # –Ф–Њ—Б—В—Г–њ–љ–Њ –Э–Х –Љ–µ–љ—П–µ—В—Б—П (—В.–Ї. –±—Л–ї–Њ —Г–ґ–µ –Ј–∞—А–µ–Ј–µ—А–≤–Є—А–Њ–≤–∞–љ–Њ)

                    materials_df.loc[materials_df["ID"] == material_id, "–Ъ–Њ–ї–Є—З–µ—Б—В–≤–Њ —И—В—Г–Ї"] = new_qty
                    materials_df.loc[materials_df["ID"] == material_id, "–Ч–∞—А–µ–Ј–µ—А–≤–Є—А–Њ–≤–∞–љ–Њ"] = new_reserved

                    # –Я–µ—А–µ—Б—З–Є—В—Л–≤–∞–µ–Љ –њ–ї–Њ—Й–∞–і—М
                    area_per_piece = float(material["–Ф–ї–Є–љ–∞"]) * float(material["–®–Є—А–Є–љ–∞"]) / 1_000_000
                    new_area = new_qty * area_per_piece
                    materials_df.loc[materials_df["ID"] == material_id, "–Ю–±—Й–∞—П –њ–ї–Њ—Й–∞–і—М"] = round(new_area, 2)

                    save_data("Materials", materials_df)
                    self.refresh_materials()

                self.refresh_reservations()
                self.refresh_writeoffs()
                self.refresh_balance()
                add_window.destroy()
                messagebox.showinfo("–£—Б–њ–µ—Е", f"вЬЕ –°–њ–Є—Б–∞–љ–Є–µ #{new_id} —Г—Б–њ–µ—И–љ–Њ —Б–Њ–Ј–і–∞–љ–Њ!\n–°–њ–Є—Б–∞–љ–Њ: {quantity} —И—В")

            except ValueError:
                messagebox.showerror("–Ю—И–Є–±–Ї–∞", "–Я—А–Њ–≤–µ—А—М—В–µ –њ—А–∞–≤–Є–ї—М–љ–Њ—Б—В—М –≤–≤–Њ–і–∞ —З–Є—Б–ї–Њ–≤—Л—Е –Ј–љ–∞—З–µ–љ–Є–є!")
            except Exception as e:
                messagebox.showerror("–Ю—И–Є–±–Ї–∞", f"–Э–µ —Г–і–∞–ї–Њ—Б—М —Б–Њ–Ј–і–∞—В—М —Б–њ–Є—Б–∞–љ–Є–µ: {e}")
                import traceback
                traceback.print_exc()

        tk.Button(add_window, text="–°–њ–Є—Б–∞—В—М", bg='#e74c3c', fg='white', font=("Arial", 12, "bold"),
                  command=save_writeoff).pack(pady=15)

    def delete_writeoff(self):
        """–£–і–∞–ї–µ–љ–Є–µ –Ј–∞–њ–Є—Б–Є –Њ —Б–њ–Є—Б–∞–љ–Є–Є (–Њ—В–Љ–µ–љ–∞ —Б–њ–Є—Б–∞–љ–Є—П)"""
        selected = self.writeoffs_tree.selection()

        if not selected:
            messagebox.showwarning("–Я—А–µ–і—Г–њ—А–µ–ґ–і–µ–љ–Є–µ", "–Т—Л–±–µ—А–Є—В–µ —Б–њ–Є—Б–∞–љ–Є–µ –і–ї—П —Г–і–∞–ї–µ–љ–Є—П!")
            return

        try:
            values = self.writeoffs_tree.item(selected[0])['values']
            writeoff_id = int(values[0])
            reserve_id = int(values[1])
            comment = values[9] if len(values) > 9 else ""

            info_msg = (
                f"–Ю—В–Љ–µ–љ–Є—В—М —Б–њ–Є—Б–∞–љ–Є–µ?\n\n"
                f"ID —Б–њ–Є—Б–∞–љ–Є—П: {writeoff_id}\n"
                f"ID —А–µ–Ј–µ—А–≤–∞: {reserve_id}\n\n"
                f"вЪ†пЄП –≠—В–Њ –і–µ–є—Б—В–≤–Є–µ:\n"
                f"вАҐ –Т–µ—А–љ—С—В –Љ–∞—В–µ—А–Є–∞–ї –≤ —А–µ–Ј–µ—А–≤\n"
                f"вАҐ –Т–µ—А–љ—С—В –Љ–∞—В–µ—А–Є–∞–ї –љ–∞ —Б–Ї–ї–∞–і\n"
                f"вАҐ –Ю–±–љ–Њ–≤–Є—В —В–∞–±–ї–Є—Ж—Г –Є–Љ–њ–Њ—А—В–∞ –Њ—В –ї–∞–Ј–µ—А—Й–Є–Ї–Њ–≤"
            )

            if not messagebox.askyesno("–Я–Њ–і—В–≤–µ—А–ґ–і–µ–љ–Є–µ", info_msg):
                return

            writeoffs_df = load_data("WriteOffs")
            reservations_df = load_data("Reservations")
            materials_df = load_data("Materials")
            order_details_df = load_data("OrderDetails")

            writeoff_row = writeoffs_df[writeoffs_df["ID —Б–њ–Є—Б–∞–љ–Є—П"] == writeoff_id]

            if writeoff_row.empty:
                messagebox.showerror("–Ю—И–Є–±–Ї–∞", f"–°–њ–Є—Б–∞–љ–Є–µ ID={writeoff_id} –љ–µ –љ–∞–є–і–µ–љ–Њ!")
                return

            writeoff_row = writeoff_row.iloc[0]

            reserve_id = int(writeoff_row["ID —А–µ–Ј–µ—А–≤–∞"])
            quantity = int(writeoff_row["–Ъ–Њ–ї–Є—З–µ—Б—В–≤–Њ"])
            material_id = int(writeoff_row["ID –Љ–∞—В–µ—А–Є–∞–ї–∞"])
            writeoff_date = str(writeoff_row["–Ф–∞—В–∞ —Б–њ–Є—Б–∞–љ–Є—П"])
            writeoff_comment = str(writeoff_row["–Ъ–Њ–Љ–Љ–µ–љ—В–∞—А–Є–є"])

            # –Ю–С–Э–Ю–Т–Ы–Х–Э–Ш–Х –†–Х–Ч–Х–†–Т–Р
            reserve_row = reservations_df[reservations_df["ID —А–µ–Ј–µ—А–≤–∞"] == reserve_id]

            if reserve_row.empty:
                messagebox.showerror("–Ю—И–Є–±–Ї–∞", f"–†–µ–Ј–µ—А–≤ ID={reserve_id} –љ–µ –љ–∞–є–і–µ–љ!")
                return

            reserve_row = reserve_row.iloc[0]
            old_written_off = int(reserve_row["–°–њ–Є—Б–∞–љ–Њ"])
            old_remainder = int(reserve_row["–Ю—Б—В–∞—В–Њ–Ї –Ї —Б–њ–Є—Б–∞–љ–Є—О"])

            new_written_off = old_written_off - quantity
            new_remainder = old_remainder + quantity

            reservations_df.loc[reservations_df["ID —А–µ–Ј–µ—А–≤–∞"] == reserve_id, "–°–њ–Є—Б–∞–љ–Њ"] = new_written_off
            reservations_df.loc[reservations_df["ID —А–µ–Ј–µ—А–≤–∞"] == reserve_id, "–Ю—Б—В–∞—В–Њ–Ї –Ї —Б–њ–Є—Б–∞–љ–Є—О"] = new_remainder

            # –Ю–С–Э–Ю–Т–Ы–Х–Э–Ш–Х –Ь–Р–Ґ–Х–†–Ш–Р–Ы–Р
            if material_id != -1:
                material = materials_df[materials_df["ID"] == material_id]

                if not material.empty:
                    material = material.iloc[0]

                    old_qty = int(material["–Ъ–Њ–ї–Є—З–µ—Б—В–≤–Њ —И—В—Г–Ї"])
                    old_reserved = int(material["–Ч–∞—А–µ–Ј–µ—А–≤–Є—А–Њ–≤–∞–љ–Њ"])

                    new_qty = old_qty + quantity
                    new_reserved = old_reserved + quantity

                    materials_df.loc[materials_df["ID"] == material_id, "–Ъ–Њ–ї–Є—З–µ—Б—В–≤–Њ —И—В—Г–Ї"] = new_qty
                    materials_df.loc[materials_df["ID"] == material_id, "–Ч–∞—А–µ–Ј–µ—А–≤–Є—А–Њ–≤–∞–љ–Њ"] = new_reserved

                    area_per_piece = float(material["–Ф–ї–Є–љ–∞"]) * float(material["–®–Є—А–Є–љ–∞"]) / 1_000_000
                    new_area = new_qty * area_per_piece
                    materials_df.loc[materials_df["ID"] == material_id, "–Ю–±—Й–∞—П –њ–ї–Њ—Й–∞–і—М"] = round(new_area, 2)

            # –Ю–С–Э–Ю–Т–Ы–Х–Э–Ш–Х –Ґ–Р–С–Ы–Ш–¶–Ђ –Ш–Ь–Я–Ю–†–Ґ–Р
            is_laser_import = "–Ы–∞–Ј–µ—А:" in writeoff_comment or "–ї–∞–Ј–µ—А—Й–Є–Ї" in writeoff_comment.lower()

            if is_laser_import and hasattr(self, 'laser_table_data') and self.laser_table_data:
                import re

                part_name = None
                parts_qty = None

                part_match = re.search(r'–Ф–µ—В–∞–ї—М:\s*([^|]+)', writeoff_comment)
                if part_match:
                    part_name = part_match.group(1).strip()

                date_match = re.search(r'–Ф–∞—В–∞ –Є–Љ–њ–Њ—А—В–∞:\s*(.+)', writeoff_comment)
                import_date_str = date_match.group(1).strip() if date_match else None

                for idx, row_data in enumerate(self.laser_table_data):
                    row_part = str(row_data.get("part", ""))

                    if part_name and (part_name.lower() in row_part.lower() or row_part.lower() in part_name.lower()):
                        row_date = str(row_data.get("–Ф–∞—В–∞ (–Ь–°–Ъ)", ""))
                        row_time = str(row_data.get("–Т—А–µ–Љ—П (–Ь–°–Ъ)", ""))
                        row_datetime = f"{row_date} {row_time}"

                        date_match_found = False
                        if import_date_str and len(row_datetime) >= 16 and len(import_date_str) >= 16:
                            if row_datetime[:16] == import_date_str[:16]:
                                date_match_found = True
                        elif not import_date_str:
                            row_writeoff_date = row_data.get("–Ф–∞—В–∞ —Б–њ–Є—Б–∞–љ–Є—П", "")

                            # –С–µ–Ј–Њ–њ–∞—Б–љ–Њ–µ –њ—А–µ–Њ–±—А–∞–Ј–Њ–≤–∞–љ–Є–µ
                            if pd.isna(row_writeoff_date) or row_writeoff_date is None:
                                row_writeoff_date = ""
                            else:
                                row_writeoff_date = str(row_writeoff_date)

                            if len(row_writeoff_date) >= 16 and len(writeoff_date) >= 16:
                                if row_writeoff_date[:16] == writeoff_date[:16]:
                                    date_match_found = True

                        if date_match_found:
                            try:
                                parts_qty = int(row_data.get("part_quantity", 0))
                                self.laser_table_data[idx]["–°–њ–Є—Б–∞–љ–Њ"] = ""
                                self.laser_table_data[idx]["–Ф–∞—В–∞ —Б–њ–Є—Б–∞–љ–Є—П"] = ""
                            except:
                                pass
                            break

                if hasattr(self, 'laser_import_tree'):
                    self.refresh_laser_import_table()
                    try:
                        self.save_laser_import_cache()
                    except:
                        pass

            # –Ю–Ґ–Ъ–Р–Ґ –Ф–Х–Ґ–Р–Ы–Х–Щ
            if is_laser_import and "–Ф–µ—В–∞–ї—М:" in writeoff_comment and parts_qty:
                try:
                    import re
                    part_match = re.search(r'–Ф–µ—В–∞–ї—М:\s*([^|]+)', writeoff_comment)

                    if part_match and parts_qty > 0:
                        part_name = part_match.group(1).strip()
                        order_id = int(writeoff_row["ID –Ј–∞–Ї–∞–Ј–∞"])

                        detail_match = order_details_df[
                            (order_details_df["ID –Ј–∞–Ї–∞–Ј–∞"] == order_id) &
                            (order_details_df["–Э–∞–Ј–≤–∞–љ–Є–µ –і–µ—В–∞–ї–Є"].str.contains(part_name, case=False, na=False))
                            ]

                        if not detail_match.empty:
                            detail_id = int(detail_match.iloc[0]["ID"])
                            old_cut = int(detail_match.iloc[0].get("–Я–Њ—А–µ–Ј–∞–љ–Њ", 0))
                            new_cut = max(0, old_cut - parts_qty)
                            order_details_df.loc[order_details_df["ID"] == detail_id, "–Я–Њ—А–µ–Ј–∞–љ–Њ"] = new_cut
                            save_data("OrderDetails", order_details_df)
                except:
                    pass

            # –£–Ф–Р–Ы–Х–Э–Ш–Х –°–Я–Ш–°–Р–Э–Ш–ѓ
            writeoffs_df = writeoffs_df[writeoffs_df["ID —Б–њ–Є—Б–∞–љ–Є—П"] != writeoff_id]

            # –°–Ю–•–†–Р–Э–Х–Э–Ш–Х
            save_data("WriteOffs", writeoffs_df)
            save_data("Reservations", reservations_df)
            save_data("Materials", materials_df)

            # –Ю–С–Э–Ю–Т–Ы–Х–Э–Ш–Х –Ш–Э–Ґ–Х–†–§–Х–Щ–°–Р
            self.refresh_writeoffs()
            self.refresh_reservations()
            self.refresh_materials()
            self.refresh_balance()

            if hasattr(self, 'refresh_orders'):
                self.refresh_orders()
            if hasattr(self, 'refresh_order_details'):
                self.refresh_order_details()

            messagebox.showinfo("–£—Б–њ–µ—Е",
                                f"вЬЕ –°–њ–Є—Б–∞–љ–Є–µ –Њ—В–Љ–µ–љ–µ–љ–Њ!\n\n"
                                f"–Т–Њ–Ј–≤—А–∞—Й–µ–љ–Њ –≤ —А–µ–Ј–µ—А–≤: {quantity} —И—В\n"
                                f"–†–µ–Ј–µ—А–≤ ID: {reserve_id}\n"
                                f"–Ю—Б—В–∞—В–Њ–Ї –Ї —Б–њ–Є—Б–∞–љ–Є—О: {new_remainder} —И—В")

        except Exception as e:
            messagebox.showerror("–Ю—И–Є–±–Ї–∞", f"–Э–µ —Г–і–∞–ї–Њ—Б—М –Њ—В–Љ–µ–љ–Є—В—М —Б–њ–Є—Б–∞–љ–Є–µ:\n{e}")
            import traceback
            traceback.print_exc()

    def find_laser_import_row_by_writeoff(self, writeoff_data):
        """
        –Я–Њ–Є—Б–Ї —Б—В—А–Њ–Ї–Є –≤ —В–∞–±–ї–Є—Ж–µ –Є–Љ–њ–Њ—А—В–∞ –њ–Њ –і–∞–љ–љ—Л–Љ —Б–њ–Є—Б–∞–љ–Є—П

        Args:
            writeoff_data: dict —Б –Ї–ї—О—З–∞–Љ–Є '–Ф–∞—В–∞ —Б–њ–Є—Б–∞–љ–Є—П', '–Ъ–Њ–Љ–Љ–µ–љ—В–∞—А–Є–є', '–Ъ–Њ–ї–Є—З–µ—Б—В–≤–Њ'

        Returns:
            list: –Є–љ–і–µ–Ї—Б—Л –љ–∞–є–і–µ–љ–љ—Л—Е —Б—В—А–Њ–Ї –≤ laser_table_data
        """
        if not hasattr(self, 'laser_table_data') or not self.laser_table_data:
            return []

        writeoff_date = writeoff_data.get('–Ф–∞—В–∞ —Б–њ–Є—Б–∞–љ–Є—П', '')
        writeoff_comment = writeoff_data.get('–Ъ–Њ–Љ–Љ–µ–љ—В–∞—А–Є–є', '')
        writeoff_qty = writeoff_data.get('–Ъ–Њ–ї–Є—З–µ—Б—В–≤–Њ', 0)

        # –Ш–Ј–≤–ї–µ–Ї–∞–µ–Љ –Є–љ—Д–Њ—А–Љ–∞—Ж–Є—О –Є–Ј –Ї–Њ–Љ–Љ–µ–љ—В–∞—А–Є—П
        # –§–Њ—А–Љ–∞—В: "–Ы–∞–Ј–µ—А: @username | –Ф–µ—В–∞–ї—М: –љ–∞–Ј–≤–∞–љ–Є–µ_–і–µ—В–∞–ї–Є"
        import re
        username_match = re.search(r'–Ы–∞–Ј–µ—А:\s*(@?\w+)', writeoff_comment)
        part_match = re.search(r'–Ф–µ—В–∞–ї—М:\s*(.+?)(?:\||$)', writeoff_comment)

        username = username_match.group(1) if username_match else None
        part_name = part_match.group(1).strip() if part_match else None

        print(f"   рЯФН –Ъ—А–Є—В–µ—А–Є–Є –њ–Њ–Є—Б–Ї–∞:")
        print(f"      –Ф–∞—В–∞: {writeoff_date}")
        print(f"      –Я–Њ–ї—М–Ј–Њ–≤–∞—В–µ–ї—М: {username}")
        print(f"      –Ф–µ—В–∞–ї—М: {part_name}")
        print(f"      –Ъ–Њ–ї–Є—З–µ—Б—В–≤–Њ: {writeoff_qty}")

        matching_indices = []

        for idx, row_data in enumerate(self.laser_table_data):
            # –Я—А–Њ–≤–µ—А—П–µ–Љ —В–Њ–ї—М–Ї–Њ —Б–њ–Є—Б–∞–љ–љ—Л–µ —Б—В—А–Њ–Ї–Є
            if row_data.get("–°–њ–Є—Б–∞–љ–Њ") not in ["вЬУ", "–Ф–∞", "Yes"]:
                continue

            match_score = 0

            # –°–Њ–њ–Њ—Б—В–∞–≤–ї–µ–љ–Є–µ –њ–Њ –і–∞—В–µ —Б–њ–Є—Б–∞–љ–Є—П (–њ—А–Є–Њ—А–Є—В–µ—В 3)
            row_writeoff_date = row_data.get("–Ф–∞—В–∞ —Б–њ–Є—Б–∞–љ–Є—П", "")
            if row_writeoff_date and writeoff_date:
                # –°—А–∞–≤–љ–Є–≤–∞–µ–Љ –њ–µ—А–≤—Л–µ 16 —Б–Є–Љ–≤–Њ–ї–Њ–≤ (–і–∞—В–∞ + –≤—А–µ–Љ—П –±–µ–Ј —Б–µ–Ї—Г–љ–і)
                if row_writeoff_date[:16] == writeoff_date[:16]:
                    match_score += 3

            # –°–Њ–њ–Њ—Б—В–∞–≤–ї–µ–љ–Є–µ –њ–Њ –њ–Њ–ї—М–Ј–Њ–≤–∞—В–µ–ї—О (–њ—А–Є–Њ—А–Є—В–µ—В 2)
            if username:
                row_username = row_data.get("username", "")
                if username.lower() in row_username.lower() or row_username.lower() in username.lower():
                    match_score += 2

            # –°–Њ–њ–Њ—Б—В–∞–≤–ї–µ–љ–Є–µ –њ–Њ –і–µ—В–∞–ї–Є (–њ—А–Є–Њ—А–Є—В–µ—В 2)
            if part_name:
                row_part = row_data.get("part", "")
                if part_name.lower() in row_part.lower() or row_part.lower() in part_name.lower():
                    match_score += 2

            # –°–Њ–њ–Њ—Б—В–∞–≤–ї–µ–љ–Є–µ –њ–Њ –Ї–Њ–ї–Є—З–µ—Б—В–≤—Г (–њ—А–Є–Њ—А–Є—В–µ—В 1)
            try:
                row_qty = int(row_data.get("metal_quantity", 0))
                if row_qty == writeoff_qty:
                    match_score += 1
            except:
                pass

            # –Х—Б–ї–Є –љ–∞–±—А–∞–ї–Є –і–Њ—Б—В–∞—В–Њ—З–љ–Њ —Б–Њ–≤–њ–∞–і–µ–љ–Є–є (–Љ–Є–љ–Є–Љ—Г–Љ 3 –±–∞–ї–ї–∞)
            if match_score >= 3:
                matching_indices.append((idx, match_score))
                print(f"      вЬУ –°—В—А–Њ–Ї–∞ #{idx + 1}: score={match_score}")

        # –°–Њ—А—В–Є—А—Г–µ–Љ –њ–Њ —Г–±—Л–≤–∞–љ–Є—О score –Є –≤–Њ–Ј–≤—А–∞—Й–∞–µ–Љ –Є–љ–і–µ–Ї—Б—Л
        matching_indices.sort(key=lambda x: x[1], reverse=True)
        return [idx for idx, score in matching_indices]
    def edit_writeoff(self):
        """–†–µ–і–∞–Ї—В–Є—А–Њ–≤–∞–љ–Є–µ —Б–њ–Є—Б–∞–љ–Є—П"""
        selected = self.writeoffs_tree.selection()
        if not selected:
            messagebox.showwarning("–Я—А–µ–і—Г–њ—А–µ–ґ–і–µ–љ–Є–µ", "–Т—Л–±–µ—А–Є—В–µ —Б–њ–Є—Б–∞–љ–Є–µ –і–ї—П —А–µ–і–∞–Ї—В–Є—А–Њ–≤–∞–љ–Є—П")
            return

        writeoff_id = self.writeoffs_tree.item(selected)["values"][0]
        writeoffs_df = load_data("WriteOffs")
        writeoff_row = writeoffs_df[writeoffs_df["ID —Б–њ–Є—Б–∞–љ–Є—П"] == writeoff_id].iloc[0]

        edit_window = tk.Toplevel(self.root)
        edit_window.title("–†–µ–і–∞–Ї—В–Є—А–Њ–≤–∞—В—М —Б–њ–Є—Б–∞–љ–Є–µ")
        edit_window.geometry("550x650")
        edit_window.configure(bg='#ecf0f1')

        tk.Label(edit_window, text=f"–†–µ–і–∞–Ї—В–Є—А–Њ–≤–∞–љ–Є–µ —Б–њ–Є—Б–∞–љ–Є—П #{writeoff_id}",
                 font=("Arial", 12, "bold"), bg='#ecf0f1').pack(pady=10)

        # –Ш–љ—Д–Њ—А–Љ–∞—Ж–Є—П –Њ —А–µ–Ј–µ—А–≤–µ (—В–Њ–ї—М–Ї–Њ –і–ї—П —З—В–µ–љ–Є—П)
        reserve_id = int(writeoff_row["ID —А–µ–Ј–µ—А–≤–∞"])
        reservations_df = load_data("Reservations")
        orders_df = load_data("Orders")

        reserve_info = f"–†–µ–Ј–µ—А–≤ #{reserve_id}"
        order_info = ""
        detail_info = ""

        if not reservations_df.empty:
            reserve_row = reservations_df[reservations_df["ID —А–µ–Ј–µ—А–≤–∞"] == reserve_id]
            if not reserve_row.empty:
                reserve_data = reserve_row.iloc[0]
                order_id = int(reserve_data["ID –Ј–∞–Ї–∞–Ј–∞"])

                if not orders_df.empty:
                    order_row = orders_df[orders_df["ID –Ј–∞–Ї–∞–Ј–∞"] == order_id]
                    if not order_row.empty:
                        customer = order_row.iloc[0]["–Ч–∞–Ї–∞–Ј—З–Є–Ї"]
                        order_name = order_row.iloc[0]["–Э–∞–Ј–≤–∞–љ–Є–µ –Ј–∞–Ї–∞–Ј–∞"]
                        order_info = f"{customer} | {order_name}"

                detail_name = reserve_data.get("–Э–∞–Ј–≤–∞–љ–Є–µ –і–µ—В–∞–ї–Є", "–С–µ–Ј –і–µ—В–∞–ї–Є")
                if pd.notna(detail_name) and detail_name != "" and detail_name != "–Э–µ —Г–Ї–∞–Ј–∞–љ–∞":
                    detail_info = f"–Ф–µ—В–∞–ї—М: {detail_name}"
                else:
                    detail_info = "–С–µ–Ј –њ—А–Є–≤—П–Ј–Ї–Є –Ї –і–µ—В–∞–ї–Є"

        info_frame = tk.LabelFrame(edit_window, text="–Ш–љ—Д–Њ—А–Љ–∞—Ж–Є—П (–љ–µ —А–µ–і–∞–Ї—В–Є—А—Г–µ—В—Б—П)",
                                   bg='#e8f4f8', font=("Arial", 9, "bold"))
        info_frame.pack(fill=tk.X, padx=20, pady=10)

        if order_info:
            tk.Label(info_frame, text=f"–Ч–∞–Ї–∞–Ј: {order_info}", bg='#e8f4f8', font=("Arial", 9)).pack(padx=10, pady=2,
                                                                                                    anchor='w')
        if detail_info:
            tk.Label(info_frame, text=detail_info, bg='#e8f4f8', font=("Arial", 9)).pack(padx=10, pady=2, anchor='w')

        material_info = f"{writeoff_row['–Ь–∞—А–Ї–∞']} {writeoff_row['–Ґ–Њ–ї—Й–Є–љ–∞']}–Љ–Љ {writeoff_row['–®–Є—А–Є–љ–∞']}x{writeoff_row['–Ф–ї–Є–љ–∞']}"
        tk.Label(info_frame, text=f"–Ь–∞—В–µ—А–Є–∞–ї: {material_info}", bg='#e8f4f8', font=("Arial", 9)).pack(padx=10, pady=2,
                                                                                                      anchor='w')
        tk.Label(info_frame, text=f"–Ф–∞—В–∞ —Б–њ–Є—Б–∞–љ–Є—П: {writeoff_row['–Ф–∞—В–∞ —Б–њ–Є—Б–∞–љ–Є—П']}",
                 bg='#e8f4f8', font=("Arial", 9)).pack(padx=10, pady=2, anchor='w')

        # –†–µ–і–∞–Ї—В–Є—А—Г–µ–Љ–Њ–µ –њ–Њ–ї–µ: –Ъ–Њ–ї–Є—З–µ—Б—В–≤–Њ
        qty_frame = tk.Frame(edit_window, bg='#ecf0f1')
        qty_frame.pack(fill=tk.X, padx=20, pady=10)
        tk.Label(qty_frame, text="–Ъ–Њ–ї–Є—З–µ—Б—В–≤–Њ (—И—В):", width=25, anchor='w',
                 bg='#ecf0f1', font=("Arial", 10, "bold")).pack(side=tk.LEFT)
        qty_entry = tk.Entry(qty_frame, font=("Arial", 10))
        qty_entry.insert(0, str(int(writeoff_row["–Ъ–Њ–ї–Є—З–µ—Б—В–≤–Њ"])))
        qty_entry.pack(side=tk.RIGHT, expand=True, fill=tk.X)

        # –†–µ–і–∞–Ї—В–Є—А—Г–µ–Љ–Њ–µ –њ–Њ–ї–µ: –Ъ–Њ–Љ–Љ–µ–љ—В–∞—А–Є–є
        comment_frame = tk.Frame(edit_window, bg='#ecf0f1')
        comment_frame.pack(fill=tk.X, padx=20, pady=10)
        tk.Label(comment_frame, text="–Ъ–Њ–Љ–Љ–µ–љ—В–∞—А–Є–є:", width=25, anchor='w',
                 bg='#ecf0f1', font=("Arial", 10)).pack(side=tk.LEFT)
        comment_entry = tk.Entry(comment_frame, font=("Arial", 10))
        comment_entry.insert(0, str(writeoff_row["–Ъ–Њ–Љ–Љ–µ–љ—В–∞—А–Є–є"]))
        comment_entry.pack(side=tk.RIGHT, expand=True, fill=tk.X)

        # –Ш–љ—Д–Њ—А–Љ–∞—Ж–Є—П –Њ —А–µ–Ј–µ—А–≤–µ
        if not reservations_df.empty and not reserve_row.empty:
            reserve_data = reserve_row.iloc[0]
            reserve_total = int(reserve_data["–Ч–∞—А–µ–Ј–µ—А–≤–Є—А–Њ–≤–∞–љ–Њ —И—В—Г–Ї"])
            reserve_written = int(reserve_data["–°–њ–Є—Б–∞–љ–Њ"])
            reserve_remainder = int(reserve_data["–Ю—Б—В–∞—В–Њ–Ї –Ї —Б–њ–Є—Б–∞–љ–Є—О"])

            stats_frame = tk.LabelFrame(edit_window, text="–°—В–∞—В–Є—Б—В–Є–Ї–∞ —А–µ–Ј–µ—А–≤–∞",
                                        bg='#fff3cd', font=("Arial", 9, "bold"))
            stats_frame.pack(fill=tk.X, padx=20, pady=10)
            tk.Label(stats_frame, text=f"–Т—Б–µ–≥–Њ –≤ —А–µ–Ј–µ—А–≤–µ: {reserve_total} —И—В",
                     bg='#fff3cd', font=("Arial", 9)).pack(anchor='w', padx=10, pady=2)
            tk.Label(stats_frame, text=f"–°–њ–Є—Б–∞–љ–Њ –≤—Б–µ–≥–Њ: {reserve_written} —И—В",
                     bg='#fff3cd', font=("Arial", 9)).pack(anchor='w', padx=10, pady=2)
            tk.Label(stats_frame, text=f"–Ю—Б—В–∞—В–Њ–Ї –Ї —Б–њ–Є—Б–∞–љ–Є—О: {reserve_remainder} —И—В",
                     bg='#fff3cd', font=("Arial", 9)).pack(anchor='w', padx=10, pady=2)

        # –Я—А–µ–і—Г–њ—А–µ–ґ–і–µ–љ–Є–µ
        warning_frame = tk.Frame(edit_window, bg='#ffcccc', relief=tk.RIDGE, borderwidth=2)
        warning_frame.pack(fill=tk.X, padx=20, pady=10)
        tk.Label(warning_frame, text="–Т–Р–Ц–Э–Ю!", font=("Arial", 9, "bold"),
                 bg='#ffcccc', fg='#c0392b').pack(anchor='w', padx=5, pady=2)
        tk.Label(warning_frame, text="вАҐ –Ш–Ј–Љ–µ–љ–µ–љ–Є–µ –Ї–Њ–ї–Є—З–µ—Б—В–≤–∞ –њ–µ—А–µ—Б—З–Є—В–∞–µ—В –±–∞–ї–∞–љ—Б –Љ–∞—В–µ—А–Є–∞–ї–Њ–≤",
                 font=("Arial", 8), bg='#ffcccc', fg='#c0392b').pack(anchor='w', padx=10)
        tk.Label(warning_frame, text="вАҐ –Ш–Ј–Љ–µ–љ–µ–љ–Є–µ –≤–ї–Є—П–µ—В –љ–∞ –Њ—Б—В–∞—В–Њ–Ї —А–µ–Ј–µ—А–≤–∞ –Ї —Б–њ–Є—Б–∞–љ–Є—О",
                 font=("Arial", 8), bg='#ffcccc', fg='#c0392b').pack(anchor='w', padx=10)

        def save_changes():
            try:
                new_qty = int(qty_entry.get().strip())
                new_comment = comment_entry.get().strip()

                if new_qty <= 0:
                    messagebox.showerror("–Ю—И–Є–±–Ї–∞", "–Ъ–Њ–ї–Є—З–µ—Б—В–≤–Њ –і–Њ–ї–ґ–љ–Њ –±—Л—В—М –±–Њ–ї—М—И–µ –љ—Г–ї—П!")
                    return

                old_qty = int(writeoff_row["–Ъ–Њ–ї–Є—З–µ—Б—В–≤–Њ"])
                difference = new_qty - old_qty

                # –Я—А–Њ–≤–µ—А—П–µ–Љ, –љ–µ –њ—А–µ–≤—Л—Б–Є—В –ї–Є –љ–Њ–≤–Њ–µ –Ї–Њ–ї–Є—З–µ—Б—В–≤–Њ –і–Њ—Б—В—Г–њ–љ—Л–є –Њ—Б—В–∞—В–Њ–Ї —А–µ–Ј–µ—А–≤–∞
                if not reservations_df.empty and not reserve_row.empty:
                    reserve_data = reserve_row.iloc[0]
                    reserve_remainder = int(reserve_data["–Ю—Б—В–∞—В–Њ–Ї –Ї —Б–њ–Є—Б–∞–љ–Є—О"])

                    # –Ф–Њ—Б—В—Г–њ–љ–Њ = —В–µ–Ї—Г—Й–Є–є –Њ—Б—В–∞—В–Њ–Ї + —Б—В–∞—А–Њ–µ —Б–њ–Є—Б–∞–љ–Є–µ
                    max_available = reserve_remainder + old_qty

                    if new_qty > max_available:
                        messagebox.showerror("–Ю—И–Є–±–Ї–∞",
                                             f"–Э–µ–ї—М–Ј—П —Б–њ–Є—Б–∞—В—М {new_qty} —И—В!\n"
                                             f"–Ь–∞–Ї—Б–Є–Љ–∞–ї—М–љ–Њ –і–Њ—Б—В—Г–њ–љ–Њ: {max_available} —И—В")
                        return

                if difference == 0 and new_comment == str(writeoff_row["–Ъ–Њ–Љ–Љ–µ–љ—В–∞—А–Є–є"]):
                    messagebox.showinfo("–Ш–љ—Д–Њ—А–Љ–∞—Ж–Є—П", "–Ш–Ј–Љ–µ–љ–µ–љ–Є–є –љ–µ –±—Л–ї–Њ")
                    edit_window.destroy()
                    return

                # –Я–Њ–і—В–≤–µ—А–ґ–і–µ–љ–Є–µ
                msg = f"–°–Њ—Е—А–∞–љ–Є—В—М –Є–Ј–Љ–µ–љ–µ–љ–Є—П?\n\n"
                if difference != 0:
                    msg += f"–Ъ–Њ–ї–Є—З–µ—Б—В–≤–Њ: {old_qty} вЖТ {new_qty} —И—В (—А–∞–Ј–љ–Є—Ж–∞: {'+' if difference > 0 else ''}{difference})\n"
                if new_comment != str(writeoff_row["–Ъ–Њ–Љ–Љ–µ–љ—В–∞—А–Є–є"]):
                    msg += f"–Ъ–Њ–Љ–Љ–µ–љ—В–∞—А–Є–є –Є–Ј–Љ–µ–љ–µ–љ"

                if not messagebox.askyesno("–Я–Њ–і—В–≤–µ—А–ґ–і–µ–љ–Є–µ", msg):
                    return

                # –Ю–±–љ–Њ–≤–ї—П–µ–Љ —Б–њ–Є—Б–∞–љ–Є–µ
                writeoffs_df.loc[writeoffs_df["ID —Б–њ–Є—Б–∞–љ–Є—П"] == writeoff_id, "–Ъ–Њ–ї–Є—З–µ—Б—В–≤–Њ"] = new_qty
                writeoffs_df.loc[writeoffs_df["ID —Б–њ–Є—Б–∞–љ–Є—П"] == writeoff_id, "–Ъ–Њ–Љ–Љ–µ–љ—В–∞—А–Є–є"] = new_comment
                save_data("WriteOffs", writeoffs_df)

                # –Х—Б–ї–Є –Ї–Њ–ї–Є—З–µ—Б—В–≤–Њ –Є–Ј–Љ–µ–љ–Є–ї–Њ—Б—М - –Њ–±–љ–Њ–≤–ї—П–µ–Љ —А–µ–Ј–µ—А–≤ –Є –Љ–∞—В–µ—А–Є–∞–ї
                if difference != 0:
                    # –Ю–±–љ–Њ–≤–ї—П–µ–Љ —А–µ–Ј–µ—А–≤
                    if not reservations_df.empty and not reserve_row.empty:
                        reserve_data = reserve_row.iloc[0]
                        current_written = int(reserve_data["–°–њ–Є—Б–∞–љ–Њ"])
                        current_remainder = int(reserve_data["–Ю—Б—В–∞—В–Њ–Ї –Ї —Б–њ–Є—Б–∞–љ–Є—О"])

                        new_written = current_written + difference
                        new_remainder = current_remainder - difference

                        reservations_df.loc[reservations_df["ID —А–µ–Ј–µ—А–≤–∞"] == reserve_id, "–°–њ–Є—Б–∞–љ–Њ"] = new_written
                        reservations_df.loc[
                            reservations_df["ID —А–µ–Ј–µ—А–≤–∞"] == reserve_id, "–Ю—Б—В–∞—В–Њ–Ї –Ї —Б–њ–Є—Б–∞–љ–Є—О"] = new_remainder
                        save_data("Reservations", reservations_df)

                    # –Ю–±–љ–Њ–≤–ї—П–µ–Љ –Љ–∞—В–µ—А–Є–∞–ї (–µ—Б–ї–Є –љ–µ –≤—А—Г—З–љ—Г—О –і–Њ–±–∞–≤–ї–µ–љ–љ—Л–є)
                    material_id = int(writeoff_row["ID –Љ–∞—В–µ—А–Є–∞–ї–∞"])
                    if material_id != -1:
                        materials_df = load_data("Materials")
                        if not materials_df[materials_df["ID"] == material_id].empty:
                            mat_row = materials_df[materials_df["ID"] == material_id].iloc[0]
                            current_qty = int(mat_row["–Ъ–Њ–ї–Є—З–µ—Б—В–≤–Њ —И—В—Г–Ї"])
                            current_reserved = int(mat_row["–Ч–∞—А–µ–Ј–µ—А–≤–Є—А–Њ–≤–∞–љ–Њ"])

                            # –†–∞–Ј–љ–Є—Ж–∞ —Б–њ–Є—Б–∞–љ–Є—П –≤–ї–Є—П–µ—В –љ–∞ –Ї–Њ–ї–Є—З–µ—Б—В–≤–Њ –Є —А–µ–Ј–µ—А–≤
                            new_mat_qty = current_qty - difference
                            new_reserved = current_reserved - difference

                            materials_df.loc[materials_df["ID"] == material_id, "–Ъ–Њ–ї–Є—З–µ—Б—В–≤–Њ —И—В—Г–Ї"] = new_mat_qty
                            materials_df.loc[materials_df["ID"] == material_id, "–Ч–∞—А–µ–Ј–µ—А–≤–Є—А–Њ–≤–∞–љ–Њ"] = new_reserved

                            # –Я–µ—А–µ—Б—З–Є—В—Л–≤–∞–µ–Љ –њ–ї–Њ—Й–∞–і—М
                            area_per_piece = float(mat_row["–Ф–ї–Є–љ–∞"]) * float(mat_row["–®–Є—А–Є–љ–∞"]) / 1_000_000
                            new_area = new_mat_qty * area_per_piece
                            materials_df.loc[materials_df["ID"] == material_id, "–Ю–±—Й–∞—П –њ–ї–Њ—Й–∞–і—М"] = round(new_area, 2)

                            save_data("Materials", materials_df)
                            self.refresh_materials()

                self.refresh_reservations()
                self.refresh_writeoffs()
                self.refresh_balance()
                edit_window.destroy()
                messagebox.showinfo("–£—Б–њ–µ—Е", f"–°–њ–Є—Б–∞–љ–Є–µ #{writeoff_id} –Њ–±–љ–Њ–≤–ї–µ–љ–Њ!")

            except ValueError:
                messagebox.showerror("–Ю—И–Є–±–Ї–∞", "–Я—А–Њ–≤–µ—А—М—В–µ –њ—А–∞–≤–Є–ї—М–љ–Њ—Б—В—М –≤–≤–Њ–і–∞ —З–Є—Б–ї–Њ–≤—Л—Е –Ј–љ–∞—З–µ–љ–Є–є!")
            except Exception as e:
                messagebox.showerror("–Ю—И–Є–±–Ї–∞", f"–Э–µ —Г–і–∞–ї–Њ—Б—М –Њ–±–љ–Њ–≤–Є—В—М —Б–њ–Є—Б–∞–љ–Є–µ: {e}")
                import traceback
                traceback.print_exc()

        tk.Button(edit_window, text="–°–Њ—Е—А–∞–љ–Є—В—М –Є–Ј–Љ–µ–љ–µ–љ–Є—П", bg='#f39c12', fg='white',
                  font=("Arial", 12, "bold"), command=save_changes).pack(pady=15)

    def setup_laser_import_tab(self):
        """–Т–Ї–ї–∞–і–Ї–∞ –Є–Љ–њ–Њ—А—В–∞ –Њ—В –ї–∞–Ј–µ—А—Й–Є–Ї–Њ–≤ - –Х–Ф–Ш–Э–°–Ґ–Т–Х–Э–Э–Р–ѓ –Т–Х–†–°–Ш–ѓ"""

        # –Ю—З–Є—Й–∞–µ–Љ —Д—А–µ–є–Љ –љ–∞ —Б–ї—Г—З–∞–є –њ–Њ–≤—В–Њ—А–љ–Њ–≥–Њ –≤—Л–Ј–Њ–≤–∞
        for widget in self.laser_import_frame.winfo_children():
            widget.destroy()

        # –Ч–∞–≥–Њ–ї–Њ–≤–Њ–Ї
        header = tk.Label(self.laser_import_frame, text="рЯУ• –Ш–Љ–њ–Њ—А—В –і–∞–љ–љ—Л—Е –Њ—В –ї–∞–Ј–µ—А—Й–Є–Ї–Њ–≤",
                          font=("Arial", 16, "bold"), bg='white', fg='#e67e22')
        header.pack(pady=10)

        # –Ш–љ—Б—В—А—Г–Ї—Ж–Є—П
        info_frame = tk.LabelFrame(self.laser_import_frame, text="вДєпЄП –Ш–љ—Д–Њ—А–Љ–∞—Ж–Є—П",
                                   bg='#d1ecf1', font=("Arial", 10, "bold"))
        info_frame.pack(fill=tk.X, padx=20, pady=10)

        instructions = """
    рЯУЛ –§–Њ—А–Љ–∞—В —Д–∞–є–ї–∞ CSV:
    вАҐ –Ъ–Њ–ї–Њ–љ–Ї–Є: –Ф–∞—В–∞ (–Ь–°–Ъ), –Т—А–µ–Љ—П (–Ь–°–Ъ), username, order, metal, metal_quantity, part, part_quantity

    рЯУМ –І—В–Њ –і–µ–ї–∞–µ—В –Є–Љ–њ–Њ—А—В:
    1. –І–Є—В–∞–µ—В —Д–∞–є–ї –Њ—В –ї–∞–Ј–µ—А—Й–Є–Ї–Њ–≤
    2. –Ю—В–Њ–±—А–∞–ґ–∞–µ—В –≤—Б–µ —Б—В—А–Њ–Ї–Є –≤ —В–∞–±–ї–Є—Ж–µ
    3. –Я–Њ–Ј–≤–Њ–ї—П–µ—В –≤—Л–±—А–∞—В—М —Б—В—А–Њ–Ї–Є –і–ї—П —Б–њ–Є—Б–∞–љ–Є—П
    4. –Р–≤—В–Њ–Љ–∞—В–Є—З–µ—Б–Ї–Є –љ–∞—Е–Њ–і–Є—В —А–µ–Ј–µ—А–≤—Л –Є —Б–њ–Є—Б—Л–≤–∞–µ—В –Љ–∞—В–µ—А–Є–∞–ї
        """

        tk.Label(info_frame, text=instructions, bg='#d1ecf1',
                 font=("Arial", 9), justify=tk.LEFT).pack(padx=10, pady=5)

        # –Ъ–љ–Њ–њ–Ї–Є —Г–њ—А–∞–≤–ї–µ–љ–Є—П
        buttons_frame = tk.Frame(self.laser_import_frame, bg='white')
        buttons_frame.pack(fill=tk.X, padx=20, pady=10)

        btn_style = {"font": ("Arial", 10, "bold"), "width": 20, "height": 2}

        tk.Button(buttons_frame, text="рЯУБ –Ш–Љ–њ–Њ—А—В —Д–∞–є–ї–∞", bg='#3498db', fg='white',
                  command=self.import_laser_table, **btn_style).pack(side=tk.LEFT, padx=5)

        tk.Button(buttons_frame, text="вЬЕ –°–њ–Є—Б–∞—В—М –≤—Л–±—А–∞–љ–љ—Л–µ", bg='#27ae60', fg='white',
                  command=self.writeoff_laser_row, **btn_style).pack(side=tk.LEFT, padx=5)
        # рЯЖХ –Э–Ю–Т–Р–ѓ –Ъ–Э–Ю–Я–Ъ–Р
        tk.Button(buttons_frame, text="рЯФµ –Я–Њ–Љ–µ—В–Є—В—М –≤—А—Г—З–љ—Г—О", bg='#2196F3', fg='white',
                  command=self.mark_manual_writeoff, **btn_style).pack(side=tk.LEFT, padx=5)

        # рЯЖХ –Ъ–Э–Ю–Я–Ъ–Р –°–Э–ѓ–Ґ–Ш–ѓ –Я–Ю–Ь–Х–Ґ–Ъ–Ш
        tk.Button(buttons_frame, text="вЖ©пЄП –°–љ—П—В—М –њ–Њ–Љ–µ—В–Ї—Г", bg='#9E9E9E', fg='white',
                  command=self.unmark_manual_writeoff, **btn_style).pack(side=tk.LEFT, padx=5)

        tk.Button(buttons_frame, text="рЯЧСпЄП –£–і–∞–ї–Є—В—М —Б—В—А–Њ–Ї–Є", bg='#e74c3c', fg='white',
                  command=self.delete_laser_row, **btn_style).pack(side=tk.LEFT, padx=5)

        tk.Button(buttons_frame, text="рЯТЊ –≠–Ї—Б–њ–Њ—А—В —В–∞–±–ї–Є—Ж—Л", bg='#9b59b6', fg='white',
                  command=self.export_laser_table, **btn_style).pack(side=tk.LEFT, padx=5)

        tk.Button(buttons_frame, text="вЬЦ –°–±—А–Њ—Б–Є—В—М —Д–Є–ї—М—В—А—Л", bg='#e67e22', fg='white',
                  command=self.clear_laser_import_filters, **btn_style).pack(side=tk.LEFT, padx=5)

        # –Ь–µ—В–Ї–∞ —В–∞–±–ї–Є—Ж—Л
        table_label = tk.Label(self.laser_import_frame,
                               text="рЯУК –Ш–Љ–њ–Њ—А—В–Є—А–Њ–≤–∞–љ–љ—Л–µ –і–∞–љ–љ—Л–µ (–≤—Л–±–µ—А–Є—В–µ —Б—В—А–Њ–Ї–Є –і–ї—П —Б–њ–Є—Б–∞–љ–Є—П)",
                               font=("Arial", 11, "bold"), bg='white', fg='#2c3e50')
        table_label.pack(pady=5)

        # –§—А–µ–є–Љ –і–ї—П —В–∞–±–ї–Є—Ж—Л
        tree_frame = tk.Frame(self.laser_import_frame, bg='white')
        tree_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)

        # Scrollbars
        scroll_y = tk.Scrollbar(tree_frame, orient=tk.VERTICAL)
        scroll_x = tk.Scrollbar(tree_frame, orient=tk.HORIZONTAL)

        # рЯЖХ –°–Ю–Ч–Ф–Р–Э–Ш–Х TREEVIEW –° –ѓ–Т–Э–Ђ–Ь–Ш –Я–Р–†–Р–Ь–Х–Ґ–†–Р–Ь–Ш
        self.laser_import_tree = ttk.Treeview(
            tree_frame,
            columns=("–Ф–∞—В–∞", "–Т—А–µ–Љ—П", "–Я–Њ–ї—М–Ј–Њ–≤–∞—В–µ–ї—М", "–Ч–∞–Ї–∞–Ј", "–Ь–µ—В–∞–ї–ї", "–Ъ–Њ–ї-–≤–Њ", "–Ф–µ—В–∞–ї—М", "–Ъ–Њ–ї-–≤–Њ –і–µ—В–∞–ї–µ–є",
                     "–°–њ–Є—Б–∞–љ–Њ", "–Ф–∞—В–∞ —Б–њ–Є—Б–∞–љ–Є—П"),
            show="headings",
            height=20,  # рЯЖХ –ѓ–Т–Э–Р–ѓ –Т–Ђ–°–Ю–Ґ–Р
            selectmode='extended',
            yscrollcommand=scroll_y.set,
            xscrollcommand=scroll_x.set
        )

        scroll_y.config(command=self.laser_import_tree.yview)
        scroll_x.config(command=self.laser_import_tree.xview)

        scroll_y.pack(side=tk.RIGHT, fill=tk.Y)
        scroll_x.pack(side=tk.BOTTOM, fill=tk.X)

        # –Э–∞—Б—В—А–Њ–є–Ї–∞ –Ї–Њ–ї–Њ–љ–Њ–Ї
        columns_config = {
            "–Ф–∞—В–∞": 100,
            "–Т—А–µ–Љ—П": 80,
            "–Я–Њ–ї—М–Ј–Њ–≤–∞—В–µ–ї—М": 120,
            "–Ч–∞–Ї–∞–Ј": 200,
            "–Ь–µ—В–∞–ї–ї": 200,
            "–Ъ–Њ–ї-–≤–Њ": 80,
            "–Ф–µ—В–∞–ї—М": 200,
            "–Ъ–Њ–ї-–≤–Њ –і–µ—В–∞–ї–µ–є": 120,
            "–°–њ–Є—Б–∞–љ–Њ": 80,
            "–Ф–∞—В–∞ —Б–њ–Є—Б–∞–љ–Є—П": 150
        }

        # –Э–∞—Б—В—А–Њ–є–Ї–∞ –Ї–Њ–ї–Њ–љ–Њ–Ї –С–Х–Ч —А–∞—Б—В—П–≥–Є–≤–∞–љ–Є—П
        for col, width in columns_config.items():
            self.laser_import_tree.heading(col, text=col)
            self.laser_import_tree.column(col, width=width, anchor=tk.CENTER, minwidth=80, stretch=False)

        self.laser_import_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # рЯЖХ –Ш–Э–Ш–¶–Ш–Р–Ы–Ш–Ч–Р–¶–Ш–ѓ EXCEL-–§–Ш–Ы–ђ–Ґ–†–Р –Ф–Ы–ѓ –Ш–Ь–Я–Ю–†–Ґ–Р
        self.laser_import_excel_filter = ExcelStyleFilter(
            tree=self.laser_import_tree,
            refresh_callback=self.refresh_laser_import_table
        )

        # рЯЖХ –Ш–Э–Ф–Ш–Ъ–Р–Ґ–Ю–† –Р–Ъ–Ґ–Ш–Т–Э–Ђ–• –§–Ш–Ы–ђ–Ґ–†–Ю–Т
        self.laser_import_filter_status = tk.Label(
            self.laser_import_frame,
            text="",
            font=("Arial", 9),
            bg='#d1ecf1',
            fg='#0c5460'
        )
        self.laser_import_filter_status.pack(pady=5)

        # –¶–≤–µ—В–Њ–≤–∞—П –Є–љ–і–Є–Ї–∞—Ж–Є—П
        self.laser_import_tree.tag_configure('written_off', background='#c8e6c9', foreground='#1b5e20')
        self.laser_import_tree.tag_configure('manual', background='#bbdefb', foreground='#0d47a1')  # –°–≤–µ—В–ї–Њ-—Б–Є–љ–Є–є
        self.laser_import_tree.tag_configure('pending', background='#fff9c4', foreground='#000000')
        self.laser_import_tree.tag_configure('error', background='#ffcccc', foreground='#b71c1c')

        # –°—В–∞—В—Г—Б–љ–∞—П —Б—В—А–Њ–Ї–∞
        self.laser_status_label = tk.Label(
            self.laser_import_frame,
            text="рЯУВ –Ш–Љ–њ–Њ—А—В–Є—А—Г–є—В–µ —Д–∞–є–ї –і–ї—П –љ–∞—З–∞–ї–∞ —А–∞–±–Њ—В—Л",
            font=("Arial", 10),
            bg='#ecf0f1',
            fg='#2c3e50',
            relief=tk.SUNKEN,
            anchor='w',
            padx=10,
            pady=5
        )
        self.laser_status_label.pack(fill=tk.X, side=tk.BOTTOM, padx=20, pady=10)

        print("вЬЕ setup_laser_import_tab() –≤—Л–њ–Њ–ї–љ–µ–љ —Г—Б–њ–µ—И–љ–Њ")

        # рЯЖХ –Р–Т–Ґ–Ю–Ч–Р–У–†–£–Ч–Ъ–Р –Ъ–≠–®–Р –Я–†–Ш –°–Ґ–Р–†–Ґ–Х
        self.load_laser_import_cache()

    def setup_details_tab(self):
        """–Т–Ї–ї–∞–і–Ї–∞ —Г—З—С—В–∞ –і–µ—В–∞–ї–µ–є"""

        # –Ч–∞–≥–Њ–ї–Њ–≤–Њ–Ї
        header = tk.Label(self.details_frame, text="рЯУР –£—З—С—В –і–µ—В–∞–ї–µ–є –њ–Њ –Ј–∞–Ї–∞–Ј–∞–Љ",
                          font=("Arial", 16, "bold"), bg='white', fg='#2c3e50')
        header.pack(pady=10)

        # –Ш–љ—Д–Њ—А–Љ–∞—Ж–Є–Њ–љ–љ–∞—П –њ–∞–љ–µ–ї—М
        info_frame = tk.LabelFrame(self.details_frame, text="вДєпЄП –Ш–љ—Д–Њ—А–Љ–∞—Ж–Є—П",
                                   bg='#d1ecf1', font=("Arial", 10, "bold"))
        info_frame.pack(fill=tk.X, padx=20, pady=10)

        info_text = """
    рЯУК –Ю—В–Њ–±—А–∞–ґ–∞—О—В—Б—П –≤—Б–µ –і–µ—В–∞–ї–Є –Є–Ј –Ј–∞–Ї–∞–Ј–Њ–≤ —Б–Њ —Б—В–∞—В—Г—Б–Њ–Љ "–Т —А–∞–±–Њ—В–µ"
    рЯЯҐ –Ч–µ–ї—С–љ—Л–є - –і–µ—В–∞–ї—М –њ–Њ–ї–љ–Њ—Б—В—М—О –њ–Њ—А–µ–Ј–∞–љ–∞
    рЯЯ° –Ц—С–ї—В—Л–є - –і–µ—В–∞–ї—М –≤ –њ—А–Њ—Ж–µ—Б—Б–µ
    вЪ™ –С–µ–ї—Л–є - –і–µ—В–∞–ї—М –љ–µ –љ–∞—З–∞—В–∞
        """

        tk.Label(info_frame, text=info_text, bg='#d1ecf1',
                 font=("Arial", 9), justify=tk.LEFT).pack(padx=10, pady=5)

        # –§—А–µ–є–Љ –і–ї—П —В–∞–±–ї–Є—Ж—Л
        tree_frame = tk.Frame(self.details_frame, bg='white')
        tree_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)

        # Scrollbars
        scroll_y = tk.Scrollbar(tree_frame, orient=tk.VERTICAL)
        scroll_x = tk.Scrollbar(tree_frame, orient=tk.HORIZONTAL)

        # –°–Њ–Ј–і–∞–љ–Є–µ —В–∞–±–ї–Є—Ж—Л
        self.details_tree = ttk.Treeview(
            tree_frame,
            columns=("ID", "–Ч–∞–Ї–∞–Ј—З–Є–Ї", "–Э–∞–Ј–≤–∞–љ–Є–µ –і–µ—В–∞–ї–Є", "–Ч–∞–Ї–∞–Ј", "–Ъ–Њ–ї–Є—З–µ—Б—В–≤–Њ", "–Я–Њ—А–µ–Ј–∞–љ–Њ", "–Я–Њ–≥–љ—Г—В–Њ", "–Ю—Б—В–∞–ї–Њ—Б—М",
                     "–Я—А–Њ–≥—А–µ—Б—Б %"),
            show="headings",
            yscrollcommand=scroll_y.set,
            xscrollcommand=scroll_x.set
        )

        scroll_y.config(command=self.details_tree.yview)
        scroll_x.config(command=self.details_tree.xview)
        scroll_y.pack(side=tk.RIGHT, fill=tk.Y)
        scroll_x.pack(side=tk.BOTTOM, fill=tk.X)

        # –Э–∞—Б—В—А–Њ–є–Ї–∞ –Ї–Њ–ї–Њ–љ–Њ–Ї
        columns_config = {
            "ID": 60,
            "–Ч–∞–Ї–∞–Ј—З–Є–Ї": 200,
            "–Э–∞–Ј–≤–∞–љ–Є–µ –і–µ—В–∞–ї–Є": 300,
            "–Ч–∞–Ї–∞–Ј": 200,
            "–Ъ–Њ–ї–Є—З–µ—Б—В–≤–Њ": 100,
            "–Я–Њ—А–µ–Ј–∞–љ–Њ": 100,
            "–Я–Њ–≥–љ—Г—В–Њ": 100,
            "–Ю—Б—В–∞–ї–Њ—Б—М": 100,
            "–Я—А–Њ–≥—А–µ—Б—Б %": 100
        }

        # –Э–∞—Б—В—А–Њ–є–Ї–∞ –Ї–Њ–ї–Њ–љ–Њ–Ї –С–Х–Ч —А–∞—Б—В—П–≥–Є–≤–∞–љ–Є—П
        for col, width in columns_config.items():
            self.details_tree.heading(col, text=col)
            self.details_tree.column(col, width=width, anchor=tk.CENTER, minwidth=80, stretch=False)

        self.details_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # рЯЖХ –Ш–Э–Ш–¶–Ш–Р–Ы–Ш–Ч–Р–¶–Ш–ѓ EXCEL-–§–Ш–Ы–ђ–Ґ–†–Р –Ф–Ы–ѓ –Ф–Х–Ґ–Р–Ы–Х–Щ
        self.details_excel_filter = ExcelStyleFilter(
            tree=self.details_tree,
            refresh_callback=self.refresh_details
        )

        # рЯЖХ –Ш–Э–Ф–Ш–Ъ–Р–Ґ–Ю–† –Р–Ъ–Ґ–Ш–Т–Э–Ђ–• –§–Ш–Ы–ђ–Ґ–†–Ю–Т
        self.details_filter_status = tk.Label(
            self.details_frame,
            text="",
            font=("Arial", 9),
            bg='#d1ecf1',
            fg='#0c5460'
        )
        self.details_filter_status.pack(pady=5)

        # –Я—А–Є–≤—П–Ј–Ї–∞ –њ—А–∞–≤–Њ–≥–Њ –Ї–ї–Є–Ї–∞ –і–ї—П –Ї–Њ–њ–Є—А–Њ–≤–∞–љ–Є—П –Є–љ—Д–Њ—А–Љ–∞—Ж–Є–Є –Њ –і–µ—В–∞–ї–Є
        self.details_tree.bind('<Button-3>', self.on_details_tab_right_click)

        # –¶–≤–µ—В–Њ–≤—Л–µ —В–µ–≥–Є
        self.details_tree.tag_configure('completed', background='#c8e6c9', foreground='#1b5e20')  # –Ч–µ–ї—С–љ—Л–є
        self.details_tree.tag_configure('in_progress', background='#fff9c4', foreground='#f57f17')  # –Ц—С–ї—В—Л–є
        self.details_tree.tag_configure('not_started', background='#ffffff', foreground='#000000')  # –С–µ–ї—Л–є
        self.details_tree.tag_configure('over_cut', background='#ffcccc',
                                        foreground='#b71c1c')  # –Ъ—А–∞—Б–љ—Л–є (–µ—Б–ї–Є –њ–Њ—А–µ–Ј–∞–љ–Њ –±–Њ–ї—М—И–µ)

        # –Я–µ—А–µ–Ї–ї—О—З–∞—В–µ–ї–Є –≤–Є–і–Є–Љ–Њ—Б—В–Є
        toggles_frame = tk.LabelFrame(self.details_frame, text="вЪЩпЄП –Э–∞—Б—В—А–Њ–є–Ї–Є –Њ—В–Њ–±—А–∞–ґ–µ–љ–Є—П",
                                      bg='#ecf0f1', font=("Arial", 10, "bold"))
        toggles_frame.pack(fill=tk.X, padx=20, pady=10)

        self.details_toggles['show_completed'] = tk.BooleanVar(value=True)
        self.details_toggles['show_not_started'] = tk.BooleanVar(value=True)
        self.details_toggles['show_in_progress'] = tk.BooleanVar(value=True)

        tk.Checkbutton(toggles_frame, text="рЯЯҐ –Я–Њ–Ї–∞–Ј–∞—В—М –Ј–∞–≤–µ—А—И—С–љ–љ—Л–µ",
                       variable=self.details_toggles['show_completed'],
                       command=self.refresh_details, bg='#ecf0f1', font=("Arial", 10),
                       activebackground='#ecf0f1').pack(side=tk.LEFT, padx=10)

        tk.Checkbutton(toggles_frame, text="рЯЯ° –Я–Њ–Ї–∞–Ј–∞—В—М –≤ –њ—А–Њ—Ж–µ—Б—Б–µ",
                       variable=self.details_toggles['show_in_progress'],
                       command=self.refresh_details, bg='#ecf0f1', font=("Arial", 10),
                       activebackground='#ecf0f1').pack(side=tk.LEFT, padx=10)

        tk.Checkbutton(toggles_frame, text="вЪ™ –Я–Њ–Ї–∞–Ј–∞—В—М –љ–µ –љ–∞—З–∞—В—Л–µ",
                       variable=self.details_toggles['show_not_started'],
                       command=self.refresh_details, bg='#ecf0f1', font=("Arial", 10),
                       activebackground='#ecf0f1').pack(side=tk.LEFT, padx=10)

        # –Ъ–љ–Њ–њ–Ї–Є —Г–њ—А–∞–≤–ї–µ–љ–Є—П
        buttons_frame = tk.Frame(self.details_frame, bg='white')
        buttons_frame.pack(fill=tk.X, padx=20, pady=10)

        btn_style = {"font": ("Arial", 10, "bold"), "width": 18, "height": 2}

        tk.Button(buttons_frame, text="рЯФД –Ю–±–љ–Њ–≤–Є—В—М", bg='#3498db', fg='white',
                  command=self.refresh_details, **btn_style).pack(side=tk.LEFT, padx=5)

        tk.Button(buttons_frame, text="вЬЦ –°–±—А–Њ—Б–Є—В—М —Д–Є–ї—М—В—А—Л", bg='#e67e22', fg='white',
                  command=self.clear_details_filters, **btn_style).pack(side=tk.LEFT, padx=5)

        tk.Button(buttons_frame, text="рЯУК –≠–Ї—Б–њ–Њ—А—В –≤ Excel", bg='#27ae60', fg='white',
                  command=self.export_details, **btn_style).pack(side=tk.LEFT, padx=5)

        tk.Button(buttons_frame, text="рЯУИ –°—В–∞—В–Є—Б—В–Є–Ї–∞", bg='#9c27b0', fg='white',
                  command=self.show_details_statistics, **btn_style).pack(side=tk.LEFT, padx=5)

        # –°—В–∞—В—Г—Б–љ–∞—П —Б—В—А–Њ–Ї–∞
        self.details_status_label = tk.Label(
            self.details_frame,
            text="",
            font=("Arial", 10),
            bg='#ecf0f1',
            fg='#2c3e50',
            relief=tk.SUNKEN,
            anchor='w',
            padx=10,
            pady=5
        )
        self.details_status_label.pack(fill=tk.X, side=tk.BOTTOM, padx=20, pady=10)

        # –Я–µ—А–≤–Њ–љ–∞—З–∞–ї—М–љ–Њ–µ –Ј–∞–њ–Њ–ї–љ–µ–љ–Є–µ
        self.refresh_details()

    def refresh_details(self):
        """–Ю–±–љ–Њ–≤–ї–µ–љ–Є–µ —В–∞–±–ї–Є—Ж—Л –і–µ—В–∞–ї–µ–є"""

        def safe_int(value, default=0):
            if value == "" or pd.isna(value) or value is None:
                return default
            try:
                return int(value)
            except (ValueError, TypeError):
                return default

        # –°–Ю–•–†–Р–Э–ѓ–Х–Ь –Р–Ъ–Ґ–Ш–Т–Э–Ђ–Х –§–Ш–Ы–ђ–Ґ–†–Ђ –Я–Х–†–Х–Ф –Ю–І–Ш–°–Ґ–Ъ–Ю–Щ
        active_filters_backup = {}
        if hasattr(self, 'details_excel_filter') and self.details_excel_filter.active_filters:
            active_filters_backup = self.details_excel_filter.active_filters.copy()

        # –Ю—З–Є—Й–∞–µ–Љ —В–∞–±–ї–Є—Ж—Г
        for item in self.details_tree.get_children():
            self.details_tree.delete(item)

        # –Ю–І–Ш–©–Р–Х–Ь –Ъ–≠–® –≠–Ы–Х–Ь–Х–Э–Ґ–Ю–Т
        if hasattr(self, 'details_excel_filter'):
            self.details_excel_filter._all_item_cache = set()

        # –Ч–∞–≥—А—Г–ґ–∞–µ–Љ –і–∞–љ–љ—Л–µ
        orders_df = load_data("Orders")
        order_details_df = load_data("OrderDetails")

        if orders_df.empty or order_details_df.empty:
            if hasattr(self, 'details_status_label'):
                self.details_status_label.config(
                    text="вЪ†пЄП –Э–µ—В –і–∞–љ–љ—Л—Е –Њ –і–µ—В–∞–ї—П—Е",
                    bg='#fff3cd',
                    fg='#856404'
                )
            return

        # рЯЖХ –Ю–І–Ш–°–Ґ–Ъ–Р –Я–£–°–Ґ–Ђ–• –Ч–Э–Р–І–Х–Э–Ш–Щ
        order_details_df["–Ъ–Њ–ї–Є—З–µ—Б—В–≤–Њ"] = order_details_df["–Ъ–Њ–ї–Є—З–µ—Б—В–≤–Њ"].replace("", 0)
        order_details_df["–Я–Њ—А–µ–Ј–∞–љ–Њ"] = order_details_df["–Я–Њ—А–µ–Ј–∞–љ–Њ"].replace("", 0)
        order_details_df["–Я–Њ–≥–љ—Г—В–Њ"] = order_details_df["–Я–Њ–≥–љ—Г—В–Њ"].replace("", 0)

        # –°–Њ—Е—А–∞–љ—П–µ–Љ –Њ—З–Є—Й–µ–љ–љ—Л–µ –і–∞–љ–љ—Л–µ
        save_data("OrderDetails", order_details_df)

        # –І–Є—В–∞–µ–Љ –њ–µ—А–µ–Ї–ї—О—З–∞—В–µ–ї–Є
        show_completed = self.details_toggles['show_completed'].get()
        show_in_progress = self.details_toggles['show_in_progress'].get()
        show_not_started = self.details_toggles['show_not_started'].get()

        # –§–Є–ї—М—В—А—Г–µ–Љ –Ј–∞–Ї–∞–Ј—Л —Б–Њ —Б—В–∞—В—Г—Б–Њ–Љ "–Т —А–∞–±–Њ—В–µ"
        active_orders = orders_df[orders_df["–°—В–∞—В—Г—Б"] == "–Т —А–∞–±–Њ—В–µ"]

        if active_orders.empty:
            if hasattr(self, 'details_status_label'):
                self.details_status_label.config(
                    text="вДєпЄП –Э–µ—В –Ј–∞–Ї–∞–Ј–Њ–≤ –≤ —А–∞–±–Њ—В–µ",
                    bg='#d1ecf1',
                    fg='#0c5460'
                )
            return

        # –°—З—С—В—З–Є–Ї–Є
        total_count = 0
        shown_count = 0
        completed_count = 0
        in_progress_count = 0
        not_started_count = 0

        # –Я—А–Њ—Е–Њ–і–Є–Љ –њ–Њ –≤—Б–µ–Љ –і–µ—В–∞–ї—П–Љ –∞–Ї—В–Є–≤–љ—Л—Е –Ј–∞–Ї–∞–Ј–Њ–≤
        for _, order_row in active_orders.iterrows():
            order_id = int(order_row["ID –Ј–∞–Ї–∞–Ј–∞"])
            order_name = order_row["–Э–∞–Ј–≤–∞–љ–Є–µ –Ј–∞–Ї–∞–Ј–∞"]
            customer_name = order_row["–Ч–∞–Ї–∞–Ј—З–Є–Ї"]

            # –Я–Њ–ї—Г—З–∞–µ–Љ –і–µ—В–∞–ї–Є —Н—В–Њ–≥–Њ –Ј–∞–Ї–∞–Ј–∞
            order_details = order_details_df[order_details_df["ID –Ј–∞–Ї–∞–Ј–∞"] == order_id]

            for _, detail_row in order_details.iterrows():
                detail_id = int(detail_row["ID"])
                detail_name = detail_row["–Э–∞–Ј–≤–∞–љ–Є–µ –і–µ—В–∞–ї–Є"]
                quantity = safe_int(detail_row["–Ъ–Њ–ї–Є—З–µ—Б—В–≤–Њ"])
                cut = safe_int(detail_row.get("–Я–Њ—А–µ–Ј–∞–љ–Њ", 0))
                bent = safe_int(detail_row.get("–Я–Њ–≥–љ—Г—В–Њ", 0))

                # –†–∞—Б—Б—З–Є—В—Л–≤–∞–µ–Љ –Њ—Б—В–∞—В–Њ–Ї –Є –њ—А–Њ–≥—А–µ—Б—Б
                remaining = quantity - cut
                progress_pct = round((cut / quantity * 100), 1) if quantity > 0 else 0

                total_count += 1

                # –Ю–њ—А–µ–і–µ–ї—П–µ–Љ —Б—В–∞—В—Г—Б
                if cut >= quantity:
                    status = 'completed'
                    completed_count += 1
                    if not show_completed:
                        continue
                elif cut > 0:
                    status = 'in_progress'
                    in_progress_count += 1
                    if not show_in_progress:
                        continue
                else:
                    status = 'not_started'
                    not_started_count += 1
                    if not show_not_started:
                        continue

                # –§–Њ—А–Љ–Є—А—Г–µ–Љ –Ј–љ–∞—З–µ–љ–Є—П
                values = (
                    detail_id,
                    customer_name,
                    detail_name,
                    order_name,
                    quantity,
                    cut,
                    bent,
                    remaining,
                    f"{progress_pct}%"
                )

                # –¶–≤–µ—В–Њ–≤–∞—П –Є–љ–і–Є–Ї–∞—Ж–Є—П —Б —Г—З—С—В–Њ–Љ –њ–µ—А–µ–њ–Њ—А–µ–Ј–Ї–Є
                if cut > quantity:
                    tag = 'over_cut'
                else:
                    tag = status

                item_id = self.details_tree.insert("", "end", values=values, tags=(tag,))
                shown_count += 1

                # –°–Ю–•–†–Р–Э–ѓ–Х–Ь item_id –Т –Ъ–≠–®
                if hasattr(self, 'details_excel_filter'):
                    if not hasattr(self.details_excel_filter, '_all_item_cache'):
                        self.details_excel_filter._all_item_cache = set()
                    self.details_excel_filter._all_item_cache.add(item_id)

        # –Р–Т–Ґ–Ю–Я–Ю–Ф–С–Ю–† –®–Ш–†–Ш–Э–Ђ –Ъ–Ю–Ы–Ю–Э–Ю–Ъ
        self.auto_resize_columns(self.details_tree, min_width=80, max_width=300)

        # –Я–Х–†–Х–Я–†–Ш–Ь–Х–Э–ѓ–Х–Ь –§–Ш–Ы–ђ–Ґ–†–Ђ
        if active_filters_backup and hasattr(self, 'details_excel_filter'):
            self.details_excel_filter.active_filters = active_filters_backup
            self.details_excel_filter.reapply_all_filters()

        # –Ю–±–љ–Њ–≤–ї—П–µ–Љ —Б—В–∞—В—Г—Б–љ—Г—О —Б—В—А–Њ–Ї—Г
        if hasattr(self, 'details_status_label'):
            status_text = (
                f"рЯУК –Ю—В–Њ–±—А–∞–ґ–µ–љ–Њ: {shown_count} –Є–Ј {total_count} | "
                f"рЯЯҐ –Ч–∞–≤–µ—А—И–µ–љ–Њ: {completed_count} | "
                f"рЯЯ° –Т –њ—А–Њ—Ж–µ—Б—Б–µ: {in_progress_count} | "
                f"вЪ™ –Э–µ –љ–∞—З–∞—В–Њ: {not_started_count}"
            )

            self.details_status_label.config(
                text=status_text,
                bg='#d4edda',
                fg='#155724'
            )

        def safe_int(value, default=0):
            if value == "" or pd.isna(value) or value is None:
                return default
            try:
                return int(value)
            except (ValueError, TypeError):
                return default

        for item in self.details_tree.get_children():
            self.details_tree.delete(item)

        order_details_df = load_data("OrderDetails")
        orders_df = load_data("Orders")

        if order_details_df.empty:
            return

        # рЯЖХ –Ю–І–Ш–°–Ґ–Ъ–Р –Я–£–°–Ґ–Ђ–• –Ч–Э–Р–І–Х–Э–Ш–Щ
        order_details_df["–Ъ–Њ–ї–Є—З–µ—Б—В–≤–Њ"] = order_details_df["–Ъ–Њ–ї–Є—З–µ—Б—В–≤–Њ"].replace("", 0)
        order_details_df["–Я–Њ—А–µ–Ј–∞–љ–Њ"] = order_details_df["–Я–Њ—А–µ–Ј–∞–љ–Њ"].replace("", 0)
        order_details_df["–Я–Њ–≥–љ—Г—В–Њ"] = order_details_df["–Я–Њ–≥–љ—Г—В–Њ"].replace("", 0)

        # –°–Њ—Е—А–∞–љ—П–µ–Љ –Њ—З–Є—Й–µ–љ–љ—Л–µ –і–∞–љ–љ—Л–µ
        save_data("OrderDetails", order_details_df)

        """–Ю–±–љ–Њ–≤–ї–µ–љ–Є–µ —В–∞–±–ї–Є—Ж—Л —Г—З—С—В–∞ –і–µ—В–∞–ї–µ–є"""

        # –Ю—З–Є—Й–∞–µ–Љ —В–∞–±–ї–Є—Ж—Г
        for item in self.details_tree.get_children():
            self.details_tree.delete(item)

        # –Ч–∞–≥—А—Г–ґ–∞–µ–Љ –і–∞–љ–љ—Л–µ
        orders_df = load_data("Orders")
        order_details_df = load_data("OrderDetails")

        if orders_df.empty or order_details_df.empty:
            if hasattr(self, 'details_status_label'):
                self.details_status_label.config(
                    text="вЪ†пЄП –Э–µ—В –і–∞–љ–љ—Л—Е –Њ –і–µ—В–∞–ї—П—Е",
                    bg='#fff3cd',
                    fg='#856404'
                )
            return

        # –І–Є—В–∞–µ–Љ –њ–µ—А–µ–Ї–ї—О—З–∞—В–µ–ї–Є
        show_completed = self.details_toggles['show_completed'].get()
        show_in_progress = self.details_toggles['show_in_progress'].get()
        show_not_started = self.details_toggles['show_not_started'].get()

        # –§–Є–ї—М—В—А—Г–µ–Љ –Ј–∞–Ї–∞–Ј—Л —Б–Њ —Б—В–∞—В—Г—Б–Њ–Љ "–Т —А–∞–±–Њ—В–µ"
        active_orders = orders_df[orders_df["–°—В–∞—В—Г—Б"] == "–Т —А–∞–±–Њ—В–µ"]

        if active_orders.empty:
            if hasattr(self, 'details_status_label'):
                self.details_status_label.config(
                    text="вДєпЄП –Э–µ—В –Ј–∞–Ї–∞–Ј–Њ–≤ –≤ —А–∞–±–Њ—В–µ",
                    bg='#d1ecf1',
                    fg='#0c5460'
                )
            return

        # –°—З—С—В—З–Є–Ї–Є
        total_count = 0
        shown_count = 0
        completed_count = 0
        in_progress_count = 0
        not_started_count = 0

        # –Я–Њ–ї—Г—З–∞–µ–Љ –∞–Ї—В–Є–≤–љ—Л–µ —Д–Є–ї—М—В—А—Л
        active_filters = {}
        if hasattr(self, 'details_filters') and self.details_filters:
            for col_name, filter_var in self.details_filters.items():
                filter_text = filter_var.get().strip().lower()
                if filter_text:
                    active_filters[col_name] = filter_text

        # –Я—А–Њ—Е–Њ–і–Є–Љ –њ–Њ –≤—Б–µ–Љ –і–µ—В–∞–ї—П–Љ –∞–Ї—В–Є–≤–љ—Л—Е –Ј–∞–Ї–∞–Ј–Њ–≤
        for _, order_row in active_orders.iterrows():
            order_id = int(order_row["ID –Ј–∞–Ї–∞–Ј–∞"])
            order_name = order_row["–Э–∞–Ј–≤–∞–љ–Є–µ –Ј–∞–Ї–∞–Ј–∞"]
            customer_name = order_row["–Ч–∞–Ї–∞–Ј—З–Є–Ї"]

            # –Я–Њ–ї—Г—З–∞–µ–Љ –і–µ—В–∞–ї–Є —Н—В–Њ–≥–Њ –Ј–∞–Ї–∞–Ј–∞
            order_details = order_details_df[order_details_df["ID –Ј–∞–Ї–∞–Ј–∞"] == order_id]

            for _, detail_row in order_details.iterrows():
                detail_id = int(detail_row["ID"])
                detail_name = detail_row["–Э–∞–Ј–≤–∞–љ–Є–µ –і–µ—В–∞–ї–Є"]
                quantity = int(detail_row["–Ъ–Њ–ї–Є—З–µ—Б—В–≤–Њ"])
                cut = int(detail_row.get("–Я–Њ—А–µ–Ј–∞–љ–Њ", 0))
                bent = int(detail_row.get("–Я–Њ–≥–љ—Г—В–Њ", 0))

                # –†–∞—Б—Б—З–Є—В—Л–≤–∞–µ–Љ –Њ—Б—В–∞—В–Њ–Ї –Є –њ—А–Њ–≥—А–µ—Б—Б
                remaining = quantity - cut
                progress_pct = round((cut / quantity * 100), 1) if quantity > 0 else 0

                total_count += 1

                # –Ю–њ—А–µ–і–µ–ї—П–µ–Љ —Б—В–∞—В—Г—Б
                if cut >= quantity:
                    status = 'completed'
                    completed_count += 1
                    if not show_completed:
                        continue
                elif cut > 0:
                    status = 'in_progress'
                    in_progress_count += 1
                    if not show_in_progress:
                        continue
                else:
                    status = 'not_started'
                    not_started_count += 1
                    if not show_not_started:
                        continue

                # –§–Њ—А–Љ–Є—А—Г–µ–Љ –Ј–љ–∞—З–µ–љ–Є—П
                values = (
                    detail_id,
                    customer_name,
                    detail_name,
                    order_name,
                    quantity,
                    cut,
                    bent,
                    remaining,
                    f"{progress_pct}%"
                )

                # –Я—А–Є–Љ–µ–љ—П–µ–Љ —Д–Є–ї—М—В—А—Л
                if active_filters:
                    skip_row = False

                    if "–Ч–∞–Ї–∞–Ј—З–Є–Ї" in active_filters:
                        if active_filters["–Ч–∞–Ї–∞–Ј—З–Є–Ї"] not in customer_name.lower():
                            skip_row = True

                    if "–Э–∞–Ј–≤–∞–љ–Є–µ –і–µ—В–∞–ї–Є" in active_filters:
                        if active_filters["–Э–∞–Ј–≤–∞–љ–Є–µ –і–µ—В–∞–ї–Є"] not in detail_name.lower():
                            skip_row = True

                    if "–Ч–∞–Ї–∞–Ј" in active_filters:
                        if active_filters["–Ч–∞–Ї–∞–Ј"] not in order_name.lower():
                            skip_row = True

                    if skip_row:
                        continue

                # –¶–≤–µ—В–Њ–≤–∞—П –Є–љ–і–Є–Ї–∞—Ж–Є—П —Б —Г—З—С—В–Њ–Љ –њ–µ—А–µ–њ–Њ—А–µ–Ј–Ї–Є
                if cut > quantity:
                    tag = 'over_cut'  # –Я–Њ—А–µ–Ј–∞–љ–Њ –±–Њ–ї—М—И–µ —З–µ–Љ –љ—Г–ґ–љ–Њ
                else:
                    tag = status

                item_id = self.details_tree.insert("", "end", values=values, tags=(tag,))
                shown_count += 1

                # –°–Ю–•–†–Р–Э–ѓ–Х–Ь item_id –Т –Ъ–≠–®
                if hasattr(self, 'details_excel_filter'):
                    if not hasattr(self.details_excel_filter, '_all_item_cache'):
                        self.details_excel_filter._all_item_cache = set()
                    self.details_excel_filter._all_item_cache.add(item_id)

        # –Р–Т–Ґ–Ю–Я–Ю–Ф–С–Ю–† –®–Ш–†–Ш–Э–Ђ –Ъ–Ю–Ы–Ю–Э–Ю–Ъ
        self.auto_resize_columns(self.details_tree, min_width=80, max_width=300)

        # –Я–Х–†–Х–Я–†–Ш–Ь–Х–Э–ѓ–Х–Ь –§–Ш–Ы–ђ–Ґ–†–Ђ
        if active_filters_backup and hasattr(self, 'details_excel_filter'):
            self.details_excel_filter.active_filters = active_filters_backup
            self.details_excel_filter.reapply_all_filters()

        # –Ю–±–љ–Њ–≤–ї—П–µ–Љ —Б—В–∞—В—Г—Б–љ—Г—О —Б—В—А–Њ–Ї—Г
        if hasattr(self, 'details_status_label'):
            status_text = (
                f"рЯУК –Ю—В–Њ–±—А–∞–ґ–µ–љ–Њ: {shown_count} –Є–Ј {total_count} | "
                f"рЯЯҐ –Ч–∞–≤–µ—А—И–µ–љ–Њ: {completed_count} | "
                f"рЯЯ° –Т –њ—А–Њ—Ж–µ—Б—Б–µ: {in_progress_count} | "
                f"вЪ™ –Э–µ –љ–∞—З–∞—В–Њ: {not_started_count}"
            )

            self.details_status_label.config(
                text=status_text,
                bg='#d4edda',
                fg='#155724'
            )

    def export_details(self):
        """–≠–Ї—Б–њ–Њ—А—В —Г—З—С—В–∞ –і–µ—В–∞–ї–µ–є –≤ Excel"""

        if not self.details_tree.get_children():
            messagebox.showwarning("–Я—А–µ–і—Г–њ—А–µ–ґ–і–µ–љ–Є–µ", "–Э–µ—В –і–∞–љ–љ—Л—Е –і–ї—П —Н–Ї—Б–њ–Њ—А—В–∞!")
            return

        file_path = filedialog.asksaveasfilename(
            title="–°–Њ—Е—А–∞–љ–Є—В—М —Г—З—С—В –і–µ—В–∞–ї–µ–є",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialfile=f"details_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )

        if not file_path:
            return

        try:
            # –°–Њ–±–Є—А–∞–µ–Љ –і–∞–љ–љ—Л–µ –Є–Ј —В–∞–±–ї–Є—Ж—Л
            data = []
            for item in self.details_tree.get_children():
                values = self.details_tree.item(item)['values']
                data.append(values)

            # –°–Њ–Ј–і–∞—С–Љ DataFrame
            columns = ["ID", "–Ч–∞–Ї–∞–Ј—З–Є–Ї", "–Э–∞–Ј–≤–∞–љ–Є–µ –і–µ—В–∞–ї–Є", "–Ч–∞–Ї–∞–Ј", "–Ъ–Њ–ї–Є—З–µ—Б—В–≤–Њ", "–Я–Њ—А–µ–Ј–∞–љ–Њ", "–Я–Њ–≥–љ—Г—В–Њ", "–Ю—Б—В–∞–ї–Њ—Б—М",
                       "–Я—А–Њ–≥—А–µ—Б—Б %"]
            df = pd.DataFrame(data, columns=columns)

            # –°–Њ—Е—А–∞–љ—П–µ–Љ –≤ Excel
            df.to_excel(file_path, index=False, engine='openpyxl')

            messagebox.showinfo("–£—Б–њ–µ—Е", f"–£—З—С—В –і–µ—В–∞–ї–µ–є —Б–Њ—Е—А–∞–љ—С–љ:\n{file_path}")

        except Exception as e:
            messagebox.showerror("–Ю—И–Є–±–Ї–∞", f"–Э–µ —Г–і–∞–ї–Њ—Б—М —Б–Њ—Е—А–∞–љ–Є—В—М —Д–∞–є–ї:\n{e}")

    def show_details_statistics(self):
        """–Я–Њ–Ї–∞–Ј–∞—В—М –і–µ—В–∞–ї—М–љ—Г—О —Б—В–∞—В–Є—Б—В–Є–Ї—Г –њ–Њ –і–µ—В–∞–ї—П–Љ"""

        orders_df = load_data("Orders")
        order_details_df = load_data("OrderDetails")

        if orders_df.empty or order_details_df.empty:
            messagebox.showinfo("–°—В–∞—В–Є—Б—В–Є–Ї–∞", "–Э–µ—В –і–∞–љ–љ—Л—Е –і–ї—П –∞–љ–∞–ї–Є–Ј–∞")
            return

        # –§–Є–ї—М—В—А—Г–µ–Љ –Ј–∞–Ї–∞–Ј—Л –≤ —А–∞–±–Њ—В–µ
        active_orders = orders_df[orders_df["–°—В–∞—В—Г—Б"] == "–Т —А–∞–±–Њ—В–µ"]

        if active_orders.empty:
            messagebox.showinfo("–°—В–∞—В–Є—Б—В–Є–Ї–∞", "–Э–µ—В –Ј–∞–Ї–∞–Ј–Њ–≤ –≤ —А–∞–±–Њ—В–µ")
            return

        # –°–Њ–±–Є—А–∞–µ–Љ —Б—В–∞—В–Є—Б—В–Є–Ї—Г
        total_details = 0
        total_qty = 0
        total_cut = 0
        total_bent = 0

        completed_details = 0
        in_progress_details = 0
        not_started_details = 0

        by_customer = {}

        for _, order_row in active_orders.iterrows():
            order_id = int(order_row["ID –Ј–∞–Ї–∞–Ј–∞"])
            customer = order_row["–Ч–∞–Ї–∞–Ј—З–Є–Ї"]

            order_details = order_details_df[order_details_df["ID –Ј–∞–Ї–∞–Ј–∞"] == order_id]

            for _, detail_row in order_details.iterrows():
                qty = int(detail_row["–Ъ–Њ–ї–Є—З–µ—Б—В–≤–Њ"])
                cut = int(detail_row.get("–Я–Њ—А–µ–Ј–∞–љ–Њ", 0))
                bent = int(detail_row.get("–Я–Њ–≥–љ—Г—В–Њ", 0))

                total_details += 1
                total_qty += qty
                total_cut += cut
                total_bent += bent

                # –Ю–њ—А–µ–і–µ–ї—П–µ–Љ —Б—В–∞—В—Г—Б
                if cut >= qty:
                    completed_details += 1
                elif cut > 0:
                    in_progress_details += 1
                else:
                    not_started_details += 1

                # –У—А—Г–њ–њ–Є—А—Г–µ–Љ –њ–Њ –Ј–∞–Ї–∞–Ј—З–Є–Ї—Г
                if customer not in by_customer:
                    by_customer[customer] = {
                        'details': 0,
                        'quantity': 0,
                        'cut': 0,
                        'bent': 0
                    }

                by_customer[customer]['details'] += 1
                by_customer[customer]['quantity'] += qty
                by_customer[customer]['cut'] += cut
                by_customer[customer]['bent'] += bent

        # –Ю–±—Й–Є–є –њ—А–Њ–≥—А–µ—Б—Б
        overall_progress = round((total_cut / total_qty * 100), 1) if total_qty > 0 else 0

        # –§–Њ—А–Љ–Є—А—Г–µ–Љ —Б–Њ–Њ–±—Й–µ–љ–Є–µ
        stats_msg = (
            f"рЯУК –Ю–С–©–Р–ѓ –°–Ґ–Р–Ґ–Ш–°–Ґ–Ш–Ъ–Р\n"
            f"{'=' * 50}\n\n"
            f"рЯУР –£–љ–Є–Ї–∞–ї—М–љ—Л—Е –і–µ—В–∞–ї–µ–є: {total_details}\n"
            f"рЯУ¶ –Т—Б–µ–≥–Њ —В—А–µ–±—Г–µ—В—Б—П –њ–Њ—А–µ–Ј–∞—В—М: {total_qty} —И—В\n"
            f"вЬВпЄП –Я–Њ—А–µ–Ј–∞–љ–Њ: {total_cut} —И—В ({overall_progress}%)\n"
            f"рЯФІ –Я–Њ–≥–љ—Г—В–Њ: {total_bent} —И—В\n"
            f"вП≥ –Ю—Б—В–∞–ї–Њ—Б—М –њ–Њ—А–µ–Ј–∞—В—М: {total_qty - total_cut} —И—В\n\n"
            f"{'=' * 50}\n\n"
            f"рЯУИ –Я–Ю –°–Ґ–Р–Ґ–£–°–Р–Ь:\n\n"
            f"рЯЯҐ –Ч–∞–≤–µ—А—И–µ–љ–Њ: {completed_details} –і–µ—В–∞–ї–µ–є\n"
            f"рЯЯ° –Т –њ—А–Њ—Ж–µ—Б—Б–µ: {in_progress_details} –і–µ—В–∞–ї–µ–є\n"
            f"вЪ™ –Э–µ –љ–∞—З–∞—В–Њ: {not_started_details} –і–µ—В–∞–ї–µ–є\n\n"
            f"{'=' * 50}\n\n"
            f"рЯС• –Я–Ю –Ч–Р–Ъ–Р–Ч–І–Ш–Ъ–Р–Ь:\n\n"
        )

        # –°–Њ—А—В–Є—А—Г–µ–Љ –Ј–∞–Ї–∞–Ј—З–Є–Ї–Њ–≤ –њ–Њ –Ї–Њ–ї–Є—З–µ—Б—В–≤—Г –і–µ—В–∞–ї–µ–є
        sorted_customers = sorted(by_customer.items(), key=lambda x: x[1]['quantity'], reverse=True)

        for customer, stats in sorted_customers[:10]:  # –Я–Њ–Ї–∞–Ј—Л–≤–∞–µ–Љ —В–Њ–њ-10
            customer_progress = round((stats['cut'] / stats['quantity'] * 100), 1) if stats['quantity'] > 0 else 0
            stats_msg += (
                f"\n{customer}:\n"
                f"  –Ф–µ—В–∞–ї–µ–є: {stats['details']}\n"
                f"  –Ґ—А–µ–±—Г–µ—В—Б—П: {stats['quantity']} —И—В\n"
                f"  –Я–Њ—А–µ–Ј–∞–љ–Њ: {stats['cut']} —И—В ({customer_progress}%)\n"
                f"  –Я–Њ–≥–љ—Г—В–Њ: {stats['bent']} —И—В\n"
            )

        if len(by_customer) > 10:
            stats_msg += f"\n... –Є –µ—Й–µ {len(by_customer) - 10} –Ј–∞–Ї–∞–Ј—З–Є–Ї–Њ–≤"

        # –°–Њ–Ј–і–∞—С–Љ –Њ–Ї–љ–Њ —Б–Њ —Б—В–∞—В–Є—Б—В–Є–Ї–Њ–є
        stats_window = tk.Toplevel(self.root)
        stats_window.title("рЯУК –°—В–∞—В–Є—Б—В–Є–Ї–∞ –њ–Њ –і–µ—В–∞–ї—П–Љ")
        stats_window.geometry("600x700")
        stats_window.configure(bg='#f0f0f0')

        # –Ґ–µ–Ї—Б—В–Њ–≤–Њ–µ –њ–Њ–ї–µ —Б–Њ —Б–Ї—А–Њ–ї–ї–Њ–Љ
        text_frame = tk.Frame(stats_window, bg='white')
        text_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        scroll = tk.Scrollbar(text_frame)
        scroll.pack(side=tk.RIGHT, fill=tk.Y)

        text_widget = tk.Text(text_frame, wrap=tk.WORD, font=("Consolas", 10),
                              yscrollcommand=scroll.set)
        text_widget.insert("1.0", stats_msg)
        text_widget.config(state=tk.DISABLED)
        text_widget.pack(fill=tk.BOTH, expand=True)

        scroll.config(command=text_widget.yview)

        # –Ъ–љ–Њ–њ–Ї–∞ –Ј–∞–Ї—А—Л—В–Є—П
        tk.Button(stats_window, text="–Ч–∞–Ї—А—Л—В—М", command=stats_window.destroy,
                  bg='#3498db', fg='white', font=("Arial", 12, "bold"),
                  width=20, height=2).pack(pady=10)

    def import_laser_writeoff_table(self):
        """–Ш–Љ–њ–Њ—А—В —В–∞–±–ї–Є—Ж—Л –Њ—В –ї–∞–Ј–µ—А—Й–Є–Ї–Њ–≤"""
        file_path = filedialog.askopenfilename(
            title="–Т—Л–±–µ—А–Є—В–µ —В–∞–±–ї–Є—Ж—Г –Њ—В –ї–∞–Ј–µ—А—Й–Є–Ї–Њ–≤",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )

        if not file_path:
            return

        try:
            import_df = pd.read_excel(file_path, engine='openpyxl')

            # –Я—А–Њ–≤–µ—А–Ї–∞ –Ї–Њ–ї–Њ–љ–Њ–Ї
            required_cols = ["–Ф–∞—В–∞ (–Ь–°–Ъ)", "–Т—А–µ–Љ—П (–Ь–°–Ъ)", "username", "order", "metal",
                             "metal_quantity", "part", "part_quantity"]
            missing = [col for col in required_cols if col not in import_df.columns]

            if missing:
                messagebox.showerror("–Ю—И–Є–±–Ї–∞", f"–Ю—В—Б—Г—В—Б—В–≤—Г—О—В –Ї–Њ–ї–Њ–љ–Ї–Є:\n{', '.join(missing)}")
                return

            # –°–Њ—Е—А–∞–љ—П–µ–Љ –і–∞–љ–љ—Л–µ
            self.laser_import_data = import_df.to_dict('records')

            # –Ю—В–Њ–±—А–∞–ґ–∞–µ–Љ
            self.refresh_laser_import_table()

            messagebox.showinfo("–£—Б–њ–µ—Е", f"–Ч–∞–≥—А—Г–ґ–µ–љ–Њ {len(self.laser_import_data)} –Ј–∞–њ–Є—Б–µ–є")

        except Exception as e:
            messagebox.showerror("–Ю—И–Є–±–Ї–∞", f"–Э–µ —Г–і–∞–ї–Њ—Б—М –Є–Љ–њ–Њ—А—В–Є—А–Њ–≤–∞—В—М:\n{e}")

    def refresh_laser_import_table(self):
        """–Ю–±–љ–Њ–≤–ї–µ–љ–Є–µ —В–∞–±–ї–Є—Ж—Л –Є–Љ–њ–Њ—А—В–∞ –Њ—В –ї–∞–Ј–µ—А—Й–Є–Ї–Њ–≤"""
        # –Ю—З–Є—Й–∞–µ–Љ —В–∞–±–ї–Є—Ж—Г
        for item in self.laser_import_tree.get_children():
            self.laser_import_tree.delete(item)

        if not hasattr(self, 'laser_table_data') or self.laser_table_data is None:
            self.laser_table_data = []
            return

        if not self.laser_table_data:
            return

        # рЯФ• –°–Ю–†–Ґ–Ш–†–Ю–Т–Ъ–Р –° –Я–†–Р–Т–Ш–Ы–ђ–Э–Ђ–Ь –§–Ю–†–Ь–Р–Ґ–Ю–Ь –Ф–Р–Ґ–Ђ
        try:
            print(f"рЯФД –°–Њ—А—В–Є—А–Њ–≤–Ї–∞ {len(self.laser_table_data)} –Ј–∞–њ–Є—Б–µ–є –њ–µ—А–µ–і –Њ—В–Њ–±—А–∞–ґ–µ–љ–Є–µ–Љ...")

            # –Я—А–µ–Њ–±—А–∞–Ј—Г–µ–Љ –≤ DataFrame
            df_display = pd.DataFrame(self.laser_table_data)

            # рЯЖХ –Я–†–Р–Т–Ш–Ы–ђ–Э–Ђ–Щ –Я–Р–†–°–Ш–Э–У –Ф–Р–Ґ–Ђ: –§–Ю–†–Ь–Р–Ґ DD.MM.YYYY
            df_display['_datetime_sort'] = pd.to_datetime(
                df_display['–Ф–∞—В–∞ (–Ь–°–Ъ)'].astype(str) + ' ' + df_display['–Т—А–µ–Љ—П (–Ь–°–Ъ)'].astype(str),
                format='%d.%m.%Y %H:%M:%S',  # вЖР –ѓ–Т–Э–Ю –£–Ъ–Р–Ч–Ђ–Т–Р–Х–Ь –§–Ю–†–Ь–Р–Ґ
                errors='coerce'
            )

            # –°–Њ—А—В–Є—А—Г–µ–Љ –њ–Њ —Г–±—Л–≤–∞–љ–Є—О (–љ–Њ–≤—Л–µ —Б–≤–µ—А—Е—Г)
            df_display = df_display.sort_values('_datetime_sort', ascending=False, na_position='last')

            # –£–і–∞–ї—П–µ–Љ –≤—А–µ–Љ–µ–љ–љ—Г—О –Ї–Њ–ї–Њ–љ–Ї—Г
            df_display = df_display.drop('_datetime_sort', axis=1)

            # –Я—А–µ–Њ–±—А–∞–Ј—Г–µ–Љ –Њ–±—А–∞—В–љ–Њ –≤ —Б–њ–Є—Б–Њ–Ї —Б–ї–Њ–≤–∞—А–µ–є
            sorted_data = df_display.to_dict('records')

            # –Я–Њ–Ї–∞–Ј—Л–≤–∞–µ–Љ –њ–µ—А–≤—Г—О –Є –њ–Њ—Б–ї–µ–і–љ—О—О –Ј–∞–њ–Є—Б—М
            if sorted_data:
                first = f"{sorted_data[0].get('–Ф–∞—В–∞ (–Ь–°–Ъ)', '')} {sorted_data[0].get('–Т—А–µ–Љ—П (–Ь–°–Ъ)', '')}"
                last = f"{sorted_data[-1].get('–Ф–∞—В–∞ (–Ь–°–Ъ)', '')} {sorted_data[-1].get('–Т—А–µ–Љ—П (–Ь–°–Ъ)', '')}"
                print(f"вЬЕ –Ю—В—Б–Њ—А—В–Є—А–Њ–≤–∞–љ–Њ: –Я–Х–†–Т–Р–ѓ (–љ–Њ–≤–∞—П) = {first}, –Я–Ю–°–Ы–Х–Ф–Э–ѓ–ѓ (—Б—В–∞—А–∞—П) = {last}")
        except Exception as e:
            print(f"вЪ†пЄП –Ю—И–Є–±–Ї–∞ —Б–Њ—А—В–Є—А–Њ–≤–Ї–Є (—Д–Њ—А–Љ–∞—В DD.MM.YYYY): {e}")

            # рЯЖХ –Я–Ю–Я–†–Ю–С–£–Х–Ь –Р–Ы–ђ–Ґ–Х–†–Э–Р–Ґ–Ш–Т–Э–Ђ–Щ –§–Ю–†–Ь–Р–Ґ
            try:
                print("рЯФД –Я—А–Њ–±—Г–µ–Љ –∞–ї—М—В–µ—А–љ–∞—В–Є–≤–љ—Л–є —Д–Њ—А–Љ–∞—В YYYY-MM-DD...")
                df_display = pd.DataFrame(self.laser_table_data)
                df_display['_datetime_sort'] = pd.to_datetime(
                    df_display['–Ф–∞—В–∞ (–Ь–°–Ъ)'].astype(str) + ' ' + df_display['–Т—А–µ–Љ—П (–Ь–°–Ъ)'].astype(str),
                    errors='coerce'
                )
                df_display = df_display.sort_values('_datetime_sort', ascending=False, na_position='last')
                df_display = df_display.drop('_datetime_sort', axis=1)
                sorted_data = df_display.to_dict('records')
                print("вЬЕ –Р–ї—М—В–µ—А–љ–∞—В–Є–≤–љ—Л–є —Д–Њ—А–Љ–∞—В —Б—А–∞–±–Њ—В–∞–ї!")
            except Exception as e2:
                print(f"вЪ†пЄП –Ш –∞–ї—М—В–µ—А–љ–∞—В–Є–≤–љ—Л–є —Д–Њ—А–Љ–∞—В –љ–µ —Б—А–∞–±–Њ—В–∞–ї: {e2}")
                import traceback
                traceback.print_exc()
                sorted_data = self.laser_table_data

        # –°–І–Б–Ґ–І–Ш–Ъ–Ш
        manual_count = 0
        auto_count = 0
        pending_count = 0

        # –Ч–∞–њ–Њ–ї–љ—П–µ–Љ —В–∞–±–ї–Є—Ж—Г –Ю–Ґ–°–Ю–†–Ґ–Ш–†–Ю–Т–Р–Э–Э–Ђ–Ь–Ш –і–∞–љ–љ—Л–Љ–Є
        for idx, row_data in enumerate(sorted_data):
            date_val = row_data.get("–Ф–∞—В–∞ (–Ь–°–Ъ)", "")
            time_val = row_data.get("–Т—А–µ–Љ—П (–Ь–°–Ъ)", "")
            username = row_data.get("username", "")
            order = row_data.get("order", "")
            metal = row_data.get("metal", "")
            metal_qty = row_data.get("metal_quantity", "")
            part = row_data.get("part", "")
            part_qty = row_data.get("part_quantity", "")
            written_off = row_data.get("–°–њ–Є—Б–∞–љ–Њ", "")
            writeoff_date = row_data.get("–Ф–∞—В–∞ —Б–њ–Є—Б–∞–љ–Є—П", "")

            # –С–Х–Ч–Ю–Я–Р–°–Э–Ю–Х –Я–†–Х–Ю–С–†–Р–Ч–Ю–Т–Р–Э–Ш–Х written_off –Т –°–Ґ–†–Ю–Ъ–£
            if pd.isna(written_off) or written_off is None:
                written_off = ""
            else:
                written_off = str(written_off).strip()

            values = (date_val, time_val, username, order, metal, metal_qty, part, part_qty, written_off, writeoff_date)

            # –¶–Т–Х–Ґ–Ю–Т–Р–ѓ –Ш–Э–Ф–Ш–Ъ–Р–¶–Ш–ѓ
            if written_off == "–Т—А—Г—З–љ—Г—О":
                tag = 'manual'
                manual_count += 1
            elif written_off in ["–Ф–∞", "вЬУ", "Yes"]:
                tag = 'written_off'
                auto_count += 1
            else:
                tag = 'pending'
                pending_count += 1

            item_id = self.laser_import_tree.insert("", "end", values=values, tags=(tag,))

            # –°–Ю–•–†–Р–Э–ѓ–Х–Ь item_id –Т –Ъ–≠–®
            if hasattr(self, 'laser_import_excel_filter'):
                if not hasattr(self.laser_import_excel_filter, '_all_item_cache'):
                    self.laser_import_excel_filter._all_item_cache = set()
                self.laser_import_excel_filter._all_item_cache.add(item_id)

        # –Р–Т–Ґ–Ю–Я–Ю–Ф–С–Ю–† –®–Ш–†–Ш–Э–Ђ –Ъ–Ю–Ы–Ю–Э–Ю–Ъ
        self.auto_resize_columns(self.laser_import_tree, min_width=80, max_width=400)

        # –Я–Х–†–Х–Я–†–Ш–Ь–Х–Э–ѓ–Х–Ь –§–Ш–Ы–ђ–Ґ–†–Ђ
        if active_filters_backup and hasattr(self, 'laser_import_excel_filter'):
            self.laser_import_excel_filter.active_filters = active_filters_backup
            self.laser_import_excel_filter.reapply_all_filters()

        print(f"рЯУК –Ю—В–Њ–±—А–∞–ґ–µ–љ–Њ: рЯФµ –°–Є–љ–Є—Е={manual_count}, рЯЯҐ –Ч–µ–ї—С–љ—Л—Е={auto_count}, рЯЯ° –Ц—С–ї—В—Л—Е={pending_count}")

    def writeoff_selected_laser_row(self):
        """–°–њ–Є—Б–∞–љ–Є–µ –≤—Л–±—А–∞–љ–љ–Њ–є —Б—В—А–Њ–Ї–Є"""
        selected = self.laser_import_tree.selection()
        if not selected:
            messagebox.showwarning("–Я—А–µ–і—Г–њ—А–µ–ґ–і–µ–љ–Є–µ", "–Т—Л–±–µ—А–Є—В–µ —Б—В—А–Њ–Ї—Г –і–ї—П —Б–њ–Є—Б–∞–љ–Є—П")
            return

        item = selected[0]
        row_index = self.laser_import_tree.index(item)

        self.process_laser_writeoff(row_index)
        self.refresh_laser_import_table()

    def writeoff_all_laser_rows(self):
        """–Ь–∞—Б—Б–Њ–≤–Њ–µ —Б–њ–Є—Б–∞–љ–Є–µ –≤—Б–µ—Е —Б—В—А–Њ–Ї"""
        if not hasattr(self, 'laser_import_data') or not self.laser_import_data:
            messagebox.showwarning("–Я—А–µ–і—Г–њ—А–µ–ґ–і–µ–љ–Є–µ", "–Ґ–∞–±–ї–Є—Ж–∞ –њ—Г—Б—В–∞")
            return

        if not messagebox.askyesno("–Я–Њ–і—В–≤–µ—А–ґ–і–µ–љ–Є–µ",
                                   f"–°–њ–Є—Б–∞—В—М –≤—Б–µ –Ј–∞–њ–Є—Б–Є ({len(self.laser_import_data)} —И—В)?"):
            return

        success = 0
        errors = 0

        for idx in range(len(self.laser_import_data)):
            if self.process_laser_writeoff(idx, silent=True):
                success += 1
            else:
                errors += 1

        self.refresh_laser_import_table()
        self.refresh_writeoffs()
        self.refresh_reservations()
        self.refresh_materials()

        messagebox.showinfo("–†–µ–Ј—Г–ї—М—В–∞—В", f"вЬЕ –°–њ–Є—Б–∞–љ–Њ: {success}\nвЭМ –Ю—И–Є–±–Њ–Ї: {errors}")

    def process_laser_writeoff(self, row_index, silent=False):
        """–Ю–±—А–∞–±–Њ—В–Ї–∞ –Њ–і–љ–Њ–є —Б—В—А–Њ–Ї–Є —Б–њ–Є—Б–∞–љ–Є—П"""
        if row_index >= len(self.laser_import_data):
            return False

        row_data = self.laser_import_data[row_index]

        # –Я—А–Њ–≤–µ—А–Ї–∞: —Г–ґ–µ —Б–њ–Є—Б–∞–љ–Њ?
        if row_data.get("_status") == "вЬЕ –°–њ–Є—Б–∞–љ–Њ":
            if not silent:
                messagebox.showwarning("–Т–љ–Є–Љ–∞–љ–Є–µ", "–£–ґ–µ —Б–њ–Є—Б–∞–љ–Њ!")
            return False

        try:
            # –Ш–Ј–≤–ї–µ–Ї–∞–µ–Љ –і–∞–љ–љ—Л–µ
            order_name = str(row_data.get("order", "")).strip()
            metal_description = str(row_data.get("metal", "")).strip()

            try:
                metal_qty = int(float(row_data.get("metal_quantity", 0)))
            except:
                row_data["_status"] = "–Ю—И–Є–±–Ї–∞: –љ–µ–Ї–Њ—А—А–µ–Ї—В–љ–Њ–µ –Ї–Њ–ї–Є—З–µ—Б—В–≤–Њ"
                return False

            part_name = str(row_data.get("part", "")).strip()
            username = str(row_data.get("username", "")).strip()
            date_str = str(row_data.get("–Ф–∞—В–∞ (–Ь–°–Ъ)", ""))
            time_str = str(row_data.get("–Т—А–µ–Љ—П (–Ь–°–Ъ)", ""))

            # –Я–Њ–Є—Б–Ї –Ј–∞–Ї–∞–Ј–∞
            orders_df = load_data("Orders")
            import re
            match = re.search(r'–£–Я-(\d+)', order_name)
            order_id = None

            if match:
                up_number = match.group(1)
                order_match = orders_df[orders_df["–Э–∞–Ј–≤–∞–љ–Є–µ –Ј–∞–Ї–∞–Ј–∞"].str.contains(
                    f"–£–Я-{up_number}", case=False, na=False, regex=False)]
                if not order_match.empty:
                    order_id = int(order_match.iloc[0]["ID –Ј–∞–Ї–∞–Ј–∞"])

            if not order_id:
                row_data["_status"] = f"–Ю—И–Є–±–Ї–∞: –Ј–∞–Ї–∞–Ј '{order_name}' –љ–µ –љ–∞–є–і–µ–љ"
                return False

            # –Я–∞—А—Б–Є–љ–≥ —А–∞–Ј–Љ–µ—А–Њ–≤
            import re
            thickness = None
            width = None
            length = None

            print(f"   рЯФН –Я–∞—А—Б–Є–љ–≥ –Љ–∞—В–µ—А–Є–∞–ї–∞: '{metal_description}'")
            print(f"   рЯУП –Ф–ї–Є–љ–∞ —Б—В—А–Њ–Ї–Є: {len(metal_description)} —Б–Є–Љ–≤–Њ–ї–Њ–≤")
            print(f"   рЯФ§ –Я–Њ–±–∞–є—В–Њ–≤–Њ: {metal_description.encode('utf-8')}")

            # –Я—А–Њ–±—Г–µ–Љ —А–∞–Ј–љ—Л–µ –њ–∞—В—В–µ—А–љ—Л
            patterns = [
                (r'(\d+(?:\.\d+)?)\s*–Љ–Љ\s*(\d+(?:\.\d+)?)\s*[xX—Е–•√Ч]\s*(\d+(?:\.\d+)?)', "–§–Њ—А–Љ–∞—В: 4.0–Љ–Љ 1500x3000"),
                (r'(\d+(?:\.\d+)?)\s*[xX—Е–•√Ч]\s*(\d+(?:\.\d+)?)\s*[xX—Е–•√Ч]\s*(\d+(?:\.\d+)?)', "–§–Њ—А–Љ–∞—В: 4x1500x3000"),
                (r'(\d+(?:\.\d+)?)\s*–Љ–Љ?\s*(\d+(?:\.\d+)?)\s*[xX—Е–•√Ч]?\s*(\d+(?:\.\d+)?)', "–§–Њ—А–Љ–∞—В –≥–Є–±–Ї–Є–є"),
            ]

            for idx, (pattern, description) in enumerate(patterns, 1):
                print(f"   рЯІ™ –Ґ–µ—Б—В –њ–∞—В—В–µ—А–љ–∞ {idx}: {description}")
                match = re.search(pattern, metal_description, re.IGNORECASE)

                if match:
                    thickness = float(match.group(1))
                    width = float(match.group(2))
                    length = float(match.group(3))
                    print(f"   вЬЕ –£–°–Я–Х–•! –Э–∞–є–і–µ–љ–Њ: {thickness} √Ч {width} √Ч {length}")
                    break
                else:
                    print(f"   вЭМ –Э–µ –њ–Њ–і–Њ—И—С–ї")

            if not thickness:
                # –Я–Њ–њ—А–Њ–±—Г–µ–Љ –љ–∞–є—В–Є —Е–Њ—В—М –Ї–∞–Ї–Є–µ-—В–Њ —З–Є—Б–ї–∞
                all_numbers = re.findall(r'\d+(?:\.\d+)?', metal_description)
                print(f"   рЯФҐ –Т—Б–µ —З–Є—Б–ї–∞ –≤ —Б—В—А–Њ–Ї–µ: {all_numbers}")

                # –Х—Б–ї–Є –љ–∞—И–ї–Є –Љ–Є–љ–Є–Љ—Г–Љ 3 —З–Є—Б–ї–∞ - –њ–Њ–њ—А–Њ–±—Г–µ–Љ –≤–Ј—П—В—М –њ–Њ—Б–ї–µ–і–љ–Є–µ 3
                if len(all_numbers) >= 3:
                    try:
                        # –Ш—Й–µ–Љ –њ–µ—А–≤–Њ–µ —З–Є—Б–ї–Њ –Ї–∞–Ї —В–Њ–ї—Й–Є–љ—Г (–Њ–±—Л—З–љ–Њ 3-10)
                        for i, num in enumerate(all_numbers):
                            val = float(num)
                            if 0.5 <= val <= 50:  # –Ґ–Њ–ї—Й–Є–љ–∞ –Њ–±—Л—З–љ–Њ –Њ—В 0.5 –і–Њ 50 –Љ–Љ
                                thickness = val
                                # –С–µ—А—С–Љ —Б–ї–µ–і—Г—О—Й–Є–µ 2 —З–Є—Б–ї–∞ –Ї–∞–Ї —А–∞–Ј–Љ–µ—А—Л
                                if i + 2 < len(all_numbers):
                                    width = float(all_numbers[i + 1])
                                    length = float(all_numbers[i + 2])
                                    print(f"   вЪ†пЄП –Ш—Б–њ–Њ–ї—М–Ј–Њ–≤–∞–љ–∞ —Н–≤—А–Є—Б—В–Є–Ї–∞: {thickness} √Ч {width} √Ч {length}")
                                    break
                    except:
                        pass

            if not thickness:
                print(f"   вЭМ –Э–Х –†–Р–°–Я–Ю–Ч–Э–Р–Э–Ю: '{metal_description}'")
                row_data["_status"] = f"–Ю—И–Є–±–Ї–∞ –њ–∞—А—Б–Є–љ–≥–∞ –Љ–∞—В–µ—А–Є–∞–ї–∞"
                return False

            print(f"   вЬЕ –Ш–Ґ–Ю–У–Ю: —В–Њ–ї—Й–Є–љ–∞={thickness}, —И–Є—А–Є–љ–∞={width}, –і–ї–Є–љ–∞={length}")

            # –Я–Њ–Є—Б–Ї —А–µ–Ј–µ—А–≤–∞
            reservations_df = load_data("Reservations")
            order_reserves = reservations_df[reservations_df["ID –Ј–∞–Ї–∞–Ј–∞"] == order_id]

            if order_reserves.empty:
                row_data["_status"] = f"–Ю—И–Є–±–Ї–∞: –љ–µ—В —А–µ–Ј–µ—А–≤–Њ–≤"
                return False

            suitable_reserve = None
            tolerance = 0.01

            for _, reserve in order_reserves.iterrows():
                thickness_match = abs(float(reserve["–Ґ–Њ–ї—Й–Є–љ–∞"]) - thickness) < tolerance

                if width and length:
                    width_match = abs(float(reserve["–®–Є—А–Є–љ–∞"]) - width) < tolerance
                    length_match = abs(float(reserve["–Ф–ї–Є–љ–∞"]) - length) < tolerance

                    if thickness_match and width_match and length_match and int(reserve["–Ю—Б—В–∞—В–Њ–Ї –Ї —Б–њ–Є—Б–∞–љ–Є—О"]) > 0:
                        suitable_reserve = reserve
                        break
                else:
                    if thickness_match and int(reserve["–Ю—Б—В–∞—В–Њ–Ї –Ї —Б–њ–Є—Б–∞–љ–Є—О"]) > 0:
                        suitable_reserve = reserve
                        break

            if suitable_reserve is None:
                row_data["_status"] = f"–Ю—И–Є–±–Ї–∞: —А–µ–Ј–µ—А–≤ –љ–µ –љ–∞–є–і–µ–љ"
                return False

            reserve_id = int(suitable_reserve["ID —А–µ–Ј–µ—А–≤–∞"])
            remainder = int(suitable_reserve["–Ю—Б—В–∞—В–Њ–Ї –Ї —Б–њ–Є—Б–∞–љ–Є—О"])

            if metal_qty > remainder:
                row_data["_status"] = f"–Ю—И–Є–±–Ї–∞: –љ–µ–і–Њ—Б—В–∞—В–Њ—З–љ–Њ ({remainder} —И—В)"
                return False

            # –°–Я–Ш–°–Р–Э–Ш–Х
            writeoffs_df = load_data("WriteOffs")
            new_writeoff_id = 1 if writeoffs_df.empty else int(writeoffs_df["ID —Б–њ–Є—Б–∞–љ–Є—П"].max()) + 1

            comment = f"–Ю–њ–µ—А–∞—В–Њ—А: {username} | –Ф–µ—В–∞–ї—М: {part_name}"
            writeoff_datetime = f"{date_str} {time_str}"

            new_writeoff = pd.DataFrame([{
                "ID —Б–њ–Є—Б–∞–љ–Є—П": new_writeoff_id,
                "ID —А–µ–Ј–µ—А–≤–∞": reserve_id,
                "ID –Ј–∞–Ї–∞–Ј–∞": order_id,
                "ID –Љ–∞—В–µ—А–Є–∞–ї–∞": int(suitable_reserve["ID –Љ–∞—В–µ—А–Є–∞–ї–∞"]),
                "–Ь–∞—А–Ї–∞": suitable_reserve["–Ь–∞—А–Ї–∞"],
                "–Ґ–Њ–ї—Й–Є–љ–∞": thickness,
                "–Ф–ї–Є–љ–∞": length,
                "–®–Є—А–Є–љ–∞": width,
                "–Ъ–Њ–ї–Є—З–µ—Б—В–≤–Њ": metal_qty,
                "–Ф–∞—В–∞ —Б–њ–Є—Б–∞–љ–Є—П": writeoff_datetime,
                "–Ъ–Њ–Љ–Љ–µ–љ—В–∞—А–Є–є": comment
            }])

            writeoffs_df = pd.concat([writeoffs_df, new_writeoff], ignore_index=True)
            save_data("WriteOffs", writeoffs_df)

            # –Ю–±–љ–Њ–≤–ї—П–µ–Љ —А–µ–Ј–µ—А–≤
            new_written_off = int(suitable_reserve["–°–њ–Є—Б–∞–љ–Њ"]) + metal_qty
            new_remainder = remainder - metal_qty

            reservations_df.loc[reservations_df["ID —А–µ–Ј–µ—А–≤–∞"] == reserve_id, "–°–њ–Є—Б–∞–љ–Њ"] = new_written_off
            reservations_df.loc[reservations_df["ID —А–µ–Ј–µ—А–≤–∞"] == reserve_id, "–Ю—Б—В–∞—В–Њ–Ї –Ї —Б–њ–Є—Б–∞–љ–Є—О"] = new_remainder
            save_data("Reservations", reservations_df)

            # –Ю–±–љ–Њ–≤–ї—П–µ–Љ —Б–Ї–ї–∞–і
            material_id = int(suitable_reserve["ID –Љ–∞—В–µ—А–Є–∞–ї–∞"])
            if material_id != -1:
                materials_df = load_data("Materials")
                if not materials_df[materials_df["ID"] == material_id].empty:
                    mat_row = materials_df[materials_df["ID"] == material_id].iloc[0]
                    old_qty = int(mat_row["–Ъ–Њ–ї–Є—З–µ—Б—В–≤–Њ —И—В—Г–Ї"])
                    new_qty = old_qty - metal_qty

                    materials_df.loc[materials_df["ID"] == material_id, "–Ъ–Њ–ї–Є—З–µ—Б—В–≤–Њ —И—В—Г–Ї"] = new_qty

                    reserved = int(mat_row["–Ч–∞—А–µ–Ј–µ—А–≤–Є—А–Њ–≤–∞–љ–Њ"])
                    new_reserved = max(0, reserved - metal_qty)
                    materials_df.loc[materials_df["ID"] == material_id, "–Ч–∞—А–µ–Ј–µ—А–≤–Є—А–Њ–≤–∞–љ–Њ"] = new_reserved
                    materials_df.loc[materials_df["ID"] == material_id, "–Ф–Њ—Б—В—Г–њ–љ–Њ"] = new_qty - new_reserved

                    save_data("Materials", materials_df)

            row_data["_status"] = "вЬЕ –°–њ–Є—Б–∞–љ–Њ"
            return True

        except Exception as e:
            row_data["_status"] = f"–Ю—И–Є–±–Ї–∞: {str(e)}"
            return False

    def clear_laser_table(self):
        """–Ю—З–Є—Б—В–Ї–∞ —В–∞–±–ї–Є—Ж—Л –Є–Љ–њ–Њ—А—В–∞"""
        if hasattr(self, 'laser_import_data'):
            self.laser_import_data = []

        for i in self.laser_import_tree.get_children():
            self.laser_import_tree.delete(i)

        messagebox.showinfo("–£—Б–њ–µ—Е", "–Ґ–∞–±–ї–Є—Ж–∞ –Њ—З–Є—Й–µ–љ–∞")

    # ==================== –Ь–Х–Ґ–Ю–Ф–Ђ –Ф–Ы–ѓ –Т–Ъ–Ы–Р–Ф–Ъ–Ш –Ш–Ь–Я–Ю–†–Ґ–Р –Ю–Ґ –Ы–Р–Ч–Х–†–©–Ш–Ъ–Ю–Т ====================

    def import_laser_table(self):
        """–Ш–Љ–њ–Њ—А—В —В–∞–±–ї–Є—Ж—Л –Њ—В –ї–∞–Ј–µ—А—Й–Є–Ї–Њ–≤ —Б —Б–Њ—Е—А–∞–љ–µ–љ–Є–µ–Љ —Б—В–∞—В—Г—Б–Њ–≤ —Б—Г—Й–µ—Б—В–≤—Г—О—Й–Є—Е –Ј–∞–њ–Є—Б–µ–є"""
        file_path = filedialog.askopenfilename(
            title="–Т—Л–±–µ—А–Є—В–µ —Д–∞–є–ї –Њ—В –ї–∞–Ј–µ—А—Й–Є–Ї–Њ–≤",
            filetypes=[("CSV files", "*.csv"), ("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )

        if not file_path:
            return

        try:
            # –Ч–∞–≥—А—Г–Ј–Ї–∞ —Д–∞–є–ї–∞
            if file_path.endswith('.csv'):
                try:
                    laser_df = pd.read_csv(file_path, sep=';', encoding='utf-8')
                except:
                    laser_df = pd.read_csv(file_path, sep=';', encoding='cp1251')
            else:
                laser_df = pd.read_excel(file_path, engine='openpyxl')

            # –Я—А–Њ–≤–µ—А–Ї–∞ –Њ–±—П–Ј–∞—В–µ–ї—М–љ—Л—Е –Ї–Њ–ї–Њ–љ–Њ–Ї
            required = ["–Ф–∞—В–∞ (–Ь–°–Ъ)", "–Т—А–µ–Љ—П (–Ь–°–Ъ)", "username", "order", "metal", "metal_quantity", "part",
                        "part_quantity"]
            missing = [col for col in required if col not in laser_df.columns]

            if missing:
                messagebox.showerror("–Ю—И–Є–±–Ї–∞", f"–Ю—В—Б—Г—В—Б—В–≤—Г—О—В –Ї–Њ–ї–Њ–љ–Ї–Є:\n{', '.join(missing)}")
                return

            # –Ф–Њ–±–∞–≤–ї—П–µ–Љ –Ї–Њ–ї–Њ–љ–Ї–Є —Б—В–∞—В—Г—Б–∞ –µ—Б–ї–Є –Є—Е –љ–µ—В
            if "–°–њ–Є—Б–∞–љ–Њ" not in laser_df.columns:
                laser_df["–°–њ–Є—Б–∞–љ–Њ"] = ""
            if "–Ф–∞—В–∞ —Б–њ–Є—Б–∞–љ–Є—П" not in laser_df.columns:
                laser_df["–Ф–∞—В–∞ —Б–њ–Є—Б–∞–љ–Є—П"] = ""

            # –°–Њ–Ј–і–∞—С–Љ —Г–љ–Є–Ї–∞–ї—М–љ—Л–є –Ї–ї—О—З –і–ї—П –Ї–∞–ґ–і–Њ–є —Б—В—А–Њ–Ї–Є
            def create_row_key(row):
                """–°–Њ–Ј–і–∞–љ–Є–µ —Г–љ–Є–Ї–∞–ї—М–љ–Њ–≥–Њ –Ї–ї—О—З–∞ –і–ї—П —Б—В—А–Њ–Ї–Є"""
                return (
                    str(row.get("–Ф–∞—В–∞ (–Ь–°–Ъ)", "")),
                    str(row.get("–Т—А–µ–Љ—П (–Ь–°–Ъ)", "")),
                    str(row.get("username", "")),
                    str(row.get("order", "")),
                    str(row.get("metal", "")),
                    str(row.get("metal_quantity", "")),
                    str(row.get("part", "")),
                    str(row.get("part_quantity", ""))
                )

            # –°–Њ–Ј–і–∞—С–Љ —Б–ї–Њ–≤–∞—А—М —Б—Г—Й–µ—Б—В–≤—Г—О—Й–Є—Е —Б—В—А–Њ–Ї —Б –Є—Е —Б—В–∞—В—Г—Б–∞–Љ–Є
            existing_rows = {}
            if hasattr(self, 'laser_table_data') and self.laser_table_data:
                for row_data in self.laser_table_data:
                    key = create_row_key(row_data)
                    existing_rows[key] = {
                        "–°–њ–Є—Б–∞–љ–Њ": row_data.get("–°–њ–Є—Б–∞–љ–Њ", ""),
                        "–Ф–∞—В–∞ —Б–њ–Є—Б–∞–љ–Є—П": row_data.get("–Ф–∞—В–∞ —Б–њ–Є—Б–∞–љ–Є—П", "")
                    }

            # –Ю–±—А–∞–±–∞—В—Л–≤–∞–µ–Љ –љ–Њ–≤—Л–є —Д–∞–є–ї
            new_rows = []
            updated_rows = 0

            for _, row in laser_df.iterrows():
                row_dict = row.to_dict()
                key = create_row_key(row_dict)

                # –Я—А–Њ–≤–µ—А—П–µ–Љ, —Б—Г—Й–µ—Б—В–≤—Г–µ—В –ї–Є —Г–ґ–µ —Н—В–∞ —Б—В—А–Њ–Ї–∞
                if key in existing_rows:
                    # –°—В—А–Њ–Ї–∞ —Г–ґ–µ –µ—Б—В—М - —Б–Њ—Е—А–∞–љ—П–µ–Љ –µ—С —Б—В–∞—В—Г—Б
                    row_dict["–°–њ–Є—Б–∞–љ–Њ"] = existing_rows[key]["–°–њ–Є—Б–∞–љ–Њ"]
                    row_dict["–Ф–∞—В–∞ —Б–њ–Є—Б–∞–љ–Є—П"] = existing_rows[key]["–Ф–∞—В–∞ —Б–њ–Є—Б–∞–љ–Є—П"]
                    updated_rows += 1
                else:
                    # –Э–Њ–≤–∞—П —Б—В—А–Њ–Ї–∞ - –Њ—Б—В–∞–≤–ї—П–µ–Љ –њ—Г—Б—В–Њ–є —Б—В–∞—В—Г—Б
                    if not row_dict.get("–°–њ–Є—Б–∞–љ–Њ"):
                        row_dict["–°–њ–Є—Б–∞–љ–Њ"] = ""
                    if not row_dict.get("–Ф–∞—В–∞ —Б–њ–Є—Б–∞–љ–Є—П"):
                        row_dict["–Ф–∞—В–∞ —Б–њ–Є—Б–∞–љ–Є—П"] = ""

                new_rows.append(row_dict)

            # –Ю–±—К–µ–і–Є–љ—П–µ–Љ –і–∞–љ–љ—Л–µ
            merged_data = []
            new_count = 0

            # –°–Њ–Ј–і–∞—С–Љ –Љ–љ–Њ–ґ–µ—Б—В–≤–Њ –Ї–ї—О—З–µ–є –Є–Ј –љ–Њ–≤–Њ–≥–Њ —Д–∞–є–ї–∞
            new_keys = set()
            for row_dict in new_rows:
                new_keys.add(create_row_key(row_dict))

            # –Ф–Њ–±–∞–≤–ї—П–µ–Љ —Б—В–∞—А—Л–µ —Б—В—А–Њ–Ї–Є, –µ—Б–ї–Є –Њ–љ–Є –µ—Б—В—М –≤ –љ–Њ–≤–Њ–Љ —Д–∞–є–ї–µ
            if hasattr(self, 'laser_table_data') and self.laser_table_data:
                for old_row in self.laser_table_data:
                    old_key = create_row_key(old_row)
                    if old_key in new_keys:
                        # –°—В—А–Њ–Ї–∞ –µ—Б—В—М –≤ –љ–Њ–≤–Њ–Љ —Д–∞–є–ї–µ - –±–µ—А—С–Љ –Є–Ј —Б—В–∞—А–Њ–є —В–∞–±–ї–Є—Ж—Л
                        merged_data.append(old_row)

            # –Ф–Њ–±–∞–≤–ї—П–µ–Љ —В–Њ–ї—М–Ї–Њ –Э–Ю–Т–Ђ–Х —Б—В—А–Њ–Ї–Є –Є–Ј –Є–Љ–њ–Њ—А—В–Є—А–Њ–≤–∞–љ–љ–Њ–≥–Њ —Д–∞–є–ї–∞
            for new_row in new_rows:
                new_key = create_row_key(new_row)
                is_new = new_key not in existing_rows

                if is_new:
                    merged_data.append(new_row)
                    new_count += 1

            # –°–Њ—Е—А–∞–љ—П–µ–Љ –Њ–±—К–µ–і–Є–љ—С–љ–љ—Л–µ –і–∞–љ–љ—Л–µ
            self.laser_table_data = merged_data

            # рЯЖХ –°–Ю–†–Ґ–Ш–†–Ю–Т–Ъ–Р: –Э–Ю–Т–Ђ–Х –Ч–Р–Я–Ш–°–Ш –Т–Т–Х–†–•–£
            try:
                print("рЯФД –°–Њ—А—В–Є—А–Њ–≤–Ї–∞ –Є–Љ–њ–Њ—А—В–Є—А–Њ–≤–∞–љ–љ—Л—Е –і–∞–љ–љ—Л—Е...")
                df_merged = pd.DataFrame(self.laser_table_data)

                df_merged['_datetime_sort'] = pd.to_datetime(
                    df_merged['–Ф–∞—В–∞ (–Ь–°–Ъ)'].astype(str) + ' ' + df_merged['–Т—А–µ–Љ—П (–Ь–°–Ъ)'].astype(str),
                    errors='coerce'
                )
                df_merged = df_merged.sort_values('_datetime_sort', ascending=False, na_position='last')
                df_merged = df_merged.drop('_datetime_sort', axis=1)

                self.laser_table_data = df_merged.to_dict('records')
                print(f"вЬЕ –Ф–∞–љ–љ—Л–µ –Њ—В—Б–Њ—А—В–Є—А–Њ–≤–∞–љ—Л: {len(self.laser_table_data)} –Ј–∞–њ–Є—Б–µ–є")
            except Exception as e:
                print(f"вЪ†пЄП –Ю—И–Є–±–Ї–∞ —Б–Њ—А—В–Є—А–Њ–≤–Ї–Є –њ–Њ—Б–ї–µ –Є–Љ–њ–Њ—А—В–∞: {e}")

            # –Ю–±–љ–Њ–≤–ї—П–µ–Љ —В–∞–±–ї–Є—Ж—Г
            self.refresh_laser_import_table()

            # –Я—А–Є–љ—Г–і–Є—В–µ–ї—М–љ–Њ–µ –Њ–±–љ–Њ–≤–ї–µ–љ–Є–µ
            self.laser_import_tree.update_idletasks()
            self.laser_import_frame.update()

            # –Р–≤—В–Њ—И–Є—А–Є–љ–∞ –Ї–Њ–ї–Њ–љ–Њ–Ї
            self.auto_resize_columns(self.laser_import_tree)

            # –Ю–±–љ–Њ–≤–ї—П–µ–Љ —Б—В–∞—В—Г—Б
            items_count = len(self.laser_import_tree.get_children())

            if hasattr(self, 'laser_status_label'):
                self.laser_status_label.config(
                    text=f"вЬЕ –Т—Б–µ–≥–Њ –Ј–∞–њ–Є—Б–µ–є: {items_count} | рЯЖХ –Э–Њ–≤—Л—Е: {new_count} | рЯФД –Ю–±–љ–Њ–≤–ї–µ–љ–Њ —Б—В–∞—В—Г—Б–Њ–≤: {updated_rows}",
                    bg='#d4edda',
                    fg='#155724'
                )

            # –§–Њ—А–Љ–Є—А—Г–µ–Љ —Б–Њ–Њ–±—Й–µ–љ–Є–µ
            result_msg = (
                f"вЬЕ –Ш–Љ–њ–Њ—А—В –Ј–∞–≤–µ—А—И—С–љ!\n\n"
                f"рЯУК –Т—Б–µ–≥–Њ –Ј–∞–њ–Є—Б–µ–є: {items_count}\n"
                f"рЯЖХ –Э–Њ–≤—Л—Е –Ј–∞–њ–Є—Б–µ–є: {new_count}\n"
                f"рЯФД –°–Њ—Е—А–∞–љ–µ–љ–Њ —Б—В–∞—В—Г—Б–Њ–≤: {updated_rows}\n\n"
            )

            # –°—З–Є—В–∞–µ–Љ —Б—В–∞—В–Є—Б—В–Є–Ї—Г –њ–Њ —Б—В–∞—В—Г—Б–∞–Љ
            if self.laser_table_data:
                auto_count = sum(1 for r in self.laser_table_data if r.get("–°–њ–Є—Б–∞–љ–Њ") in ["вЬУ", "–Ф–∞", "Yes"])
                manual_count = sum(1 for r in self.laser_table_data if r.get("–°–њ–Є—Б–∞–љ–Њ") == "–Т—А—Г—З–љ—Г—О")
                pending_count = sum(1 for r in self.laser_table_data if not r.get("–°–њ–Є—Б–∞–љ–Њ"))

                result_msg += (
                    f"рЯУИ –°—В–∞—В–Є—Б—В–Є–Ї–∞:\n"
                    f"  вАҐ вЬЕ –Р–≤—В–Њ–Љ–∞—В–Є—З–µ—Б–Ї–Є —Б–њ–Є—Б–∞–љ–Њ: {auto_count}\n"
                    f"  вАҐ рЯФµ –Я–Њ–Љ–µ—З–µ–љ–Њ –≤—А—Г—З–љ—Г—О: {manual_count}\n"
                    f"  вАҐ рЯЯ° –Ю–ґ–Є–і–∞–µ—В —Б–њ–Є—Б–∞–љ–Є—П: {pending_count}\n"
                )

            # –°–Њ—Е—А–∞–љ—П–µ–Љ –≤ –Ї—Н—И
            try:
                self.save_laser_import_cache()
            except Exception as cache_err:
                print(f"вЪ†пЄП –Э–µ —Г–і–∞–ї–Њ—Б—М —Б–Њ—Е—А–∞–љ–Є—В—М –Ї—Н—И: {cache_err}")

            messagebox.showinfo("–£—Б–њ–µ—Е", result_msg)

        except Exception as e:
            messagebox.showerror("–Ю—И–Є–±–Ї–∞ –Є–Љ–њ–Њ—А—В–∞", f"–Э–µ —Г–і–∞–ї–Њ—Б—М –Є–Љ–њ–Њ—А—В–Є—А–Њ–≤–∞—В—М —Д–∞–є–ї:\n\n{str(e)}")
            print(f"вЭМ –Ю—И–Є–±–Ї–∞ –Є–Љ–њ–Њ—А—В–∞: {e}")
            import traceback
            traceback.print_exc()

    def refresh_laser_import_table(self):
        """–Ю–±–љ–Њ–≤–ї–µ–љ–Є–µ —В–∞–±–ї–Є—Ж—Л –Є–Љ–њ–Њ—А—В–∞ –Њ—В –ї–∞–Ј–µ—А—Й–Є–Ї–Њ–≤"""
        # –Ю—З–Є—Й–∞–µ–Љ —В–∞–±–ї–Є—Ж—Г
        for item in self.laser_import_tree.get_children():
            self.laser_import_tree.delete(item)

        if not hasattr(self, 'laser_table_data') or self.laser_table_data is None:
            self.laser_table_data = []
            return

        if not self.laser_table_data:
            return

        # –°–Ю–†–Ґ–Ш–†–Ю–Т–Ъ–Р: –Э–Ю–Т–Ђ–Х –Ч–Р–Я–Ш–°–Ш –Т–Т–Х–†–•–£
        try:
            # –Я—А–µ–Њ–±—А–∞–Ј—Г–µ–Љ –≤ DataFrame
            df_display = pd.DataFrame(self.laser_table_data)

            # –Я–∞—А—Б–Є–љ–≥ –і–∞—В—Л –≤ —Д–Њ—А–Љ–∞—В–µ DD.MM.YYYY HH:MM:SS
            df_display['_datetime_sort'] = pd.to_datetime(
                df_display['–Ф–∞—В–∞ (–Ь–°–Ъ)'].astype(str) + ' ' + df_display['–Т—А–µ–Љ—П (–Ь–°–Ъ)'].astype(str),
                format='%d.%m.%Y %H:%M:%S',
                errors='coerce'
            )

            # –°–Њ—А—В–Є—А—Г–µ–Љ –њ–Њ —Г–±—Л–≤–∞–љ–Є—О (–љ–Њ–≤—Л–µ –≤–≤–µ—А—Е—Г)
            df_display = df_display.sort_values('_datetime_sort', ascending=False, na_position='last')

            # –£–і–∞–ї—П–µ–Љ –≤—А–µ–Љ–µ–љ–љ—Г—О –Ї–Њ–ї–Њ–љ–Ї—Г
            df_display = df_display.drop('_datetime_sort', axis=1)

            # –Я—А–µ–Њ–±—А–∞–Ј—Г–µ–Љ –Њ–±—А–∞—В–љ–Њ –≤ —Б–њ–Є—Б–Њ–Ї —Б–ї–Њ–≤–∞—А–µ–є
            sorted_data = df_display.to_dict('records')

        except Exception as e:
            print(f"вЪ†пЄП –Ю—И–Є–±–Ї–∞ —Б–Њ—А—В–Є—А–Њ–≤–Ї–Є: {e}")
            sorted_data = self.laser_table_data

        # рЯЖХ –°–Ю–•–†–Р–Э–ѓ–Х–Ь –Ю–Ґ–°–Ю–†–Ґ–Ш–†–Ю–Т–Р–Э–Э–Ђ–Х –Ф–Р–Э–Э–Ђ–Х –Ю–С–†–Р–Ґ–Э–Ю
        self.laser_table_data = sorted_data

        # –°–І–Б–Ґ–І–Ш–Ъ–Ш
        manual_count = 0
        auto_count = 0
        pending_count = 0

        # –Ч–∞–њ–Њ–ї–љ—П–µ–Љ —В–∞–±–ї–Є—Ж—Г –Ю–Ґ–°–Ю–†–Ґ–Ш–†–Ю–Т–Р–Э–Э–Ђ–Ь–Ш –і–∞–љ–љ—Л–Љ–Є
        for idx, row_data in enumerate(sorted_data):
            date_val = row_data.get("–Ф–∞—В–∞ (–Ь–°–Ъ)", "")
            time_val = row_data.get("–Т—А–µ–Љ—П (–Ь–°–Ъ)", "")
            username = row_data.get("username", "")
            order = row_data.get("order", "")
            metal = row_data.get("metal", "")
            metal_qty = row_data.get("metal_quantity", "")
            part = row_data.get("part", "")
            part_qty = row_data.get("part_quantity", "")
            written_off = row_data.get("–°–њ–Є—Б–∞–љ–Њ", "")
            writeoff_date = row_data.get("–Ф–∞—В–∞ —Б–њ–Є—Б–∞–љ–Є—П", "")

            # –С–Х–Ч–Ю–Я–Р–°–Э–Ю–Х –Я–†–Х–Ю–С–†–Р–Ч–Ю–Т–Р–Э–Ш–Х written_off –Т –°–Ґ–†–Ю–Ъ–£
            if pd.isna(written_off) or written_off is None:
                written_off = ""
            else:
                written_off = str(written_off).strip()

            values = (date_val, time_val, username, order, metal, metal_qty, part, part_qty, written_off, writeoff_date)

            # –¶–Т–Х–Ґ–Ю–Т–Р–ѓ –Ш–Э–Ф–Ш–Ъ–Р–¶–Ш–ѓ
            if written_off == "–Т—А—Г—З–љ—Г—О":
                tag = 'manual'
                manual_count += 1
            elif written_off in ["–Ф–∞", "вЬУ", "Yes"]:
                tag = 'written_off'
                auto_count += 1
            else:
                tag = 'pending'
                pending_count += 1

            self.laser_import_tree.insert("", "end", values=values, tags=(tag,))

        self.auto_resize_columns(self.laser_import_tree)

    def test_add_rows(self):
        """–Ґ–µ—Б—В–Њ–≤–∞—П —Д—Г–љ–Ї—Ж–Є—П –і–ї—П –њ—А–Њ–≤–µ—А–Ї–Є –Њ—В–Њ–±—А–∞–ґ–µ–љ–Є—П —Б—В—А–Њ–Ї"""
        print("\nрЯІ™ –Ґ–Х–°–Ґ: –Ф–Њ–±–∞–≤–ї–µ–љ–Є–µ —В–µ—Б—В–Њ–≤—Л—Е —Б—В—А–Њ–Ї...")

        # –Ю—З–Є—Й–∞–µ–Љ
        for item in self.laser_import_tree.get_children():
            self.laser_import_tree.delete(item)

        # –Ф–Њ–±–∞–≤–ї—П–µ–Љ 3 —В–µ—Б—В–Њ–≤—Л–µ —Б—В—А–Њ–Ї–Є
        test_data = [
            ("01.01.2026", "10:00", "@test1", "–Ґ–µ—Б—В –Ј–∞–Ї–∞–Ј 1", "–°—В3 10—Е1500—Е3000", "5", "–Ф–µ—В–∞–ї—М A", "100", "", ""),
            ("02.01.2026", "11:00", "@test2", "–Ґ–µ—Б—В –Ј–∞–Ї–∞–Ј 2", "–°—В3 12—Е1500—Е3000", "3", "–Ф–µ—В–∞–ї—М B", "50", "", ""),
            ("03.01.2026", "12:00", "@test3", "–Ґ–µ—Б—В –Ј–∞–Ї–∞–Ј 3", "09–У2–° 8—Е1500—Е3000", "2", "–Ф–µ—В–∞–ї—М C", "75", "", "")
        ]

        for idx, values in enumerate(test_data, 1):
            item_id = self.laser_import_tree.insert("", "end", values=values, tags=('pending',))
            print(f"  вЬУ –Ґ–µ—Б—В–Њ–≤–∞—П —Б—В—А–Њ–Ї–∞ {idx} –і–Њ–±–∞–≤–ї–µ–љ–∞: ID={item_id}")

        # –Я—А–Њ–≤–µ—А–Ї–∞
        items_count = len(self.laser_import_tree.get_children())
        print(f"вЬЕ –Ґ–Х–°–Ґ: –Т —В–∞–±–ї–Є—Ж–µ {items_count} —Н–ї–µ–Љ–µ–љ—В–Њ–≤")

        # –Я—А–Є–љ—Г–і–Є—В–µ–ї—М–љ–Њ–µ –Њ–±–љ–Њ–≤–ї–µ–љ–Є–µ
        self.laser_import_tree.update_idletasks()

        messagebox.showinfo("–Ґ–µ—Б—В", f"–Ф–Њ–±–∞–≤–ї–µ–љ–Њ —В–µ—Б—В–Њ–≤—Л—Е —Б—В—А–Њ–Ї: {items_count}")

    def writeoff_laser_row(self):
        """–°–њ–Є—Б–∞–љ–Є–µ –≤—Л–±—А–∞–љ–љ—Л—Е —Б—В—А–Њ–Ї —Б —В–Њ—З–љ—Л–Љ —Б–Њ–њ–Њ—Б—В–∞–≤–ї–µ–љ–Є–µ–Љ –Ј–∞–Ї–∞–Ј–∞, –Љ–∞—В–µ—А–Є–∞–ї–∞ –Є –і–µ—В–∞–ї–Є"""
        selected_items = self.laser_import_tree.selection()

        if not selected_items:
            messagebox.showwarning("–Я—А–µ–і—Г–њ—А–µ–ґ–і–µ–љ–Є–µ", "–Т—Л–±–µ—А–Є—В–µ —Б—В—А–Њ–Ї–Є –і–ї—П —Б–њ–Є—Б–∞–љ–Є—П!")
            return

        # –Я—А–Њ–≤–µ—А—П–µ–Љ, —З—В–Њ –≤—Л–±—А–∞–љ–љ—Л–µ —Б—В—А–Њ–Ї–Є –µ—Й–µ –љ–µ —Б–њ–Є—Б–∞–љ—Л
        rows_to_writeoff = []
        already_written_off = []

        for item in selected_items:
            values = self.laser_import_tree.item(item)['values']
            if values[8] in ["–Ф–∞", "вЬУ", "Yes"]:  # –Ъ–Њ–ї–Њ–љ–Ї–∞ "–°–њ–Є—Б–∞–љ–Њ"
                already_written_off.append(values[3])  # order
            else:
                rows_to_writeoff.append((item, values))

        if already_written_off:
            messagebox.showinfo("–Ш–љ—Д–Њ—А–Љ–∞—Ж–Є—П",
                                f"–Э–µ–Ї–Њ—В–Њ—А—Л–µ —Б—В—А–Њ–Ї–Є —Г–ґ–µ —Б–њ–Є—Б–∞–љ—Л:\n" + "\n".join(already_written_off[:5]))

        if not rows_to_writeoff:
            messagebox.showwarning("–Я—А–µ–і—Г–њ—А–µ–ґ–і–µ–љ–Є–µ", "–Э–µ—В —Б—В—А–Њ–Ї –і–ї—П —Б–њ–Є—Б–∞–љ–Є—П!")
            return

        # –Я–Њ–і—В–≤–µ—А–ґ–і–µ–љ–Є–µ
        if not messagebox.askyesno("–Я–Њ–і—В–≤–µ—А–ґ–і–µ–љ–Є–µ",
                                   f"–°–њ–Є—Б–∞—В—М –≤—Л–±—А–∞–љ–љ—Л–µ —Б—В—А–Њ–Ї–Є ({len(rows_to_writeoff)} —И—В)?"):
            return

        try:
            # –Ч–∞–≥—А—Г–ґ–∞–µ–Љ –і–∞–љ–љ—Л–µ
            orders_df = load_data("Orders")
            reservations_df = load_data("Reservations")
            materials_df = load_data("Materials")
            writeoffs_df = load_data("WriteOffs")
            order_details_df = load_data("OrderDetails")

            success_count = 0
            errors = []

            print(f"\n{'=' * 80}")
            print(f"рЯФµ –Э–Р–І–Р–Ы–Ю –°–Я–Ш–°–Р–Э–Ш–ѓ: {len(rows_to_writeoff)} —Б—В—А–Њ–Ї(–Є)")
            print(f"{'=' * 80}")

            for item, values in rows_to_writeoff:
                try:
                    date_val, time_val, username, order_name, metal_desc, metal_qty_str, part_name, part_qty = values[
                        :8]

                    print(f"\nрЯУЛ –Ю–±—А–∞–±–Њ—В–Ї–∞ —Б—В—А–Њ–Ї–Є:")
                    print(f"   –Ч–∞–Ї–∞–Ј: {order_name}")
                    print(f"   –Ь–µ—В–∞–ї–ї: {metal_desc}")
                    print(f"   –Ъ–Њ–ї–Є—З–µ—Б—В–≤–Њ –Љ–µ—В–∞–ї–ї–∞: {metal_qty_str}")
                    print(f"   –Ф–µ—В–∞–ї—М: {part_name}")

                    # ========== –®–Р–У 1: –Я–Ю–Ш–°–Ъ –Ч–Р–Ъ–Р–Ч–Р ==========
                    # –Ш—Й–µ–Љ –њ–Њ —В–Њ—З–љ–Њ–Љ—Г —Б–Њ–≤–њ–∞–і–µ–љ–Є—О –Є–ї–Є –њ–Њ –љ–Њ–Љ–µ—А—Г –£–Я-XXX
                    order_match = None

                    # –Я—А–Њ–±—Г–µ–Љ –љ–∞–є—В–Є –љ–Њ–Љ–µ—А –£–Я-XXX
                    import re
                    up_match = re.search(r'–£–Я-(\d+)', order_name)
                    if up_match:
                        up_number = up_match.group(1)
                        order_match = orders_df[
                            orders_df["–Э–∞–Ј–≤–∞–љ–Є–µ –Ј–∞–Ї–∞–Ј–∞"].str.contains(f"–£–Я-{up_number}", case=False, na=False)]
                        print(f"   рЯФН –Я–Њ–Є—Б–Ї –њ–Њ –£–Я-{up_number}")

                    # –Х—Б–ї–Є –љ–µ –љ–∞—И–ї–Є, –Є—Й–µ–Љ –њ–Њ —З–∞—Б—В–Є—З–љ–Њ–Љ—Г —Б–Њ–≤–њ–∞–і–µ–љ–Є—О –љ–∞–Ј–≤–∞–љ–Є—П
                    if order_match is None or order_match.empty:
                        order_match = orders_df[
                            orders_df["–Э–∞–Ј–≤–∞–љ–Є–µ –Ј–∞–Ї–∞–Ј–∞"].str.contains(order_name, case=False, na=False)]
                        print(f"   рЯФН –Я–Њ–Є—Б–Ї –њ–Њ –љ–∞–Ј–≤–∞–љ–Є—О: {order_name}")

                    if order_match.empty:
                        errors.append(f"вЭМ –Ч–∞–Ї–∞–Ј '{order_name}' –љ–µ –љ–∞–є–і–µ–љ –≤ –±–∞–Ј–µ")
                        print(f"   вЭМ –Ч–∞–Ї–∞–Ј –љ–µ –љ–∞–є–і–µ–љ")
                        continue

                    order_id = int(order_match.iloc[0]["ID –Ј–∞–Ї–∞–Ј–∞"])
                    print(f"   вЬЕ –Ч–∞–Ї–∞–Ј –љ–∞–є–і–µ–љ: ID={order_id}")

                    # ========== –®–Р–У 2: –Я–Р–†–°–Ш–Э–У –Ь–Р–Ґ–Х–†–Ш–Р–Ы–Р ==========
                    # –Я—А–Є–Љ–µ—А—Л:
                    # "–У–Ъ –°—В.3 4.0–Љ–Љ 1500x3000" вЖТ –Љ–∞—А–Ї–∞="–У–Ъ –°—В.3", —В–Њ–ї—Й–Є–љ–∞=4.0, —И–Є—А–Є–љ–∞=1500, –і–ї–Є–љ–∞=3000
                    # "–У–Ъ –°—В.3 6—Е1500—Е3000" вЖТ –Љ–∞—А–Ї–∞="–У–Ъ –°—В.3", —В–Њ–ї—Й–Є–љ–∞=6, —И–Є—А–Є–љ–∞=1500, –і–ї–Є–љ–∞=3000

                    thickness = None
                    width = None
                    length = None
                    marka = None

                    print(f"   рЯФН –Я–∞—А—Б–Є–љ–≥ –Љ–∞—В–µ—А–Є–∞–ї–∞: '{metal_desc}'")

                    # рЯЖХ –Я–Р–Ґ–Ґ–Х–†–Э 1: –§–Њ—А–Љ–∞—В —Б "–Љ–Љ" (4.0–Љ–Љ 1500x3000)
                    pattern1 = r'(\d+(?:\.\d+)?)\s*–Љ–Љ\s*(\d+(?:\.\d+)?)\s*[xX—Е–•√Ч]\s*(\d+(?:\.\d+)?)'
                    match1 = re.search(pattern1, metal_desc, re.IGNORECASE)

                    if match1:
                        thickness = float(match1.group(1))
                        width = float(match1.group(2))
                        length = float(match1.group(3))
                        # –Ь–∞—А–Ї–∞ - –≤—Б—С –і–Њ —А–∞–Ј–Љ–µ—А–Њ–≤
                        marka = metal_desc.split(match1.group(0))[0].strip()
                        print(f"   вЬЕ –Я–∞—В—В–µ—А–љ 1 (—Б –Љ–Љ): {thickness}–Љ–Љ {width}x{length}, –Љ–∞—А–Ї–∞='{marka}'")

                    # –Я–Р–Ґ–Ґ–Х–†–Э 2: –Ъ–ї–∞—Б—Б–Є—З–µ—Б–Ї–Є–є —Д–Њ—А–Љ–∞—В (6—Е1500—Е3000)
                    if not thickness:
                        pattern2 = r'(\d+(?:\.\d+)?)\s*[xX—Е–•√Ч]\s*(\d+(?:\.\d+)?)\s*[xX—Е–•√Ч]\s*(\d+(?:\.\d+)?)'
                        match2 = re.search(pattern2, metal_desc)

                        if match2:
                            thickness = float(match2.group(1))
                            width = float(match2.group(2))
                            length = float(match2.group(3))
                            # –Ь–∞—А–Ї–∞ - –≤—Б—С –і–Њ —А–∞–Ј–Љ–µ—А–Њ–≤
                            marka = metal_desc.split(match2.group(0))[0].strip()
                            print(f"   вЬЕ –Я–∞—В—В–µ—А–љ 2 (–±–µ–Ј –Љ–Љ): {thickness}—Е{width}—Е{length}, –Љ–∞—А–Ї–∞='{marka}'")

                    if not thickness or not marka:
                        errors.append(f"вЭМ –Э–µ —Г–і–∞–ї–Њ—Б—М —А–∞—Б–њ–∞—А—Б–Є—В—М –Љ–∞—В–µ—А–Є–∞–ї: {metal_desc}")
                        print(f"   вЭМ –Ю—И–Є–±–Ї–∞ –њ–∞—А—Б–Є–љ–≥–∞ –Љ–∞—В–µ—А–Є–∞–ї–∞: '{metal_desc}'")
                        continue

                    print(f"   рЯУ¶ –†–∞—Б–њ–∞—А—Б–µ–љ–љ—Л–є –Љ–∞—В–µ—А–Є–∞–ї:")
                    print(f"      –Ь–∞—А–Ї–∞: {marka}")
                    print(f"      –Ґ–Њ–ї—Й–Є–љ–∞: {thickness} –Љ–Љ")
                    print(f"      –†–∞–Ј–Љ–µ—А: {width}x{length}")

                    # ========== –®–Р–У 3: –Я–Ю–Ш–°–Ъ –Ф–Х–Ґ–Р–Ы–Ш –Т –Ч–Р–Ъ–Р–Ч–Х ==========
                    detail_id = None
                    detail_match = order_details_df[
                        (order_details_df["ID –Ј–∞–Ї–∞–Ј–∞"] == order_id) &
                        (order_details_df["–Э–∞–Ј–≤–∞–љ–Є–µ –і–µ—В–∞–ї–Є"].str.contains(part_name, case=False, na=False))
                        ]

                    if not detail_match.empty:
                        detail_id = int(detail_match.iloc[0]["ID"])
                        print(f"   рЯФІ –Ф–µ—В–∞–ї—М –љ–∞–є–і–µ–љ–∞: ID={detail_id}, –Э–∞–Ј–≤–∞–љ–Є–µ='{part_name}'")
                    else:
                        print(f"   вЪ†пЄП –Ф–µ—В–∞–ї—М '{part_name}' –љ–µ –љ–∞–є–і–µ–љ–∞ –≤ –Ј–∞–Ї–∞–Ј–µ (—Б–њ–Є—Б–∞–љ–Є–µ –±–µ–Ј –њ—А–Є–≤—П–Ј–Ї–Є)")

                    # ========== –®–Р–У 4: –Я–Ю–Ш–°–Ъ –†–Х–Ч–Х–†–Т–Р –° –£–І–Х–Ґ–Ю–Ь –Ь–Р–Ґ–Х–†–Ш–Р–Ы–Р –Ш –Ф–Х–Ґ–Р–Ы–Ш ==========
                    # –Ш—Й–µ–Љ —А–µ–Ј–µ—А–≤—Л —Н—В–Њ–≥–Њ –Ј–∞–Ї–∞–Ј–∞
                    order_reserves = reservations_df[
                        (reservations_df["ID –Ј–∞–Ї–∞–Ј–∞"] == order_id) &
                        (reservations_df["–Ю—Б—В–∞—В–Њ–Ї –Ї —Б–њ–Є—Б–∞–љ–Є—О"] > 0)
                        ]

                    if order_reserves.empty:
                        errors.append(f"вЭМ –Э–µ—В –і–Њ—Б—В—Г–њ–љ—Л—Е —А–µ–Ј–µ—А–≤–Њ–≤ –і–ї—П –Ј–∞–Ї–∞–Ј–∞ '{order_name}'")
                        print(f"   вЭМ –†–µ–Ј–µ—А–≤—Л –љ–µ –љ–∞–є–і–µ–љ—Л")
                        continue

                    print(f"   рЯФН –Э–∞–є–і–µ–љ–Њ —А–µ–Ј–µ—А–≤–Њ–≤ –і–ї—П –Ј–∞–Ї–∞–Ј–∞: {len(order_reserves)}")

                    # –§–Є–ї—М—В—А—Г–µ–Љ —А–µ–Ј–µ—А–≤—Л –њ–Њ –Љ–∞—В–µ—А–Є–∞–ї—Г
                    suitable_reserves = order_reserves[
                        (order_reserves["–Ь–∞—А–Ї–∞"].str.contains(marka, case=False, na=False)) &
                        (order_reserves["–Ґ–Њ–ї—Й–Є–љ–∞"] == thickness)
                        ]

                    # –Х—Б–ї–Є —Г–Ї–∞–Ј–∞–љ—Л —А–∞–Ј–Љ–µ—А—Л, —Д–Є–ї—М—В—А—Г–µ–Љ –Є –њ–Њ –љ–Є–Љ
                    if width and length:
                        suitable_reserves = suitable_reserves[
                            (suitable_reserves["–®–Є—А–Є–љ–∞"] == width) &
                            (suitable_reserves["–Ф–ї–Є–љ–∞"] == length)
                            ]

                    print(f"   рЯФН –Я–Њ–і—Е–Њ–і—П—Й–Є—Е –њ–Њ –Љ–∞—В–µ—А–Є–∞–ї—Г: {len(suitable_reserves)}")

                    # –Х—Б–ї–Є –љ–∞–є–і–µ–љ–∞ –і–µ—В–∞–ї—М, —Д–Є–ї—М—В—А—Г–µ–Љ –њ–Њ –і–µ—В–∞–ї–Є
                    if detail_id:
                        detail_reserves = suitable_reserves[suitable_reserves["ID –і–µ—В–∞–ї–Є"] == detail_id]
                        if not detail_reserves.empty:
                            suitable_reserves = detail_reserves
                            print(f"   вЬЕ –†–µ–Ј–µ—А–≤—Л —Б –њ—А–Є–≤—П–Ј–Ї–Њ–є –Ї –і–µ—В–∞–ї–Є ID={detail_id}: {len(suitable_reserves)}")

                    if suitable_reserves.empty:
                        errors.append(
                            f"вЭМ –Э–µ –љ–∞–є–і–µ–љ —А–µ–Ј–µ—А–≤ –і–ї—П:\n"
                            f"   –Ч–∞–Ї–∞–Ј: {order_name}\n"
                            f"   –Ь–∞—В–µ—А–Є–∞–ї: {marka} {thickness}–Љ–Љ {width}x{length}\n"
                            f"   –Ф–µ—В–∞–ї—М: {part_name}"
                        )
                        print(f"   вЭМ –Я–Њ–і—Е–Њ–і—П—Й–Є–є —А–µ–Ј–µ—А–≤ –љ–µ –љ–∞–є–і–µ–љ")
                        continue

                    # –С–µ—А—С–Љ –њ–µ—А–≤—Л–є –њ–Њ–і—Е–Њ–і—П—Й–Є–є —А–µ–Ј–µ—А–≤
                    reserve_row = suitable_reserves.iloc[0]
                    reserve_id = int(reserve_row["ID —А–µ–Ј–µ—А–≤–∞"])
                    remainder = int(reserve_row["–Ю—Б—В–∞—В–Њ–Ї –Ї —Б–њ–Є—Б–∞–љ–Є—О"])

                    print(f"   вЬЕ –Т—Л–±—А–∞–љ —А–µ–Ј–µ—А–≤ ID={reserve_id}, –Њ—Б—В–∞—В–Њ–Ї={remainder} —И—В")

                    # ========== –®–Р–У 5: –Ъ–Ю–Ы–Ш–І–Х–°–Ґ–Т–Ю –Ф–Ы–ѓ –°–Я–Ш–°–Р–Э–Ш–ѓ ==========
                    try:
                        qty_to_writeoff = int(metal_qty_str)
                    except:
                        qty_to_writeoff = 1

                    if qty_to_writeoff > remainder:
                        errors.append(
                            f"вЪ†пЄП –Э–µ–і–Њ—Б—В–∞—В–Њ—З–љ–Њ –Љ–∞—В–µ—А–Є–∞–ї–∞ –≤ —А–µ–Ј–µ—А–≤–µ #{reserve_id}:\n"
                            f"   –Ч–∞–њ—А–Њ—И–µ–љ–Њ: {qty_to_writeoff}, –Ф–Њ—Б—В—Г–њ–љ–Њ: {remainder}"
                        )
                        print(f"   вЪ†пЄП –Э–µ–і–Њ—Б—В–∞—В–Њ—З–љ–Њ –Љ–∞—В–µ—А–Є–∞–ї–∞: –љ—Г–ґ–љ–Њ {qty_to_writeoff}, –µ—Б—В—М {remainder}")
                        # –°–њ–Є—Б—Л–≤–∞–µ–Љ —Б–Ї–Њ–ї—М–Ї–Њ –µ—Б—В—М
                        qty_to_writeoff = remainder

                    print(f"   рЯУЭ –С—Г–і–µ—В —Б–њ–Є—Б–∞–љ–Њ: {qty_to_writeoff} —И—В")

                    # ========== –®–Р–У 6: –°–Ю–Ч–Ф–Р–Э–Ш–Х –°–Я–Ш–°–Р–Э–Ш–ѓ ==========
                    new_writeoff_id = 1 if writeoffs_df.empty else int(writeoffs_df["ID —Б–њ–Є—Б–∞–љ–Є—П"].max()) + 1

                    # рЯЖХ –£–Ы–£–І–®–Х–Э–Э–Ђ–Щ –Ъ–Ю–Ь–Ь–Х–Э–Ґ–Р–†–Ш–Щ –і–ї—П —Б–≤—П–Ј–Є —Б —В–∞–±–ї–Є—Ж–µ–є –Є–Љ–њ–Њ—А—В–∞
                    comment_text = (
                        f"–Ы–∞–Ј–µ—А: {username} | "
                        f"–Ф–µ—В–∞–ї—М: {part_name} | "
                        f"–Ф–∞—В–∞ –Є–Љ–њ–Њ—А—В–∞: {date_val} {time_val}"
                    )

                    new_writeoff = pd.DataFrame([{
                        "ID —Б–њ–Є—Б–∞–љ–Є—П": new_writeoff_id,
                        "ID —А–µ–Ј–µ—А–≤–∞": reserve_id,
                        "ID –Ј–∞–Ї–∞–Ј–∞": reserve_row["ID –Ј–∞–Ї–∞–Ј–∞"],
                        "ID –Љ–∞—В–µ—А–Є–∞–ї–∞": reserve_row["ID –Љ–∞—В–µ—А–Є–∞–ї–∞"],
                        "–Ь–∞—А–Ї–∞": reserve_row["–Ь–∞—А–Ї–∞"],
                        "–Ґ–Њ–ї—Й–Є–љ–∞": reserve_row["–Ґ–Њ–ї—Й–Є–љ–∞"],
                        "–Ф–ї–Є–љ–∞": reserve_row["–Ф–ї–Є–љ–∞"],
                        "–®–Є—А–Є–љ–∞": reserve_row["–®–Є—А–Є–љ–∞"],
                        "–Ъ–Њ–ї–Є—З–µ—Б—В–≤–Њ": qty_to_writeoff,
                        "–Ф–∞—В–∞ —Б–њ–Є—Б–∞–љ–Є—П": f"{date_val} {time_val}",  # рЯЖХ –°–Ю–•–†–Р–Э–ѓ–Х–Ь –Ш–°–•–Ю–Ф–Э–£–Ѓ –Ф–Р–Ґ–£
                        "–Ъ–Њ–Љ–Љ–µ–љ—В–∞—А–Є–є": comment_text  # рЯЖХ –†–Р–°–®–Ш–†–Х–Э–Э–Ђ–Щ –Ъ–Ю–Ь–Ь–Х–Э–Ґ–Р–†–Ш–Щ
                    }])

                    writeoffs_df = pd.concat([writeoffs_df, new_writeoff], ignore_index=True)

                    # ========== –®–Р–У 7: –Ю–С–Э–Ю–Т–Ы–Х–Э–Ш–Х –†–Х–Ч–Х–†–Т–Р ==========
                    new_written_off = int(reserve_row["–°–њ–Є—Б–∞–љ–Њ"]) + qty_to_writeoff
                    new_remainder = int(reserve_row["–Ч–∞—А–µ–Ј–µ—А–≤–Є—А–Њ–≤–∞–љ–Њ —И—В—Г–Ї"]) - new_written_off

                    reservations_df.loc[reservations_df["ID —А–µ–Ј–µ—А–≤–∞"] == reserve_id, "–°–њ–Є—Б–∞–љ–Њ"] = new_written_off
                    reservations_df.loc[
                        reservations_df["ID —А–µ–Ј–µ—А–≤–∞"] == reserve_id, "–Ю—Б—В–∞—В–Њ–Ї –Ї —Б–њ–Є—Б–∞–љ–Є—О"] = new_remainder

                    print(f"   вЬЕ –†–µ–Ј–µ—А–≤ –Њ–±–љ–Њ–≤–ї–µ–љ: –°–њ–Є—Б–∞–љ–Њ={new_written_off}, –Ю—Б—В–∞—В–Њ–Ї={new_remainder}")

                    # ========== –®–Р–У 8: –Ю–С–Э–Ю–Т–Ы–Х–Э–Ш–Х –Ь–Р–Ґ–Х–†–Ш–Р–Ы–Р –Э–Р –°–Ъ–Ы–Р–Ф–Х ==========
                    material_id = int(reserve_row["ID –Љ–∞—В–µ—А–Є–∞–ї–∞"])
                    if material_id != -1:
                        material = materials_df[materials_df["ID"] == material_id]
                        if not material.empty:
                            material = material.iloc[0]

                            new_qty = int(material["–Ъ–Њ–ї–Є—З–µ—Б—В–≤–Њ —И—В—Г–Ї"]) - qty_to_writeoff
                            new_reserved = int(material["–Ч–∞—А–µ–Ј–µ—А–≤–Є—А–Њ–≤–∞–љ–Њ"]) - qty_to_writeoff

                            materials_df.loc[materials_df["ID"] == material_id, "–Ъ–Њ–ї–Є—З–µ—Б—В–≤–Њ —И—В—Г–Ї"] = new_qty
                            materials_df.loc[materials_df["ID"] == material_id, "–Ч–∞—А–µ–Ј–µ—А–≤–Є—А–Њ–≤–∞–љ–Њ"] = new_reserved

                            # –Я–µ—А–µ—Б—З–Є—В—Л–≤–∞–µ–Љ –њ–ї–Њ—Й–∞–і—М
                            area_per_piece = float(material["–Ф–ї–Є–љ–∞"]) * float(material["–®–Є—А–Є–љ–∞"]) / 1_000_000
                            new_area = new_qty * area_per_piece
                            materials_df.loc[materials_df["ID"] == material_id, "–Ю–±—Й–∞—П –њ–ї–Њ—Й–∞–і—М"] = round(new_area, 2)

                            print(f"   вЬЕ –°–Ї–ї–∞–і –Њ–±–љ–Њ–≤–ї–µ–љ: –Т—Б–µ–≥–Њ={new_qty}, –Ч–∞—А–µ–Ј–µ—А–≤–Є—А–Њ–≤–∞–љ–Њ={new_reserved}")

                    # ========== –®–Р–У 9: –Ю–С–Э–Ю–Т–Ы–Х–Э–Ш–Х –Ф–Х–Ґ–Р–Ы–Ш –Т –Ч–Р–Ъ–Р–Ч–Х (–Я–Ю–†–Х–Ч–Р–Э–Ю) ==========
                    if detail_id:
                        try:
                            # –Ч–∞–≥—А—Г–ґ–∞–µ–Љ –і–µ—В–∞–ї–Є –Ј–∞–Ї–∞–Ј–∞ (–µ—Б–ї–Є –µ—Й—С –љ–µ –Ј–∞–≥—А—Г–ґ–µ–љ—Л)
                            if 'order_details_df' not in locals():
                                order_details_df = load_data("OrderDetails")

                            detail_row = order_details_df[order_details_df["ID"] == detail_id]

                            if not detail_row.empty:
                                detail_row = detail_row.iloc[0]
                                detail_name_full = detail_row["–Э–∞–Ј–≤–∞–љ–Є–µ –і–µ—В–∞–ї–Є"]

                                old_cut = int(detail_row.get("–Я–Њ—А–µ–Ј–∞–љ–Њ", 0))

                                # –Ъ–Њ–ї–Є—З–µ—Б—В–≤–Њ –і–µ—В–∞–ї–µ–є –Є–Ј –Є–Љ–њ–Њ—А—В–∞
                                try:
                                    parts_qty = int(part_qty)
                                except:
                                    parts_qty = 0

                                new_cut = old_cut + parts_qty

                                # –Ю–±–љ–Њ–≤–ї—П–µ–Љ –Ї–Њ–ї–Є—З–µ—Б—В–≤–Њ –њ–Њ—А–µ–Ј–∞–љ–љ—Л—Е –і–µ—В–∞–ї–µ–є
                                order_details_df.loc[order_details_df["ID"] == detail_id, "–Я–Њ—А–µ–Ј–∞–љ–Њ"] = new_cut

                                # –Я—А–Њ–≤–µ—А—П–µ–Љ –Њ–±—Й–µ–µ –Ї–Њ–ї–Є—З–µ—Б—В–≤–Њ
                                total_qty = int(detail_row.get("–Ъ–Њ–ї–Є—З–µ—Б—В–≤–Њ", 0))

                                print(f"   рЯУР –Ф–µ—В–∞–ї—М '{detail_name_full}' –Њ–±–љ–Њ–≤–ї–µ–љ–∞:")
                                print(f"      ID –і–µ—В–∞–ї–Є: {detail_id}")
                                print(f"      –Т—Б–µ–≥–Њ —В—А–µ–±—Г–µ—В—Б—П: {total_qty}")
                                print(f"      –С—Л–ї–Њ –њ–Њ—А–µ–Ј–∞–љ–Њ: {old_cut}")
                                print(f"      –Ф–Њ–±–∞–≤–ї–µ–љ–Њ: +{parts_qty}")
                                print(f"      –°—В–∞–ї–Њ –њ–Њ—А–µ–Ј–∞–љ–Њ: {new_cut}")

                                # –°–Њ—Е—А–∞–љ—П–µ–Љ –Є–Ј–Љ–µ–љ–µ–љ–Є—П
                                save_data("OrderDetails", order_details_df)

                                print(f"      рЯТЊ OrderDetails —Б–Њ—Е—А–∞–љ—С–љ")

                                # –Х—Б–ї–Є –њ–Њ—А–µ–Ј–∞–љ–Њ –±–Њ–ї—М—И–µ –Є–ї–Є —А–∞–≤–љ–Њ —В—А–µ–±—Г–µ–Љ–Њ–Љ—Г - –њ–Њ–Ї–∞–Ј—Л–≤–∞–µ–Љ —Г–≤–µ–і–Њ–Љ–ї–µ–љ–Є–µ
                                if new_cut >= total_qty:
                                    print(f"      вЬЕ –Ф–µ—В–∞–ї—М –њ–Њ–ї–љ–Њ—Б—В—М—О –њ–Њ—А–µ–Ј–∞–љ–∞! ({new_cut}/{total_qty})")
                                else:
                                    remaining = total_qty - new_cut
                                    print(f"      вП≥ –Ю—Б—В–∞–ї–Њ—Б—М –њ–Њ—А–µ–Ј–∞—В—М: {remaining} —И—В")
                            else:
                                print(f"   вЪ†пЄП –Ф–µ—В–∞–ї—М ID={detail_id} –љ–µ –љ–∞–є–і–µ–љ–∞ –≤ OrderDetails")

                        except Exception as e:
                            print(f"   вЪ†пЄП –Ю—И–Є–±–Ї–∞ –Њ–±–љ–Њ–≤–ї–µ–љ–Є—П –і–µ—В–∞–ї–Є: {e}")
                            import traceback
                            traceback.print_exc()
                    else:
                        print(f"   вДєпЄП –Ф–µ—В–∞–ї—М –љ–µ –љ–∞–є–і–µ–љ–∞ –≤ –±–∞–Ј–µ, –њ—А–Њ–њ—Г—Б–Ї–∞–µ–Љ –Њ–±–љ–Њ–≤–ї–µ–љ–Є–µ '–Я–Њ—А–µ–Ј–∞–љ–Њ'")

                    # ========== –®–Р–У 10: –Ю–С–Э–Ю–Т–Ы–Х–Э–Ш–Х –°–Ґ–Р–Ґ–£–°–Р –Т –Ґ–Р–С–Ы–Ш–¶–Х –Ш–Ь–Я–Ю–†–Ґ–Р ==========
                    item_index = self.laser_import_tree.index(item)
                    self.laser_table_data[item_index]["–°–њ–Є—Б–∞–љ–Њ"] = "вЬУ"
                    self.laser_table_data[item_index]["–Ф–∞—В–∞ —Б–њ–Є—Б–∞–љ–Є—П"] = datetime.now().strftime("%Y-%m-%d %H:%M")

                    # ========== –®–Р–У 9: –Ю–С–Э–Ю–Т–Ы–Х–Э–Ш–Х –°–Ґ–Р–Ґ–£–°–Р –Т –Ґ–Р–С–Ы–Ш–¶–Х –Ш–Ь–Я–Ю–†–Ґ–Р ==========
                    item_index = self.laser_import_tree.index(item)
                    self.laser_table_data[item_index]["–°–њ–Є—Б–∞–љ–Њ"] = "вЬУ"
                    self.laser_table_data[item_index]["–Ф–∞—В–∞ —Б–њ–Є—Б–∞–љ–Є—П"] = datetime.now().strftime("%Y-%m-%d %H:%M")

                    success_count += 1
                    print(f"   вЬЕ –°–Я–Ш–°–Р–Э–Ш–Х –Т–Ђ–Я–Ю–Ы–Э–Х–Э–Ю –£–°–Я–Х–®–Э–Ю")

                except Exception as e:
                    error_msg = f"вЭМ –Ю—И–Є–±–Ї–∞ –Њ–±—А–∞–±–Њ—В–Ї–Є —Б—В—А–Њ–Ї–Є '{order_name}': {str(e)}"
                    errors.append(error_msg)
                    print(f"   {error_msg}")
                    import traceback
                    traceback.print_exc()

            # ========== –°–Ю–•–†–Р–Э–Х–Э–Ш–Х –Ш–Ч–Ь–Х–Э–Х–Э–Ш–Щ ==========
            print(f"\n{'=' * 80}")
            print(f"рЯТЊ –°–Ю–•–†–Р–Э–Х–Э–Ш–Х –Ш–Ч–Ь–Х–Э–Х–Э–Ш–Щ –Т –С–Р–Ч–£ –Ф–Р–Э–Э–Ђ–•")
            print(f"{'=' * 80}")

            save_data("WriteOffs", writeoffs_df)
            save_data("Reservations", reservations_df)
            save_data("Materials", materials_df)

            print(f"вЬЕ –Ф–∞–љ–љ—Л–µ —Б–Њ—Е—А–∞–љ–µ–љ—Л")

            # –Ю–С–Э–Ю–Т–Ы–Х–Э–Ш–Х –Ш–Э–Ґ–Х–†–§–Х–Щ–°–Р
            self.refresh_laser_import_table()
            self.refresh_materials()
            self.refresh_reservations()
            self.refresh_writeoffs()
            self.refresh_balance()

            if hasattr(self, 'refresh_orders'):
                self.refresh_orders()
            if hasattr(self, 'refresh_order_details'):
                self.refresh_order_details()

            # рЯЖХ –Ю–С–Э–Ю–Т–Ы–ѓ–Х–Ь –Т–Ъ–Ы–Р–Ф–Ъ–£ –£–І–Б–Ґ–Р –Ф–Х–Ґ–Р–Ы–Х–Щ
            if hasattr(self, 'refresh_details'):
                self.refresh_details()


            print(f"вЬЕ –Ш–љ—В–µ—А—Д–µ–є—Б –Њ–±–љ–Њ–≤–ї–µ–љ")

            # ========== –†–Х–Ч–£–Ы–ђ–Ґ–Р–Ґ ==========
            print(f"\n{'=' * 80}")
            print(f"вЬЕ –°–Я–Ш–°–Р–Э–Ш–Х –Ч–Р–Т–Х–†–®–Х–Э–Ю")
            print(f"   –£—Б–њ–µ—И–љ–Њ: {success_count}")
            print(f"   –Ю—И–Є–±–Њ–Ї: {len(errors)}")
            print(f"{'=' * 80}\n")

            result_msg = f"вЬЕ –£—Б–њ–µ—И–љ–Њ —Б–њ–Є—Б–∞–љ–Њ: {success_count} –Ј–∞–њ–Є—Б–µ–є"
            if errors:
                result_msg += f"\n\nвЪ† –Ю—И–Є–±–Ї–Є ({len(errors)}):\n" + "\n".join(errors[:10])
                if len(errors) > 10:
                    result_msg += f"\n... –Є –µ—Й–µ {len(errors) - 10}"

            messagebox.showinfo("–†–µ–Ј—Г–ї—М—В–∞—В —Б–њ–Є—Б–∞–љ–Є—П", result_msg)

            # рЯЖХ –Р–Т–Ґ–Ю–°–Ю–•–†–Р–Э–Х–Э–Ш–Х –Я–Ю–°–Ы–Х –°–Я–Ш–°–Р–Э–Ш–ѓ
            self.save_laser_import_cache()

        except Exception as e:
            print(f"\nрЯТ• –Ъ–†–Ш–Ґ–Ш–І–Х–°–Ъ–Р–ѓ –Ю–®–Ш–С–Ъ–Р: {e}")
            import traceback
            traceback.print_exc()
            messagebox.showerror("–Ю—И–Є–±–Ї–∞", f"–Э–µ —Г–і–∞–ї–Њ—Б—М –≤—Л–њ–Њ–ї–љ–Є—В—М —Б–њ–Є—Б–∞–љ–Є–µ:\n{e}")

    def mark_manual_writeoff(self):
        """–Я–Њ–Љ–µ—В–Ї–∞ —Б—В—А–Њ–Ї –Ї–∞–Ї '—Б–њ–Є—Б–∞–љ–Њ –≤—А—Г—З–љ—Г—О' –±–µ–Ј —Д–∞–Ї—В–Є—З–µ—Б–Ї–Њ–≥–Њ —Б–њ–Є—Б–∞–љ–Є—П"""
        selected_items = self.laser_import_tree.selection()

        if not selected_items:
            messagebox.showwarning("–Я—А–µ–і—Г–њ—А–µ–ґ–і–µ–љ–Є–µ", "–Т—Л–±–µ—А–Є—В–µ —Б—В—А–Њ–Ї–Є –і–ї—П –њ–Њ–Љ–µ—В–Ї–Є!")
            return

        # –Я—А–Њ–≤–µ—А—П–µ–Љ, —З—В–Њ —Б—В—А–Њ–Ї–Є –µ—Й–µ –љ–µ —Б–њ–Є—Б–∞–љ—Л
        rows_to_mark = []
        already_marked = []

        for item in selected_items:
            values = self.laser_import_tree.item(item)['values']
            status = values[8]  # –Ъ–Њ–ї–Њ–љ–Ї–∞ "–°–њ–Є—Б–∞–љ–Њ"

            if status in ["вЬУ", "–Ф–∞", "Yes"]:
                already_marked.append(f"{values[3]} (–∞–≤—В–Њ–Љ–∞—В–Є—З–µ—Б–Ї–Є)")
            elif status == "–Т—А—Г—З–љ—Г—О":
                already_marked.append(f"{values[3]} (—Г–ґ–µ –њ–Њ–Љ–µ—З–µ–љ–Њ –≤—А—Г—З–љ—Г—О)")
            else:
                rows_to_mark.append((item, values))

        if already_marked:
            messagebox.showinfo("–Ш–љ—Д–Њ—А–Љ–∞—Ж–Є—П",
                                f"–Э–µ–Ї–Њ—В–Њ—А—Л–µ —Б—В—А–Њ–Ї–Є —Г–ґ–µ –Њ–±—А–∞–±–Њ—В–∞–љ—Л:\n" + "\n".join(already_marked[:5]))

        if not rows_to_mark:
            messagebox.showwarning("–Я—А–µ–і—Г–њ—А–µ–ґ–і–µ–љ–Є–µ", "–Э–µ—В —Б—В—А–Њ–Ї –і–ї—П –њ–Њ–Љ–µ—В–Ї–Є!")
            return

        # –Я–Њ–і—В–≤–µ—А–ґ–і–µ–љ–Є–µ
        confirm_msg = (
            f"–Я–Њ–Љ–µ—В–Є—В—М {len(rows_to_mark)} —Б—В—А–Њ–Ї(–Є) –Ї–∞–Ї '—Б–њ–Є—Б–∞–љ–Њ –≤—А—Г—З–љ—Г—О'?\n\n"
            f"вЪ†пЄП –≠—В–Њ –Э–Х —Б–њ–Є—И–µ—В –Љ–∞—В–µ—А–Є–∞–ї —Б —А–µ–Ј–µ—А–≤–Њ–≤!\n"
            f"–≠—В–Њ —В–Њ–ї—М–Ї–Њ –њ–Њ–Љ–µ—В–Є—В —Б—В—А–Њ–Ї–Є –і–ї—П –њ–Њ—Б–ї–µ–і—Г—О—Й–µ–≥–Њ —А—Г—З–љ–Њ–≥–Њ —Б–њ–Є—Б–∞–љ–Є—П.\n\n"
            f"–°—В—А–Њ–Ї–Є –Њ–Ї—А–∞—Б—П—В—Б—П –≤ —Б–≤–µ—В–ї–Њ-—Б–Є–љ–Є–є —Ж–≤–µ—В."
        )

        if not messagebox.askyesno("–Я–Њ–і—В–≤–µ—А–ґ–і–µ–љ–Є–µ", confirm_msg):
            return

        try:
            marked_count = 0

            for item, values in rows_to_mark:
                # –Ю–±–љ–Њ–≤–ї—П–µ–Љ —Б—В–∞—В—Г—Б –≤ —В–∞–±–ї–Є—Ж–µ –і–∞–љ–љ—Л—Е
                item_index = self.laser_import_tree.index(item)

                if item_index < len(self.laser_table_data):
                    self.laser_table_data[item_index]["–°–њ–Є—Б–∞–љ–Њ"] = "–Т—А—Г—З–љ—Г—О"
                    self.laser_table_data[item_index]["–Ф–∞—В–∞ —Б–њ–Є—Б–∞–љ–Є—П"] = datetime.now().strftime("%Y-%m-%d %H:%M")

                    # –Ю–±–љ–Њ–≤–ї—П–µ–Љ –≤–Є–Ј—Г–∞–ї—М–љ–Њ–µ –Њ—В–Њ–±—А–∞–ґ–µ–љ–Є–µ
                    new_values = list(values)
                    new_values[8] = "–Т—А—Г—З–љ—Г—О"  # –Ъ–Њ–ї–Њ–љ–Ї–∞ "–°–њ–Є—Б–∞–љ–Њ"
                    new_values[9] = datetime.now().strftime("%Y-%m-%d %H:%M")  # –Ъ–Њ–ї–Њ–љ–Ї–∞ "–Ф–∞—В–∞ —Б–њ–Є—Б–∞–љ–Є—П"

                    self.laser_import_tree.item(item, values=new_values, tags=('manual',))
                    marked_count += 1

            messagebox.showinfo("–£—Б–њ–µ—Е",
                                f"вЬЕ –Я–Њ–Љ–µ—З–µ–љ–Њ —Б—В—А–Њ–Ї: {marked_count}\n\n"
                                f"рЯФµ –°—В—А–Њ–Ї–Є –Њ–Ї—А–∞—И–µ–љ—Л –≤ —Б–≤–µ—В–ї–Њ-—Б–Є–љ–Є–є —Ж–≤–µ—В\n"
                                f"рЯУЭ –Э–µ –Ј–∞–±—Г–і—М—В–µ —Б–њ–Є—Б–∞—В—М –Љ–∞—В–µ—А–Є–∞–ї –≤—А—Г—З–љ—Г—О!")

        except Exception as e:
            messagebox.showerror("–Ю—И–Є–±–Ї–∞", f"–Э–µ —Г–і–∞–ї–Њ—Б—М –њ–Њ–Љ–µ—В–Є—В—М —Б—В—А–Њ–Ї–Є:\n{e}")
            import traceback
            traceback.print_exc()

    def unmark_manual_writeoff(self):
        """–°–љ—П—В–Є–µ –њ–Њ–Љ–µ—В–Ї–Є '—Б–њ–Є—Б–∞–љ–Њ –≤—А—Г—З–љ—Г—О'"""
        selected_items = self.laser_import_tree.selection()

        if not selected_items:
            messagebox.showwarning("–Я—А–µ–і—Г–њ—А–µ–ґ–і–µ–љ–Є–µ", "–Т—Л–±–µ—А–Є—В–µ —Б—В—А–Њ–Ї–Є –і–ї—П —Б–љ—П—В–Є—П –њ–Њ–Љ–µ—В–Ї–Є!")
            return

        # –Я—А–Њ–≤–µ—А—П–µ–Љ, —З—В–Њ —Б—В—А–Њ–Ї–Є –њ–Њ–Љ–µ—З–µ–љ—Л –≤—А—Г—З–љ—Г—О
        rows_to_unmark = []

        for item in selected_items:
            values = self.laser_import_tree.item(item)['values']
            status = values[8]  # –Ъ–Њ–ї–Њ–љ–Ї–∞ "–°–њ–Є—Б–∞–љ–Њ"

            if status == "–Т—А—Г—З–љ—Г—О":
                rows_to_unmark.append((item, values))

        if not rows_to_unmark:
            messagebox.showwarning("–Я—А–µ–і—Г–њ—А–µ–ґ–і–µ–љ–Є–µ",
                                   "–Т—Л–±—А–∞–љ–љ—Л–µ —Б—В—А–Њ–Ї–Є –љ–µ –њ–Њ–Љ–µ—З–µ–љ—Л –≤—А—Г—З–љ—Г—О!\n\n"
                                   "–°–љ—П—В—М –Љ–Њ–ґ–љ–Њ —В–Њ–ї—М–Ї–Њ –њ–Њ–Љ–µ—В–Ї—Г '–Т—А—Г—З–љ—Г—О'.\n"
                                   "–Р–≤—В–Њ–Љ–∞—В–Є—З–µ—Б–Ї–Є–µ —Б–њ–Є—Б–∞–љ–Є—П —Г–і–∞–ї—П—О—В—Б—П —З–µ—А–µ–Ј –≤–Ї–ї–∞–і–Ї—Г '–°–њ–Є—Б–∞–љ–Є–µ –Љ–∞—В–µ—А–Є–∞–ї–Њ–≤'.")
            return

        # –Я–Њ–і—В–≤–µ—А–ґ–і–µ–љ–Є–µ
        if not messagebox.askyesno("–Я–Њ–і—В–≤–µ—А–ґ–і–µ–љ–Є–µ",
                                   f"–°–љ—П—В—М –њ–Њ–Љ–µ—В–Ї—Г —Б {len(rows_to_unmark)} —Б—В—А–Њ–Ї(–Є)?"):
            return

        try:
            unmarked_count = 0

            for item, values in rows_to_unmark:
                # –Ю–±–љ–Њ–≤–ї—П–µ–Љ —Б—В–∞—В—Г—Б –≤ —В–∞–±–ї–Є—Ж–µ –і–∞–љ–љ—Л—Е
                item_index = self.laser_import_tree.index(item)

                if item_index < len(self.laser_table_data):
                    self.laser_table_data[item_index]["–°–њ–Є—Б–∞–љ–Њ"] = ""
                    self.laser_table_data[item_index]["–Ф–∞—В–∞ —Б–њ–Є—Б–∞–љ–Є—П"] = ""

                    # –Ю–±–љ–Њ–≤–ї—П–µ–Љ –≤–Є–Ј—Г–∞–ї—М–љ–Њ–µ –Њ—В–Њ–±—А–∞–ґ–µ–љ–Є–µ
                    new_values = list(values)
                    new_values[8] = ""  # –Ъ–Њ–ї–Њ–љ–Ї–∞ "–°–њ–Є—Б–∞–љ–Њ"
                    new_values[9] = ""  # –Ъ–Њ–ї–Њ–љ–Ї–∞ "–Ф–∞—В–∞ —Б–њ–Є—Б–∞–љ–Є—П"

                    self.laser_import_tree.item(item, values=new_values, tags=('pending',))
                    unmarked_count += 1

            messagebox.showinfo("–£—Б–њ–µ—Е", f"вЬЕ –°–љ—П—В–Њ –њ–Њ–Љ–µ—В–Њ–Ї: {unmarked_count}")

        except Exception as e:
            messagebox.showerror("–Ю—И–Є–±–Ї–∞", f"–Э–µ —Г–і–∞–ї–Њ—Б—М —Б–љ—П—В—М –њ–Њ–Љ–µ—В–Ї—Г:\n{e}")

    def edit_laser_row(self):
        """–†–µ–і–∞–Ї—В–Є—А–Њ–≤–∞–љ–Є–µ –≤—Л–±—А–∞–љ–љ–Њ–є —Б—В—А–Њ–Ї–Є –Є–Љ–њ–Њ—А—В–∞"""
        selected = self.laser_import_tree.selection()
        if not selected or len(selected) != 1:
            messagebox.showwarning("–Я—А–µ–і—Г–њ—А–µ–ґ–і–µ–љ–Є–µ", "–Т—Л–±–µ—А–Є—В–µ –Њ–і–љ—Г —Б—В—А–Њ–Ї—Г –і–ї—П —А–µ–і–∞–Ї—В–Є—А–Њ–≤–∞–љ–Є—П!")
            return

        item_index = self.laser_import_tree.index(selected[0])
        row_data = self.laser_table_data[item_index]

        # –Ю–Ї–љ–Њ —А–µ–і–∞–Ї—В–Є—А–Њ–≤–∞–љ–Є—П
        edit_window = tk.Toplevel(self.root)
        edit_window.title("–†–µ–і–∞–Ї—В–Є—А–Њ–≤–∞–љ–Є–µ –Ј–∞–њ–Є—Б–Є")
        edit_window.geometry("500x400")
        edit_window.configure(bg='#ecf0f1')

        tk.Label(edit_window, text="–†–µ–і–∞–Ї—В–Є—А–Њ–≤–∞–љ–Є–µ –Ј–∞–њ–Є—Б–Є –Њ—В –ї–∞–Ј–µ—А—Й–Є–Ї–Њ–≤",
                 font=("Arial", 12, "bold"), bg='#ecf0f1').pack(pady=10)

        # –Я–Њ–ї—П –і–ї—П —А–µ–і–∞–Ї—В–Є—А–Њ–≤–∞–љ–Є—П
        fields = [
            ("–Ч–∞–Ї–∞–Ј:", "order"),
            ("–Ь–µ—В–∞–ї–ї:", "metal"),
            ("–Ъ–Њ–ї-–≤–Њ –Љ–µ—В–∞–ї–ї–∞:", "metal_quantity"),
            ("–Ф–µ—В–∞–ї—М:", "part"),
            ("–Ъ–Њ–ї-–≤–Њ –і–µ—В–∞–ї–µ–є:", "part_quantity")
        ]

        entries = {}
        for label_text, key in fields:
            frame = tk.Frame(edit_window, bg='#ecf0f1')
            frame.pack(fill=tk.X, padx=20, pady=5)
            tk.Label(frame, text=label_text, width=20, anchor='w', bg='#ecf0f1', font=("Arial", 10)).pack(side=tk.LEFT)
            entry = tk.Entry(frame, font=("Arial", 10))
            entry.insert(0, str(row_data.get(key, "")))
            entry.pack(side=tk.RIGHT, expand=True, fill=tk.X)
            entries[key] = entry

        def save_changes():
            for key, entry in entries.items():
                self.laser_table_data[item_index][key] = entry.get()

            self.refresh_laser_import_table()
            edit_window.destroy()
            messagebox.showinfo("–£—Б–њ–µ—Е", "–Ч–∞–њ–Є—Б—М –Њ–±–љ–Њ–≤–ї–µ–љ–∞!")

        tk.Button(edit_window, text="рЯТЊ –°–Њ—Е—А–∞–љ–Є—В—М", bg='#3498db', fg='white',
                  font=("Arial", 12, "bold"), command=save_changes).pack(pady=20)

    def delete_laser_row(self):
        """–£–і–∞–ї–µ–љ–Є–µ –≤—Л–±—А–∞–љ–љ—Л—Е —Б—В—А–Њ–Ї"""
        selected_items = self.laser_import_tree.selection()

        if not selected_items:
            messagebox.showwarning("–Я—А–µ–і—Г–њ—А–µ–ґ–і–µ–љ–Є–µ", "–Т—Л–±–µ—А–Є—В–µ —Б—В—А–Њ–Ї–Є –і–ї—П —Г–і–∞–ї–µ–љ–Є—П!")
            return

        if not messagebox.askyesno("–Я–Њ–і—В–≤–µ—А–ґ–і–µ–љ–Є–µ",
                                   f"–£–і–∞–ї–Є—В—М –≤—Л–±—А–∞–љ–љ—Л–µ —Б—В—А–Њ–Ї–Є ({len(selected_items)} —И—В)?"):
            return

        # –£–і–∞–ї—П–µ–Љ –≤ –Њ–±—А–∞—В–љ–Њ–Љ –њ–Њ—А—П–і–Ї–µ, —З—В–Њ–±—Л –Є–љ–і–µ–Ї—Б—Л –љ–µ —Б–±–Є–≤–∞–ї–Є—Б—М
        indices_to_delete = sorted([self.laser_import_tree.index(item) for item in selected_items], reverse=True)

        for index in indices_to_delete:
            del self.laser_table_data[index]

        self.refresh_laser_import_table()
        messagebox.showinfo("–£—Б–њ–µ—Е", f"–£–і–∞–ї–µ–љ–Њ –Ј–∞–њ–Є—Б–µ–є: {len(indices_to_delete)}")

        # рЯЖХ –Р–Т–Ґ–Ю–°–Ю–•–†–Р–Э–Х–Э–Ш–Х –Я–Ю–°–Ы–Х –£–Ф–Р–Ы–Х–Э–Ш–ѓ
        self.save_laser_import_cache()

    def export_laser_table(self):
        """–≠–Ї—Б–њ–Њ—А—В —В–∞–±–ї–Є—Ж—Л –Њ–±—А–∞—В–љ–Њ –≤ Excel"""
        if not self.laser_table_data:
            messagebox.showwarning("–Я—А–µ–і—Г–њ—А–µ–ґ–і–µ–љ–Є–µ", "–Э–µ—В –і–∞–љ–љ—Л—Е –і–ї—П —Н–Ї—Б–њ–Њ—А—В–∞!")
            return

        file_path = filedialog.asksaveasfilename(
            title="–°–Њ—Е—А–∞–љ–Є—В—М —В–∞–±–ї–Є—Ж—Г",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("CSV files", "*.csv"), ("All files", "*.*")],
            initialfile=f"laser_log_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )

        if not file_path:
            return

        try:
            df = pd.DataFrame(self.laser_table_data)

            if file_path.endswith('.csv'):
                df.to_csv(file_path, index=False, sep=';', encoding='utf-8')
            else:
                df.to_excel(file_path, index=False, engine='openpyxl')

            messagebox.showinfo("–£—Б–њ–µ—Е", f"–Ґ–∞–±–ї–Є—Ж–∞ —Б–Њ—Е—А–∞–љ–µ–љ–∞:\n{file_path}")
        except Exception as e:
            messagebox.showerror("–Ю—И–Є–±–Ї–∞", f"–Э–µ —Г–і–∞–ї–Њ—Б—М —Б–Њ—Е—А–∞–љ–Є—В—М —Д–∞–є–ї:\n{e}")

    def export_laser_table(self):
        """–≠–Ї—Б–њ–Њ—А—В —В–∞–±–ї–Є—Ж—Л –Њ–±—А–∞—В–љ–Њ –≤ Excel"""
        if not self.laser_table_data:
            messagebox.showwarning("–Я—А–µ–і—Г–њ—А–µ–ґ–і–µ–љ–Є–µ", "–Э–µ—В –і–∞–љ–љ—Л—Е –і–ї—П —Н–Ї—Б–њ–Њ—А—В–∞!")
            return

        file_path = filedialog.asksaveasfilename(
            title="–°–Њ—Е—А–∞–љ–Є—В—М —В–∞–±–ї–Є—Ж—Г",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("CSV files", "*.csv"), ("All files", "*.*")],
            initialfile=f"laser_log_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )

        if not file_path:
            return

        try:
            df = pd.DataFrame(self.laser_table_data)

            if file_path.endswith('.csv'):
                df.to_csv(file_path, index=False, sep=';', encoding='utf-8')
            else:
                df.to_excel(file_path, index=False, engine='openpyxl')

            messagebox.showinfo("–£—Б–њ–µ—Е", f"–Ґ–∞–±–ї–Є—Ж–∞ —Б–Њ—Е—А–∞–љ–µ–љ–∞:\n{file_path}")
        except Exception as e:
            messagebox.showerror("–Ю—И–Є–±–Ї–∞", f"–Э–µ —Г–і–∞–ї–Њ—Б—М —Б–Њ—Е—А–∞–љ–Є—В—М —Д–∞–є–ї:\n{e}")

    # рЯЖХ –Э–Ю–Т–Ђ–Щ –Ь–Х–Ґ–Ю–Ф - –°–Ю–•–†–Р–Э–Х–Э–Ш–Х –Ъ–≠–®–Р
    def save_laser_import_cache(self):
        """–Р–≤—В–Њ–Љ–∞—В–Є—З–µ—Б–Ї–Њ–µ —Б–Њ—Е—А–∞–љ–µ–љ–Є–µ —В–∞–±–ї–Є—Ж—Л –Є–Љ–њ–Њ—А—В–∞ –≤ –Ї—Н—И-—Д–∞–є–ї"""
        if not hasattr(self, 'laser_table_data') or not self.laser_table_data:
            return

        try:
            script_dir = os.path.dirname(os.path.abspath(__file__))
            cache_file = os.path.join(script_dir, "laser_import_cache.xlsx")

            print(f"рЯТЊ –°–Њ—Е—А–∞–љ–µ–љ–Є–µ {len(self.laser_table_data)} –Ј–∞–њ–Є—Б–µ–є –≤ –Ї—Н—И...")

            df = pd.DataFrame(self.laser_table_data)
            df.to_excel(cache_file, index=False, engine='openpyxl')

            print(f"вЬЕ –Ъ—Н—И –Є–Љ–њ–Њ—А—В–∞ —Б–Њ—Е—А–∞–љ—С–љ: {len(self.laser_table_data)} –Ј–∞–њ–Є—Б–µ–є")
        except Exception as e:
            print(f"вЪ†пЄП –Ю—И–Є–±–Ї–∞ —Б–Њ—Е—А–∞–љ–µ–љ–Є—П –Ї—Н—И–∞: {e}")

    # рЯЖХ –Э–Ю–Т–Ђ–Щ –Ь–Х–Ґ–Ю–Ф - –Ч–Р–У–†–£–Ч–Ъ–Р –Ъ–≠–®–Р
    def load_laser_import_cache(self):
        """–Р–≤—В–Њ–Љ–∞—В–Є—З–µ—Б–Ї–∞—П –Ј–∞–≥—А—Г–Ј–Ї–∞ —В–∞–±–ї–Є—Ж—Л –Є–Љ–њ–Њ—А—В–∞ –Є–Ј –Ї—Н—И-—Д–∞–є–ї–∞"""
        try:
            # –Ш—Б–њ–Њ–ї—М–Ј—Г–µ–Љ —В–µ–Ї—Г—Й—Г—О –і–Є—А–µ–Ї—В–Њ—А–Є—О —Б–Ї—А–Є–њ—В–∞
            script_dir = os.path.dirname(os.path.abspath(__file__))
            cache_file = os.path.join(script_dir, "laser_import_cache.xlsx")

            if not os.path.exists(cache_file):
                print(f"вДєпЄП –Ъ—Н—И –Є–Љ–њ–Њ—А—В–∞ –љ–µ –љ–∞–є–і–µ–љ: {cache_file}")
                return

            # –Ч–∞–≥—А—Г–ґ–∞–µ–Љ –Є–Ј Excel
            df = pd.read_excel(cache_file, engine='openpyxl')

            if df.empty:
                print("вДєпЄП –Ъ—Н—И –Є–Љ–њ–Њ—А—В–∞ –њ—Г—Б—В")
                return

            # –Я—А–Њ–≤–µ—А—П–µ–Љ –љ–∞–ї–Є—З–Є–µ –љ–µ–Њ–±—Е–Њ–і–Є–Љ—Л—Е –Ї–Њ–ї–Њ–љ–Њ–Ї
            required = ["–Ф–∞—В–∞ (–Ь–°–Ъ)", "–Т—А–µ–Љ—П (–Ь–°–Ъ)", "username", "order", "metal", "metal_quantity", "part",
                        "part_quantity"]

            if all(col in df.columns for col in required):
                # рЯЖХ –Т–Р–Ц–Э–Ю: –Я—А–µ–Њ–±—А–∞–Ј—Г–µ–Љ NaN –≤ –њ—Г—Б—В—Л–µ —Б—В—А–Њ–Ї–Є –њ–µ—А–µ–і –Ї–Њ–љ–≤–µ—А—В–∞—Ж–Є–µ–є
                df = df.fillna("")

                def load_laser_import_cache(self):
                    """–Р–≤—В–Њ–Љ–∞—В–Є—З–µ—Б–Ї–∞—П –Ј–∞–≥—А—Г–Ј–Ї–∞ —В–∞–±–ї–Є—Ж—Л –Є–Љ–њ–Њ—А—В–∞ –Є–Ј –Ї—Н—И-—Д–∞–є–ї–∞"""
                    try:
                        script_dir = os.path.dirname(os.path.abspath(__file__))
                        cache_file = os.path.join(script_dir, "laser_import_cache.xlsx")

                        if not os.path.exists(cache_file):
                            print(f"вДєпЄП –Ъ—Н—И –Є–Љ–њ–Њ—А—В–∞ –љ–µ –љ–∞–є–і–µ–љ: {cache_file}")
                            return

                        df = pd.read_excel(cache_file, engine='openpyxl')

                        if df.empty:
                            print("вДєпЄП –Ъ—Н—И –Є–Љ–њ–Њ—А—В–∞ –њ—Г—Б—В")
                            return

                        required = ["–Ф–∞—В–∞ (–Ь–°–Ъ)", "–Т—А–µ–Љ—П (–Ь–°–Ъ)", "username", "order", "metal", "metal_quantity", "part",
                                    "part_quantity"]

                        if all(col in df.columns for col in required):
                            df = df.fillna("")

                            # рЯЖХ –°–Ю–†–Ґ–Ш–†–Ю–Т–Ъ–Р: –Э–Ю–Т–Ђ–Х –Ч–Р–Я–Ш–°–Ш –Т–Т–Х–†–•–£
                            try:
                                print("рЯФД –°–Њ—А—В–Є—А–Њ–≤–Ї–∞ –Ї—Н—И–∞...")
                                df['_datetime_sort'] = pd.to_datetime(
                                    df['–Ф–∞—В–∞ (–Ь–°–Ъ)'].astype(str) + ' ' + df['–Т—А–µ–Љ—П (–Ь–°–Ъ)'].astype(str),
                                    errors='coerce'
                                )
                                df = df.sort_values('_datetime_sort', ascending=False, na_position='last')
                                df = df.drop('_datetime_sort', axis=1)

                                if not df.empty:
                                    first = f"{df.iloc[0]['–Ф–∞—В–∞ (–Ь–°–Ъ)']} {df.iloc[0]['–Т—А–µ–Љ—П (–Ь–°–Ъ)']}"
                                    last = f"{df.iloc[-1]['–Ф–∞—В–∞ (–Ь–°–Ъ)']} {df.iloc[-1]['–Т—А–µ–Љ—П (–Ь–°–Ъ)']}"
                                    print(f"вЬЕ –Ю—В—Б–Њ—А—В–Є—А–Њ–≤–∞–љ–Њ: –њ–µ—А–≤–∞—П={first}, –њ–Њ—Б–ї–µ–і–љ—П—П={last}")
                            except Exception as e:
                                print(f"вЪ†пЄП –Ю—И–Є–±–Ї–∞ —Б–Њ—А—В–Є—А–Њ–≤–Ї–Є: {e}")

                            # –Я—А–µ–Њ–±—А–∞–Ј—Г–µ–Љ –≤ —Б–њ–Є—Б–Њ–Ї —Б–ї–Њ–≤–∞—А–µ–є
                            self.laser_table_data = df.to_dict('records')

                            # –Ф–Њ–њ–Њ–ї–љ–Є—В–µ–ї—М–љ–∞—П –Њ—З–Є—Б—В–Ї–∞
                            for row in self.laser_table_data:
                                if "–°–њ–Є—Б–∞–љ–Њ" in row:
                                    if pd.isna(row["–°–њ–Є—Б–∞–љ–Њ"]) or row["–°–њ–Є—Б–∞–љ–Њ"] is None:
                                        row["–°–њ–Є—Б–∞–љ–Њ"] = ""
                                    else:
                                        row["–°–њ–Є—Б–∞–љ–Њ"] = str(row["–°–њ–Є—Б–∞–љ–Њ"]).strip()

                                if "–Ф–∞—В–∞ —Б–њ–Є—Б–∞–љ–Є—П" in row:
                                    if pd.isna(row["–Ф–∞—В–∞ —Б–њ–Є—Б–∞–љ–Є—П"]) or row["–Ф–∞—В–∞ —Б–њ–Є—Б–∞–љ–Є—П"] is None:
                                        row["–Ф–∞—В–∞ —Б–њ–Є—Б–∞–љ–Є—П"] = ""
                                    else:
                                        row["–Ф–∞—В–∞ —Б–њ–Є—Б–∞–љ–Є—П"] = str(row["–Ф–∞—В–∞ —Б–њ–Є—Б–∞–љ–Є—П"]).strip()

                            print(f"вЬЕ –Ч–∞–≥—А—Г–ґ–µ–љ –Ї—Н—И –Є–Љ–њ–Њ—А—В–∞: {len(self.laser_table_data)} –Ј–∞–њ–Є—Б–µ–є –Є–Ј {cache_file}")

                            # –Ю–±–љ–Њ–≤–ї—П–µ–Љ —В–∞–±–ї–Є—Ж—Г
                            if hasattr(self, 'laser_import_tree'):
                                self.refresh_laser_import_table()

                                if hasattr(self, 'laser_status_label'):
                                    items_count = len(self.laser_import_tree.get_children())
                                    auto_count = sum(1 for r in self.laser_table_data if
                                                     r.get("–°–њ–Є—Б–∞–љ–Њ", "").strip() in ["вЬУ", "–Ф–∞", "Yes"])
                                    manual_count = sum(
                                        1 for r in self.laser_table_data if r.get("–°–њ–Є—Б–∞–љ–Њ", "").strip() == "–Т—А—Г—З–љ—Г—О")
                                    pending_count = sum(
                                        1 for r in self.laser_table_data if not r.get("–°–њ–Є—Б–∞–љ–Њ", "").strip())

                                    status_text = (
                                        f"рЯУВ –Ч–∞–≥—А—Г–ґ–µ–љ–Њ –Є–Ј –Ї—Н—И–∞: {items_count} | "
                                        f"вЬЕ –°–њ–Є—Б–∞–љ–Њ: {auto_count} | "
                                        f"рЯФµ –Т—А—Г—З–љ—Г—О: {manual_count} | "
                                        f"рЯЯ° –Ю–ґ–Є–і–∞–µ—В: {pending_count}"
                                    )
                                    self.laser_status_label.config(text=status_text, bg='#d1ecf1', fg='#0c5460')

                    except Exception as e:
                        pass
                # –Я—А–µ–Њ–±—А–∞–Ј—Г–µ–Љ –≤ —Б–њ–Є—Б–Њ–Ї —Б–ї–Њ–≤–∞—А–µ–є
                self.laser_table_data = df.to_dict('records')

                # рЯЖХ –Ф–Ю–Я–Ю–Ы–Э–Ш–Ґ–Х–Ы–ђ–Э–Р–ѓ –Ю–І–Ш–°–Ґ–Ъ–Р: —Г–±–µ–і–Є–Љ—Б—П —З—В–Њ –≤—Б–µ –Ј–љ–∞—З–µ–љ–Є—П - —Б—В—А–Њ–Ї–Є –Є–ї–Є —З–Є—Б–ї–∞
                for row in self.laser_table_data:
                    # –Я—А–µ–Њ–±—А–∞–Ј—Г–µ–Љ "–°–њ–Є—Б–∞–љ–Њ" –≤ —Б—В—А–Њ–Ї—Г
                    if "–°–њ–Є—Б–∞–љ–Њ" in row:
                        if pd.isna(row["–°–њ–Є—Б–∞–љ–Њ"]) or row["–°–њ–Є—Б–∞–љ–Њ"] is None:
                            row["–°–њ–Є—Б–∞–љ–Њ"] = ""
                        else:
                            row["–°–њ–Є—Б–∞–љ–Њ"] = str(row["–°–њ–Є—Б–∞–љ–Њ"]).strip()

                    # –Я—А–µ–Њ–±—А–∞–Ј—Г–µ–Љ "–Ф–∞—В–∞ —Б–њ–Є—Б–∞–љ–Є—П" –≤ —Б—В—А–Њ–Ї—Г
                    if "–Ф–∞—В–∞ —Б–њ–Є—Б–∞–љ–Є—П" in row:
                        if pd.isna(row["–Ф–∞—В–∞ —Б–њ–Є—Б–∞–љ–Є—П"]) or row["–Ф–∞—В–∞ —Б–њ–Є—Б–∞–љ–Є—П"] is None:
                            row["–Ф–∞—В–∞ —Б–њ–Є—Б–∞–љ–Є—П"] = ""
                        else:
                            row["–Ф–∞—В–∞ —Б–њ–Є—Б–∞–љ–Є—П"] = str(row["–Ф–∞—В–∞ —Б–њ–Є—Б–∞–љ–Є—П"]).strip()

                print(f"вЬЕ –Ч–∞–≥—А—Г–ґ–µ–љ –Ї—Н—И –Є–Љ–њ–Њ—А—В–∞: {len(self.laser_table_data)} –Ј–∞–њ–Є—Б–µ–є –Є–Ј {cache_file}")

                # рЯЖХ –Ф–Ш–Р–У–Э–Ю–°–Ґ–Ш–Ъ–Р: –≤—Л–≤–µ–і–µ–Љ –њ–µ—А–≤—Л–µ —Б—В—А–Њ–Ї–Є –і–ї—П –њ—А–Њ–≤–µ—А–Ї–Є
                if self.laser_table_data:
                    print("\nрЯФН –Я—А–Њ–≤–µ—А–Ї–∞ –Ј–∞–≥—А—Г–ґ–µ–љ–љ—Л—Е –і–∞–љ–љ—Л—Е:")
                    for i, row in enumerate(self.laser_table_data[:3]):
                        status = row.get("–°–њ–Є—Б–∞–љ–Њ", "")
                        print(f"   –°—В—А–Њ–Ї–∞ {i + 1}: –°–њ–Є—Б–∞–љ–Њ = '{status}' (—В–Є–њ: {type(status).__name__})")

                # –Ю–±–љ–Њ–≤–ї—П–µ–Љ —В–∞–±–ї–Є—Ж—Г
                if hasattr(self, 'laser_import_tree'):
                    self.refresh_laser_import_table()

                    # –Ю–±–љ–Њ–≤–ї—П–µ–Љ —Б—В–∞—В—Г—Б
                    if hasattr(self, 'laser_status_label'):
                        items_count = len(self.laser_import_tree.get_children())

                        # –°—З–Є—В–∞–µ–Љ —Б—В–∞—В–Є—Б—В–Є–Ї—Г
                        auto_count = sum(
                            1 for r in self.laser_table_data if r.get("–°–њ–Є—Б–∞–љ–Њ", "").strip() in ["вЬУ", "–Ф–∞", "Yes"])
                        manual_count = sum(
                            1 for r in self.laser_table_data if r.get("–°–њ–Є—Б–∞–љ–Њ", "").strip() == "–Т—А—Г—З–љ—Г—О")
                        pending_count = sum(1 for r in self.laser_table_data if not r.get("–°–њ–Є—Б–∞–љ–Њ", "").strip())

                        status_text = (
                            f"рЯУВ –Ч–∞–≥—А—Г–ґ–µ–љ–Њ –Є–Ј –Ї—Н—И–∞: {items_count} | "
                            f"вЬЕ –°–њ–Є—Б–∞–љ–Њ: {auto_count} | "
                            f"рЯФµ –Т—А—Г—З–љ—Г—О: {manual_count} | "
                            f"рЯЯ° –Ю–ґ–Є–і–∞–µ—В: {pending_count}"
                        )

                        self.laser_status_label.config(
                            text=status_text,
                            bg='#d1ecf1',
                            fg='#0c5460'
                        )
            else:
                print("вЪ†пЄП –Ъ—Н—И –Є–Љ–њ–Њ—А—В–∞ –Є–Љ–µ–µ—В –љ–µ–њ—А–∞–≤–Є–ї—М–љ—Г—О —Б—В—А—Г–Ї—В—Г—А—Г")

        except Exception as e:
            print(f"вЪ†пЄП –Ю—И–Є–±–Ї–∞ –Ј–∞–≥—А—Г–Ј–Ї–Є –Ї—Н—И–∞: {e}")
            import traceback
            traceback.print_exc()

    def setup_balance_tab(self):
        """–Т–Ї–ї–∞–і–Ї–∞ –±–∞–ї–∞–љ—Б–∞ –Љ–∞—В–µ—А–Є–∞–ї–Њ–≤"""

        # –Ч–∞–≥–Њ–ї–Њ–≤–Њ–Ї
        header = tk.Label(self.balance_frame, text="–С–∞–ї–∞–љ—Б –Љ–∞—В–µ—А–Є–∞–ї–Њ–≤ –њ–Њ –Љ–∞—А–Ї–∞–Љ –Є —В–Њ–ї—Й–Є–љ–∞–Љ",
                          font=("Arial", 16, "bold"), bg='white', fg='#2c3e50')
        header.pack(pady=10)

        # рЯЖХ –§—А–µ–є–Љ —В–∞–±–ї–Є—Ж—Л –Э–Р –Т–°–Х–Щ –®–Ш–†–Ш–Э–Х (—Г–±—А–∞–љ–Њ —Ж–µ–љ—В—А–Є—А–Њ–≤–∞–љ–Є–µ)
        tree_frame = tk.Frame(self.balance_frame, bg='white')
        tree_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)

        scroll_y = tk.Scrollbar(tree_frame, orient=tk.VERTICAL)
        scroll_x = tk.Scrollbar(tree_frame, orient=tk.HORIZONTAL)

        # –Ґ–Р–С–Ы–Ш–¶–Р
        self.balance_tree = ttk.Treeview(tree_frame,
                                         columns=("–Ь–∞—А–Ї–∞", "–Ґ–Њ–ї—Й–Є–љ–∞", "–†–∞–Ј–Љ–µ—А", "–Т—Б–µ–≥–Њ", "–Ч–∞—А–µ–Ј–µ—А–≤–Є—А–Њ–≤–∞–љ–Њ", "–Ф–Њ—Б—В—Г–њ–љ–Њ"),
                                         show="headings", yscrollcommand=scroll_y.set, xscrollcommand=scroll_x.set)

        scroll_y.config(command=self.balance_tree.yview)
        scroll_x.config(command=self.balance_tree.xview)
        scroll_y.pack(side=tk.RIGHT, fill=tk.Y)
        scroll_x.pack(side=tk.BOTTOM, fill=tk.X)

        # –Э–Р–°–Ґ–†–Ю–Щ–Ъ–Р –Ъ–Ю–Ы–Ю–Э–Ю–Ъ –С–Х–Ч –†–Р–°–Ґ–ѓ–У–Ш–Т–Р–Э–Ш–ѓ
        for col in self.balance_tree["columns"]:
            self.balance_tree.heading(col, text=col)
            self.balance_tree.column(col, anchor=tk.CENTER, width=100, minwidth=80, stretch=False)

        # –Ґ–Р–С–Ы–Ш–¶–Р –Ч–Р–Я–Ю–Ы–Э–ѓ–Х–Ґ –Т–°–Ѓ –Т–Ђ–°–Ю–Ґ–£ –Ш –®–Ш–†–Ш–Э–£ –§–†–Х–Щ–Ь–Р
        self.balance_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # –Ш–Э–Ш–¶–Ш–Р–Ы–Ш–Ч–Р–¶–Ш–ѓ –§–Ш–Ы–ђ–Ґ–†–Р –Т –°–Ґ–Ш–Ы–Х EXCEL
        self.balance_excel_filter = ExcelStyleFilter(
            tree=self.balance_tree,
            refresh_callback=self.refresh_balance
        )

        # –¶–Т–Х–Ґ–Ю–Т–Р–ѓ –Ш–Э–Ф–Ш–Ъ–Р–¶–Ш–ѓ
        self.balance_tree.tag_configure('negative', background='#f8d7da', foreground='#721c24')
        self.balance_tree.tag_configure('available', background='#d4edda', foreground='#155724')
        self.balance_tree.tag_configure('fully_reserved', background='#fff3cd', foreground='#856404')
        self.balance_tree.tag_configure('empty', background='#d1ecf1', foreground='#0c5460')

        # –Ш–Э–Ф–Ш–Ъ–Р–Ґ–Ю–† –Р–Ъ–Ґ–Ш–Т–Э–Ђ–• –§–Ш–Ы–ђ–Ґ–†–Ю–Т
        self.balance_filter_status = tk.Label(
            self.balance_frame,
            text="",
            font=("Arial", 9),
            bg='#d1ecf1',
            fg='#0c5460'
        )
        self.balance_filter_status.pack(pady=5)

        # –Ы–Х–У–Х–Э–Ф–Р –¶–Т–Х–Ґ–Ю–Т
        legend_frame = tk.Frame(self.balance_frame, bg='white')
        legend_frame.pack(pady=5)

        tk.Label(legend_frame, text="–Ы–µ–≥–µ–љ–і–∞:", font=("Arial", 10, "bold"), bg='white').pack(side=tk.LEFT, padx=5)

        # –Ю—В—А–Є—Ж–∞—В–µ–ї—М–љ–Њ–µ –Ј–љ–∞—З–µ–љ–Є–µ (–њ—А–Њ–±–ї–µ–Љ–∞)
        negative_label = tk.Label(legend_frame, text="  –Ю—В—А–Є—Ж–∞—В–µ–ї—М–љ–Њ–µ  ", bg='#f8d7da', fg='#721c24',
                                  font=("Arial", 9, "bold"), relief=tk.RAISED, borderwidth=1)
        negative_label.pack(side=tk.LEFT, padx=3)

        # –Я–Њ–ї–љ–Њ—Б—В—М—О –Ј–∞—А–µ–Ј–µ—А–≤–Є—А–Њ–≤–∞–љ–Њ
        reserved_label = tk.Label(legend_frame, text="  –Я–Њ–ї–љ–Њ—Б—В—М—О –Ј–∞—А–µ–Ј–µ—А–≤–Є—А–Њ–≤–∞–љ–Њ  ", bg='#fff3cd', fg='#856404',
                                  font=("Arial", 9, "bold"), relief=tk.RAISED, borderwidth=1)
        reserved_label.pack(side=tk.LEFT, padx=3)

        # –Х—Б—В—М –і–Њ—Б—В—Г–њ–љ–Њ
        available_label = tk.Label(legend_frame, text="  –Х—Б—В—М –і–Њ—Б—В—Г–њ–љ–Њ  ", bg='#d4edda', fg='#155724',
                                   font=("Arial", 9, "bold"), relief=tk.RAISED, borderwidth=1)
        available_label.pack(side=tk.LEFT, padx=3)

        # –Э–µ—В –≤ –љ–∞–ї–Є—З–Є–Є
        empty_label = tk.Label(legend_frame, text="  –Э–µ—В –≤ –љ–∞–ї–Є—З–Є–Є  ", bg='#d1ecf1', fg='#0c5460',
                               font=("Arial", 9, "bold"), relief=tk.RAISED, borderwidth=1)
        empty_label.pack(side=tk.LEFT, padx=3)

        # –Ъ–љ–Њ–њ–Ї–Є —Г–њ—А–∞–≤–ї–µ–љ–Є—П
        buttons_frame = tk.Frame(self.balance_frame, bg='white')
        buttons_frame.pack(fill=tk.X, padx=10, pady=5)

        tk.Button(buttons_frame, text="рЯФД –Ю–±–љ–Њ–≤–Є—В—М", bg='#3498db', fg='white',
                  font=("Arial", 10), command=self.refresh_balance).pack(side=tk.LEFT, padx=5)

        tk.Button(buttons_frame, text="вЬЦ –°–±—А–Њ—Б–Є—В—М –≤—Б–µ —Д–Є–ї—М—В—А—Л", bg='#e67e22', fg='white',
                  font=("Arial", 10), command=self.clear_balance_filters).pack(side=tk.LEFT, padx=5)

        # –Я–µ—А–≤–Є—З–љ–∞—П –Ј–∞–≥—А—Г–Ј–Ї–∞ –і–∞–љ–љ—Л—Е
        self.refresh_balance()

    def clear_balance_filters(self):
        """–°–±—А–Њ—Б–Є—В—М –≤—Б–µ —Д–Є–ї—М—В—А—Л –±–∞–ї–∞–љ—Б–∞"""
        if hasattr(self, 'balance_excel_filter'):
            self.balance_excel_filter.clear_all_filters()

    def refresh_balance(self):
        """–Ю–±–љ–Њ–≤–ї–µ–љ–Є–µ —В–∞–±–ї–Є—Ж—Л –±–∞–ї–∞–љ—Б–∞ –Љ–∞—В–µ—А–Є–∞–ї–Њ–≤"""

        # –°–Ю–•–†–Р–Э–ѓ–Х–Ь –Р–Ъ–Ґ–Ш–Т–Э–Ђ–Х –§–Ш–Ы–ђ–Ґ–†–Ђ –Я–Х–†–Х–Ф –Ю–І–Ш–°–Ґ–Ъ–Ю–Щ
        active_filters_backup = {}
        if hasattr(self, 'balance_excel_filter') and self.balance_excel_filter.active_filters:
            active_filters_backup = self.balance_excel_filter.active_filters.copy()
            print(f"рЯФН –°–Њ—Е—А–∞–љ–µ–љ—Л —Д–Є–ї—М—В—А—Л: {list(active_filters_backup.keys())}")

        # –Я–Ю–Ы–Э–Ю–°–Ґ–ђ–Ѓ –Ю–І–Ш–©–Р–Х–Ь –Ф–Х–†–Х–Т–Ю
        for i in self.balance_tree.get_children():
            self.balance_tree.delete(i)

        # –Ю–І–Ш–©–Р–Х–Ь –Ъ–≠–® –≠–Ы–Х–Ь–Х–Э–Ґ–Ю–Т
        if hasattr(self, 'balance_excel_filter'):
            self.balance_excel_filter._all_item_cache = set()

        df = load_data("Materials")

        if not df.empty:
            # –У—А—Г–њ–њ–Є—А—Г–µ–Љ –њ–Њ –Љ–∞—А–Ї–µ, —В–Њ–ї—Й–Є–љ–µ –Є —А–∞–Ј–Љ–µ—А—Г
            balance_data = {}

            for index, row in df.iterrows():
                marka = row["–Ь–∞—А–Ї–∞"]
                thickness = row["–Ґ–Њ–ї—Й–Є–љ–∞"]
                length = row["–Ф–ї–Є–љ–∞"]
                width = row["–®–Є—А–Є–љ–∞"]
                size_key = f"{int(length)}x{int(width)}"

                key = (marka, thickness, size_key)

                if key not in balance_data:
                    balance_data[key] = {
                        "total": 0,
                        "reserved": 0,
                        "available": 0
                    }

                balance_data[key]["total"] += int(row["–Ъ–Њ–ї–Є—З–µ—Б—В–≤–Њ —И—В—Г–Ї"])
                balance_data[key]["reserved"] += int(row["–Ч–∞—А–µ–Ј–µ—А–≤–Є—А–Њ–≤–∞–љ–Њ"])
                balance_data[key]["available"] += int(row["–Ф–Њ—Б—В—Г–њ–љ–Њ"])

            # –Ч–∞–њ–Њ–ї–љ—П–µ–Љ —В–∞–±–ї–Є—Ж—Г —Б —Ж–≤–µ—В–Њ–≤–Њ–є –Є–љ–і–Є–Ї–∞—Ж–Є–µ–є
            for (marka, thickness, size), data in sorted(balance_data.items()):
                total = data["total"]
                reserved = data["reserved"]
                available = data["available"]

                # –Ю–С–Ђ–І–Э–Ђ–Х VALUES –С–Х–Ч –Я–£–°–Ґ–Ю–Щ –Ъ–Ю–Ы–Ю–Э–Ъ–Ш
                values = (marka, thickness, size, total, reserved, available)

                # –¶–Т–Х–Ґ–Ю–Т–Р–ѓ –Ш–Э–Ф–Ш–Ъ–Р–¶–Ш–ѓ
                if available < 0:
                    tag = 'negative'
                elif available > 0:
                    tag = 'available'
                elif available == 0 and total > 0:
                    tag = 'fully_reserved'
                else:
                    tag = 'empty'

                item_id = self.balance_tree.insert("", "end", values=values, tags=(tag,))

                # –°–Ю–•–†–Р–Э–ѓ–Х–Ь item_id –Т –Ъ–≠–®
                if hasattr(self, 'balance_excel_filter'):
                    if not hasattr(self.balance_excel_filter, '_all_item_cache'):
                        self.balance_excel_filter._all_item_cache = set()
                    self.balance_excel_filter._all_item_cache.add(item_id)

        # –Р–Т–Ґ–Ю–Я–Ю–Ф–С–Ю–† –®–Ш–†–Ш–Э–Ђ –Ъ–Ю–Ы–Ю–Э–Ю–Ъ –° –Ю–У–†–Р–Э–Ш–І–Х–Э–Ш–ѓ–Ь–Ш
        self.auto_resize_columns(self.balance_tree, min_width=100, max_width=300)

        # –Я–Х–†–Х–Я–†–Ш–Ь–Х–Э–ѓ–Х–Ь –§–Ш–Ы–ђ–Ґ–†–Ђ –Я–Ю–°–Ы–Х –Ч–Р–У–†–£–Ч–Ъ–Ш –Ф–Р–Э–Э–Ђ–•
        if active_filters_backup and hasattr(self, 'balance_excel_filter'):
            print(f"рЯФД –Я–µ—А–µ–њ—А–Є–Љ–µ–љ—П—О —Д–Є–ї—М—В—А—Л: {list(active_filters_backup.keys())}")
            self.balance_excel_filter.active_filters = active_filters_backup
            self.balance_excel_filter.reapply_all_filters()

    def export_balance(self):
        """–≠–Ї—Б–њ–Њ—А—В –±–∞–ї–∞–љ—Б–∞ –≤ Excel"""
        file_path = filedialog.asksaveasfilename(
            title="–°–Њ—Е—А–∞–љ–Є—В—М –±–∞–ї–∞–љ—Б",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialfile=f"balance_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )

        if not file_path:
            return

        try:
            # –°–Њ–±–Є—А–∞–µ–Љ –і–∞–љ–љ—Л–µ –Є–Ј —В–∞–±–ї–Є—Ж—Л
            data = []
            for item in self.balance_tree.get_children():
                values = self.balance_tree.item(item)['values']
                data.append(values)

            df = pd.DataFrame(data, columns=self.balance_tree['columns'])
            df.to_excel(file_path, index=False, engine='openpyxl')

            messagebox.showinfo("–£—Б–њ–µ—Е", f"–С–∞–ї–∞–љ—Б —Б–Њ—Е—А–∞–љ–µ–љ:\n{file_path}")
        except Exception as e:
            messagebox.showerror("–Ю—И–Є–±–Ї–∞", f"–Э–µ —Г–і–∞–ї–Њ—Б—М —Б–Њ—Е—А–∞–љ–Є—В—М —Д–∞–є–ї:\n{e}")


if __name__ == "__main__":
    try:
        initialize_database()
        root = tk.Tk()
        app = ProductionApp(root)
        root.mainloop()
    except Exception as e:
        print(f"–Ъ—А–Є—В–Є—З–µ—Б–Ї–∞—П –Њ—И–Є–±–Ї–∞: {e}")
        import traceback

        traceback.print_exc()
        messagebox.showerror("–Ъ—А–Є—В–Є—З–µ—Б–Ї–∞—П –Њ—И–Є–±–Ї–∞", str(e))
# -*- coding: utf-8 -*-
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import pandas as pd
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os

DATABASE_FILE = "production_database.xlsx"

def initialize_database():
    if not os.path.exists(DATABASE_FILE):
        wb = Workbook()
        materials_sheet = wb.active
        materials_sheet.title = "Materials"
        materials_sheet.append([
            "ID", "Марка", "Толщина", "Длина", "Ширина",
            "Количество штук", "Общая площадь", "Зарезервировано", "Доступно", "Дата добавления"
        ])
        orders_sheet = wb.create_sheet("Orders")
        orders_sheet.append(["ID заказа", "Название заказа", "Заказчик", "Дата создания", "Статус", "Примечания"])
        order_details_sheet = wb.create_sheet("OrderDetails")
        order_details_sheet.append(["ID", "ID заказа", "Название детали", "Количество"])
        reservations_sheet = wb.create_sheet("Reservations")
        reservations_sheet.append([
            "ID резерва", "ID заказа", "ID материала", "Марка", "Толщина", "Длина", "Ширина",
            "Зарезервировано штук", "Списано", "Остаток к списанию", "Дата резерва"
        ])
        writeoffs_sheet = wb.create_sheet("WriteOffs")
        writeoffs_sheet.append([
            "ID списания", "ID резерва", "ID заказа", "ID материала", "Марка", "Толщина", "Длина", "Ширина",
            "Количество", "Дата списания", "Комментарий"
        ])
        wb.save(DATABASE_FILE)
        print(f"База данных '{DATABASE_FILE}' создана!")

def load_data(sheet_name):
    try:
        df = pd.read_excel(DATABASE_FILE, sheet_name=sheet_name, engine='openpyxl')
        if df.empty:
            return df
        df = df.fillna("")
        return df
    except Exception as e:
        print(f"Ошибка загрузки данных из {sheet_name}: {e}")
        return pd.DataFrame()

def save_data(sheet_name, dataframe):
    try:
        book = load_workbook(DATABASE_FILE)
        if sheet_name in book.sheetnames:
            del book[sheet_name]
        sheet = book.create_sheet(sheet_name)
        for col_num, column_title in enumerate(dataframe.columns, 1):
            sheet.cell(row=1, column=col_num).value = str(column_title)
        for row_num, row_data in enumerate(dataframe.values, 2):
            for col_num, cell_value in enumerate(row_data, 1):
                sheet.cell(row=row_num, column=col_num).value = cell_value
        book.save(DATABASE_FILE)
        book.close()
    except Exception as e:
        print(f"Ошибка сохранения данных в {sheet_name}: {e}")
        messagebox.showerror("Ошибка", f"Невозможно сохранить данные: {e}")

class ProductionApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Система учета производства")
        self.root.geometry("1400x800")
        self.root.configure(bg='#f0f0f0')
        self.notebook = ttk.Notebook(root)
        self.notebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        self.materials_frame = tk.Frame(self.notebook, bg='white')
        self.notebook.add(self.materials_frame, text='Материалы на складе')
        self.setup_materials_tab()
        self.orders_frame = tk.Frame(self.notebook, bg='white')
        self.notebook.add(self.orders_frame, text='Заказы')
        self.setup_orders_tab()
        self.reservations_frame = tk.Frame(self.notebook, bg='white')
        self.notebook.add(self.reservations_frame, text='Резервирование')
        self.setup_reservations_tab()
        self.writeoffs_frame = tk.Frame(self.notebook, bg='white')
        self.notebook.add(self.writeoffs_frame, text='Списание материалов')
        self.setup_writeoffs_tab()
        self.balance_frame = tk.Frame(self.notebook, bg='white')
        self.notebook.add(self.balance_frame, text='Баланс материалов')
        self.setup_balance_tab()

    def create_filter_panel(self, parent_frame, tree_widget, columns_to_filter, refresh_callback):
        """Создание панели фильтрации для любой таблицы"""
        filter_frame = tk.LabelFrame(parent_frame, text="🔍 Фильтры", bg='#e8f4f8', font=("Arial", 10, "bold"))
        filter_frame.pack(fill=tk.X, padx=10, pady=5)

        # Словарь для хранения Entry виджетов фильтров
        filter_entries = {}

        # Создаём поля фильтрации для каждой колонки
        row = 0
        col = 0
        max_cols = 4  # Количество фильтров в одной строке

        for column_name in columns_to_filter:
            filter_container = tk.Frame(filter_frame, bg='#e8f4f8')
            filter_container.grid(row=row, column=col, padx=5, pady=3, sticky='w')

            tk.Label(filter_container, text=f"{column_name}:", bg='#e8f4f8', font=("Arial", 9)).pack(side=tk.LEFT)

            entry = tk.Entry(filter_container, width=15, font=("Arial", 9))
            entry.pack(side=tk.LEFT, padx=5)

            filter_entries[column_name] = entry

            # Привязываем событие изменения текста к функции фильтрации
            entry.bind('<KeyRelease>', lambda e, tree=tree_widget, filters=filter_entries, cb=refresh_callback:
            self.apply_filters(tree, filters, cb))

            col += 1
            if col >= max_cols:
                col = 0
                row += 1

        # Кнопки управления фильтрами
        buttons_container = tk.Frame(filter_frame, bg='#e8f4f8')
        buttons_container.grid(row=row + 1, column=0, columnspan=max_cols, pady=5)

        tk.Button(buttons_container, text="🗑️ Очистить фильтры", bg='#95a5a6', fg='white',
                  font=("Arial", 9),
                  command=lambda: self.clear_filters(filter_entries, tree_widget, refresh_callback)).pack(side=tk.LEFT,
                                                                                                          padx=5)

        tk.Button(buttons_container, text="🔄 Обновить", bg='#3498db', fg='white',
                  font=("Arial", 9), command=refresh_callback).pack(side=tk.LEFT, padx=5)

        return filter_entries

    def apply_filters(self, tree, filter_entries, refresh_callback):
        """Применить фильтры к таблице"""
        # Собираем активные фильтры
        active_filters = {}
        for col_name, entry in filter_entries.items():
            filter_text = entry.get().strip().lower()
            if filter_text:
                active_filters[col_name] = filter_text

        # Если нет фильтров - показываем всё
        if not active_filters:
            refresh_callback()
            return

        # Сохраняем текущие данные
        all_items = []
        for item in tree.get_children():
            all_items.append(tree.item(item)['values'])

        # Очищаем таблицу
        for item in tree.get_children():
            tree.delete(item)

        # Фильтруем и добавляем обратно
        columns = tree['columns']
        for item_values in all_items:
            match = True
            for col_name, filter_text in active_filters.items():
                try:
                    col_index = columns.index(col_name)
                    cell_value = str(item_values[col_index]).lower()
                    if filter_text not in cell_value:
                        match = False
                        break
                except (ValueError, IndexError):
                    continue

            if match:
                tree.insert("", "end", values=item_values)

    def clear_filters(self, filter_entries, tree, refresh_callback):
        """Очистить все фильтры"""
        for entry in filter_entries.values():
            entry.delete(0, tk.END)
        refresh_callback()

    def setup_materials_tab(self):
        header = tk.Label(self.materials_frame, text="Учет листового проката на складе",
                         font=("Arial", 16, "bold"), bg='white', fg='#2c3e50')
        header.pack(pady=10)
        tree_frame = tk.Frame(self.materials_frame, bg='white')
        tree_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        scroll_y = tk.Scrollbar(tree_frame, orient=tk.VERTICAL)
        scroll_x = tk.Scrollbar(tree_frame, orient=tk.HORIZONTAL)
        self.materials_tree = ttk.Treeview(tree_frame,
            columns=("ID", "Марка", "Толщина", "Длина", "Ширина", "Кол-во шт", "Площадь", "Резерв", "Доступно", "Дата"),
            show="headings", yscrollcommand=scroll_y.set, xscrollcommand=scroll_x.set)
        scroll_y.config(command=self.materials_tree.yview)
        scroll_x.config(command=self.materials_tree.xview)
        scroll_y.pack(side=tk.RIGHT, fill=tk.Y)
        scroll_x.pack(side=tk.BOTTOM, fill=tk.X)
        columns_config = {"ID": 50, "Марка": 100, "Толщина": 80, "Длина": 80, "Ширина": 80,
            "Кол-во шт": 80, "Площадь": 100, "Резерв": 80, "Доступно": 80, "Дата": 100}
        for col, width in columns_config.items():
            self.materials_tree.heading(col, text=col)
            self.materials_tree.column(col, width=width, anchor=tk.CENTER)
        self.materials_tree.pack(fill=tk.BOTH, expand=True)
        # Панель фильтрации
        self.materials_filters = self.create_filter_panel(
            self.materials_frame,
            self.materials_tree,
            ["ID", "Марка", "Толщина", "Длина", "Ширина", "Кол-во шт", "Резерв", "Доступно"],
            self.refresh_materials
        )
        buttons_frame = tk.Frame(self.materials_frame, bg='white')
        buttons_frame.pack(fill=tk.X, padx=10, pady=10)
        btn_style = {"font": ("Arial", 10), "width": 15, "height": 2}
        tk.Button(buttons_frame, text="Добавить", bg='#27ae60', fg='white', command=self.add_material, **btn_style).pack(side=tk.LEFT, padx=5)
        tk.Button(buttons_frame, text="Импорт из Excel", bg='#9b59b6', fg='white', command=self.import_materials, **btn_style).pack(side=tk.LEFT, padx=5)
        tk.Button(buttons_frame, text="Скачать шаблон", bg='#3498db', fg='white', command=self.download_template, **btn_style).pack(side=tk.LEFT, padx=5)
        tk.Button(buttons_frame, text="Редактировать", bg='#f39c12', fg='white', command=self.edit_material, **btn_style).pack(side=tk.LEFT, padx=5)
        tk.Button(buttons_frame, text="Удалить", bg='#e74c3c', fg='white', command=self.delete_material, **btn_style).pack(side=tk.LEFT, padx=5)
        tk.Button(buttons_frame, text="Обновить", bg='#95a5a6', fg='white', command=self.refresh_materials, **btn_style).pack(side=tk.LEFT, padx=5)
        self.refresh_materials()

    def refresh_materials(self):
        for i in self.materials_tree.get_children():
            self.materials_tree.delete(i)
        df = load_data("Materials")
        if not df.empty:
            for index, row in df.iterrows():
                values = [row["ID"], row["Марка"], row["Толщина"], row["Длина"], row["Ширина"],
                         row["Количество штук"], row["Общая площадь"], row["Зарезервировано"],
                         row["Доступно"], row["Дата добавления"]]
                self.materials_tree.insert("", "end", values=values)

    def download_template(self):
        file_path = filedialog.asksaveasfilename(title="Сохранить шаблон", defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")], initialfile="template_materials.xlsx")
        if not file_path:
            return
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Материалы"
            ws.append(["Марка", "Толщина", "Длина", "Ширина", "Количество штук"])
            examples = [["09Г2С", 10, 6000, 1500, 5], ["Ст3", 12, 6000, 1500, 3], ["40Х", 8, 3000, 1250, 10]]
            for example in examples:
                ws.append(example)
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                ws.column_dimensions[column].width = max_length + 2
            wb.save(file_path)
            messagebox.showinfo("Успех", f"Шаблон сохранен в:\n{file_path}")
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось создать шаблон: {e}")

    def import_materials(self):
        file_path = filedialog.askopenfilename(title="Выберите файл Excel с материалами",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")])
        if not file_path:
            return
        try:
            import_df = pd.read_excel(file_path, engine='openpyxl')
            required_columns = ["Марка", "Толщина", "Длина", "Ширина", "Количество штук"]
            missing_columns = [col for col in required_columns if col not in import_df.columns]
            if missing_columns:
                messagebox.showerror("Ошибка", f"В файле отсутствуют колонки:\n{', '.join(missing_columns)}")
                return
            materials_df = load_data("Materials")
            current_max_id = 0 if materials_df.empty else int(materials_df["ID"].max())
            imported_count = 0
            errors = []
            for idx, row in import_df.iterrows():
                try:
                    if pd.isna(row["Марка"]) or row["Марка"] == "":
                        continue
                    marka = str(row["Марка"]).strip()
                    thickness = float(row["Толщина"])
                    length = float(row["Длина"])
                    width = float(row["Ширина"])
                    quantity = int(row["Количество штук"])
                    duplicate = materials_df[(materials_df["Марка"] == marka) & (materials_df["Толщина"] == thickness) &
                        (materials_df["Длина"] == length) & (materials_df["Ширина"] == width)]
                    if not duplicate.empty:
                        material_id = duplicate.iloc[0]["ID"]
                        old_qty = int(duplicate.iloc[0]["Количество штук"])
                        new_qty = old_qty + quantity
                        reserved = int(duplicate.iloc[0]["Зарезервировано"])
                        area = (length * width * new_qty) / 1000000
                        materials_df.loc[materials_df["ID"] == material_id, "Количество штук"] = new_qty
                        materials_df.loc[materials_df["ID"] == material_id, "Общая площадь"] = round(area, 2)
                        materials_df.loc[materials_df["ID"] == material_id, "Доступно"] = new_qty - reserved
                    else:
                        current_max_id += 1
                        area = (length * width * quantity) / 1000000
                        new_row = pd.DataFrame([{"ID": current_max_id, "Марка": marka, "Толщина": thickness,
                            "Длина": length, "Ширина": width, "Количество штук": quantity,
                            "Общая площадь": round(area, 2), "Зарезервировано": 0, "Доступно": quantity,
                            "Дата добавления": datetime.now().strftime("%Y-%m-%d")}])
                        materials_df = pd.concat([materials_df, new_row], ignore_index=True)
                    imported_count += 1
                except Exception as e:
                    errors.append(f"Строка {idx + 2}: {str(e)}")
            save_data("Materials", materials_df)
            self.refresh_materials()
            self.refresh_balance()
            result_msg = f"Успешно импортировано: {imported_count} материалов"
            if errors:
                result_msg += f"\n\nОшибки:\n" + "\n".join(errors[:10])
            messagebox.showinfo("Результат импорта", result_msg)
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось импортировать данные:\n{e}")

    def add_material(self):
        add_window = tk.Toplevel(self.root)
        add_window.title("Добавить материал")
        add_window.geometry("450x500")
        add_window.configure(bg='#ecf0f1')
        tk.Label(add_window, text="Добавление листового проката", font=("Arial", 12, "bold"), bg='#ecf0f1').pack(pady=10)
        fields = [("Марка стали:", "marka"), ("Толщина (мм):", "thickness"), ("Длина (мм):", "length"),
            ("Ширина (мм):", "width"), ("Количество штук:", "quantity")]
        entries = {}
        for label_text, key in fields:
            frame = tk.Frame(add_window, bg='#ecf0f1')
            frame.pack(fill=tk.X, padx=20, pady=5)
            tk.Label(frame, text=label_text, width=20, anchor='w', bg='#ecf0f1', font=("Arial", 10)).pack(side=tk.LEFT)
            entry = tk.Entry(frame, font=("Arial", 10))
            entry.pack(side=tk.RIGHT, expand=True, fill=tk.X)
            entries[key] = entry
        def save_material():
            try:
                marka = entries["marka"].get().strip()
                thickness = float(entries["thickness"].get().strip())
                length = float(entries["length"].get().strip())
                width = float(entries["width"].get().strip())
                quantity = int(entries["quantity"].get().strip())
                if not marka:
                    messagebox.showwarning("Предупреждение", "Заполните марку стали!")
                    return
                area = (length * width * quantity) / 1000000
                df = load_data("Materials")
                new_id = 1 if df.empty else int(df["ID"].max()) + 1
                new_row = pd.DataFrame([{"ID": new_id, "Марка": marka, "Толщина": thickness, "Длина": length,
                    "Ширина": width, "Количество штук": quantity, "Общая площадь": round(area, 2),
                    "Зарезервировано": 0, "Доступно": quantity, "Дата добавления": datetime.now().strftime("%Y-%m-%d")}])
                df = pd.concat([df, new_row], ignore_index=True)
                save_data("Materials", df)
                self.refresh_materials()
                self.refresh_balance()
                add_window.destroy()
                messagebox.showinfo("Успех", "Материал успешно добавлен!")
            except ValueError:
                messagebox.showerror("Ошибка", "Проверьте правильность ввода числовых значений!")
            except Exception as e:
                messagebox.showerror("Ошибка", f"Не удалось добавить материал: {e}")
        tk.Button(add_window, text="Сохранить", bg='#27ae60', fg='white', font=("Arial", 12, "bold"), command=save_material).pack(pady=20)

    def edit_material(self):
        selected = self.materials_tree.selection()
        if not selected:
            messagebox.showwarning("Предупреждение", "Выберите материал для редактирования")
            return
        item_id = self.materials_tree.item(selected)["values"][0]
        df = load_data("Materials")
        row = df[df["ID"] == item_id].iloc[0]
        edit_window = tk.Toplevel(self.root)
        edit_window.title("Редактировать материал")
        edit_window.geometry("450x500")
        edit_window.configure(bg='#ecf0f1')
        tk.Label(edit_window, text="Редактирование материала", font=("Arial", 12, "bold"), bg='#ecf0f1').pack(pady=10)
        fields = [("Марка стали:", "Марка"), ("Толщина (мм):", "Толщина"), ("Длина (мм):", "Длина"),
            ("Ширина (мм):", "Ширина"), ("Количество штук:", "Количество штук")]
        entries = {}
        for label_text, key in fields:
            frame = tk.Frame(edit_window, bg='#ecf0f1')
            frame.pack(fill=tk.X, padx=20, pady=5)
            tk.Label(frame, text=label_text, width=20, anchor='w', bg='#ecf0f1', font=("Arial", 10)).pack(side=tk.LEFT)
            entry = tk.Entry(frame, font=("Arial", 10))
            entry.insert(0, str(row[key]))
            entry.pack(side=tk.RIGHT, expand=True, fill=tk.X)
            entries[key] = entry
        def save_changes():
            try:
                thickness = float(entries["Толщина"].get())
                length = float(entries["Длина"].get())
                width = float(entries["Ширина"].get())
                quantity = int(entries["Количество штук"].get())
                reserved = int(row["Зарезервировано"])
                area = (length * width * quantity) / 1000000
                df.loc[df["ID"] == item_id, "Марка"] = entries["Марка"].get()
                df.loc[df["ID"] == item_id, "Толщина"] = thickness
                df.loc[df["ID"] == item_id, "Длина"] = length
                df.loc[df["ID"] == item_id, "Ширина"] = width
                df.loc[df["ID"] == item_id, "Количество штук"] = quantity
                df.loc[df["ID"] == item_id, "Общая площадь"] = round(area, 2)
                df.loc[df["ID"] == item_id, "Доступно"] = quantity - reserved
                save_data("Materials", df)
                self.refresh_materials()
                self.refresh_balance()
                edit_window.destroy()
                messagebox.showinfo("Успех", "Материал успешно обновлен!")
            except Exception as e:
                messagebox.showerror("Ошибка", f"Не удалось обновить материал: {e}")
        tk.Button(edit_window, text="Сохранить", bg='#3498db', fg='white', font=("Arial", 12, "bold"), command=save_changes).pack(pady=20)

    def delete_material(self):
        selected = self.materials_tree.selection()
        if not selected:
            messagebox.showwarning("Предупреждение", "Выберите материалы для удаления")
            return
        count = len(selected)
        if messagebox.askyesno("Подтверждение", f"Удалить выбранные материалы ({count} шт)?"):
            df = load_data("Materials")
            for item in selected:
                item_id = self.materials_tree.item(item)["values"][0]
                df = df[df["ID"] != item_id]
            save_data("Materials", df)
            self.refresh_materials()
            self.refresh_balance()
            messagebox.showinfo("Успех", f"Удалено материалов: {count}")

    def setup_orders_tab(self):
        header = tk.Label(self.orders_frame, text="Управление заказами", font=("Arial", 16, "bold"), bg='white', fg='#2c3e50')
        header.pack(pady=10)
        orders_label = tk.Label(self.orders_frame, text="Список заказов", font=("Arial", 12, "bold"), bg='white')
        orders_label.pack(pady=5)
        tree_frame = tk.Frame(self.orders_frame, bg='white')
        tree_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        scroll_y = tk.Scrollbar(tree_frame, orient=tk.VERTICAL)
        self.orders_tree = ttk.Treeview(tree_frame, columns=("ID", "Название", "Заказчик", "Дата", "Статус", "Примечания"),
            show="headings", yscrollcommand=scroll_y.set, height=8)
        scroll_y.config(command=self.orders_tree.yview)
        scroll_y.pack(side=tk.RIGHT, fill=tk.Y)
        columns_config = {"ID": 80, "Название": 200, "Заказчик": 150, "Дата": 100, "Статус": 100, "Примечания": 200}
        for col, width in columns_config.items():
            self.orders_tree.heading(col, text=col)
            self.orders_tree.column(col, width=width, anchor=tk.CENTER)
        self.orders_tree.pack(fill=tk.BOTH, expand=True)
        # Панель фильтрации заказов
        self.orders_filters = self.create_filter_panel(
            self.orders_frame,
            self.orders_tree,
            ["ID", "Название", "Заказчик", "Статус"],
            self.refresh_orders
        )
        self.orders_tree.bind('<<TreeviewSelect>>', self.on_order_select)
        buttons_frame = tk.Frame(self.orders_frame, bg='white')
        buttons_frame.pack(fill=tk.X, padx=10, pady=5)
        btn_style = {"font": ("Arial", 10), "width": 15, "height": 2}
        tk.Button(buttons_frame, text="Добавить заказ", bg='#27ae60', fg='white', command=self.add_order, **btn_style).pack(side=tk.LEFT, padx=5)
        tk.Button(buttons_frame, text="Импорт из Excel", bg='#9b59b6', fg='white', command=self.import_orders, **btn_style).pack(side=tk.LEFT, padx=5)
        tk.Button(buttons_frame, text="Скачать шаблон", bg='#3498db', fg='white', command=self.download_orders_template, **btn_style).pack(side=tk.LEFT, padx=5)
        tk.Button(buttons_frame, text="Редактировать", bg='#f39c12', fg='white', command=self.edit_order, **btn_style).pack(side=tk.LEFT, padx=5)
        tk.Button(buttons_frame, text="Удалить заказ", bg='#e74c3c', fg='white', command=self.delete_order, **btn_style).pack(side=tk.LEFT, padx=5)
        tk.Button(buttons_frame, text="Обновить", bg='#95a5a6', fg='white', command=self.refresh_orders, **btn_style).pack(side=tk.LEFT, padx=5)
        details_label = tk.Label(self.orders_frame, text="Детали выбранного заказа", font=("Arial", 12, "bold"), bg='white')
        details_label.pack(pady=5)
        details_tree_frame = tk.Frame(self.orders_frame, bg='white')
        details_tree_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        scroll_y2 = tk.Scrollbar(details_tree_frame, orient=tk.VERTICAL)
        self.order_details_tree = ttk.Treeview(details_tree_frame, columns=("ID", "ID заказа", "Название детали", "Количество"),
            show="headings", yscrollcommand=scroll_y2.set, height=6)
        scroll_y2.config(command=self.order_details_tree.yview)
        scroll_y2.pack(side=tk.RIGHT, fill=tk.Y)
        for col in ["ID", "ID заказа", "Название детали", "Количество"]:
            self.order_details_tree.heading(col, text=col)
            self.order_details_tree.column(col, width=150, anchor=tk.CENTER)
        self.order_details_tree.pack(fill=tk.BOTH, expand=True)
        # Панель фильтрации деталей
        self.order_details_filters = self.create_filter_panel(
            self.orders_frame,
            self.order_details_tree,
            ["Название детали", "Количество"],
            self.refresh_order_details
        )
        details_buttons_frame = tk.Frame(self.orders_frame, bg='white')
        details_buttons_frame.pack(fill=tk.X, padx=10, pady=5)
        tk.Button(details_buttons_frame, text="Добавить деталь", bg='#27ae60', fg='white', command=self.add_order_detail, **btn_style).pack(side=tk.LEFT, padx=5)
        tk.Button(details_buttons_frame, text="Удалить деталь", bg='#e74c3c', fg='white', command=self.delete_order_detail, **btn_style).pack(side=tk.LEFT, padx=5)
        self.refresh_orders()

    def on_order_select(self, event):
        self.refresh_order_details()

    def refresh_orders(self):
        for i in self.orders_tree.get_children():
            self.orders_tree.delete(i)
        df = load_data("Orders")
        if not df.empty:
            for index, row in df.iterrows():
                values = [row["ID заказа"], row["Название заказа"], row["Заказчик"], row["Дата создания"], row["Статус"], row["Примечания"]]
                self.orders_tree.insert("", "end", values=values)

    def refresh_order_details(self):
        for i in self.order_details_tree.get_children():
            self.order_details_tree.delete(i)
        selected = self.orders_tree.selection()
        if not selected:
            return
        order_id = self.orders_tree.item(selected)["values"][0]
        df = load_data("OrderDetails")
        if not df.empty:
            order_details = df[df["ID заказа"] == order_id]
            for index, row in order_details.iterrows():
                self.order_details_tree.insert("", "end", values=tuple(row))

    def download_orders_template(self):
        file_path = filedialog.asksaveasfilename(title="Сохранить шаблон", defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")], initialfile="template_orders.xlsx")
        if not file_path:
            return
        try:
            wb = Workbook()
            ws_orders = wb.active
            ws_orders.title = "Заказы"
            headers_orders = ["Название заказа", "Заказчик", "Статус", "Примечания"]
            ws_orders.append(headers_orders)
            examples_orders = [
                ["Заказ №1 - Металлоконструкции", "ООО Стройтех", "Новый", "Срочный заказ"],
                ["Заказ №2 - Лестница", "ИП Иванов", "В работе", ""],
                ["Заказ №3 - Ограждение", "ООО Метпром", "Новый", "Требуется предоплата"]
            ]
            for example in examples_orders:
                ws_orders.append(example)
            for col in ws_orders.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                ws_orders.column_dimensions[column].width = max_length + 2
            ws_details = wb.create_sheet("Детали")
            headers_details = ["Название заказа", "Название детали", "Количество"]
            ws_details.append(headers_details)
            examples_details = [
                ["Заказ №1 - Металлоконструкции", "Балка двутавровая 20", 15],
                ["Заказ №1 - Металлоконструкции", "Швеллер 16", 8],
                ["Заказ №2 - Лестница", "Ступень 300x250", 12],
                ["Заказ №2 - Лестница", "Поручень", 2],
                ["Заказ №3 - Ограждение", "Стойка 50x50", 20]
            ]
            for example in examples_details:
                ws_details.append(example)
            for col in ws_details.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                ws_details.column_dimensions[column].width = max_length + 2
            wb.save(file_path)
            messagebox.showinfo("Успех", f"Шаблон сохранен в:\n{file_path}\n\n📋 ИНСТРУКЦИЯ:\n\nЛист 'Заказы':\n• Название заказа - уникальное имя\n• Заказчик - обязательно\n• Статус: Новый, В работе, Завершен, Отменен\n• Примечания - опционально\n\nЛист 'Детали':\n• Название заказа - должно совпадать с листом 'Заказы'\n• Название детали - обязательно\n• Количество - число")
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось создать шаблон: {e}")

    def import_orders(self):
        file_path = filedialog.askopenfilename(title="Выберите файл Excel с заказами",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")])
        if not file_path:
            return
        try:
            try:
                orders_import_df = pd.read_excel(file_path, sheet_name="Заказы", engine='openpyxl')
            except:
                messagebox.showerror("Ошибка", "В файле отсутствует лист 'Заказы'!\n\nИспользуйте шаблон.")
                return
            try:
                details_import_df = pd.read_excel(file_path, sheet_name="Детали", engine='openpyxl')
                has_details = True
            except:
                details_import_df = pd.DataFrame()
                has_details = False
            required_columns_orders = ["Название заказа", "Заказчик"]
            missing_columns = [col for col in required_columns_orders if col not in orders_import_df.columns]
            if missing_columns:
                messagebox.showerror("Ошибка", f"В листе 'Заказы' отсутствуют колонки:\n{', '.join(missing_columns)}\n\nИспользуйте кнопку 'Скачать шаблон'.")
                return
            if has_details and not details_import_df.empty:
                required_columns_details = ["Название заказа", "Название детали", "Количество"]
                missing_details = [col for col in required_columns_details if col not in details_import_df.columns]
                if missing_details:
                    messagebox.showwarning("Предупреждение", f"В листе 'Детали' отсутствуют колонки:\n{', '.join(missing_details)}\n\nДетали не будут импортированы.")
                    has_details = False
            orders_df = load_data("Orders")
            current_max_order_id = 1000 if orders_df.empty else int(orders_df["ID заказа"].max())
            order_details_df = load_data("OrderDetails")
            current_max_detail_id = 0 if order_details_df.empty else int(order_details_df["ID"].max())
            imported_orders = 0
            imported_details = 0
            errors = []
            valid_statuses = ["Новый", "В работе", "Завершен", "Отменен"]
            order_name_to_id = {}
            for idx, row in orders_import_df.iterrows():
                try:
                    if pd.isna(row["Название заказа"]) or str(row["Название заказа"]).strip() == "":
                        continue
                    if pd.isna(row["Заказчик"]) or str(row["Заказчик"]).strip() == "":
                        errors.append(f"Заказы, строка {idx + 2}: Отсутствует заказчик")
                        continue
                    order_name = str(row["Название заказа"]).strip()
                    customer = str(row["Заказчик"]).strip()
                    status = "Новый"
                    if "Статус" in orders_import_df.columns and not pd.isna(row["Статус"]):
                        status_input = str(row["Статус"]).strip()
                        if status_input in valid_statuses:
                            status = status_input
                        else:
                            errors.append(f"Заказы, строка {idx + 2}: Неверный статус '{status_input}', установлен 'Новый'")
                    notes = ""
                    if "Примечания" in orders_import_df.columns and not pd.isna(row["Примечания"]):
                        notes = str(row["Примечания"]).strip()
                    current_max_order_id += 1
                    new_order_id = current_max_order_id
                    new_row = pd.DataFrame([{
                        "ID заказа": new_order_id,
                        "Название заказа": order_name,
                        "Заказчик": customer,
                        "Дата создания": datetime.now().strftime("%Y-%m-%d"),
                        "Статус": status,
                        "Примечания": notes
                    }])
                    orders_df = pd.concat([orders_df, new_row], ignore_index=True)
                    imported_orders += 1
                    order_name_to_id[order_name] = new_order_id
                except Exception as e:
                    errors.append(f"Заказы, строка {idx + 2}: {str(e)}")
            if has_details and not details_import_df.empty:
                for idx, row in details_import_df.iterrows():
                    try:
                        if pd.isna(row["Название заказа"]) or str(row["Название заказа"]).strip() == "":
                            continue
                        order_name = str(row["Название заказа"]).strip()
                        if order_name not in order_name_to_id:
                            errors.append(f"Детали, строка {idx + 2}: Заказ '{order_name}' не найден в листе 'Заказы'")
                            continue
                        if pd.isna(row["Название детали"]) or str(row["Название детали"]).strip() == "":
                            errors.append(f"Детали, строка {idx + 2}: Отсутствует название детали")
                            continue
                        detail_name = str(row["Название детали"]).strip()
                        if pd.isna(row["Количество"]):
                            errors.append(f"Детали, строка {idx + 2}: Отсутствует количество для детали '{detail_name}'")
                            continue
                        try:
                            quantity = float(row["Количество"])
                            quantity = int(quantity)
                            if quantity <= 0:
                                errors.append(f"Детали, строка {idx + 2}: Количество должно быть больше нуля для детали '{detail_name}'")
                                continue
                        except (ValueError, TypeError):
                            errors.append(f"Детали, строка {idx + 2}: Неверное количество '{row['Коли��ество']}' для детали '{detail_name}'")
                            continue
                        current_max_detail_id += 1
                        order_id = order_name_to_id[order_name]
                        new_detail = pd.DataFrame([{
                            "ID": current_max_detail_id,
                            "ID заказа": order_id,
                            "Название детали": detail_name,
                            "Количество": quantity
                        }])
                        order_details_df = pd.concat([order_details_df, new_detail], ignore_index=True)
                        imported_details += 1
                    except Exception as e:
                        errors.append(f"Детали, строка {idx + 2}: {str(e)}")
            save_data("Orders", orders_df)
            if imported_details > 0:
                save_data("OrderDetails", order_details_df)
            self.refresh_orders()
            result_msg = f"✅ Успешно импортировано:\n• Заказов: {imported_orders}\n• Деталей: {imported_details}"
            if errors:
                result_msg += f"\n\n⚠ Ошибки ({len(errors)}):\n" + "\n".join(errors[:15])
                if len(errors) > 15:
                    result_msg += f"\n... и еще {len(errors) - 15} ошибок"
            messagebox.showinfo("Результат импорта", result_msg)
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось импортировать данные:\n{e}")

    def add_order(self):
        add_window = tk.Toplevel(self.root)
        add_window.title("Добавить заказ")
        add_window.geometry("450x450")
        add_window.configure(bg='#ecf0f1')
        tk.Label(add_window, text="Создание нового заказа", font=("Arial", 12, "bold"), bg='#ecf0f1').pack(pady=10)
        fields = [("Название заказа:", "name"), ("Заказчик:", "customer"), ("Примечания:", "notes")]
        entries = {}
        for label_text, key in fields:
            frame = tk.Frame(add_window, bg='#ecf0f1')
            frame.pack(fill=tk.X, padx=20, pady=5)
            tk.Label(frame, text=label_text, width=20, anchor='w', bg='#ecf0f1', font=("Arial", 10)).pack(side=tk.LEFT)
            entry = tk.Entry(frame, font=("Arial", 10))
            entry.pack(side=tk.RIGHT, expand=True, fill=tk.X)
            entries[key] = entry
        status_frame = tk.Frame(add_window, bg='#ecf0f1')
        status_frame.pack(fill=tk.X, padx=20, pady=5)
        tk.Label(status_frame, text="Статус:", width=20, anchor='w', bg='#ecf0f1', font=("Arial", 10)).pack(side=tk.LEFT)
        status_var = tk.StringVar(value="Новый")
        status_combo = ttk.Combobox(status_frame, textvariable=status_var, values=["Новый", "В работе", "Завершен", "Отменен"],
            font=("Arial", 10), state="readonly")
        status_combo.pack(side=tk.RIGHT, expand=True, fill=tk.X)
        def save_order():
            try:
                name = entries["name"].get().strip()
                customer = entries["customer"].get().strip()
                if not name or not customer:
                    messagebox.showwarning("Предупреждение", "Заполните название и заказчика!")
                    return
                df = load_data("Orders")
                new_id = 1001 if df.empty else int(df["ID заказа"].max()) + 1
                new_row = pd.DataFrame([{"ID заказа": new_id, "Название заказа": name, "Заказчик": customer,
                    "Дата создания": datetime.now().strftime("%Y-%m-%d"), "Статус": status_var.get(), "Примечания": entries["notes"].get()}])
                df = pd.concat([df, new_row], ignore_index=True)
                save_data("Orders", df)
                self.refresh_orders()
                add_window.destroy()
                messagebox.showinfo("Успех", f"Заказ #{new_id} успешно создан!")
            except Exception as e:
                messagebox.showerror("Ошибка", f"Не удалось создать заказ: {e}")
        tk.Button(add_window, text="Создать заказ", bg='#27ae60', fg='white', font=("Arial", 12, "bold"), command=save_order).pack(pady=20)

    def edit_order(self):
        selected = self.orders_tree.selection()
        if not selected:
            messagebox.showwarning("Предупреждение", "Выберите заказ для редактирования")
            return
        item_id = self.orders_tree.item(selected)["values"][0]
        df = load_data("Orders")
        row = df[df["ID заказа"] == item_id].iloc[0]
        edit_window = tk.Toplevel(self.root)
        edit_window.title("Редактировать заказ")
        edit_window.geometry("450x450")
        edit_window.configure(bg='#ecf0f1')
        tk.Label(edit_window, text=f"Редактирование заказа #{item_id}", font=("Arial", 12, "bold"), bg='#ecf0f1').pack(pady=10)
        fields = [("Название заказа:", "Название заказа"), ("Заказчик:", "Заказчик"), ("Примечания:", "Примечания")]
        entries = {}
        for label_text, key in fields:
            frame = tk.Frame(edit_window, bg='#ecf0f1')
            frame.pack(fill=tk.X, padx=20, pady=5)
            tk.Label(frame, text=label_text, width=20, anchor='w', bg='#ecf0f1', font=("Arial", 10)).pack(side=tk.LEFT)
            entry = tk.Entry(frame, font=("Arial", 10))
            entry.insert(0, str(row[key]))
            entry.pack(side=tk.RIGHT, expand=True, fill=tk.X)
            entries[key] = entry
        status_frame = tk.Frame(edit_window, bg='#ecf0f1')
        status_frame.pack(fill=tk.X, padx=20, pady=5)
        tk.Label(status_frame, text="Статус:", width=20, anchor='w', bg='#ecf0f1', font=("Arial", 10)).pack(side=tk.LEFT)
        status_var = tk.StringVar(value=row["Статус"])
        status_combo = ttk.Combobox(status_frame, textvariable=status_var, values=["Новый", "В работе", "Завершен", "Отменен"],
            font=("Arial", 10), state="readonly")
        status_combo.pack(side=tk.RIGHT, expand=True, fill=tk.X)
        def save_changes():
            try:
                df.loc[df["ID заказа"] == item_id, "Название заказа"] = entries["Название заказа"].get()
                df.loc[df["ID заказа"] == item_id, "Заказчик"] = entries["Заказчик"].get()
                df.loc[df["ID заказа"] == item_id, "Статус"] = status_var.get()
                df.loc[df["ID заказа"] == item_id, "Примечания"] = entries["Примечания"].get()
                save_data("Orders", df)
                self.refresh_orders()
                edit_window.destroy()
                messagebox.showinfo("Успех", "Заказ успешно обновлен!")
            except Exception as e:
                messagebox.showerror("Ошибка", f"Не удалось обновить заказ: {e}")
        tk.Button(edit_window, text="Сохранить", bg='#3498db', fg='white', font=("Arial", 12, "bold"), command=save_changes).pack(pady=20)

    def delete_order(self):
        selected = self.orders_tree.selection()
        if not selected:
            messagebox.showwarning("Предупреждение", "Выберите заказы для удаления")
            return
        count = len(selected)
        if messagebox.askyesno("Подтверждение", f"Удалить выбранные заказы ({count} шт)?"):
            df = load_data("Orders")
            details_df = load_data("OrderDetails")
            for item in selected:
                item_id = self.orders_tree.item(item)["values"][0]
                df = df[df["ID заказа"] != item_id]
                if not details_df.empty:
                    details_df = details_df[details_df["ID заказа"] != item_id]
            save_data("Orders", df)
            if not details_df.empty or len(selected) > 0:
                save_data("OrderDetails", details_df)
            self.refresh_orders()
            self.refresh_order_details()
            messagebox.showinfo("Успех", f"Удалено заказов: {count}")

    def add_order_detail(self):
        selected = self.orders_tree.selection()
        if not selected:
            messagebox.showwarning("Предупреждение", "Сначала выберите заказ!")
            return
        order_id = self.orders_tree.item(selected)["values"][0]
        add_window = tk.Toplevel(self.root)
        add_window.title("Добавить деталь")
        add_window.geometry("400x300")
        add_window.configure(bg='#ecf0f1')
        tk.Label(add_window, text=f"Добавление детали к заказу #{order_id}", font=("Arial", 12, "bold"), bg='#ecf0f1').pack(pady=10)
        name_frame = tk.Frame(add_window, bg='#ecf0f1')
        name_frame.pack(fill=tk.X, padx=20, pady=5)
        tk.Label(name_frame, text="Название детали:", width=20, anchor='w', bg='#ecf0f1', font=("Arial", 10)).pack(side=tk.LEFT)
        name_entry = tk.Entry(name_frame, font=("Arial", 10))
        name_entry.pack(side=tk.RIGHT, expand=True, fill=tk.X)
        qty_frame = tk.Frame(add_window, bg='#ecf0f1')
        qty_frame.pack(fill=tk.X, padx=20, pady=5)
        tk.Label(qty_frame, text="Количество:", width=20, anchor='w', bg='#ecf0f1', font=("Arial", 10)).pack(side=tk.LEFT)
        qty_entry = tk.Entry(qty_frame, font=("Arial", 10))
        qty_entry.pack(side=tk.RIGHT, expand=True, fill=tk.X)
        def save_detail():
            try:
                detail_name = name_entry.get().strip()
                quantity = int(qty_entry.get().strip())
                if not detail_name:
                    messagebox.showwarning("Предупреждение", "Введите название детали!")
                    return
                df = load_data("OrderDetails")
                new_id = 1 if df.empty else int(df["ID"].max()) + 1
                new_row = pd.DataFrame([{"ID": new_id, "ID заказа": order_id, "Название детали": detail_name, "Количество": quantity}])
                df = pd.concat([df, new_row], ignore_index=True)
                save_data("OrderDetails", df)
                self.refresh_order_details()
                add_window.destroy()
                messagebox.showinfo("Успех", "Деталь добавлена!")
            except ValueError:
                messagebox.showerror("Ошибка", "Количество должно быть числом!")
            except Exception as e:
                messagebox.showerror("Ошибка", f"Не удалось добавить деталь: {e}")
        tk.Button(add_window, text="Добавить", bg='#27ae60', fg='white', font=("Arial", 12, "bold"), command=save_detail).pack(pady=20)

    def delete_order_detail(self):
        selected = self.order_details_tree.selection()
        if not selected:
            messagebox.showwarning("Предупреждение", "Выберите детали для удаления")
            return
        count = len(selected)
        if messagebox.askyesno("Подтверждение", f"Удалить выбранные детали ({count} шт)?"):
            df = load_data("OrderDetails")
            for item in selected:
                detail_id = self.order_details_tree.item(item)["values"][0]
                df = df[df["ID"] != detail_id]
            save_data("OrderDetails", df)
            self.refresh_order_details()
            messagebox.showinfo("Успех", f"Удалено деталей: {count}")

    def setup_reservations_tab(self):
        header = tk.Label(self.reservations_frame, text="Резервирование материалов", font=("Arial", 16, "bold"), bg='white', fg='#2c3e50')
        header.pack(pady=10)
        tree_frame = tk.Frame(self.reservations_frame, bg='white')
        tree_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        scroll_y = tk.Scrollbar(tree_frame, orient=tk.VERTICAL)
        scroll_x = tk.Scrollbar(tree_frame, orient=tk.HORIZONTAL)
        self.reservations_tree = ttk.Treeview(tree_frame,
            columns=("ID", "Заказ", "Материал", "Марка", "Толщина", "Размер", "Резерв", "Списано", "Остаток", "Дата"),
            show="headings", yscrollcommand=scroll_y.set, xscrollcommand=scroll_x.set)
        scroll_y.config(command=self.reservations_tree.yview)
        scroll_x.config(command=self.reservations_tree.xview)
        scroll_y.pack(side=tk.RIGHT, fill=tk.Y)
        scroll_x.pack(side=tk.BOTTOM, fill=tk.X)
        for col in self.reservations_tree["columns"]:
            self.reservations_tree.heading(col, text=col)
            self.reservations_tree.column(col, width=110, anchor=tk.CENTER)
        self.reservations_tree.pack(fill=tk.BOTH, expand=True)
        # Панель фильтрации
        self.reservations_filters = self.create_filter_panel(
            self.reservations_frame,
            self.reservations_tree,
            ["ID", "Заказ", "Марка", "Толщина", "Резерв", "Списано", "Остаток"],
            self.refresh_reservations
        )
        buttons_frame = tk.Frame(self.reservations_frame, bg='white')
        buttons_frame.pack(fill=tk.X, padx=10, pady=10)
        btn_style = {"font": ("Arial", 10), "width": 18, "height": 2}
        tk.Button(buttons_frame, text="Зарезервировать", bg='#27ae60', fg='white', command=self.add_reservation, **btn_style).pack(side=tk.LEFT, padx=5)
        tk.Button(buttons_frame, text="Удалить резерв", bg='#e74c3c', fg='white', command=self.delete_reservation, **btn_style).pack(side=tk.LEFT, padx=5)
        tk.Button(buttons_frame, text="Обновить", bg='#95a5a6', fg='white', command=self.refresh_reservations, **btn_style).pack(side=tk.LEFT, padx=5)
        self.refresh_reservations()

    def refresh_reservations(self):
        for i in self.reservations_tree.get_children():
            self.reservations_tree.delete(i)
        df = load_data("Reservations")
        if not df.empty:
            for index, row in df.iterrows():
                size_str = f"{row['Ширина']}x{row['Длина']}"
                values = [row["ID резерва"], row["ID заказа"], row["ID материала"], row["Марка"], row["Толщина"],
                         size_str, row["Зарезервировано штук"], row["Списано"], row["Остаток к списанию"], row["Дата резерва"]]
                self.reservations_tree.insert("", "end", values=values)

    def add_reservation(self):
        orders_df = load_data("Orders")
        if orders_df.empty:
            messagebox.showwarning("Предупреждение", "Сначала создайте заказы!")
            return
        add_window = tk.Toplevel(self.root)
        add_window.title("Создать резерв")
        add_window.geometry("550x500")
        add_window.configure(bg='#ecf0f1')
        tk.Label(add_window, text="Резервирование материала под заказ", font=("Arial", 12, "bold"), bg='#ecf0f1').pack(pady=10)
        order_frame = tk.Frame(add_window, bg='#ecf0f1')
        order_frame.pack(fill=tk.X, padx=20, pady=5)
        tk.Label(order_frame, text="Заказ:", width=20, anchor='w', bg='#ecf0f1', font=("Arial", 10)).pack(side=tk.LEFT)
        order_options = [f"{int(row['ID заказа'])} - {row['Название заказа']}" for _, row in orders_df.iterrows()]
        order_var = tk.StringVar()
        order_combo = ttk.Combobox(order_frame, textvariable=order_var, values=order_options, font=("Arial", 10), state="readonly", width=35)
        order_combo.pack(side=tk.RIGHT, expand=True, fill=tk.X)
        material_frame = tk.Frame(add_window, bg='#ecf0f1')
        material_frame.pack(fill=tk.X, padx=20, pady=5)
        tk.Label(material_frame, text="Материал:", width=20, anchor='w', bg='#ecf0f1', font=("Arial", 10)).pack(side=tk.LEFT)
        materials_df = load_data("Materials")
        material_options = ["[Добавить вручную]"]
        if not materials_df.empty:
            material_options.extend([f"{int(row['ID'])} - {row['Марка']} {row['Толщина']}мм {row['Ширина']}x{row['Длина']} (доступно: {int(row['Доступно'])} шт)"
                           for _, row in materials_df.iterrows()])
        material_var = tk.StringVar()
        material_combo = ttk.Combobox(material_frame, textvariable=material_var, values=material_options, font=("Arial", 10), state="readonly", width=35)
        material_combo.pack(side=tk.RIGHT, expand=True, fill=tk.X)
        material_combo.current(0)
        manual_frame = tk.LabelFrame(add_window, text="Параметры материала (для ручного ввода)", bg='#ecf0f1', font=("Arial", 10, "bold"))
        manual_frame.pack(fill=tk.X, padx=20, pady=10)
        manual_entries = {}
        manual_fields = [("Марка стали:", "marka"), ("Толщина (мм):", "thickness"), ("Длина (мм):", "length"), ("Ширина (мм):", "width")]
        for label_text, key in manual_fields:
            frame = tk.Frame(manual_frame, bg='#ecf0f1')
            frame.pack(fill=tk.X, padx=10, pady=3)
            tk.Label(frame, text=label_text, width=18, anchor='w', bg='#ecf0f1', font=("Arial", 9)).pack(side=tk.LEFT)
            entry = tk.Entry(frame, font=("Arial", 9))
            entry.pack(side=tk.RIGHT, expand=True, fill=tk.X)
            manual_entries[key] = entry
        qty_frame = tk.Frame(add_window, bg='#ecf0f1')
        qty_frame.pack(fill=tk.X, padx=20, pady=5)
        tk.Label(qty_frame, text="Количество (шт):", width=20, anchor='w', bg='#ecf0f1', font=("Arial", 10, "bold")).pack(side=tk.LEFT)
        qty_entry = tk.Entry(qty_frame, font=("Arial", 10))
        qty_entry.pack(side=tk.RIGHT, expand=True, fill=tk.X)
        def save_reservation():
            try:
                if not order_var.get():
                    messagebox.showwarning("Предупреждение", "Выберите заказ!")
                    return
                if not material_var.get():
                    messagebox.showwarning("Предупреждение", "Выберите материал!")
                    return
                order_id = int(order_var.get().split(" - ")[0])
                quantity = int(qty_entry.get())
                if material_var.get() == "[Добавить вручную]":
                    marka = manual_entries["marka"].get().strip()
                    thickness = float(manual_entries["thickness"].get().strip())
                    length = float(manual_entries["length"].get().strip())
                    width = float(manual_entries["width"].get().strip())
                    if not marka:
                        messagebox.showwarning("Предупреждение", "Заполните марку стали!")
                        return
                    material_id = -1
                else:
                    material_id = int(material_var.get().split(" - ")[0])
                    material_row = materials_df[materials_df["ID"] == material_id].iloc[0]
                    marka = material_row["Марка"]
                    thickness = material_row["Толщина"]
                    length = material_row["Длина"]
                    width = material_row["Ширина"]
                reservations_df = load_data("Reservations")
                new_id = 1 if reservations_df.empty else int(reservations_df["ID резерва"].max()) + 1
                new_row = pd.DataFrame([{"ID резерва": new_id, "ID заказа": order_id, "ID материала": material_id,
                    "Марка": marka, "Толщина": thickness, "Длина": length, "Ширина": width,
                    "Зарезервировано штук": quantity, "Списано": 0, "Остаток к списанию": quantity,
                    "Дата резерва": datetime.now().strftime("%Y-%m-%d")}])
                reservations_df = pd.concat([reservations_df, new_row], ignore_index=True)
                save_data("Reservations", reservations_df)
                if material_id != -1:
                    materials_df.loc[materials_df["ID"] == material_id, "Зарезервировано"] = int(material_row["Зарезервировано"]) + quantity
                    materials_df.loc[materials_df["ID"] == material_id, "Доступно"] = int(material_row["Доступно"]) - quantity
                    save_data("Materials", materials_df)
                    self.refresh_materials()
                self.refresh_reservations()
                self.refresh_balance()
                add_window.destroy()
                messagebox.showinfo("Успех", f"Резерв #{new_id} успешно создан!")
            except ValueError:
                messagebox.showerror("Ошибка", "Проверьте правильность ввода числовых значений!")
            except Exception as e:
                messagebox.showerror("Ошибка", f"Не удалось создать резерв: {e}")
        tk.Button(add_window, text="Зарезервировать", bg='#27ae60', fg='white', font=("Arial", 12, "bold"), command=save_reservation).pack(pady=15)

    def delete_reservation(self):
        selected = self.reservations_tree.selection()
        if not selected:
            messagebox.showwarning("Предупреждение", "Выберите резервы для удаления")
            return
        count = len(selected)
        if messagebox.askyesno("Подтверждение", f"Удалить выбранные резервы ({count} шт)?\n\nМатериалы вернутся на склад!"):
            reservations_df = load_data("Reservations")
            materials_df = load_data("Materials")
            for item in selected:
                reserve_id = self.reservations_tree.item(item)["values"][0]
                reserve_row = reservations_df[reservations_df["ID резерва"] == reserve_id].iloc[0]
                material_id = reserve_row["ID материала"]
                if material_id != -1:
                    quantity_to_return = int(reserve_row["Остаток к списанию"])
                    if not materials_df[materials_df["ID"] == material_id].empty:
                        mat_row = materials_df[materials_df["ID"] == material_id].iloc[0]
                        materials_df.loc[materials_df["ID"] == material_id, "Зарезервировано"] = int(mat_row["Зарезервировано"]) - quantity_to_return
                        materials_df.loc[materials_df["ID"] == material_id, "Доступно"] = int(mat_row["Доступно"]) + quantity_to_return
                reservations_df = reservations_df[reservations_df["ID резерва"] != reserve_id]
            save_data("Reservations", reservations_df)
            save_data("Materials", materials_df)
            self.refresh_materials()
            self.refresh_reservations()
            self.refresh_balance()
            messagebox.showinfo("Успех", f"Удалено резервов: {count}")

    def setup_writeoffs_tab(self):
        header = tk.Label(self.writeoffs_frame, text="Списание зарезервированных материалов", font=("Arial", 16, "bold"), bg='white', fg='#2c3e50')
        header.pack(pady=10)
        tree_frame = tk.Frame(self.writeoffs_frame, bg='white')
        tree_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        scroll_y = tk.Scrollbar(tree_frame, orient=tk.VERTICAL)
        scroll_x = tk.Scrollbar(tree_frame, orient=tk.HORIZONTAL)
        self.writeoffs_tree = ttk.Treeview(tree_frame,
            columns=("ID", "ID резерва", "Заказ", "Материал", "Марка", "Толщина", "Размер", "Количество", "Дата", "Комментарий"),
            show="headings", yscrollcommand=scroll_y.set, xscrollcommand=scroll_x.set)
        scroll_y.config(command=self.writeoffs_tree.yview)
        scroll_x.config(command=self.writeoffs_tree.xview)
        scroll_y.pack(side=tk.RIGHT, fill=tk.Y)
        scroll_x.pack(side=tk.BOTTOM, fill=tk.X)
        columns_config = {"ID": 50, "ID резерва": 80, "Заказ": 70, "Материал": 80, "Марка": 90, "Толщина": 70,
            "Размер": 110, "Количество": 90, "Дата": 140, "Комментарий": 180}
        for col, width in columns_config.items():
            self.writeoffs_tree.heading(col, text=col)
            self.writeoffs_tree.column(col, width=width, anchor=tk.CENTER)
        self.writeoffs_tree.pack(fill=tk.BOTH, expand=True)
        # Панель фильтрации
        self.writeoffs_filters = self.create_filter_panel(
            self.writeoffs_frame,
            self.writeoffs_tree,
            ["ID", "ID резерва", "Заказ", "Марка", "Т��лщина", "Количество"],
            self.refresh_writeoffs
        )
        buttons_frame = tk.Frame(self.writeoffs_frame, bg='white')
        buttons_frame.pack(fill=tk.X, padx=10, pady=10)
        btn_style = {"font": ("Arial", 10), "width": 18, "height": 2}
        tk.Button(buttons_frame, text="Списать материал", bg='#e67e22', fg='white', command=self.add_writeoff, **btn_style).pack(side=tk.LEFT, padx=5)
        tk.Button(buttons_frame, text="Удалить списание", bg='#e74c3c', fg='white', command=self.delete_writeoff, **btn_style).pack(side=tk.LEFT, padx=5)
        tk.Button(buttons_frame, text="Обновить", bg='#95a5a6', fg='white', command=self.refresh_writeoffs, **btn_style).pack(side=tk.LEFT, padx=5)
        self.refresh_writeoffs()

    def refresh_writeoffs(self):
        for i in self.writeoffs_tree.get_children():
            self.writeoffs_tree.delete(i)
        df = load_data("WriteOffs")
        if not df.empty:
            for index, row in df.iterrows():
                size_str = f"{row['Ширина']}x{row['Длина']}"
                values = [row["ID списания"], row["ID резерва"], row["ID заказа"], row["ID материала"], row["Марка"],
                         row["Толщина"], size_str, row["Количество"], row["Дата списания"], row["Комментарий"]]
                self.writeoffs_tree.insert("", "end", values=values)

    def add_writeoff(self):
        reservations_df = load_data("Reservations")
        if reservations_df.empty:
            messagebox.showwarning("Предупреждение", "Нет зарезервированных материалов для списания!")
            return
        available_reserves = reservations_df[reservations_df["Остаток к списанию"] > 0]
        if available_reserves.empty:
            messagebox.showwarning("Предупреждение", "Все зарезервированные материалы уже списаны!")
            return
        add_window = tk.Toplevel(self.root)
        add_window.title("Списать материал")
        add_window.geometry("600x450")
        add_window.configure(bg='#fff3e0')
        tk.Label(add_window, text="Списание зарезервированного материала", font=("Arial", 12, "bold"), bg='#fff3e0', fg='#e67e22').pack(pady=10)
        reserve_frame = tk.Frame(add_window, bg='#fff3e0')
        reserve_frame.pack(fill=tk.X, padx=20, pady=5)
        tk.Label(reserve_frame, text="Резерв:", width=20, anchor='w', bg='#fff3e0', font=("Arial", 10)).pack(side=tk.LEFT)
        reserve_options = []
        for _, row in available_reserves.iterrows():
            reserve_text = f"ID:{int(row['ID резерва'])} | Заказ:{int(row['ID заказа'])} | {row['Марка']} {row['Толщина']}мм {row['Ширина']}x{row['Длина']} | Доступно:{int(row['Остаток к списанию'])} шт"
            reserve_options.append(reserve_text)
        reserve_var = tk.StringVar()
        reserve_combo = ttk.Combobox(reserve_frame, textvariable=reserve_var, values=reserve_options, font=("Arial", 9), state="readonly", width=60)
        reserve_combo.pack(side=tk.RIGHT, expand=True, fill=tk.X)
        if reserve_options:
            reserve_combo.current(0)
        qty_frame = tk.Frame(add_window, bg='#fff3e0')
        qty_frame.pack(fill=tk.X, padx=20, pady=5)
        tk.Label(qty_frame, text="Количество (шт):", width=20, anchor='w', bg='#fff3e0', font=("Arial", 10, "bold")).pack(side=tk.LEFT)
        qty_entry = tk.Entry(qty_frame, font=("Arial", 10))
        qty_entry.pack(side=tk.RIGHT, expand=True, fill=tk.X)
        comment_frame = tk.Frame(add_window, bg='#fff3e0')
        comment_frame.pack(fill=tk.X, padx=20, pady=5)
        tk.Label(comment_frame, text="Комментарий:", width=20, anchor='w', bg='#fff3e0', font=("Arial", 10)).pack(side=tk.LEFT)
        comment_entry = tk.Entry(comment_frame, font=("Arial", 10))
        comment_entry.pack(side=tk.RIGHT, expand=True, fill=tk.X)
        info_label = tk.Label(add_window, text="⚠ Списание уменьшит резерв и количество материала на складе!",
                             font=("Arial", 9, "italic"), bg='#fff3e0', fg='#d35400')
        info_label.pack(pady=10)
        def save_writeoff():
            try:
                if not reserve_var.get():
                    messagebox.showwarning("Предупреждение", "Выберите резерв!")
                    return
                reserve_id = int(reserve_var.get().split("ID:")[1].split(" |")[0])
                quantity = int(qty_entry.get())
                comment = comment_entry.get().strip()
                reserve_row = reservations_df[reservations_df["ID резерва"] == reserve_id].iloc[0]
                available_qty = int(reserve_row["Остаток к списанию"])
                if quantity <= 0:
                    messagebox.showerror("Ошибка", "Количество должно быть больше нуля!")
                    return
                if quantity > available_qty:
                    messagebox.showerror("Ошибка", f"Недостаточно зарезервированного материала!\nДоступно: {available_qty} шт\nЗапрошено: {quantity} шт")
                    return
                writeoffs_df = load_data("WriteOffs")
                new_id = 1 if writeoffs_df.empty else int(writeoffs_df["ID списания"].max()) + 1
                new_row = pd.DataFrame([{
                    "ID списания": new_id,
                    "ID резерва": reserve_id,
                    "ID заказа": reserve_row["ID заказа"],
                    "ID материала": reserve_row["ID материала"],
                    "Марка": reserve_row["Марка"],
                    "Толщина": reserve_row["Толщина"],
                    "Длина": reserve_row["Длина"],
                    "Ширина": reserve_row["Ширина"],
                    "Количество": quantity,
                    "Дата списания": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "Комментарий": comment
                }])
                writeoffs_df = pd.concat([writeoffs_df, new_row], ignore_index=True)
                save_data("WriteOffs", writeoffs_df)
                reservations_df.loc[reservations_df["ID резерва"] == reserve_id, "Списано"] = int(reserve_row["Списано"]) + quantity
                reservations_df.loc[reservations_df["ID резерва"] == reserve_id, "Остаток к списанию"] = available_qty - quantity
                save_data("Reservations", reservations_df)
                material_id = reserve_row["ID материала"]
                if material_id != -1:
                    materials_df = load_data("Materials")
                    if not materials_df[materials_df["ID"] == material_id].empty:
                        mat_row = materials_df[materials_df["ID"] == material_id].iloc[0]
                        current_qty = int(mat_row["Количество штук"])
                        current_reserved = int(mat_row["Зарезервировано"])
                        new_qty = current_qty - quantity
                        new_reserved = current_reserved - quantity
                        area = (float(mat_row["Длина"]) * float(mat_row["Ширина"]) * new_qty) / 1000000
                        materials_df.loc[materials_df["ID"] == material_id, "Количество штук"] = new_qty
                        materials_df.loc[materials_df["ID"] == material_id, "Зарезервировано"] = new_reserved
                        materials_df.loc[materials_df["ID"] == material_id, "Общая площадь"] = round(area, 2)
                        materials_df.loc[materials_df["ID"] == material_id, "Доступно"] = new_qty - new_reserved
                        save_data("Materials", materials_df)
                        self.refresh_materials()
                self.refresh_reservations()
                self.refresh_writeoffs()
                self.refresh_balance()
                add_window.destroy()
                messagebox.showinfo("Успех", f"Списание #{new_id} успешно выполнено!\n\nСписано: {quantity} шт\nОстаток в резерве: {available_qty - quantity} шт")
            except ValueError:
                messagebox.showerror("Ошибка", "Количество должно быть числом!")
            except Exception as e:
                messagebox.showerror("Ошибка", f"Не удалось выполнить списание: {e}")
        tk.Button(add_window, text="Списать", bg='#e67e22', fg='white', font=("Arial", 12, "bold"), command=save_writeoff).pack(pady=15)

    def delete_writeoff(self):
        selected = self.writeoffs_tree.selection()
        if not selected:
            messagebox.showwarning("Предупреждение", "Выберите списания для удаления")
            return
        count = len(selected)
        if messagebox.askyesno("Подтверждение", f"Удалить выбранные списания ({count} шт)?\n\nВнимание: Материал вернется в резерв и на склад!"):
            writeoffs_df = load_data("WriteOffs")
            reservations_df = load_data("Reservations")
            materials_df = load_data("Materials")
            for item in selected:
                writeoff_id = self.writeoffs_tree.item(item)["values"][0]
                writeoff_row = writeoffs_df[writeoffs_df["ID списания"] == writeoff_id].iloc[0]
                reserve_id = writeoff_row["ID резерва"]
                material_id = writeoff_row["ID материала"]
                quantity_to_return = int(writeoff_row["Количество"])
                if not reservations_df[reservations_df["ID резерва"] == reserve_id].empty:
                    res_row = reservations_df[reservations_df["ID резерва"] == reserve_id].iloc[0]
                    reservations_df.loc[reservations_df["ID резерва"] == reserve_id, "Списано"] = int(res_row["Списано"]) - quantity_to_return
                    reservations_df.loc[reservations_df["ID резерва"] == reserve_id, "Остаток к списанию"] = int(res_row["Остаток к списанию"]) + quantity_to_return
                if material_id != -1:
                    if not materials_df[materials_df["ID"] == material_id].empty:
                        mat_row = materials_df[materials_df["ID"] == material_id].iloc[0]
                        current_qty = int(mat_row["Количество штук"])
                        current_reserved = int(mat_row["Зарезервировано"])
                        new_qty = current_qty + quantity_to_return
                        new_reserved = current_reserved + quantity_to_return
                        area = (float(mat_row["Длина"]) * float(mat_row["Ширина"]) * new_qty) / 1000000
                        materials_df.loc[materials_df["ID"] == material_id, "Количество штук"] = new_qty
                        materials_df.loc[materials_df["ID"] == material_id, "Зарезервировано"] = new_reserved
                        materials_df.loc[materials_df["ID"] == material_id, "Общая площадь"] = round(area, 2)
                        materials_df.loc[materials_df["ID"] == material_id, "Доступно"] = new_qty - new_reserved
                writeoffs_df = writeoffs_df[writeoffs_df["ID списания"] != writeoff_id]
            save_data("WriteOffs", writeoffs_df)
            save_data("Reservations", reservations_df)
            save_data("Materials", materials_df)
            self.refresh_materials()
            self.refresh_reservations()
            self.refresh_writeoffs()
            self.refresh_balance()
            messagebox.showinfo("Успех", f"Отменено списаний: {count}")

    def setup_balance_tab(self):
        header = tk.Label(self.balance_frame, text="Баланс материалов", font=("Arial", 16, "bold"), bg='white', fg='#2c3e50')
        header.pack(pady=10)
        info_label = tk.Label(self.balance_frame, text="Красный - не хватает | Желтый - на нуле | Зеленый - в наличии",
                             font=("Arial", 10), bg='white', fg='#7f8c8d')
        info_label.pack(pady=5)
        tree_frame = tk.Frame(self.balance_frame, bg='white')
        tree_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        scroll_y = tk.Scrollbar(tree_frame, orient=tk.VERTICAL)
        scroll_x = tk.Scrollbar(tree_frame, orient=tk.HORIZONTAL)
        self.balance_tree = ttk.Treeview(tree_frame,
            columns=("Материал", "Марка", "Толщина", "Размер", "В наличии", "Зарезервировано", "Итого"),
            show="headings", yscrollcommand=scroll_y.set, xscrollcommand=scroll_x.set)
        scroll_y.config(command=self.balance_tree.yview)
        scroll_x.config(command=self.balance_tree.xview)
        scroll_y.pack(side=tk.RIGHT, fill=tk.Y)
        scroll_x.pack(side=tk.BOTTOM, fill=tk.X)
        columns_config = {"Материал": 100, "Марка": 120, "Толщина": 100, "Размер": 150,
            "В наличии": 100, "Зарезервировано": 130, "Итого": 100}
        for col, width in columns_config.items():
            self.balance_tree.heading(col, text=col)
            self.balance_tree.column(col, width=width, anchor=tk.CENTER)
        self.balance_tree.pack(fill=tk.BOTH, expand=True)
        # Панель фильтрации
        self.balance_filters = self.create_filter_panel(
            self.balance_frame,
            self.balance_tree,
            ["Марка", "Толщина", "Размер", "В наличии", "Зарезервировано"],
            self.refresh_balance
        )
        self.balance_tree.tag_configure('negative', background='#ffcccc')
        self.balance_tree.tag_configure('zero', background='#fff9c4')
        self.balance_tree.tag_configure('positive', background='#c8e6c9')
        buttons_frame = tk.Frame(self.balance_frame, bg='white')
        buttons_frame.pack(fill=tk.X, padx=10, pady=10)
        btn_style = {"font": ("Arial", 10), "width": 15, "height": 2}
        tk.Button(buttons_frame, text="Обновить", bg='#95a5a6', fg='white', command=self.refresh_balance, **btn_style).pack(side=tk.LEFT, padx=5)
        self.refresh_balance()

    def refresh_balance(self):
        for i in self.balance_tree.get_children():
            self.balance_tree.delete(i)
        materials_df = load_data("Materials")
        reservations_df = load_data("Reservations")
        balance_dict = {}
        if not materials_df.empty:
            for index, row in materials_df.iterrows():
                key = (row["Марка"], float(row["Толщина"]), float(row["Длина"]), float(row["Ширина"]))
                if key not in balance_dict:
                    balance_dict[key] = {"material_id": row["ID"], "in_stock": int(row["Количество штук"]), "reserved": 0}
        if not reservations_df.empty:
            for index, row in reservations_df.iterrows():
                key = (row["Марка"], float(row["Толщина"]), float(row["Длина"]), float(row["Ширина"]))
                reserved_qty = int(row["Зарезервировано штук"])
                if key not in balance_dict:
                    balance_dict[key] = {"material_id": -1, "in_stock": 0, "reserved": reserved_qty}
                else:
                    balance_dict[key]["reserved"] += reserved_qty
        for key, data in sorted(balance_dict.items()):
            marka, thickness, length, width = key
            in_stock = data["in_stock"]
            reserved = data["reserved"]
            total = in_stock - reserved
            size_str = f"{width} x {length}"
            material_id = data["material_id"]
            material_label = f"ID: {material_id}" if material_id != -1 else "Вручную"
            values = [material_label, marka, f"{thickness} мм", size_str, in_stock, reserved, total]
            if total < 0:
                tag = 'negative'
            elif total == 0:
                tag = 'zero'
            else:
                tag = 'positive'
            self.balance_tree.insert("", "end", values=values, tags=(tag,))

if __name__ == "__main__":
    try:
        initialize_database()
        root = tk.Tk()
        app = ProductionApp(root)
        root.mainloop()
    except Exception as e:
        print(f"Критическая ошибка: {e}")
        import traceback
        traceback.print_exc()
        messagebox.showerror("Критическая ошибка", str(e))